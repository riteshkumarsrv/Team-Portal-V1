"""
One-off generator: April 2026 Team 1 leave worksheet → Excel (from manager grid snapshot).
Columns: NAME, No. of days, Leave Type, Date From, Date To.
"No. of days" = weekdays (Mon–Fri) inclusive in range. CompOFF rows included; total sums Leave only.
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    raise SystemExit("openpyxl required: pip install openpyxl")


def weekdays_in_range(d0: date, d1: date) -> int:
    n = 0
    d = d0
    while d <= d1:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


def row(name: str, d0: date, d1: date, leave_type: str) -> tuple[str, int, str, date, date]:
    return (name, weekdays_in_range(d0, d1), leave_type, d0, d1)


def main() -> Path:
    y = 2026
    m = 4
    d = lambda day: date(y, m, day)

    rows: list[tuple[str, int, str, date, date]] = [
        # Sangjukta
        row("Sangjukta", d(15), d(15), "Leave"),
        row("Sangjukta", d(20), d(24), "Leave"),
        row("Sangjukta", d(27), d(28), "Leave"),
        row("Sangjukta", d(30), d(30), "Leave"),
        # Sumit Patra — CompOFF
        row("Sumit Patra", d(7), d(7), "CompOFF"),
        row("Sumit Patra", d(17), d(17), "CompOFF"),
        row("Sumit Patra", d(29), d(29), "CompOFF"),
        # Anas P
        row("Anas P", d(9), d(9), "Leave"),
        row("Anas P", d(10), d(10), "Leave"),
        row("Anas P", d(14), d(15), "Leave"),
        row("Anas P", d(29), d(29), "Leave"),
        # Jancy M — Apr 13–21 spans weekend (Sat 18, Sun 19 excluded from weekday count)
        row("Jancy M", d(2), d(2), "Leave"),
        row("Jancy M", d(6), d(6), "Leave"),
        row("Jancy M", d(9), d(10), "Leave"),
        row("Jancy M", d(13), d(21), "Leave"),
        # Sasi Sampath
        row("Sasi Sampath", d(10), d(10), "Leave"),
        row("Sasi Sampath", d(15), d(15), "Leave"),
        row("Sasi Sampath", d(27), d(30), "Leave"),
        # Sataysai Maddala
        row("Sataysai Maddala", d(6), d(6), "Leave"),
        row("Sataysai Maddala", d(20), d(21), "Leave"),
        row("Sataysai Maddala", d(29), d(29), "Leave"),
        # Bharath Krishna
        row("Bharath Krishna", d(28), d(28), "Leave"),
        # Dabbiru Sehshal
        row("Dabbiru Sehshal", d(6), d(6), "Leave"),
        row("Dabbiru Sehshal", d(21), d(23), "Leave"),
        row("Dabbiru Sehshal", d(27), d(30), "Leave"),
        # Mure JC — Apr 17–20 spans weekend
        row("Mure JC", d(17), d(20), "Leave"),
        row("Mure JC", d(23), d(24), "CompOFF"),
        # Sasi Penchala
        row("Sasi Penchala", d(2), d(2), "Leave"),
        row("Sasi Penchala", d(24), d(24), "Leave"),
        # Shaista Anjum (image: Shaista; roster often Shaishta — using image spelling)
        row("Shaista Anjum", d(2), d(2), "Leave"),
        row("Shaista Anjum", d(6), d(7), "Leave"),
        # Shruthi
        row("Shruthi", d(22), d(22), "Leave"),
        # Rashmi
        row("Rashmi", d(7), d(7), "Leave"),
        row("Rashmi", d(29), d(29), "Leave"),
        # Varshitha
        row("Varshitha", d(23), d(24), "Leave"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "April 2026 Leaves"

    title = "Leave Management worksheet — April 2026 — Team 1"
    ws.append([title])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    hdr = ["NAME", "No. of days", "Leave Type", "Date From", "Date To"]
    ws.append(hdr)
    thin = Side(style="thin", color="888888")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)
    for c in range(1, 6):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = grid
        cell.alignment = Alignment(horizontal="center", vertical="center")

    leave_total = 0
    for name, nd, ltype, df, dt in rows:
        ws.append([name, nd, ltype, df.isoformat(), dt.isoformat()])
        r = ws.max_row
        if ltype == "Leave":
            leave_total += nd
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = grid
            if c == 2:
                cell.alignment = Alignment(horizontal="center")
            if c >= 4:
                cell.alignment = Alignment(horizontal="center")

    ws.append([])
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="Total leaves taken (weekdays, Leave type only)")
    ws.cell(row=r, column=2, value=leave_total)
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    for c in (1, 2):
        ws.cell(row=r, column=c).border = grid

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    out = Path(__file__).resolve().parent.parent / "Leave_Management_April_2026_Team1.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p = main()
    print(p)
