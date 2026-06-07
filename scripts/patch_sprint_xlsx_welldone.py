"""One-off: trim Daily tasks Details after row N, append appreciation rows from DB."""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--db", required=True, type=Path)
    ap.add_argument("--sprint-id", type=int, default=14)
    ap.add_argument("--sheet", default="Daily tasks Details")
    ap.add_argument("--keep-rows", type=int, default=42, help="Clear everything strictly after this row")
    args = ap.parse_args()

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    team = (
        conn.execute(
            """
            SELECT COALESCE(NULLIF(TRIM(t.name), ''), 'Team') AS team_name
            FROM scrum_sprint s
            LEFT JOIN teams t ON t.id = s.team_id
            WHERE s.id = ?
            """,
            (args.sprint_id,),
        ).fetchone()
    )
    team_name = str(team["team_name"]) if team else "Team"
    rows = conn.execute(
        """
        SELECT i.id AS sticky_id, i.assignee, i.title, a.author, a.comment, a.created_at
        FROM scrum_item_appreciation a
        JOIN scrum_sprint_item i ON i.id = a.item_id
        WHERE i.sprint_id = ?
        ORDER BY i.assignee, i.id, a.id
        """,
        (args.sprint_id,),
    ).fetchall()
    conn.close()

    wb = load_workbook(args.xlsx)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet!r}; have {wb.sheetnames}")
    ws = wb[args.sheet]
    max_r = ws.max_row
    if max_r > args.keep_rows:
        ws.delete_rows(args.keep_rows + 1, max_r - args.keep_rows)

    rz = args.keep_rows + 1
    sub_font = Font(bold=True, size=11)
    ws.cell(row=rz, column=1, value="Well Done — task appreciations (all notes)").font = sub_font
    rz += 1
    ws.cell(row=rz, column=1, value="Team")
    ws.cell(row=rz, column=2, value="Member (assignee)")
    ws.cell(row=rz, column=3, value="Sticky ID")
    ws.cell(row=rz, column=4, value="Sticky title")
    ws.cell(row=rz, column=5, value="Appreciation author")
    ws.cell(row=rz, column=6, value="Appreciation comment")
    ws.cell(row=rz, column=7, value="Recorded at (UTC)")
    for c in range(1, 8):
        ws.cell(row=rz, column=c).font = Font(bold=True)
    rz += 1
    if not rows:
        ws.cell(row=rz, column=1, value="(No appreciation comments recorded for this sprint.)")
        rz += 1
    else:
        for r in rows:
            ws.cell(row=rz, column=1, value=team_name)
            ws.cell(row=rz, column=2, value=r["assignee"])
            ws.cell(row=rz, column=3, value=r["sticky_id"])
            ws.cell(row=rz, column=4, value=r["title"])
            ws.cell(row=rz, column=5, value=r["author"])
            ws.cell(row=rz, column=6, value=r["comment"])
            ws.cell(row=rz, column=7, value=r["created_at"])
            rz += 1

    wb.save(args.xlsx)
    print(f"Saved {args.xlsx} — {len(rows)} appreciation row(s), sheet {args.sheet!r} trimmed after row {args.keep_rows}.")


if __name__ == "__main__":
    main()
