"""
Team management portal — production web app (Flask + SQLite).
"""

from __future__ import annotations

import calendar
import csv
import difflib
import hashlib
import hmac
import re
import base64
import io
import json
import math
import requests
import logging
import os
import secrets
import shutil
import smtplib
import subprocess
import threading
import sqlite3
import unicodedata
from calendar import monthrange
from datetime import date, datetime, time, timedelta, timezone

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None  # type: ignore[misc, assignment]
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Iterable, Sequence
from urllib.parse import urlparse

from werkzeug.utils import secure_filename

from config import apply_flask_config_from_environ, load_application_environment
from database.session import get_db
from team_portal.views import register_blueprints

from nokia_portal_roster import NOKIA_PORTAL_DIRECTORY, lookup_portal_directory, normalize_email
from flask import (
    Flask,
    Response,
    flash,
    g,
    has_request_context,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_wtf import CSRFProtect
from flask_wtf.csrf import validate_csrf

_log = logging.getLogger(__name__)

# Windows: try winget once if Tesseract is missing (see _resolve_tesseract_cmd).
_tesseract_win_install_lock = threading.Lock()
_tesseract_win_install_tried = False

# Canonical roster — aligned with Nokia e-tool team grid (names as used on leave tracker / meet / audit).
EMPLOYEES: tuple[str, ...] = (
    "Akanksha Jha",
    "Anas P",
    "Archit Sugha",
    "Bharath G Krishna",
    "Dabbiru Siva Seshasai",
    "Farhan Thumalla",
    "Gajendra Singh Thakur",
    "Gumparlapati Penchala Sasi Kumar",
    "Harshitha K",
    "Jayachandra Reddy Mure",
    "Jancy Mariam Jose",
    "Maddala Satyasai",
    "Mubarak Palagiri",
    "Ramya Ure",
    "Rashmi K Nazare",
    "Sangjukta Giri",
    "Sasikumar Sampath",
    "Shaishta Anjum",
    "Shyam Bhaskar Katuri",
    "Siddhant Mandal",
    "Siya Chugh",
    "Shruthi K",
    "Sumit Patra",
    "Varshini Raj",
    "Varshitha S",
    "Zaiba Nousheen Khanum",
)

LEAVE_REASONS = [
    ("pl", "PL — Planned leave"),
    ("ul", "UL — Unplanned leave"),
    ("sl", "SL — Sick leave"),
    ("ll", "LL — Long leave"),
    ("compoff", "CompOFF"),
]

# Portal apply form maps UI labels to stored `reason` codes (manager grid uses same labels dict).
PORTAL_LEAVE_FORM_CHOICES = [
    ("pl", "Annual leave"),
    ("sl", "Sick leave"),
    ("ul", "Casual leave"),
    ("compoff", "CompOFF"),
]

PORTAL_DOD_CHECKLIST_LABELS: tuple[str, ...] = (
    "Code reviewed",
    "Unit tests written and passing",
    "No open blockers",
    "Acceptance criteria verified",
    "Documentation updated (if applicable)",
)

# Single calendar day: full or half; multi-day ranges use duration_type "multi".
DAY_PART_CHOICES = [
    ("full", "Full day"),
    ("half_am", "Half day — morning"),
    ("half_pm", "Half day — afternoon"),
]

DURATION_CHOICES = DAY_PART_CHOICES + [
    ("multi", "Multiple days (date range)"),
]

ATTENDANCE_STATUS = [
    ("present", "In office"),
    ("remote", "Remote / WFH"),
    ("leave", "On approved leave"),
    ("absent", "Absent (unplanned)"),
]

SCRUM_STATUS_CHOICES: tuple[tuple[str, str], ...] = (
    ("open", "Open"),
    ("doing", "In progress"),
    ("done", "Done"),
    ("blocked", "Blocked"),
)

# Kanban columns (sticky board). Sticky TYPE is the fixed set in SCRUM_BUILTIN_TASK_KIND_ROWS (`scrum_team_task_kind` rows).
SCRUM_KANBAN_COLUMNS: tuple[str, ...] = ("backlog", "do", "doing", "done")
SCRUM_LEGACY_TASK_KIND_MAP: dict[str, str] = {"story": "code", "task": "ndy", "bug": "fsy"}
_SCRUM_HEX_COLOR = re.compile(r"^#[0-9a-fA-F]{6}$")

SCRUM_HOUR_EPS = 0.05
SCRUM_SPRINT_READONLY_FLASH = (
    "This sprint is closed and cannot be edited. On the sprint team page, click “Open sprint” to unlock it."
)
SCRUM_SPRINT_AFTER_END_FLASH = (
    "This sprint’s end date has passed — the board is read-only, burndown counts activity only through the "
    "last sprint day, and you cannot change sprint dates or names here. Create or open a new sprint for new work."
)


def _sprint_clock_utc() -> datetime:
    """UTC wall clock for sprint window end and auto-close (tests may monkeypatch ``app._sprint_clock_utc``)."""
    return datetime.now(timezone.utc)


def _sprint_close_zone():
    """IANA zone for sprint calendar boundaries (``TEAM_TRACKER_SCRUM_SPRINT_CLOSE_TZ``, default UTC)."""
    if ZoneInfo is None:
        return None
    key = (os.environ.get("TEAM_TRACKER_SCRUM_SPRINT_CLOSE_TZ") or "UTC").strip()
    try:
        return ZoneInfo(key)
    except Exception:
        try:
            return ZoneInfo("UTC")
        except Exception:
            return None


def _sprint_inclusive_calendar_window_ended(end_date_iso: str, *, now_utc: datetime | None = None) -> bool:
    """
    True when ``now_utc`` is at or after the first instant of the calendar day *after* the sprint's inclusive
    ``end_date`` in ``TEAM_TRACKER_SCRUM_SPRINT_CLOSE_TZ`` (default UTC). That marks the end of the last sprint day
    in that timezone.
    """
    try:
        ed = date.fromisoformat(str(end_date_iso or "")[:10])
    except ValueError:
        return False
    now = now_utc or _sprint_clock_utc()
    zi = _sprint_close_zone()
    if zi is None:
        return now.date() > ed
    cutoff_local = datetime.combine(ed + timedelta(days=1), time.min, tzinfo=zi)
    return now >= cutoff_local.astimezone(timezone.utc)


SCRUM_STICKY_AREA_MAX_LEN = 500
SCRUM_KANBAN_WEEKDAY_HOURS = 8.0
SCRUM_TEAM_KIND_BURN_BREAKDOWN_BELOW_PCT = 67.0
SCRUM_DONE_ARTIFACT_MAX = 20
SCRUM_DONE_ARTIFACT_URL_MAX = 600
SCRUM_DONE_ARTIFACT_LABEL_MAX = 120
SCRUM_VERIFICATION_URL_MAX = 12
SCRUM_APPRECIATION_BODY_MAX = 2000
# Sticky file attachments (stored under `<db-dir>/scrum_item_attachments/`).
SCRUM_ITEM_ATTACHMENT_MAX_BYTES = 10 * 1024 * 1024
SCRUM_ITEM_ATTACHMENTS_MAX_PER_STICKY = 20
SCRUM_ITEM_ATTACHMENT_SUFFIXES: frozenset[str] = frozenset(
    {
        ".pdf",
        ".png",
        ".jpg",
        ".jpeg",
        ".gif",
        ".webp",
        ".txt",
        ".csv",
        ".md",
        ".json",
        ".xlsx",
        ".xls",
        ".docx",
        ".pptx",
        ".zip",
        ".gz",
        ".log",
        ".patch",
        ".svg",
        ".drawio",
        ".xml",
        ".html",
        ".htm",
    }
)
# Fixed sprint window on create / date save: 14 calendar days inclusive of the start day (2 weeks).
SCRUM_SPRINT_DEFAULT_CALENDAR_DAYS_INCLUSIVE = 14


def scrum_sprint_default_end_date(start: date) -> date:
    """Last day of the default sprint window: N calendar days inclusive of start (N=14 → end = start + 13)."""
    return start + timedelta(days=SCRUM_SPRINT_DEFAULT_CALENDAR_DAYS_INCLUSIVE - 1)


def _next_sprint_start_after_latest_end(conn: sqlite3.Connection, team_id: int) -> date | None:
    """Calendar day after the team's latest sprint end (MAX(end_date)); None if there are no sprints."""
    row = conn.execute(
        "SELECT MAX(substr(end_date, 1, 10)) AS m FROM scrum_sprint WHERE team_id = ?",
        (int(team_id),),
    ).fetchone()
    raw = (row["m"] if row else None) or ""
    raw = str(raw).strip()[:10]
    if not raw:
        return None
    try:
        d_end = date.fromisoformat(raw)
    except ValueError:
        return None
    return d_end + timedelta(days=1)


def suggest_next_sprint_name_from_previous(previous_name: str) -> str | None:
    """
    If the previous sprint name ends with a digit block (e.g. FRONTIER-FB2612), return the same
    prefix with that integer incremented (FRONTIER-FB2613). Preserves zero-padding width when the
    incremented value fits; otherwise uses the natural string form (e.g. ...099 -> ...100).
    Returns None when there is no trailing digit run to increment.
    """
    s = (previous_name or "").strip()
    if not s:
        return None
    m = re.match(r"^(.*?)(\d+)$", s)
    if not m:
        return None
    prefix, digits = m.group(1), m.group(2)
    try:
        n = int(digits, 10)
    except ValueError:
        return None
    nxt = n + 1
    s_next = str(nxt)
    if len(s_next) < len(digits):
        s_next = s_next.zfill(len(digits))
    return f"{prefix}{s_next}"


# Default rows for scrum_team_task_kind (code, label, color_hex, sort_order). One row per team; no ad-hoc types.
SCRUM_BUILTIN_TASK_KIND_ROWS: tuple[tuple[str, str, str, int], ...] = (
    ("ndy", "NDY", "#7c3aed", 0),
    ("code", "CODE", "#eab308", 1),
    ("fsy", "FSY", "#f97316", 2),
    ("improvement", "Improvement", "#22c55e", 3),
    ("process_tools", "Process&Tools", "#06b6d4", 4),
)
SCRUM_TASK_KIND_CODES: frozenset[str] = frozenset(r[0] for r in SCRUM_BUILTIN_TASK_KIND_ROWS)

# Sprint team hero: stacked bar chart column order (labels from `scrum_team_task_kind`).
SPRINT_TEAM_KIND_STACK_ORDER: tuple[str, ...] = (
    "ndy",
    "fsy",
    "code",
    "process_tools",
    "improvement",
)

# Sprint team: area stack chart — only roll into "Other" above this many distinct areas (safety cap).
SCRUM_AREA_STACK_MAX_BUCKETS: int = 120
SCRUM_AREA_STACK_NO_AREA_LABEL: str = "(no area)"
SCRUM_AREA_STACK_OTHER_LABEL: str = "Other"

# HPPM-style summary: fixed “Work type” order; sticky `task_kind` maps into these (Absences = leave debit hours).
SCRUM_HPPM_SUMMARY_ROWS: tuple[tuple[str | None, str], ...] = (
    ("fsy", "Feature Support (Y)"),
    ("ndy", "New Development"),
    (None, "Absences"),
    ("code", "Verification"),
    ("improvement", "Internal Improvement"),
    ("process_tools", "Process & Tools"),
)
SCRUM_HPPM_LABEL_BY_CODE: dict[str, str] = {c: lab for c, lab in SCRUM_HPPM_SUMMARY_ROWS if c is not None}

# Stacked “est vs burnt” bar colors: burnt≤est, burnt>est, remaining (HPPM view uses a cooler contrast set).
SCRUM_STACK_SEGMENT_FILLS_DEFAULT: tuple[str, str, str] = ("#fb7185", "#f59e0b", "#38bdf8")
SCRUM_STACK_SEGMENT_FILLS_HPPM: tuple[str, str, str] = ("#a78bfa", "#fb923c", "#2dd4bf")


def _stack_segment_fills(chart_palette: str) -> tuple[str, str, str]:
    if (chart_palette or "").strip().lower() == "hppm":
        return SCRUM_STACK_SEGMENT_FILLS_HPPM
    return SCRUM_STACK_SEGMENT_FILLS_DEFAULT


# Sprint export Summary charts: kind order for clustered est vs burnt (includes CODE).
_SUMMARY_EXPORT_KIND_CHART_CODES: tuple[str, ...] = (
    "ndy",
    "fsy",
    "code",
    "improvement",
    "process_tools",
)


def _export_coerce_task_kind(code: object) -> str:
    raw = str(code or "ndy").strip().lower()
    if raw in SCRUM_LEGACY_TASK_KIND_MAP:
        raw = SCRUM_LEGACY_TASK_KIND_MAP[raw]
    if raw not in SCRUM_TASK_KIND_CODES:
        return "ndy"
    return raw


def _export_roll_kind_est_burnt(items: list[sqlite3.Row] | list[dict]) -> tuple[dict[str, float], dict[str, float]]:
    est: dict[str, float] = {k: 0.0 for k in _SUMMARY_EXPORT_KIND_CHART_CODES}
    brn: dict[str, float] = {k: 0.0 for k in _SUMMARY_EXPORT_KIND_CHART_CODES}
    for it in items:
        k = _export_coerce_task_kind(it["task_kind"])
        if k not in est:
            k = "ndy"
        est[k] += float(it["estimate_hours"] or 0)
        brn[k] += float(it["total_burnt_hours"] or 0)
    return est, brn


def _export_roll_kind_est_burnt_assignee(items: list[sqlite3.Row] | list[dict], assignee: str) -> tuple[dict[str, float], dict[str, float]]:
    emp = (assignee or "").strip()
    sub = [it for it in items if str(it["assignee"] or "").strip() == emp]
    return _export_roll_kind_est_burnt(sub)


def _safe_done_artifact_url(u: str | None) -> str | None:
    u = (u or "").strip()
    if not u or len(u) > SCRUM_DONE_ARTIFACT_URL_MAX:
        return None
    p = urlparse(u)
    if p.scheme not in ("http", "https") or not p.netloc:
        return None
    return u


def _normalize_done_artifacts_from_api(raw) -> tuple[str | None, str | None]:
    """Validate API payload; returns (json array string, error code or None)."""
    if raw is None:
        return "[]", None
    if not isinstance(raw, list):
        return None, "bad_artifacts"
    if len(raw) > SCRUM_DONE_ARTIFACT_MAX:
        return None, "artifacts_limit"
    out: list[dict[str, str]] = []
    for ent in raw:
        if isinstance(ent, str):
            url = _safe_done_artifact_url(ent)
            if not url:
                return None, "bad_artifact_url"
            out.append({"url": url, "label": ""})
        elif isinstance(ent, dict):
            url = _safe_done_artifact_url(str(ent.get("url") or ent.get("u") or ""))
            if not url:
                return None, "bad_artifact_url"
            label = str(ent.get("label") or ent.get("l") or "").strip()[:SCRUM_DONE_ARTIFACT_LABEL_MAX]
            out.append({"url": url, "label": label})
        else:
            return None, "bad_artifacts"
    return json.dumps(out), None


def _normalize_done_artifacts_from_lines(text: str) -> tuple[str | None, str | None]:
    raw_list: list[dict[str, str]] = []
    for line in (text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if "|" in line:
            label, _, rest = line.partition("|")
            raw_list.append({"label": label.strip(), "url": rest.strip()})
        else:
            raw_list.append({"label": "", "url": line})
    return _normalize_done_artifacts_from_api(raw_list)


def _normalize_verification_urls_mixed(raw) -> list[str]:
    """Normalize optional proof URLs from API (list of strings) or pasted lines (single string)."""
    out: list[str] = []
    if raw is None:
        return out
    if isinstance(raw, list):
        for ent in raw:
            u = _safe_done_artifact_url(str(ent).strip())
            if u and u not in out:
                out.append(u)
    elif isinstance(raw, str):
        for line in raw.splitlines():
            u = _safe_done_artifact_url(line.strip())
            if u and u not in out:
                out.append(u)
    return out[:SCRUM_VERIFICATION_URL_MAX]


def _parse_done_artifacts_db(raw: str | None) -> list[dict[str, str]]:
    if not raw or not str(raw).strip():
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    out: list[dict[str, str]] = []
    for ent in data:
        if not isinstance(ent, dict):
            continue
        url = _safe_done_artifact_url(str(ent.get("url") or ""))
        if not url:
            continue
        label = str(ent.get("label") or "").strip()[:SCRUM_DONE_ARTIFACT_LABEL_MAX]
        out.append({"url": url, "label": label})
    return out


def _done_artifacts_to_lines(items: list[dict[str, str]]) -> str:
    lines: list[str] = []
    for it in items:
        u = (it.get("url") or "").strip()
        if not u:
            continue
        lb = (it.get("label") or "").strip()
        if lb:
            lines.append(f"{lb} | {u}")
        else:
            lines.append(u)
    return "\n".join(lines)


def _norm(s: str) -> str:
    return unicodedata.normalize("NFKC", s or "").strip().lower()


def fuzzy_employee_matches(query: str, limit: int = 5, roster: Sequence[str] | None = None) -> list[str]:
    q = _norm(query)
    if len(q) < 1:
        return []
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    scored: list[tuple[float, str]] = []
    for e in roster_t:
        en = _norm(e)
        if not en:
            continue
        if q in en:
            scored.append((0.0, e))
            continue
        ratio = difflib.SequenceMatcher(None, q, en).ratio()
        if ratio >= 0.35:
            scored.append((1.0 - ratio, e))
    scored.sort(key=lambda x: (x[0], x[1].lower()))
    out: list[str] = []
    seen: set[str] = set()
    for _, name in scored:
        if name not in seen:
            seen.add(name)
            out.append(name)
        if len(out) >= limit:
            break
    return out


def resolve_employee_name(raw: str, roster: Sequence[str] | None = None) -> str | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    if raw in roster_t:
        return raw
    raw_n = _norm(raw)
    for e in roster_t:
        if _norm(e) == raw_n:
            return e
    m = difflib.get_close_matches(raw, roster_t, n=1, cutoff=0.55)
    if m:
        return m[0]
    m2 = fuzzy_employee_matches(raw, 1, roster=roster_t)
    return m2[0] if m2 else None


def _exact_roster_name_match(raw: str, roster: Sequence[str]) -> str | None:
    """Exact roster name only (case/normalized equality). Used for leave-tracker grids so other teams' rows are not fuzzy-matched."""
    raw = (raw or "").strip()
    if not raw:
        return None
    roster_t = tuple(roster)
    if raw in roster_t:
        return raw
    raw_n = _norm(raw)
    for e in roster_t:
        if _norm(e) == raw_n:
            return e
    return None


def _bucket_leave_requests_by_roster(
    leaves: Sequence[sqlite3.Row], roster: Sequence[str]
) -> dict[str, list[sqlite3.Row]]:
    """Group leave rows under roster display names without fuzzy cross-team matching."""
    roster_t = tuple(roster)
    by_emp: dict[str, list[sqlite3.Row]] = {e: [] for e in roster_t}
    for row in leaves:
        db_name = str(row["employee_name"] or "").strip()
        if not db_name:
            continue
        if db_name in by_emp:
            by_emp[db_name].append(row)
            continue
        resolved = _exact_roster_name_match(db_name, roster_t)
        if resolved:
            by_emp[resolved].append(row)
    return by_emp


# Pasted Nokia / Excel grid: treat these cell values as "no leave" in Nokia.
NOKIA_EMPTY_MARKERS = frozenset({".", "-", "—", "", "0", "x", "X"})


def _resolve_nokia_row_name(name_raw: str, roster: Sequence[str] | None = None) -> str | None:
    name_raw = (name_raw or "").strip()
    if not name_raw:
        return None
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    if "," in name_raw:
        last, first = name_raw.split(",", 1)
        last, first = last.strip(), first.strip()
        if first and last:
            for candidate in (f"{first} {last}", f"{last} {first}", first, last):
                r2 = resolve_employee_name(candidate, roster=roster_t)
                if r2:
                    return r2
            m = fuzzy_employee_matches(f"{first} {last}", 1, roster=roster_t)
            if m:
                return m[0]
    r = resolve_employee_name(name_raw, roster=roster_t)
    if r:
        return r
    m = fuzzy_employee_matches(name_raw.replace(",", " "), 1, roster=roster_t)
    return m[0] if m else None


def _find_nokia_day_start_index(header_cells: list[str]) -> int | None:
    """Column index where day-of-month 1 starts (header row from Nokia / Excel)."""
    raw: list[str | None] = []
    for c in header_cells:
        s = c.strip()
        if s.isdigit():
            raw.append(str(int(s)))  # normalize 01 -> 1
        else:
            raw.append(None)

    for i, cell in enumerate(raw):
        if cell != "1":
            continue
        ok = 1
        for day in range(2, 32):
            j = i + day - 1
            if j >= len(raw):
                break
            c = raw[j]
            if c is None:
                break
            try:
                if int(c) != day:
                    break
            except ValueError:
                break
            ok += 1
        if ok >= 10:
            return i
    return None


def _grid_looks_tab_separated(text: str) -> bool:
    for line in (text or "").splitlines():
        if line.count("\t") >= 8:
            return True
    return False


def _grid_looks_csv(text: str) -> bool:
    """Comma-separated export (not tab-pasted Excel). First row may be a title line."""
    if _grid_looks_tab_separated(text):
        return False
    n = 0
    for ln in (text or "").splitlines():
        s = ln.strip()
        if not s:
            continue
        if s.count(",") >= 4:
            return True
        n += 1
        if n >= 24:
            break
    return False


def _csv_to_tsv(text: str) -> str:
    """Normalize CSV (possibly quoted) to tab-separated lines for parse_nokia_grid_tsv."""
    raw = (text or "").strip()
    if not raw:
        return ""
    sample = raw[:16384] if len(raw) > 16384 else raw
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    buf = io.StringIO(raw)
    reader = csv.reader(buf, dialect)
    out_lines: list[str] = []
    for row in reader:
        if not row or not any((c or "").strip() for c in row):
            continue
        out_lines.append("\t".join((c or "").strip() for c in row))
    return "\n".join(out_lines)


def _ocr_header_score(text: str) -> int:
    """Higher = more likely to contain a 1…31 style header for parse_nokia_grid_whitespace."""
    best = 0
    for ln in (text or "").replace("\r\n", "\n").split("\n"):
        parts = ln.split()
        if len(parts) < 5:
            continue
        if _find_nokia_day_start_index(parts) is not None:
            return 2000 + len(parts)
        best = max(best, sum(1 for p in parts if p.isdigit() and 1 <= int(p) <= 31))
    return best


def parse_nokia_grid_whitespace(
    text: str, year: int, month: int, roster: Sequence[str] | None = None
) -> tuple[dict[str, set[int]] | None, list[tuple[str, str]], str | None]:
    """
    Parse Nokia-like grid when cells are separated by spaces (e.g. OCR from a screenshot).
    Expects a header line containing 1, 2, 3… and data rows starting with numeric Emp ID.
    """
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    unmatched: list[tuple[str, str]] = []
    _, last_day = monthrange(year, month)
    lines = [ln for ln in (text or "").replace("\r\n", "\n").split("\n") if ln.strip()]
    if not lines:
        return None, [], "Paste is empty."

    data_start = 0
    found_header = False
    for li, line in enumerate(lines):
        parts = line.split()
        if len(parts) < 10:
            continue
        if _find_nokia_day_start_index(parts) is not None:
            data_start = li + 1
            found_header = True
            break

    if not found_header:
        for li in range(len(lines) - 1):
            merged = f"{lines[li]} {lines[li + 1]}"
            parts = merged.split()
            if len(parts) < 10:
                continue
            if _find_nokia_day_start_index(parts) is not None:
                data_start = li + 2
                found_header = True
                break

    if not found_header:
        return None, [], (
            "Could not find day columns (1, 2, 3…) in the text read from the image. "
            "Try a larger or sharper screenshot, or paste / upload a CSV export of the same grid (comma-separated, from Excel)."
        )

    nokia_days: dict[str, set[int]] = {e: set() for e in roster_t}

    for line in lines[data_start:]:
        parts = line.split()
        if len(parts) < 3:
            continue
        if not parts[0].isdigit():
            continue
        emp_id = parts[0]
        best_pick: tuple[int, int, str, str] | None = None  # (name_len, ds, resolved, name_raw)
        max_ds = min(len(parts), max(2, len(parts) - last_day + 6))
        for ds in range(1, max_ds):
            name_raw = " ".join(parts[1:ds]).strip()
            if not name_raw:
                continue
            resolved = _resolve_nokia_row_name(name_raw, roster=roster_t)
            if not resolved:
                continue
            remainder = len(parts) - ds
            if remainder < last_day or remainder > last_day + 10:
                continue
            pick = (len(name_raw), ds, resolved, name_raw)
            if best_pick is None or pick[:2] > best_pick[:2]:
                best_pick = pick

        if best_pick is None:
            unmatched.append((emp_id, " ".join(parts[1 : min(4, len(parts))])))
            continue
        _, ds, resolved, _name_raw = best_pick
        day_cells = parts[ds:]
        for dom in range(1, last_day + 1):
            if dom - 1 >= len(day_cells):
                break
            val = day_cells[dom - 1].strip()
            if not val or val in NOKIA_EMPTY_MARKERS:
                continue
            if val.lower() in ("w", "wo", "na", "n/a"):
                continue
            nokia_days[resolved].add(dom)

    return nokia_days, unmatched, None


def parse_nokia_grid_tsv(
    text: str, year: int, month: int, roster: Sequence[str] | None = None
) -> tuple[dict[str, set[int]] | None, list[tuple[str, str]], str | None]:
    """
    Parse tab-separated grid copied from Excel (Nokia e-tool export view).
    Returns (nokia_leave_by_roster_name -> set of day-of-month), unmatched (emp_id, raw_name), error_message.
    """
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    unmatched: list[tuple[str, str]] = []
    _, last_day = monthrange(year, month)
    lines = [ln for ln in (text or "").replace("\r\n", "\n").split("\n") if ln.strip()]
    if not lines:
        return None, [], "Paste is empty."

    day_start_idx: int | None = None
    data_start = 0
    for li, line in enumerate(lines):
        cells = line.split("\t")
        idx = _find_nokia_day_start_index(cells)
        if idx is not None:
            day_start_idx = idx
            data_start = li + 1
            break

    if day_start_idx is None:
        for li, line in enumerate(lines):
            cells = line.split("\t")
            low0 = cells[0].strip().lower() if cells else ""
            if "emp" in low0 and "id" in low0.replace(" ", "") and len(cells) >= 10:
                day_start_idx = 2
                data_start = li + 1
                if data_start < len(lines):
                    nxt = lines[data_start].split("\t")
                    if len(nxt) > day_start_idx and not (nxt[day_start_idx].strip().isdigit()):
                        data_start += 1
                break

    if day_start_idx is None:
        return None, [], (
            "Could not find day columns (1…31). The grid text must include a header row with day numbers 1, 2, 3…"
        )

    while data_start < len(lines):
        cells = lines[data_start].split("\t")
        if len(cells) > day_start_idx + 1:
            c0, c1 = cells[0].strip(), cells[1].strip() if len(cells) > 1 else ""
            if c0.isdigit() or "," in c1 or resolve_employee_name(c1, roster=roster_t) or fuzzy_employee_matches(c1, 1, roster=roster_t):
                break
        data_start += 1

    nokia_days: dict[str, set[int]] = {e: set() for e in roster_t}

    for line in lines[data_start:]:
        cells = line.split("\t")
        if len(cells) <= day_start_idx:
            continue
        emp_id = cells[0].strip() if cells else ""
        name_raw = cells[1].strip() if len(cells) > 1 else ""
        if not name_raw and not emp_id:
            continue
        resolved = _resolve_nokia_row_name(name_raw, roster=roster_t)
        if not resolved:
            unmatched.append((emp_id or "—", name_raw or "—"))
            continue
        for dom in range(1, last_day + 1):
            ci = day_start_idx + dom - 1
            if ci >= len(cells):
                break
            val = cells[ci].strip()
            if not val or val in NOKIA_EMPTY_MARKERS:
                continue
            if val.lower() in ("w", "wo", "na", "n/a"):
                continue
            nokia_days[resolved].add(dom)

    return nokia_days, unmatched, None


def parse_nokia_grid_combined(
    text: str, year: int, month: int, roster: Sequence[str] | None = None
) -> tuple[dict[str, set[int]] | None, list[tuple[str, str]], str | None]:
    """Try tab-separated, then CSV, then whitespace / OCR-style lines."""
    t = (text or "").strip()
    if not t:
        return None, [], "Paste is empty."
    errs: list[str] = []
    if _grid_looks_tab_separated(t):
        m, u, e = parse_nokia_grid_tsv(t, year, month, roster=roster)
        if e is None:
            return m, u, e
        errs.append(e)
    if _grid_looks_csv(t):
        tsv = _csv_to_tsv(t)
        if tsv.strip():
            m, u, e = parse_nokia_grid_tsv(tsv, year, month, roster=roster)
            if e is None:
                return m, u, e
            errs.append(e)
    m2, u2, e2 = parse_nokia_grid_whitespace(t, year, month, roster=roster)
    if e2 is None:
        return m2, u2, e2
    errs.append(e2)
    return None, [], errs[0] if errs else e2


# Stored on leave_requests.description for rows created from Nokia e-tool → Mark Approved; grid shows green "A".
NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER = "[nokia-audit-approved]"
# Flask session key: last successful "Show Approved" parse so "Mark Approved" applies the same segments
# when paste text + employee match (fingerprint).
NOKIA_AUDIT_SESSION_APPROVED_SEGS = "nokia_audit_approved_segments_v1"
# Last successful Show Approved / Show DSM tables (same employee) for Compare merge.
NOKIA_AUDIT_SESSION_LAST_APPROVED_PREVIEW = "nokia_audit_last_approved_preview_v1"
NOKIA_AUDIT_SESSION_LAST_DSM = "nokia_audit_last_dsm_v1"


def _leave_reason_counts_toward_day_units(reason: str | None) -> bool:
    """CompOFF is shown on grids but excluded from leave day-units, capacity debits, and HPPM-style absence totals."""
    return (reason or "").strip().casefold() != "compoff"


def _leave_row_display_tiebreak(row: sqlite3.Row) -> tuple[int, int]:
    """Prefer Nokia-audit tagged rows when several leaves overlap the same calendar day (then highest id)."""
    desc = str(row["description"] or "")
    has_marker = 1 if NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in desc else 0
    return (has_marker, int(row["id"]))


def _nokia_reason_and_label_from_line(low: str) -> tuple[str, str]:
    """Map a Nokia summary line (lowercased) to a leave_requests.reason code and display label."""
    rl = dict(LEAVE_REASONS)
    if "sick" in low or "medical" in low or "hospital" in low or "health" in low or "doctor" in low:
        return "sl", rl["sl"]
    if "birthday" in low or "anniversary" in low or "b'day" in low or "bday" in low:
        return "ll", rl["ll"]
    if re.search(r"\bcomp[\s-]*off\b", low) or ("compensatory" in low and "off" in low):
        return "compoff", rl["compoff"]
    if "casual" in low or "unplanned" in low or "emergency" in low or "personal" in low or "compassionate" in low:
        return "ul", rl["ul"]
    if "work from home" in low or re.search(r"\bwfh\b", low):
        return "ul", rl["ul"]
    if "maternity" in low or "paternity" in low or "bereavement" in low:
        return "ul", rl["ul"]
    return "pl", rl["pl"]


def _nokia_segment_tuples_from_filtered_approved_lines(filtered: str) -> list[tuple[date, date, str, str, bool]]:
    """
    Parse lines that already contain 'approved' (case-insensitive).
    Returns one tuple per parseable row: (start, end, reason_code, reason_label, half_day_hint), unclipped.
    """
    patt = re.compile(r"\b(\d{2})[-/](\d{2})[-/](\d{4})\b")

    def to_date(m: re.Match[str]) -> date | None:
        try:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, d)
        except ValueError:
            return None

    out: list[tuple[date, date, str, str, bool]] = []
    for line in (filtered or "").splitlines():
        s = line.strip()
        if not s:
            continue
        low = s.lower()
        if not any(
            k in low
            for k in (
                "approved",
                "leave entry",
                "pending",
                "rejected",
            )
        ):
            continue
        matches = list(patt.finditer(s))
        if not matches:
            continue
        dates: list[date] = []
        for m in matches:
            dd = to_date(m)
            if dd:
                dates.append(dd)
        if not dates:
            continue
        if len(dates) >= 2:
            d0, d1 = dates[0], dates[1]
        else:
            d0 = d1 = dates[0]
        if d1 < d0:
            d0, d1 = d1, d0
        reason, label = _nokia_reason_and_label_from_line(low)
        half_hint = bool(re.search(r"(-\s*0\.5|0\.5\s*day|\b0\.5\b)", low, re.I))
        out.append((d0, d1, reason, label, half_hint))
    return out


def iter_nokia_employee_approved_segments_clipped_to_month(
    text: str, year: int, month: int
) -> tuple[list[tuple[date, date, str, str, bool]], str | None]:
    """
    Parse Nokia *per-employee leave summary* (Annual / Sick tables): lines with status keywords
    and From/To dates as DD-MM-YYYY or DD/MM/YYYY. Returns segments clipped to [year, month] as
    (start, end, reason_code, reason_label, half_day_hint).
    """
    raw = (text or "").strip()
    if not raw:
        return [], "No text to parse (paste OCR text or CSV, or use a screenshot so OCR can run)."
    patt = re.compile(r"\b(\d{2})[-/](\d{2})[-/](\d{4})\b")

    def to_date(m: re.Match[str]) -> date | None:
        try:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, d)
        except ValueError:
            return None

    first = date(year, month, 1)
    _, last_d = monthrange(year, month)
    last = date(year, month, last_d)

    segments: list[tuple[date, date, str, str, bool]] = []

    for line in raw.splitlines():
        s = line.strip()
        if not s:
            continue
        low = s.lower()
        if not any(
            k in low
            for k in (
                "approved",
                "leave entry",
                "pending",
                "rejected",
            )
        ):
            continue
        matches = list(patt.finditer(s))
        if not matches:
            continue
        dates: list[date] = []
        for m in matches:
            dd = to_date(m)
            if dd:
                dates.append(dd)
        if not dates:
            continue
        if len(dates) >= 2:
            d0, d1 = dates[0], dates[1]
        else:
            d0 = d1 = dates[0]
        if d1 < d0:
            d0, d1 = d1, d0
        ov_lo = max(d0, first)
        ov_hi = min(d1, last)
        if ov_lo > ov_hi:
            continue
        reason, label = _nokia_reason_and_label_from_line(low)
        half_hint = bool(re.search(r"(-\s*0\.5|0\.5\s*day|\b0\.5\b)", low, re.I))
        segments.append((ov_lo, ov_hi, reason, label, half_hint))

    if not segments:
        return [], (
            f"No leave rows with dates in {calendar.month_name[month]} {year} were found. "
            "Use lines that include Approved (or Leave entry) and From/To dates like 15-05-2026, "
            "or switch to team calendar grid mode."
        )
    return segments, None


def parse_nokia_employee_summary_leave_dates(
    text: str, year: int, month: int
) -> tuple[set[int] | None, str | None]:
    """
    Parse Nokia *per-employee leave summary* (Annual / Sick tables): lines with Approved leave
    and From/To dates as DD-MM-YYYY or DD/MM/YYYY. Returns day-of-month numbers in [year, month]
    that Nokia marks as leave (full-day presence; half-days still count as that calendar day).
    """
    segs, err = iter_nokia_employee_approved_segments_clipped_to_month(text, year, month)
    if err:
        return None, err
    out: set[int] = set()
    for lo, hi, _, _, _ in segs:
        cur = lo
        while cur <= hi:
            out.add(cur.day)
            cur += timedelta(days=1)
    return out, None


def _employee_tracker_leave_dates_covered(
    conn: sqlite3.Connection, employee_name: str, range_start: date, range_end: date
) -> set[date]:
    """Calendar dates in [range_start, range_end] where employee already has pending or approved leave."""
    covered: set[date] = set()
    rows = conn.execute(
        """
        SELECT start_date, end_date FROM leave_requests
        WHERE employee_name = ?
          AND status IN ('pending', 'approved')
          AND start_date <= ? AND end_date >= ?
        """,
        (employee_name, range_end.isoformat(), range_start.isoformat()),
    ).fetchall()
    for r in rows:
        sd = date.fromisoformat(str(r["start_date"])[:10])
        ed = date.fromisoformat(str(r["end_date"])[:10])
        a = max(sd, range_start)
        b = min(ed, range_end)
        if a > b:
            continue
        cur = a
        while cur <= b:
            covered.add(cur)
            cur += timedelta(days=1)
    return covered


def _nokia_tag_tracker_leave_description_marker_for_days(
    conn: sqlite3.Connection, employee_name: str, cover_dates: Iterable[date]
) -> int:
    """
    For each calendar day in cover_dates, append the Nokia marker to pending/approved leave rows
    for that employee that overlap the day but do not already contain the marker.
    Returns how many distinct leave_requests rows were updated.
    """
    marker = NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER
    updated_ids: set[int] = set()
    for d in cover_dates:
        d_iso = d.isoformat()
        rows = conn.execute(
            """
            SELECT id, description FROM leave_requests
            WHERE employee_name = ?
              AND status IN ('pending', 'approved')
              AND start_date <= ? AND end_date >= ?
            """,
            (employee_name, d_iso, d_iso),
        ).fetchall()
        for r in rows:
            rid = int(r["id"])
            if rid in updated_ids:
                continue
            desc = str(r["description"] or "")
            if marker in desc:
                continue
            new_desc = f"{desc} {marker}".strip() if desc else marker
            conn.execute("UPDATE leave_requests SET description = ? WHERE id = ?", (new_desc, rid))
            updated_ids.add(rid)
    return len(updated_ids)


def _nokia_paste_fingerprint(raw_text: str) -> str:
    return hashlib.sha256((raw_text or "").strip().encode("utf-8")).hexdigest()


PREVIEW_LEAVE_TYPE_LABEL = {
    "sl": "Sick leave",
    "pl": "Annual leave",
    "ul": "Casual leave",
    "ll": "Birthday / anniversary leave",
    "compoff": "CompOFF",
}


def _nokia_eleavetool_type_display(type_label: str, days_val: Any) -> str:
    """
    Nokia eLeave tool tables: show planned/annual leave as **A** (1 day) or **A1/2** (0.5 day)
    from day-units. Other leave types keep their full label.
    Recognizes preview text ``Annual leave`` and segment label ``PL — Planned leave``.
    """
    t = (type_label or "").strip()
    if t not in ("Annual leave", "PL — Planned leave"):
        return t or "—"
    try:
        d = float(str(days_val).strip().replace(",", "."))
    except (TypeError, ValueError):
        d = 1.0
    if abs(d - 0.5) < 1e-6:
        return "A1/2"
    return "A"


def _nokia_segs_to_store(segs: list[tuple[date, date, str, str, bool]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for lo, hi, reason, label, half in segs:
        out.append(
            {
                "lo": lo.isoformat(),
                "hi": hi.isoformat(),
                "reason": reason,
                "label": label,
                "half": bool(half),
            }
        )
    return out


def _nokia_segs_from_store(items: Any) -> list[tuple[date, date, str, str, bool]] | None:
    if not isinstance(items, list):
        return None
    try:
        segs: list[tuple[date, date, str, str, bool]] = []
        for it in items:
            if not isinstance(it, dict):
                return None
            segs.append(
                (
                    date.fromisoformat(str(it["lo"])[:10]),
                    date.fromisoformat(str(it["hi"])[:10]),
                    str(it["reason"]),
                    str(it["label"]),
                    bool(it.get("half")),
                )
            )
        return segs
    except (KeyError, TypeError, ValueError):
        return None


def _nokia_all_cover_weekdays_from_segs(segs: list[tuple[date, date, str, str, bool]]) -> set[date]:
    """Weekdays (Mon–Fri) in each Nokia segment range — used when tagging existing tracker rows."""
    all_cover: set[date] = set()
    for lo, hi, _, _, _ in segs:
        for cur in _weekdays_in_range_inclusive(lo, hi):
            all_cover.add(cur)
    return all_cover


def _nokia_plan_inserts_for_segments(
    conn: sqlite3.Connection,
    employee_name: str,
    segs: list[tuple[date, date, str, str, bool]],
) -> list[dict[str, Any]]:
    """
    Build planned INSERT rows for the given segments against the current DB (same connection),
    including rows inserted earlier in this transaction. Re-queries covered dates at the start
    of each segment so later Nokia lines are not dropped by in-memory simulation of earlier lines.
    Only **Monday–Friday** dates in each segment range are considered (weekends never get new rows).
    """
    overall_lo = min(t[0] for t in segs)
    overall_hi = max(t[1] for t in segs)
    plans: list[dict[str, Any]] = []

    for lo, hi, reason, label, half_hint in segs:
        have_dates = _employee_tracker_leave_dates_covered(conn, employee_name, overall_lo, overall_hi)
        missing: list[date] = []
        dates_to_consider = _weekdays_in_range_inclusive(lo, hi)
        for cur in dates_to_consider:
            if cur not in have_dates:
                missing.append(cur)
        if not missing:
            continue
        i = 0
        while i < len(missing):
            j = i
            while j + 1 < len(missing) and missing[j + 1] == missing[j] + timedelta(days=1):
                j += 1
            run_start, run_end = missing[i], missing[j]
            n_days = (run_end - run_start).days + 1
            use_half = bool(half_hint and n_days == 1 and lo == hi)
            if use_half:
                dur = "half_am"
                days_label = "0.5"
            elif n_days == 1:
                dur = "full"
                days_label = "1"
            else:
                dur = "multi"
                days_label = str(n_days)
            plans.append(
                {
                    "run_start": run_start,
                    "run_end": run_end,
                    "dur": dur,
                    "reason": reason,
                    "label": label,
                    "days_label": days_label,
                }
            )
            i = j + 1

    return plans


def _nokia_preview_rows_from_segments(
    employee_name: str, segs: list[tuple[date, date, str, str, bool]]
) -> list[dict[str, Any]]:
    """
    One preview table row per parsed Nokia approved line (segment), before tracker deduplication.
    Day counts follow the same weekday rules as marking for multi-calendar-day ranges.
    """
    rows: list[dict[str, Any]] = []
    for d0, d1, reason, label, half_hint in segs:
        wd_in_span = len(_weekdays_in_range_inclusive(d0, d1))
        if d0 == d1:
            if d0.weekday() >= 5:
                days_s = "0"
            elif half_hint:
                days_s = "0.5"
            else:
                days_s = "1"
        else:
            days_s = str(wd_in_span)
        typ = PREVIEW_LEAVE_TYPE_LABEL.get(reason, str(label or reason))
        typ = _nokia_eleavetool_type_display(typ, days_s)
        rows.append(
            {
                "name": employee_name,
                "days": days_s,
                "type": typ,
                "status": "Approved",
                "dates_range": f"{d0.strftime('%d-%m-%Y')} → {d1.strftime('%d-%m-%Y')}",
            }
        )
    return rows


def _nokia_preview_rows_from_plans(employee_name: str, plans: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for p in plans:
        rs: date = p["run_start"]
        re: date = p["run_end"]
        reason = str(p["reason"])
        typ = PREVIEW_LEAVE_TYPE_LABEL.get(reason, str(p.get("label") or reason))
        typ = _nokia_eleavetool_type_display(typ, p["days_label"])
        rows.append(
            {
                "name": employee_name,
                "days": p["days_label"],
                "type": typ,
                "status": "Approved",
                "dates_range": f"{rs.strftime('%d-%m-%Y')} → {re.strftime('%d-%m-%Y')}",
            }
        )
    return rows


def _nokia_marked_summary_from_plan_row(employee_name: str, p: dict[str, Any]) -> dict[str, Any]:
    rs: date = p["run_start"]
    re: date = p["run_end"]
    return {
        "name": employee_name,
        "days": p["days_label"],
        "type": _nokia_eleavetool_type_display(str(p.get("label") or ""), p.get("days_label")),
        "status": "Approved in tracker",
        "period": f"{rs.isoformat()} — {re.isoformat()}",
        "dates_range": f"{rs.strftime('%d-%m-%Y')} → {re.strftime('%d-%m-%Y')}",
        "_sort_start": rs.isoformat(),
    }


def _nokia_mark_approved_leaves(
    app: Flask,
    employee_name: str,
    year: int,
    month: int,
    text: str,
    *,
    precomputed_segments: list[tuple[date, date, str, str, bool]] | None = None,
) -> tuple[list[dict[str, Any]], int, str | None]:
    """
    For Nokia segments, only **Monday–Friday** are written to the tracker (Saturday/Sunday are always
    omitted), including single-day weekend lines from the paste (no row is created for those days).
    Rows are tagged so the monthly grid shows green "A".
    Days that already have tracker leave get the same marker appended to overlapping rows so the
    calendar shows green "A" without duplicating days.
    When ``precomputed_segments`` is set (same paste + employee as the last successful
    **Show Approved Leaves** in this session), inserts follow that plan so they match the preview table.
    Otherwise segments are parsed from ``text``. year/month are kept for the route signature.
    """
    if precomputed_segments is None:
        filtered = "\n".join(L for L in (text or "").splitlines() if "approved" in L.lower())
        if not filtered.strip():
            return [], 0, (
                'No lines containing "Approved" were found. Copy the approved rows from Nokia eLeave '
                "(one employee at a time), then try again."
            )
        segs = _nokia_segment_tuples_from_filtered_approved_lines(filtered.strip())
        if not segs:
            return [], 0, (
                "No leave rows with recognizable dates were found in the pasted text. "
                "Use lines that include Approved (or Leave entry) and dates like 15-05-2026."
            )
    else:
        segs = precomputed_segments
        if not segs:
            return [], 0, (
                "No approved leave rows to apply. Click Show Approved Leaves first, "
                "then Mark Approved leave in Leave tracker."
            )

    all_cover = _nokia_all_cover_weekdays_from_segs(segs)

    created = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    ip = client_ip()
    conn = get_db(app)
    marked: list[dict[str, Any]] = []

    try:
        for seg in segs:
            plans_part = _nokia_plan_inserts_for_segments(conn, employee_name, [seg])
            for p in plans_part:
                conn.execute(
                    """
                    INSERT INTO leave_requests
                    (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
                    VALUES (?, ?, ?, ?, ?, ?, 'approved', ?, ?)
                    """,
                    (
                        employee_name,
                        str(p["reason"]),
                        NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER,
                        p["run_start"].isoformat(),
                        p["run_end"].isoformat(),
                        str(p["dur"]),
                        created,
                        ip,
                    ),
                )
                marked.append(_nokia_marked_summary_from_plan_row(employee_name, p))
        tagged_existing = _nokia_tag_tracker_leave_description_marker_for_days(conn, employee_name, all_cover)
        conn.commit()
    except Exception as ex:  # noqa: BLE001
        conn.rollback()
        conn.close()
        return [], 0, str(ex)
    conn.close()
    marked.sort(key=lambda r: (r.get("_sort_start") or "", str(r.get("period") or "")))
    for r in marked:
        r.pop("_sort_start", None)
    return marked, tagged_existing, None


def _nokia_audit_tracker_approved_leave_rows(
    app: Flask,
    employee_name: str,
    year: int,
    month: int,
) -> list[dict[str, Any]]:
    """Approved leave_requests overlapping [year, month] for one roster employee (manager view)."""
    first = date(year, month, 1)
    _, ld = monthrange(year, month)
    last = date(year, month, ld)
    start_iso = first.isoformat()
    end_iso = last.isoformat()
    reason_labels = dict(LEAVE_REASONS)
    conn = get_db(app)
    rows = list(
        conn.execute(
            """
            SELECT employee_name, reason, start_date, end_date, duration_type, status
            FROM leave_requests
            WHERE employee_name = ?
              AND status = 'approved'
              AND start_date <= ? AND end_date >= ?
            ORDER BY start_date ASC, id ASC
            """,
            (employee_name, end_iso, start_iso),
        )
    )
    conn.close()
    out: list[dict[str, Any]] = []
    for r in rows:
        sd = date.fromisoformat(str(r["start_date"])[:10])
        ed = date.fromisoformat(str(r["end_date"])[:10])
        span = (ed - sd).days + 1
        dur = str(r["duration_type"] or "")
        if dur in ("half_am", "half_pm"):
            days_s = "0.5"
        elif dur == "multi" or span > 1:
            days_s = str(span)
        else:
            days_s = "1"
        out.append(
            {
                "name": r["employee_name"],
                "from_date": str(r["start_date"])[:10],
                "to_date": str(r["end_date"])[:10],
                "type": reason_labels.get(str(r["reason"]), str(r["reason"])),
                "days": days_s,
                "status": "Approved",
            }
        )
    return out


def _dsm_audit_days_token_is_half(days_val: Any) -> bool:
    t = str(days_val or "").strip().replace(",", ".")
    if not t:
        return False
    try:
        return abs(float(t) - 0.5) < 1e-9
    except (ValueError, TypeError):
        return False


def _nokia_audit_dsm_row_sort_key(r: dict[str, Any]) -> tuple[str, date, date]:
    pr = _parse_nokia_audit_dd_mm_range(str(r.get("dates_range") or ""))
    nm = str(r.get("name") or "")
    if not pr:
        return (nm, date.max, date.max)
    return (nm, pr[0], pr[1])


def _merge_contiguous_nokia_audit_dsm_leave_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Merge consecutive DSM audit rows when the tracker stored adjacent single-day (or abutting)
    ranges for the same person, status, and grid type — so Nokia **Compare** sees one DSM span
    matching one Nokia line instead of two fragments (e.g. 26–26 + 27–27 → 26–27).
    Half-day rows are not merged with neighbours.
    """
    if not rows or len(rows) < 2:
        return list(rows or [])
    expanded: list[tuple[tuple[date, date], dict[str, Any]]] = []
    unparsed: list[dict[str, Any]] = []
    for r in rows:
        pr = _parse_nokia_audit_dd_mm_range(str(r.get("dates_range") or ""))
        if not pr:
            unparsed.append(dict(r))
            continue
        sd, ed = pr
        if ed < sd:
            sd, ed = ed, sd
        expanded.append(((sd, ed), dict(r)))
    expanded.sort(key=lambda it: (str(it[1].get("name") or ""), it[0][0], it[0][1]))
    merged: list[dict[str, Any]] = []
    idx = 0
    while idx < len(expanded):
        (bsd, bed), base = expanded[idx]
        last_end = bed
        j = idx
        while j + 1 < len(expanded):
            (nsd, ned), nxt = expanded[j + 1]
            same = (
                str(base.get("name") or "") == str(nxt.get("name") or "")
                and str(base.get("status") or "") == str(nxt.get("status") or "")
                and str(base.get("type") or "") == str(nxt.get("type") or "")
            )
            contiguous = last_end + timedelta(days=1) == nsd
            halves = _dsm_audit_days_token_is_half(base.get("days")) or _dsm_audit_days_token_is_half(nxt.get("days"))
            if not (same and contiguous and not halves):
                break
            last_end = ned
            j += 1
        csd, ced = bsd, last_end
        if csd == ced:
            days_s = (
                str(base.get("days")).strip()
                if _dsm_audit_days_token_is_half(base.get("days"))
                else "1"
            )
        else:
            days_s = str(len(_weekdays_in_range_inclusive(csd, ced)))
        merged.append(
            {
                "name": base.get("name"),
                "days": days_s,
                "type": base.get("type"),
                "status": base.get("status"),
                "dates_range": f"{csd.strftime('%d-%m-%Y')} → {ced.strftime('%d-%m-%Y')}",
            }
        )
        idx = j + 1
    merged.extend(unparsed)
    merged.sort(key=_nokia_audit_dsm_row_sort_key)
    return merged


def _nokia_audit_dsm_leave_rows(
    app: Flask,
    employee_name: str,
    year: int,
) -> list[dict[str, Any]]:
    """
    Pending + approved ``leave_requests`` overlapping the **calendar year** ``year`` (1 Jan–31 Dec),
    for the DSM / Leave tracker view. **Type** uses the same codes as the month grid
    (``leave_cell_code``: PL/UL/SL/LL/CO, half-day ½, Nokia A).
    """
    year_start = date(year, 1, 1).isoformat()
    year_end = date(year, 12, 31).isoformat()
    conn = get_db(app)
    rows = list(
        conn.execute(
            """
            SELECT employee_name, reason, description, start_date, end_date, duration_type, status
            FROM leave_requests
            WHERE employee_name = ?
              AND status IN ('pending', 'approved')
              AND start_date <= ? AND end_date >= ?
            ORDER BY start_date ASC, id ASC
            """,
            (employee_name, year_end, year_start),
        )
    )
    conn.close()
    out: list[dict[str, Any]] = []
    for r in rows:
        sd = date.fromisoformat(str(r["start_date"])[:10])
        ed = date.fromisoformat(str(r["end_date"])[:10])
        dur = str(r["duration_type"] or "")
        reason = str(r["reason"] or "")
        status = str(r["status"] or "")
        desc = str(r["description"] or "")

        if dur in ("half_am", "half_pm"):
            days_s = "0.5"
        elif sd == ed:
            days_s = "1"
        else:
            days_s = str(len(_weekdays_in_range_inclusive(sd, ed)))

        type_cell = leave_cell_code(reason, dur, status, desc)
        if status == "pending":
            status_label = "Pending"
        elif status == "approved":
            status_label = "Approved"
        else:
            status_label = status.replace("_", " ").strip().title() or "—"

        out.append(
            {
                "name": r["employee_name"],
                "days": days_s,
                "type": type_cell,
                "status": status_label,
                "dates_range": f"{sd.strftime('%d-%m-%Y')} → {ed.strftime('%d-%m-%Y')}",
            }
        )
    return _merge_contiguous_nokia_audit_dsm_leave_rows(out)


def _nokia_paste_approved_preview_rows(
    _app: Flask,
    text: str,
    employee_name: str,
) -> tuple[list[dict[str, Any]], str | None, list[tuple[date, date, str, str, bool]] | None]:
    """
    Build one preview table row per parsed Nokia **approved** line (all types), including lines whose
    calendar days overlap others — the table lists every Nokia row; **Mark Approved** still skips
    days already on the tracker when inserting.
    Returns ``(rows, error, segments)``; ``segments`` is set on success for the next Mark action
    when paste text and employee are unchanged. ``_app`` is unused but kept for a stable call signature.
    """
    filtered = "\n".join(L for L in (text or "").splitlines() if "approved" in L.lower())
    if not filtered.strip():
        return [], "Paste Nokia eLeave text (approved rows), then click Show Approved Leaves.", None

    segs = _nokia_segment_tuples_from_filtered_approved_lines(filtered.strip())
    if not segs:
        return [], (
            "No leave rows with recognizable dates were found in the pasted text. "
            "Use lines that include Approved (or Leave entry) and dates like 15-05-2026."
        ), None

    # Chronological order (not lexicographic on dd-mm-yyyy strings).
    segs.sort(key=lambda t: (t[0], t[1]))
    rows = _nokia_preview_rows_from_segments(employee_name, segs)
    return rows, None, segs


def _sum_nokia_audit_day_column(rows: list[dict[str, Any]]) -> float:
    """Sum numeric 'days' cells from Nokia preview / marked summary rows (handles '1', '0.5', '3')."""
    total = 0.0
    for r in rows or []:
        raw = r.get("days")
        try:
            total += float(str(raw).strip().replace(",", "."))
        except (TypeError, ValueError):
            continue
    return round(total, 2)


def _parse_nokia_audit_dd_mm_range(dates_range: str) -> tuple[date, date] | None:
    """
    Parse an inclusive date range from a Nokia / DSM table cell.

    Accepts ``dd-mm-yyyy → dd-mm-yyyy`` (Unicode arrow), ``dd-mm-yyyy -> dd-mm-yyyy``,
    slashes, and any text where the **first** and **last** ``dd-mm-yyyy`` / ``dd/mm/yyyy``
    tokens define the span (single-day cells often repeat the same date twice).
    """
    s = (dates_range or "").strip()
    if not s:
        return None
    pat = re.compile(r"\b(\d{2})[-/](\d{2})[-/](\d{4})\b")

    def _one(m: re.Match[str]) -> date | None:
        for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(m.group(0), fmt).date()
            except ValueError:
                continue
        return None

    matches = list(pat.finditer(s))
    if not matches:
        return None
    if len(matches) == 1:
        d0 = _one(matches[0])
        return (d0, d0) if d0 else None
    d_lo = _one(matches[0])
    d_hi = _one(matches[-1])
    if not d_lo or not d_hi:
        return None
    if d_hi < d_lo:
        d_lo, d_hi = d_hi, d_lo
    return d_lo, d_hi


def _nokia_audit_row_days_cell(val: Any) -> float:
    if val in (None, "", "—"):
        return 0.0
    try:
        return float(str(val).strip().replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _nokia_audit_type_key_for_dedupe(type_s: str) -> str:
    """Normalize eLeave / planned leave labels so duplicate rows still match after display renames."""
    t = (type_s or "").strip()
    if t in ("A", "A1/2", "Annual leave", "PL — Planned leave"):
        return "__pl__"
    return t.casefold()


def _nokia_audit_dedupe_elv_rows_for_compare(nokia_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Nokia paste sometimes repeats the same approved line (same calendar span, day count, and type).

    Each duplicate would otherwise be paired greedily against the same DSM leave, leaving later
    copies with empty DSM columns while the tracker still has a single row — confusing in Compare.
    Keep the first occurrence in paste order and drop identical follow-ups.
    """
    seen: set[tuple[Any, ...]] = set()
    out: list[dict[str, Any]] = []
    for r in nokia_rows or []:
        pr = _parse_nokia_audit_dd_mm_range(str(r.get("dates_range") or ""))
        days_s = str(r.get("days") or "").strip()
        type_s = str(r.get("type") or "").strip()
        norm_type = _nokia_audit_type_key_for_dedupe(type_s)
        if pr:
            key = (pr[0], pr[1], days_s, norm_type)
        else:
            key = (str(r.get("dates_range") or "").strip(), days_s, norm_type)
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def _calendar_overlap_days(a0: date, a1: date, b0: date, b1: date) -> int:
    lo = max(a0, b0)
    hi = min(a1, b1)
    if lo > hi:
        return 0
    return (hi - lo).days + 1


def _nokia_audit_session_safe_row_dicts(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for r in rows or []:
        out.append({str(k): "" if v is None else str(v) for k, v in r.items()})
    return out


def _compare_row_from_pair(nokia: dict[str, Any] | None, dsm: dict[str, Any] | None) -> dict[str, Any]:
    src = nokia or dsm or {}
    return {
        "name": str(src.get("name") or "").strip() or "—",
        "days_elv": (nokia or {}).get("days") if nokia else "—",
        "days_dsm": (dsm or {}).get("days") if dsm else "—",
        "type_elv": (
            _nokia_eleavetool_type_display(str((nokia or {}).get("type") or ""), (nokia or {}).get("days"))
            if nokia
            else "—"
        ),
        "type_dsm": (dsm or {}).get("type") if dsm else "—",
        "dates_elv": (nokia or {}).get("dates_range") if nokia else "—",
        "dates_dsm": (dsm or {}).get("dates_range") if dsm else "—",
    }


def _nokia_audit_build_compare_rows(
    nokia_rows: list[dict[str, Any]],
    dsm_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Merge Nokia **Show Approved** preview rows with **Show DSM** tracker rows by overlapping
    calendar date ranges.

    Identical Nokia preview rows (same parsed dates, day count, and type) are deduplicated first
    so one tracker row is not “consumed” only by the first copy, leaving duplicates with blank DSM.

    Each remaining Nokia row takes the best-overlapping **unused** DSM row (prefer exact span,
    then closest day count, then stable order).
    """
    nokia_rows = _nokia_audit_dedupe_elv_rows_for_compare(list(nokia_rows or []))
    out: list[dict[str, Any]] = []
    used_dsm: set[int] = set()
    for nr in nokia_rows or []:
        pr = _parse_nokia_audit_dd_mm_range(str(nr.get("dates_range") or ""))
        if not pr:
            out.append(_compare_row_from_pair(nr, None))
            continue
        n0, n1 = pr
        best_j: int | None = None
        best_key: tuple[int, int, float, int] | None = None
        nd = _nokia_audit_row_days_cell(nr.get("days"))
        for j, dr in enumerate(dsm_rows or []):
            if j in used_dsm:
                continue
            prd = _parse_nokia_audit_dd_mm_range(str(dr.get("dates_range") or ""))
            if not prd:
                continue
            ov = _calendar_overlap_days(n0, n1, prd[0], prd[1])
            if ov <= 0:
                continue
            exact = 1 if (n0, n1) == (prd[0], prd[1]) else 0
            dd = _nokia_audit_row_days_cell(dr.get("days"))
            cand: tuple[int, int, float, int] = (
                -ov,
                -exact,
                abs(nd - dd),
                j,
            )
            if best_key is None or cand < best_key:
                best_key = cand
                best_j = j
        if best_j is not None:
            used_dsm.add(best_j)
            out.append(_compare_row_from_pair(nr, dsm_rows[best_j]))
        else:
            out.append(_compare_row_from_pair(nr, None))
    for j, dr in enumerate(dsm_rows or []):
        if j not in used_dsm:
            out.append(_compare_row_from_pair(None, dr))

    def _sort_key(row: dict[str, Any]) -> tuple[date, str]:
        pr = _parse_nokia_audit_dd_mm_range(str(row.get("dates_elv") or ""))
        if not pr:
            pr = _parse_nokia_audit_dd_mm_range(str(row.get("dates_dsm") or ""))
        if not pr:
            return date.max, ""
        return pr[0], str(row.get("name") or "")

    out.sort(key=_sort_key)
    out = _merge_compare_rows_contiguous_elv_shared_dsm(out)
    return out


def _merge_compare_rows_contiguous_elv_shared_dsm(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    After pairing, merge adjacent compare rows when the first row holds the DSM span and the next
    row has **no** DSM column (already paired to the same leave in the tracker) but **contiguous**
    eTool dates that extend the first row's eTool span **within** that DSM range — one visual row
    instead of a fragment + orphan (e.g. eTool 23–23 + 24–25 with DSM 23–25 on the first row only).
    """
    if len(rows) < 2:
        return list(rows or [])

    def _dsm_cell_empty(v: Any) -> bool:
        s = str(v or "").strip()
        return not s or s in ("—", "-")

    def _same_dsm_or_nxt_empty(cur_dsm: Any, nxt_dsm: Any) -> bool:
        if _dsm_cell_empty(nxt_dsm):
            return True
        a = _parse_nokia_audit_dd_mm_range(str(cur_dsm or ""))
        b = _parse_nokia_audit_dd_mm_range(str(nxt_dsm or ""))
        return bool(a and b and a == b)

    out: list[dict[str, Any]] = []
    i = 0
    nlen = len(rows)
    while i < nlen:
        cur = dict(rows[i])
        j = i + 1
        while j < nlen:
            nxt = rows[j]
            if str(cur.get("name") or "") != str(nxt.get("name") or ""):
                break
            pr_d_cur = _parse_nokia_audit_dd_mm_range(str(cur.get("dates_dsm") or ""))
            if not pr_d_cur or _dsm_cell_empty(cur.get("dates_dsm")):
                break
            if not _same_dsm_or_nxt_empty(cur.get("dates_dsm"), nxt.get("dates_dsm")):
                break
            pe_c = _parse_nokia_audit_dd_mm_range(str(cur.get("dates_elv") or ""))
            pe_n = _parse_nokia_audit_dd_mm_range(str(nxt.get("dates_elv") or ""))
            if not pe_c or not pe_n:
                break
            c0, c1 = pe_c
            n0, n1 = pe_n
            if c1 < c0:
                c0, c1 = c1, c0
            if n1 < n0:
                n0, n1 = n1, n0
            if c1 + timedelta(days=1) != n0:
                break
            if str(cur.get("type_elv") or "") != str(nxt.get("type_elv") or ""):
                break
            if not _dsm_cell_empty(nxt.get("dates_dsm")):
                if str(cur.get("type_dsm") or "") != str(nxt.get("type_dsm") or ""):
                    break
            d0, d1 = pr_d_cur
            m0, m1 = min(c0, n0), max(c1, n1)
            if m0 < d0 or m1 > d1:
                break
            cur["dates_elv"] = f"{m0.strftime('%d-%m-%Y')} → {m1.strftime('%d-%m-%Y')}"
            if m0 == m1:
                cur["days_elv"] = "1"
            else:
                cur["days_elv"] = str(len(_weekdays_in_range_inclusive(m0, m1)))
            t1 = str(cur.get("type_elv") or "")
            t2 = str(nxt.get("type_elv") or "")
            if t1 in ("A", "A1/2") and t2 in ("A", "A1/2"):
                cur["type_elv"] = _nokia_eleavetool_type_display("Annual leave", cur["days_elv"])
            j += 1
        out.append(cur)
        i = j
    return out


def _nokia_audit_compare_rows_from_session(employee_name: str, year: int) -> tuple[list[dict[str, Any]] | None, str | None]:
    """
    If the session holds a successful **Show Approved Leaves** preview and **Show DSM Leaves** load
    for ``employee_name`` and calendar ``year``, return merged compare rows. Otherwise return
    ``(None, error_message)``.
    """
    ap = session.get(NOKIA_AUDIT_SESSION_LAST_APPROVED_PREVIEW)
    dp = session.get(NOKIA_AUDIT_SESSION_LAST_DSM)
    ap_ok = (
        isinstance(ap, dict)
        and str(ap.get("employee_name") or "").strip() == employee_name
        and isinstance(ap.get("rows"), list)
    )
    dp_ok = (
        isinstance(dp, dict)
        and str(dp.get("employee_name") or "").strip() == employee_name
        and int(dp.get("year") or -1) == int(year)
        and isinstance(dp.get("rows"), list)
    )
    if not ap_ok:
        return None, (
            "Run Show Approved Leaves successfully first for this employee "
            "(so the eLeave tool preview is saved), then Show DSM Leaves for this calendar year, "
            "then Compare."
        )
    if not dp_ok:
        return None, (
            f"Run Show DSM Leaves first for calendar year {year} with this employee selected, "
            "then Compare."
        )
    nokia_list = list(ap.get("rows") or [])
    dsm_list = list(dp.get("rows") or [])
    return _nokia_audit_build_compare_rows(nokia_list, dsm_list), None


def _sum_nokia_audit_compare_days(rows: list[dict[str, Any]], key: str) -> float:
    total = 0.0
    for r in rows or []:
        raw = r.get(key)
        if raw in (None, "", "—"):
            continue
        try:
            total += float(str(raw).strip().replace(",", "."))
        except (TypeError, ValueError):
            continue
    return round(total, 2)


def build_nokia_compare_xlsx_bytes(
    rows: list[dict[str, Any]],
    *,
    employee_name: str,
    year: int,
) -> tuple[bytes | None, str | None]:
    """Build a one-sheet .xlsx matching the Compare table (eLeave vs DSM)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return None, "Excel export requires openpyxl (install dependencies)."

    headers = (
        "NAME",
        "No. of Days (eTool)",
        "No. of Days (DSM)",
        "eLeavetool Type",
        "Type of leave DSM",
        "Leave dates (eTool)",
        "Leave dates (DSM)",
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Compare"
    ws.cell(row=1, column=1, value="Compare — eLeave tool vs DSM")
    ws.cell(row=2, column=1, value="Employee")
    ws.cell(row=2, column=2, value=employee_name)
    ws.cell(row=2, column=3, value="Calendar year")
    ws.cell(row=2, column=4, value=int(year))
    start_row = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    r = start_row + 1
    for row in rows or []:
        ws.cell(row=r, column=1, value=row.get("name"))
        ws.cell(row=r, column=2, value=row.get("days_elv"))
        ws.cell(row=r, column=3, value=row.get("days_dsm"))
        ws.cell(row=r, column=4, value=row.get("type_elv"))
        ws.cell(row=r, column=5, value=row.get("type_dsm"))
        ws.cell(row=r, column=6, value=row.get("dates_elv"))
        ws.cell(row=r, column=7, value=row.get("dates_dsm"))
        r += 1
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=2, value=_sum_nokia_audit_compare_days(rows, "days_elv"))
    ws.cell(row=r, column=3, value=_sum_nokia_audit_compare_days(rows, "days_dsm"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


def _nokia_tsv_last_nonempty_cell(line: str) -> str | None:
    """Last non-empty tab-separated cell (Nokia / eLeave TSV exports often put the employee there)."""
    if "\t" not in (line or ""):
        return None
    cells = [c.strip() for c in line.split("\t")]
    for c in reversed(cells):
        if c:
            return " ".join(c.split())
    return None


def _nokia_roster_cell_exact_match(cell: str, roster: Sequence[str]) -> str | None:
    """If ``cell`` normalizes to a roster display name, return that roster string; else None."""
    if not (cell or "").strip():
        return None
    cf = " ".join(cell.split()).casefold()
    for n in roster:
        nn = " ".join((n or "").split()).strip()
        if nn and nn.casefold() == cf:
            return n
    return None


def _nokia_paste_has_nokia_approved_leave_date_line(text: str) -> bool:
    """True if pasted text has at least one line that looks like Nokia approved leave (status + date)."""
    date_pat = re.compile(r"\b(\d{2})[-/](\d{2})[-/](\d{4})\b")
    for line in (text or "").splitlines():
        s = line.strip()
        if not s or "approved" not in s.lower():
            continue
        if date_pat.search(s):
            return True
    return False


def _nokia_paste_trailing_roster_name_hints(text: str, roster: Sequence[str]) -> set[str]:
    """
    From lines that look like Nokia approved leave (contains 'approved' and a DD-MM-YYYY date),
    detect roster names: prefer the **last TSV column** when the line uses tabs (eLeave copy/paste),
    else fall back to the longest roster name matching a **line suffix** (free-text Nokia rows).
    """
    roster_sorted = sorted((n.strip() for n in roster if (n or "").strip()), key=len, reverse=True)
    date_pat = re.compile(r"\b(\d{2})[-/](\d{2})[-/](\d{4})\b")
    hints: set[str] = set()
    for line in (text or "").splitlines():
        s = line.strip()
        if not s or "approved" not in s.lower():
            continue
        if not date_pat.search(s):
            continue
        last_cell = _nokia_tsv_last_nonempty_cell(s)
        if last_cell is not None:
            hit = _nokia_roster_cell_exact_match(last_cell, roster)
            if hit is not None:
                hints.add(hit)
                continue
        tail = s.rstrip(" \t.,;:")
        low_tail = tail.casefold()
        for name in roster_sorted:
            n = name.strip()
            if low_tail.endswith(n.casefold()):
                hints.add(n)
                break
    return hints


def _nokia_paste_employee_must_match_selected_or_error(
    selected: str,
    text: str,
    roster: Sequence[str],
    *,
    require_roster_name_in_paste: bool = False,
) -> str | None:
    """
    If pasted Nokia lines clearly name one or more roster employees (TSV last column or suffix
    heuristic), ensure they match the dropdown selection. Returns an error message to show, or None
    when OK / not verifiable.

    When ``require_roster_name_in_paste`` is True (Show Approved, Show DSM Leaves, and Mark approved actions),
    any pasted line that looks like Nokia approved leave with a date must also yield a detectable roster name;
    otherwise the user must fix the copy so the employee can be verified against the dropdown.
    """
    sel = " ".join((selected or "").split()).strip()
    if not sel:
        return None
    hints = _nokia_paste_trailing_roster_name_hints(text, roster)
    if require_roster_name_in_paste and _nokia_paste_has_nokia_approved_leave_date_line(text) and not hints:
        return (
            "Could not find a roster employee name in the pasted Nokia approved rows. "
            "Each line should end with the same name as Employee (roster name), or use "
            "tab-separated eLeave copy with that employee in the last column, then try again."
        )
    if not hints:
        return None
    distinct = {" ".join(h.split()) for h in hints}
    sel_cf = sel.casefold()
    if len(distinct) > 1:
        names_s = ", ".join(sorted(distinct))
        return (
            f"Pasted text contains more than one employee name ({names_s}). "
            "Select one person in Employee (roster name), paste only their Nokia rows "
            "(each line should end with that employee name), then try again."
        )
    only = next(iter(distinct))
    if only.casefold() != sel_cf:
        return (
            f"Employee name in the pasted Nokia text ({only}) does not match "
            f"Employee (roster name) ({sel}). Copy rows for {sel} from Nokia "
            "(lines should end with that name), then try again."
        )
    return None


def _try_winget_install_tesseract_windows(max_wait_s: int = 600) -> bool:
    """Best-effort silent install of Tesseract via winget (Windows only)."""
    try:
        kwargs: dict = dict(
            capture_output=True,
            text=True,
            timeout=max_wait_s,
        )
        if os.name == "nt" and hasattr(subprocess, "CREATE_NO_WINDOW"):
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        r = subprocess.run(
            [
                "winget",
                "install",
                "-e",
                "--id",
                "UB-Mannheim.TesseractOCR",
                "--accept-package-agreements",
                "--accept-source-agreements",
                "--disable-interactivity",
            ],
            **kwargs,
        )
    except (OSError, subprocess.TimeoutExpired) as e:
        _log.warning("winget install Tesseract failed: %s", e)
        return False
    if r.returncode != 0:
        _log.warning(
            "winget install Tesseract exit %s: %s",
            r.returncode,
            ((r.stderr or "") + (r.stdout or ""))[:800],
        )
    return r.returncode == 0


def _resolve_tesseract_cmd() -> str | None:
    """Path to the tesseract executable: TESSERACT_CMD, PATH, then common Windows install locations."""
    env = (os.environ.get("TESSERACT_CMD") or "").strip()
    if env:
        return env
    found = shutil.which("tesseract")
    if found:
        return found

    def _windows_candidates() -> list[str]:
        cands: list[str] = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ]
        local = (os.environ.get("LOCALAPPDATA") or "").strip()
        if local:
            cands.append(str(Path(local) / "Programs" / "Tesseract-OCR" / "tesseract.exe"))
        pf = (os.environ.get("ProgramFiles") or "").strip()
        if pf:
            cands.append(str(Path(pf) / "Tesseract-OCR" / "tesseract.exe"))
        pfx86 = (os.environ.get("ProgramFiles(x86)") or "").strip()
        if pfx86:
            cands.append(str(Path(pfx86) / "Tesseract-OCR" / "tesseract.exe"))
        return cands

    if os.name == "nt":
        for p in _windows_candidates():
            if Path(p).is_file():
                return str(Path(p))
        auto = (os.environ.get("TEAM_TRACKER_AUTO_TESSERACT") or "1").lower() in ("1", "true", "yes")
        if auto:
            global _tesseract_win_install_tried
            with _tesseract_win_install_lock:
                if not _tesseract_win_install_tried:
                    _tesseract_win_install_tried = True
                    _try_winget_install_tesseract_windows()
            for p in _windows_candidates():
                if Path(p).is_file():
                    return str(Path(p))
    return None


def _ocr_image_to_text(raw: bytes) -> tuple[str | None, str | None]:
    """Run Tesseract on image bytes. Returns (text, None) or (None, user-facing error)."""
    try:
        from PIL import Image, ImageOps
        import pytesseract
    except ImportError:
        return None, (
            "Nokia compare needs Pillow and pytesseract on the server (pip install pillow pytesseract)."
        )
    tess = _resolve_tesseract_cmd()
    if not tess:
        return None, (
            "Tesseract OCR was not found and could not be installed automatically. On Windows with winget, "
            "run: winget install -e --id UB-Mannheim.TesseractOCR --accept-package-agreements --accept-source-agreements "
            "(or install from https://github.com/tesseract-ocr/tesseract), then set TESSERACT_CMD to tesseract.exe if needed. "
            "If automatic winget is blocked, set TEAM_TRACKER_AUTO_TESSERACT=0 and install Tesseract manually."
        )
    pytesseract.pytesseract.tesseract_cmd = tess
    try:
        img = Image.open(io.BytesIO(raw)).convert("RGB")
    except OSError as e:
        return None, f"Could not read image: {e}"
    w, h = img.size
    if w and w < 1000:
        scale = min(3, max(2, 1000 // w))
        img = img.resize((w * scale, h * scale), Image.Resampling.LANCZOS)
    gray = ImageOps.autocontrast(img.convert("L"))
    configs = [
        "--oem 3 --psm 6",
        "--oem 3 --psm 11",
        "--oem 3 --psm 4",
        "--oem 3 --psm 12",
        "--oem 3 --psm 3",
    ]
    best_txt = ""
    best_sc = -1
    last_err: str | None = None
    for cfg in configs:
        try:
            cand = (pytesseract.image_to_string(gray, config=cfg) or "").strip()
        except Exception as e:
            msg = str(e).lower()
            if "tesseract" in msg or "not installed" in msg or "the path" in msg:
                return None, (
                    "Tesseract failed to run. Check that it is installed correctly, or set TESSERACT_CMD to the full path "
                    r"of tesseract.exe (e.g. C:\Program Files\Tesseract-OCR\tesseract.exe). "
                    "Details: https://github.com/tesseract-ocr/tesseract"
                )
            last_err = str(e)
            continue
        if not cand:
            continue
        sc = _ocr_header_score(cand)
        if sc > best_sc:
            best_sc = sc
            best_txt = cand
    if best_txt:
        return best_txt, None
    if last_err:
        return None, last_err
    return None, (
        "OCR returned no readable text. Try a larger, sharper screenshot, or paste / upload a CSV export of the grid. "
        "Also verify Tesseract language data is installed (eng)."
    )


def app_leave_days_by_employee_month(
    app: Flask, year: int, month: int, roster: Sequence[str] | None = None
) -> dict[str, set[int]]:
    """Calendar days in month (1…last) where the leave tracker would show leave (pending/approved, meet-day rules)."""
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    first = date(year, month, 1)
    _, last_day = monthrange(year, month)
    last = date(year, month, last_day)
    start_iso = first.isoformat()
    end_iso = last.isoformat()
    day_dec = _load_meet_leave_day_map(app, start_iso, end_iso)

    conn = get_db(app)
    leaves = list(
        conn.execute(
            """
            SELECT * FROM leave_requests
            WHERE start_date <= ? AND end_date >= ?
              AND status IN ('pending', 'approved')
            ORDER BY id ASC
            """,
            (end_iso, start_iso),
        )
    )
    conn.close()

    by_emp: dict[str, list[sqlite3.Row]] = {e: [] for e in roster_t}
    for row in leaves:
        if row["employee_name"] in by_emp:
            by_emp[row["employee_name"]].append(row)

    out: dict[str, set[int]] = {e: set() for e in roster_t}
    for emp in roster_t:
        for dom in range(1, last_day + 1):
            d = date(year, month, dom)
            d_iso = d.isoformat()
            overlapping = _overlapping_leaves_for_day(by_emp[emp], d, day_dec)
            if not overlapping:
                continue
            pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
            pool = pending_only or overlapping
            row = max(pool, key=_leave_row_display_tiebreak)
            eff = _effective_leave_status(row, d_iso, day_dec)
            if eff not in ("pending", "approved"):
                continue
            if _leave_reason_counts_toward_day_units(row["reason"]):
                out[emp].add(dom)
    return out


def _leave_tracker_day_rows_for_range(
    app: Flask, roster: Sequence[str], sd: date, ed: date
) -> tuple[list[dict[str, Any]], dict[str, float]]:
    """Per calendar day in [sd, ed], rows and totals using the same rules as the leave tracker / worksheet
    (pending + approved only, meet_leave_day removals and per-day approvals, winning row by id, half days = 0.5).
    """
    roster_t = tuple(roster)
    if not roster_t:
        return [], {}
    start_iso = sd.isoformat()
    end_iso = ed.isoformat()
    day_dec = _load_meet_leave_day_map(app, start_iso, end_iso)
    ph = ",".join("?" * len(roster_t))
    conn = get_db(app)
    leaves = list(
        conn.execute(
            f"""
            SELECT * FROM leave_requests
            WHERE start_date <= ? AND end_date >= ?
              AND status IN ('pending', 'approved')
              AND employee_name IN ({ph})
            ORDER BY id ASC
            """,
            (end_iso, start_iso, *roster_t),
        )
    )
    conn.close()

    by_emp: dict[str, list[sqlite3.Row]] = {e: [] for e in roster_t}
    for row in leaves:
        if row["employee_name"] in by_emp:
            by_emp[row["employee_name"]].append(row)

    reason_l = dict(LEAVE_REASONS)
    dur_l = dict(DURATION_CHOICES)
    detail: list[dict[str, Any]] = []
    totals: dict[str, float] = {e: 0.0 for e in roster_t}

    for emp in roster_t:
        for d in _daterange_inclusive(sd, ed):
            d_iso = d.isoformat()
            overlapping = _overlapping_leaves_for_day(by_emp[emp], d, day_dec)
            if not overlapping:
                continue
            pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
            pool = pending_only or overlapping
            row = max(pool, key=_leave_row_display_tiebreak)
            eff = _effective_leave_status(row, d_iso, day_dec)
            if eff not in ("pending", "approved"):
                continue
            dur = (row["duration_type"] or "full").strip()
            if dur in ("half_am", "half_pm"):
                unit = 0.5
            else:
                unit = 1.0
            if _leave_reason_counts_toward_day_units(row["reason"]):
                totals[emp] += unit
            else:
                unit = 0.0
            detail.append(
                {
                    "employee_name": emp,
                    "leave_date": d_iso,
                    "day_units": unit,
                    "reason": row["reason"],
                    "reason_label": reason_l.get(row["reason"], row["reason"]),
                    "duration_type": dur,
                    "duration_label": dur_l.get(dur, dur),
                    "status": eff,
                    "leave_request_id": int(row["id"]),
                    "request_start": row["start_date"],
                    "request_end": row["end_date"],
                }
            )

    detail.sort(key=lambda r: (r["employee_name"], r["leave_date"], r["leave_request_id"]))
    return detail, totals


def _leave_report_overlap_window(
    leave_start_iso: str, leave_end_iso: str, win_start: date, win_end: date
) -> tuple[date, date] | None:
    """Intersection of [leave_start, leave_end] with [win_start, win_end], inclusive, or None if empty."""
    try:
        ls = date.fromisoformat(str(leave_start_iso)[:10])
        le = date.fromisoformat(str(leave_end_iso)[:10])
    except ValueError:
        return None
    if le < ls:
        ls, le = le, ls
    lo = max(ls, win_start)
    hi = min(le, win_end)
    if lo > hi:
        return None
    return lo, hi


def compare_nokia_vs_app(
    nokia_days: dict[str, set[int]],
    app_days: dict[str, set[int]],
    year: int,
    month: int,
    roster: Sequence[str] | None = None,
) -> tuple[list[dict], list[dict]]:
    """Returns (in_app_not_nokia, in_nokia_not_app) as list of {employee, day, date}."""
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    _, last_day = monthrange(year, month)
    in_app_not_nokia: list[dict] = []
    in_nokia_not_app: list[dict] = []
    for emp in roster_t:
        for dom in range(1, last_day + 1):
            in_app = dom in app_days.get(emp, set())
            in_nokia = dom in nokia_days.get(emp, set())
            d_iso = date(year, month, dom).isoformat()
            if in_app and not in_nokia:
                in_app_not_nokia.append({"employee": emp, "day": dom, "date": d_iso})
            if in_nokia and not in_app:
                in_nokia_not_app.append({"employee": emp, "day": dom, "date": d_iso})
    return in_app_not_nokia, in_nokia_not_app


def client_ip() -> str:
    if not has_request_context():
        return ""
    xff = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    if xff:
        return xff[:200]
    return (request.remote_addr or "")[:200]


def _portal_otp_smtp_ready(app: Flask) -> bool:
    host = (app.config.get("PORTAL_OTP_SMTP_HOST") or "").strip()
    from_addr = (app.config.get("PORTAL_OTP_FROM") or "").strip()
    return bool(host and from_addr)


def _portal_otp_mail_configured(app: Flask) -> bool:
    """True when email OTP is available (SMTP + From, or local dev console delivery)."""
    if _portal_otp_smtp_ready(app):
        return True
    return bool(app.config.get("PORTAL_OTP_DEV_CONSOLE"))


def _portal_otp_code_hmac(app: Flask, code: str) -> str:
    secret = (app.config.get("SECRET_KEY") or "dev").encode("utf-8")
    return hmac.new(secret, code.strip().encode("utf-8"), hashlib.sha256).hexdigest()


def _portal_employee_access_code_hmac(app: Flask, code: str) -> str:
    secret = (app.config.get("SECRET_KEY") or "dev").encode("utf-8")
    payload = f"portal-employee-access:{code.strip()}".encode("utf-8")
    return hmac.new(secret, payload, hashlib.sha256).hexdigest()


def _normalize_portal_access_code(raw: str) -> str | None:
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) != 6:
        return None
    return digits


def _portal_directory_hit_for_roster(roster_name: str) -> dict[str, str] | None:
    target = (roster_name or "").strip()
    if not target:
        return None
    target_n = _norm(target)
    for em, display, roster in NOKIA_PORTAL_DIRECTORY:
        if roster == target or display == target or _norm(roster) == target_n or _norm(display) == target_n:
            return {"email": em, "display_name": display, "roster_name": roster}
    return None


def _portal_access_codes_configured(app: Flask) -> bool:
    conn = get_db(app)
    row = conn.execute("SELECT 1 FROM portal_employee_access_code LIMIT 1").fetchone()
    conn.close()
    return row is not None


def _portal_employee_access_code_status_for_roster(
    app: Flask, roster: Sequence[str]
) -> list[dict[str, Any]]:
    names = [str(n).strip() for n in roster if str(n or "").strip()]
    if not names:
        return []
    ph = ",".join("?" * len(names))
    conn = get_db(app)
    rows = conn.execute(
        f"""
        SELECT employee_name, employee_email, updated_at, updated_by_manager_email
        FROM portal_employee_access_code
        WHERE employee_name IN ({ph})
        """,
        names,
    ).fetchall()
    email_rows = conn.execute(
        f"""
        SELECT employee_name, employee_email
        FROM team_roster
        WHERE employee_name IN ({ph})
        """,
        names,
    ).fetchall()
    conn.close()
    by_name = {str(r["employee_name"]): r for r in rows}
    roster_email_by_name: dict[str, str] = {}
    for r in email_rows:
        nm = str(r["employee_name"] or "").strip()
        em = normalize_email(str(r["employee_email"] or ""))
        if nm and em and nm not in roster_email_by_name:
            roster_email_by_name[nm] = em
    out: list[dict[str, Any]] = []
    for name in names:
        row = by_name.get(name)
        roster_email = roster_email_by_name.get(name, "")
        code_email = normalize_email(str(row["employee_email"] or "")) if row else ""
        out.append(
            {
                "employee_name": name,
                "employee_email": roster_email or code_email,
                "is_set": row is not None,
                "updated_at": str(row["updated_at"] or "") if row else "",
                "updated_by": str(row["updated_by_manager_email"] or "") if row else "",
            }
        )
    return out


def _roster_email_for_employee(app: Flask, employee_name: str) -> str:
    canonical = (employee_name or "").strip()
    if not canonical:
        return ""
    conn = get_db(app)
    row = conn.execute(
        """
        SELECT employee_email FROM team_roster
        WHERE employee_name = ? COLLATE NOCASE AND trim(employee_email) != ''
        ORDER BY sort_order
        LIMIT 1
        """,
        (canonical,),
    ).fetchone()
    conn.close()
    return normalize_email(str(row["employee_email"] or "")) if row else ""


def _lookup_team_roster_by_email(app: Flask, email: str) -> dict[str, str] | None:
    em = normalize_email(email)
    if not em:
        return None
    conn = get_db(app)
    row = conn.execute(
        """
        SELECT tr.employee_name, tr.employee_email, t.name AS team_name
        FROM team_roster tr
        JOIN teams t ON t.id = tr.team_id
        WHERE lower(trim(tr.employee_email)) = ?
        ORDER BY tr.sort_order, tr.employee_name COLLATE NOCASE
        LIMIT 1
        """,
        (em,),
    ).fetchone()
    conn.close()
    if not row:
        return None
    name = str(row["employee_name"] or "").strip()
    if not name:
        return None
    return {
        "employee_name": name,
        "employee_email": normalize_email(str(row["employee_email"] or "")) or em,
        "team_name": str(row["team_name"] or "").strip(),
    }


def _canonical_roster_name_on_manager_roster(roster: Sequence[str], employee_name: str) -> str | None:
    target = (employee_name or "").strip()
    if not target:
        return None
    target_n = _norm(target)
    for name in roster:
        nm = (name or "").strip()
        if nm == target or _norm(nm) == target_n:
            return nm
    resolved = resolve_employee_name(target, roster=roster)
    return resolved


def _set_portal_employee_access_code(
    app: Flask, employee_name: str, code: str, manager_email: str = ""
) -> str | None:
    """Store a 6-digit access code for one roster employee. Returns error message or None."""
    normalized = _normalize_portal_access_code(code)
    if not normalized:
        return "Enter a 6-digit numeric code."
    canonical = (employee_name or "").strip()
    if not canonical:
        return "Employee name is required."
    ts = _utc_stamp()
    digest = _portal_employee_access_code_hmac(app, normalized)
    roster_email = _roster_email_for_employee(app, canonical)
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO portal_employee_access_code
            (employee_name, employee_email, code_hmac, updated_at, updated_by_manager_email)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(employee_name) DO UPDATE SET
            employee_email = excluded.employee_email,
            code_hmac = excluded.code_hmac,
            updated_at = excluded.updated_at,
            updated_by_manager_email = excluded.updated_by_manager_email
        """,
        (canonical, roster_email, digest, ts, (manager_email or "").strip()),
    )
    conn.commit()
    conn.close()
    return None


def _clear_portal_employee_access_code(app: Flask, employee_name: str) -> None:
    canonical = (employee_name or "").strip()
    if not canonical:
        return
    conn = get_db(app)
    conn.execute("DELETE FROM portal_employee_access_code WHERE employee_name = ?", (canonical,))
    conn.commit()
    conn.close()


def _verify_portal_employee_access_code(
    app: Flask, employee_name: str, code: str, *, login_email: str = ""
) -> bool:
    canonical = (employee_name or "").strip()
    normalized = _normalize_portal_access_code(code)
    if not canonical or not normalized:
        return False
    digest = _portal_employee_access_code_hmac(app, normalized)
    conn = get_db(app)
    row = conn.execute(
        """
        SELECT code_hmac, employee_email FROM portal_employee_access_code
        WHERE employee_name = ?
        """,
        (canonical,),
    ).fetchone()
    conn.close()
    if not row or str(row["code_hmac"] or "") != digest:
        return False
    login_em = normalize_email(login_email)
    if login_em:
        roster_em = _roster_email_for_employee(app, canonical)
        code_em = normalize_email(str(row["employee_email"] or ""))
        expected = roster_em or code_em
        if expected and login_em != expected:
            return False
    return True


def _resolve_portal_access_code_identity(app: Flask, identifier: str) -> dict | None:
    """Map email or roster name to sign-in identity when a manager access code is used."""
    raw = (identifier or "").strip()
    if not raw:
        return None
    union = get_union_roster_for_app(app)
    if "@" in raw:
        em = normalize_email(raw)
        roster_hit = _lookup_team_roster_by_email(app, em)
        if roster_hit:
            name = roster_hit["employee_name"]
            if name in union or _match_roster_employee_name(name, union):
                return {
                    "email": roster_hit["employee_email"] or em,
                    "display_name": name,
                    "roster_name": name,
                }
        hit = lookup_portal_directory(em)
        if hit:
            roster_name = _resolve_portal_employee_roster_name(
                app, email=em, display_name=hit["display_name"], roster_name=hit["roster_name"]
            )
            if roster_name and roster_name in union:
                return {
                    "email": em,
                    "display_name": hit["display_name"],
                    "roster_name": roster_name,
                }
        return None
    roster_name = resolve_employee_name(raw, roster=union)
    if not roster_name:
        roster_name = _match_roster_employee_name(raw, union)
    if not roster_name:
        return None
    hit = _portal_directory_hit_for_roster(roster_name)
    if hit:
        return hit
    return {
        "email": f"{roster_name.replace(' ', '.').lower()}@nokia.com",
        "display_name": roster_name,
        "roster_name": roster_name,
    }


def send_portal_otp_smtp(app: Flask, to_addr: str, code: str) -> None:
    """Send a one-time sign-in code (raises on SMTP failure)."""
    host = (app.config.get("PORTAL_OTP_SMTP_HOST") or "").strip()
    if not host:
        raise RuntimeError("PORTAL_OTP_SMTP_HOST is not set")
    port = int(app.config.get("PORTAL_OTP_SMTP_PORT") or 587)
    user = (app.config.get("PORTAL_OTP_SMTP_USER") or "").strip()
    password = (app.config.get("PORTAL_OTP_SMTP_PASSWORD") or "").strip()
    from_addr = (app.config.get("PORTAL_OTP_FROM") or "").strip()
    from_name = (app.config.get("PORTAL_OTP_FROM_NAME") or "TEAM MANAGEMENT PORTAL").strip()
    use_tls = bool(app.config.get("PORTAL_OTP_USE_TLS", True))
    minutes = int(app.config.get("PORTAL_OTP_TTL_MINUTES") or 15)

    msg = EmailMessage()
    msg["Subject"] = "Your TEAM MANAGEMENT PORTAL sign-in code"
    msg["From"] = f"{from_name} <{from_addr}>"
    msg["To"] = to_addr
    msg.set_content(
        f"Your one-time sign-in code is: {code}\n\n"
        f"It expires in {minutes} minutes. If you did not request this, you can ignore this email.\n"
    )

    if port == 465:
        with smtplib.SMTP_SSL(host, port, timeout=30) as smtp:
            if user:
                smtp.login(user, password)
            smtp.send_message(msg)
        return
    with smtplib.SMTP(host, port, timeout=30) as smtp:
        if use_tls:
            smtp.starttls()
        if user:
            smtp.login(user, password)
        smtp.send_message(msg)


def leave_cell_code(reason: str, duration_type: str, status: str, description: str = "") -> str:
    """Leave tracker cell: PL/UL/SL/LL/CO with optional ½ for half-day; Nokia-audit rows show A."""
    if NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in (description or ""):
        half = duration_type in ("half_am", "half_pm")
        suf = "\u00bd" if half else ""
        return "A" + suf
    half = duration_type in ("half_am", "half_pm")
    suf = "\u00bd" if half else ""
    letter = {"pl": "PL", "ul": "UL", "sl": "SL", "ll": "LL", "compoff": "CO"}.get(reason, "UL")
    return letter + suf


def cell_css_class(reason: str, status: str, description: str = "") -> str:
    if NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in (description or ""):
        base = "cell-nokia-approved"
        return f"{base} cell-pending" if status == "pending" else base
    base = {
        "pl": "cell-pl",
        "ul": "cell-ul",
        "sl": "cell-sl",
        "ll": "cell-ll",
        "compoff": "cell-compoff",
    }.get(reason, "cell-ul")
    return f"{base} cell-pending" if status == "pending" else base


def _seed_default_team_if_empty(conn: sqlite3.Connection) -> None:
    n = int(conn.execute("SELECT COUNT(*) AS c FROM teams").fetchone()["c"])
    if n > 0:
        return
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    cur = conn.execute("INSERT INTO teams (name, created_at) VALUES (?, ?)", ("Default", ts))
    tid = int(cur.lastrowid)
    for i, name in enumerate(EMPLOYEES):
        conn.execute(
            "INSERT INTO team_roster (team_id, employee_name, employee_email, sort_order) VALUES (?, ?, '', ?)",
            (tid, name, i),
        )


def _migrate_scrum_team_task_kinds(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_team_task_kind_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_team_task_kind (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER NOT NULL REFERENCES teams(id) ON DELETE CASCADE,
            code TEXT NOT NULL,
            label TEXT NOT NULL,
            color_hex TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            UNIQUE(team_id, code)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_task_kind_team ON scrum_team_task_kind(team_id, sort_order)"
    )
    for t in conn.execute("SELECT id FROM teams"):
        tid = int(t["id"])
        for code, label, color, so in SCRUM_BUILTIN_TASK_KIND_ROWS:
            conn.execute(
                """
                INSERT OR IGNORE INTO scrum_team_task_kind (team_id, code, label, color_hex, sort_order)
                VALUES (?, ?, ?, ?, ?)
                """,
                (tid, code, label, color, so),
            )
    conn.execute("UPDATE scrum_sprint_item SET task_kind = 'code' WHERE lower(trim(task_kind)) = 'story'")
    conn.execute("UPDATE scrum_sprint_item SET task_kind = 'ndy' WHERE lower(trim(task_kind)) = 'task'")
    conn.execute("UPDATE scrum_sprint_item SET task_kind = 'fsy' WHERE lower(trim(task_kind)) = 'bug'")
    for r in conn.execute(
        """
        SELECT i.id AS iid, i.task_kind AS tc, s.team_id AS team_id
        FROM scrum_sprint_item i
        JOIN scrum_sprint s ON s.id = i.sprint_id
        """
    ):
        code = (r["tc"] or "").strip().lower()
        tid = int(r["team_id"])
        hit = conn.execute(
            "SELECT 1 FROM scrum_team_task_kind WHERE team_id = ? AND code = ?",
            (tid, code),
        ).fetchone()
        if not hit:
            conn.execute("UPDATE scrum_sprint_item SET task_kind = 'ndy' WHERE id = ?", (int(r["iid"]),))
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_team_task_kind_v1')")


def _migrate_scrum_task_kind_short_labels(conn: sqlite3.Connection) -> None:
    """Rename built-in kind labels from 'NDY — purple' style to short 'NDY' (one-time)."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_task_kind_short_labels_v1",)).fetchone():
        return
    if conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='scrum_team_task_kind'"
    ).fetchone():
        pairs: tuple[tuple[str, str, str], ...] = (
            ("ndy", "NDY", "NDY — purple"),
            ("ndy", "NDY", "NDY-purple"),
            ("ndy", "NDY", "NDY - purple"),
            ("code", "CoDe", "CoDe — yellow"),
            ("code", "CoDe", "CoDe-yellow"),
            ("code", "CoDe", "CoDe - yellow"),
            ("fsy", "FSY", "FSY — orange"),
            ("fsy", "FSY", "FSY-orange"),
            ("fsy", "FSY", "FSY - orange"),
        )
        for code, new_label, old_label in pairs:
            conn.execute(
                "UPDATE scrum_team_task_kind SET label = ? WHERE code = ? AND label = ?",
                (new_label, code, old_label),
            )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_task_kind_short_labels_v1')")


def _migrate_scrum_task_kind_fixed_three_v1(conn: sqlite3.Connection) -> None:
    """Keep only NDY, CODE, and FSY task types; remap stray item kinds to NDY."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_task_kind_fixed_three_v1",)).fetchone():
        return
    if not conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='scrum_team_task_kind'"
    ).fetchone():
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_task_kind_fixed_three_v1')")
        return
    conn.execute(
        """
        UPDATE scrum_sprint_item SET task_kind = 'ndy'
        WHERE task_kind IS NULL OR trim(task_kind) = ''
           OR lower(trim(task_kind)) NOT IN ('ndy', 'fsy', 'code')
        """
    )
    conn.execute(
        "UPDATE scrum_sprint_item SET sticky_color_hex = NULL WHERE sticky_color_hex IS NOT NULL AND trim(sticky_color_hex) != ''"
    )
    conn.execute("DELETE FROM scrum_team_task_kind WHERE code NOT IN ('ndy', 'fsy', 'code')")
    for t in conn.execute("SELECT id FROM teams"):
        tid = int(t["id"])
        for code, label, color, so in SCRUM_BUILTIN_TASK_KIND_ROWS:
            conn.execute(
                """
                INSERT OR IGNORE INTO scrum_team_task_kind (team_id, code, label, color_hex, sort_order)
                VALUES (?, ?, ?, ?, ?)
                """,
                (tid, code, label, color, so),
            )
            conn.execute(
                """
                UPDATE scrum_team_task_kind SET label = ?, color_hex = ?, sort_order = ?
                WHERE team_id = ? AND code = ?
                """,
                (label, color, so, tid, code),
            )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_task_kind_fixed_three_v1')")


def _migrate_scrum_task_kind_builtin_five_v1(conn: sqlite3.Connection) -> None:
    """Add Improvement and Process&Tools built-in kinds; keep only SCRUM_TASK_KIND_CODES in DB."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_task_kind_builtin_five_v1",)).fetchone():
        return
    if not conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='scrum_team_task_kind'"
    ).fetchone():
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_task_kind_builtin_five_v1')")
        return
    allowed = tuple(sorted(SCRUM_TASK_KIND_CODES))
    ph = ",".join("?" * len(allowed))
    conn.execute(
        f"""
        UPDATE scrum_sprint_item SET task_kind = 'ndy'
        WHERE task_kind IS NULL OR trim(task_kind) = ''
           OR lower(trim(task_kind)) NOT IN ({ph})
        """,
        allowed,
    )
    conn.execute(f"DELETE FROM scrum_team_task_kind WHERE code NOT IN ({ph})", allowed)
    for t in conn.execute("SELECT id FROM teams"):
        tid = int(t["id"])
        for code, label, color, so in SCRUM_BUILTIN_TASK_KIND_ROWS:
            conn.execute(
                """
                INSERT OR IGNORE INTO scrum_team_task_kind (team_id, code, label, color_hex, sort_order)
                VALUES (?, ?, ?, ?, ?)
                """,
                (tid, code, label, color, so),
            )
            conn.execute(
                """
                UPDATE scrum_team_task_kind SET label = ?, color_hex = ?, sort_order = ?
                WHERE team_id = ? AND code = ?
                """,
                (label, color, so, tid, code),
            )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_task_kind_builtin_five_v1')")


def _migrate_scrum_sprint_team_capacity_hours_v1(conn: sqlite3.Connection) -> None:
    """Store leave-adjusted team capacity (sum of roster Mon–Fri hours minus leave) on each sprint."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_sprint_team_capacity_v1",)).fetchone():
        return
    if not conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='scrum_sprint'"
    ).fetchone():
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_sprint_team_capacity_v1')")
        return
    cols = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint)")}
    if "team_capacity_hours" not in cols:
        conn.execute("ALTER TABLE scrum_sprint ADD COLUMN team_capacity_hours REAL")
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_sprint_team_capacity_v1')")


def _migrate_scrum_sprint_is_closed_v1(conn: sqlite3.Connection) -> None:
    """Sprint lock: when is_closed=1, board and sprint metadata are read-only until a manager re-opens."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_sprint_is_closed_v1",)).fetchone():
        return
    if not conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='scrum_sprint'"
    ).fetchone():
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_sprint_is_closed_v1')")
        return
    cols = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint)")}
    if "is_closed" not in cols:
        conn.execute("ALTER TABLE scrum_sprint ADD COLUMN is_closed INTEGER NOT NULL DEFAULT 0")
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_sprint_is_closed_v1')")


def _migrate_scrum_item_appreciation_v1(conn: sqlite3.Connection) -> None:
    """Manager 'Well Done' appreciation comments per sticky (scrum_sprint_item)."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_item_appreciation_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_item_appreciation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            author TEXT NOT NULL DEFAULT '',
            comment TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_appreciation_item ON scrum_item_appreciation(item_id, created_at DESC)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_appreciation_v1')")


def _migrate_portal_email_otp_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("portal_email_otp_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS portal_email_otp (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            code_hmac TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            client_ip TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_portal_email_otp_email_created ON portal_email_otp(email, created_at DESC)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('portal_email_otp_v1')")


def _migrate_portal_employee_access_code_v1(conn: sqlite3.Connection) -> None:
    if conn.execute(
        "SELECT 1 FROM app_migrations WHERE id = ?", ("portal_employee_access_code_v1",)
    ).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS portal_employee_access_code (
            employee_name TEXT PRIMARY KEY COLLATE NOCASE,
            code_hmac TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            updated_by_manager_email TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('portal_employee_access_code_v1')")


def _migrate_roster_email_v1(conn: sqlite3.Connection) -> None:
    """Add employee_email to team_roster."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("roster_email_v1",)).fetchone():
        return
    existing = {row[1] for row in conn.execute("PRAGMA table_info(team_roster)").fetchall()}
    if "employee_email" not in existing:
        conn.execute("ALTER TABLE team_roster ADD COLUMN employee_email TEXT NOT NULL DEFAULT ''")
    conn.execute("INSERT INTO app_migrations (id) VALUES ('roster_email_v1')")


def _migrate_portal_employee_access_code_email_v1(conn: sqlite3.Connection) -> None:
    """Add employee_email to portal access-code rows."""
    if conn.execute(
        "SELECT 1 FROM app_migrations WHERE id = ?", ("portal_employee_access_code_email_v1",)
    ).fetchone():
        return
    existing = {row[1] for row in conn.execute("PRAGMA table_info(portal_employee_access_code)").fetchall()}
    if "employee_email" not in existing:
        conn.execute(
            "ALTER TABLE portal_employee_access_code ADD COLUMN employee_email TEXT NOT NULL DEFAULT ''"
        )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('portal_employee_access_code_email_v1')")


def _migrate_scrum_portal_proposal_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_portal_proposal_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_portal_proposal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER NOT NULL,
            sprint_id INTEGER NOT NULL,
            item_id INTEGER REFERENCES scrum_sprint_item(id) ON DELETE SET NULL,
            proposer_roster_name TEXT NOT NULL DEFAULT '',
            action TEXT NOT NULL,
            payload_json TEXT NOT NULL DEFAULT '{}',
            status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL,
            resolved_at TEXT,
            resolution_note TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_portal_proposal_team_status ON scrum_portal_proposal(team_id, status, created_at DESC)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_portal_proposal_sprint ON scrum_portal_proposal(sprint_id, status)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_portal_proposal_v1')")


def _migrate_scrum_item_activity_committed_hours_nullable_v1(conn: sqlite3.Connection) -> None:
    """Allow NULL committed_hours so a Do → In progress move can omit burnt hours (unset vs zero)."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_item_activity_hours_nullable_v1",)).fetchone():
        return
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'scrum_item_activity'"
    ).fetchone()
    if not row:
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_activity_hours_nullable_v1')")
        return
    notnull = 1
    for col in conn.execute("PRAGMA table_info(scrum_item_activity)"):
        if str(col[1]) == "committed_hours":
            notnull = int(col[3])
            break
    if notnull == 0:
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_activity_hours_nullable_v1')")
        return
    conn.executescript(
        """
        CREATE TABLE scrum_item_activity__m (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            body TEXT NOT NULL DEFAULT '',
            committed_hours REAL,
            from_column TEXT NOT NULL DEFAULT '',
            to_column TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );
        INSERT INTO scrum_item_activity__m (id, item_id, body, committed_hours, from_column, to_column, created_at)
            SELECT id, item_id, body, committed_hours, from_column, to_column, created_at FROM scrum_item_activity;
        DROP TABLE scrum_item_activity;
        ALTER TABLE scrum_item_activity__m RENAME TO scrum_item_activity;
        CREATE INDEX IF NOT EXISTS idx_scrum_activity_item ON scrum_item_activity(item_id, created_at DESC);
        """
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_activity_hours_nullable_v1')")


def _migrate_managers_table_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("managers_table_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS managers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE COLLATE NOCASE,
            password_hash TEXT NOT NULL,
            team_name TEXT NOT NULL DEFAULT 'Default',
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_managers_email ON managers(email)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('managers_table_v1')")


def _migrate_teams_owner_email_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("teams_owner_email_v1",)).fetchone():
        return
    cols = {row[1] for row in conn.execute("PRAGMA table_info(teams)")}
    if "owner_email" not in cols:
        conn.execute("ALTER TABLE teams ADD COLUMN owner_email TEXT NOT NULL DEFAULT ''")
        conn.execute(
            """
            UPDATE teams
            SET owner_email = (
                SELECT email FROM managers WHERE team_name = teams.name LIMIT 1
            )
            WHERE owner_email = '' AND EXISTS (
                SELECT 1 FROM managers WHERE team_name = teams.name
            )
            """
        )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('teams_owner_email_v1')")


def _migrate_lpo_manager_emails_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("lpo_manager_emails_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS lpo_manager_emails (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE COLLATE NOCASE,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_lpo_manager_emails_email ON lpo_manager_emails(email)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('lpo_manager_emails_v1')")


def _migrate_strip_approved_employee_change_prefix_v1(conn: sqlite3.Connection) -> None:
    """Strip the obsolete '[approved employee change] ' prefix from all activity bodies."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("strip_approved_employee_change_prefix_v1",)).fetchone():
        return
    prefix = "[approved employee change] "
    conn.execute(
        """
        UPDATE scrum_item_activity
        SET body = SUBSTR(body, ?)
        WHERE body LIKE ?
        """,
        (len(prefix) + 1, prefix + "%"),
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('strip_approved_employee_change_prefix_v1')")


def _migrate_lpo_manager_team_access_v1(conn: sqlite3.Connection) -> None:
    """Per-LPO team access: which teams each LPO email is allowed to manage."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("lpo_manager_team_access_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS lpo_manager_team_access (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lpo_email TEXT NOT NULL COLLATE NOCASE,
            team_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE (lpo_email, team_id)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_lpo_team_access_email ON lpo_manager_team_access(lpo_email)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('lpo_manager_team_access_v1')")


def _migrate_scrum_item_linked_files_v1(conn: sqlite3.Connection) -> None:
    """URL-linked reference files per sprint item (no upload needed)."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_item_linked_files_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_item_linked_file (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            sort_order INTEGER NOT NULL DEFAULT 0,
            display_name TEXT NOT NULL DEFAULT '',
            url TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT '',
            updated_by TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_item_linked_file_item ON scrum_item_linked_file(item_id, sort_order)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_linked_files_v1')")


def _migrate_scrum_item_linked_file_auth_v1(conn: sqlite3.Connection) -> None:
    """Add optional stored credentials (auth_user, auth_pass) for proxied SharePoint links."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_item_linked_file_auth_v1",)).fetchone():
        return
    cols = {row[1] for row in conn.execute("PRAGMA table_info(scrum_item_linked_file)")}
    if "auth_user" not in cols:
        conn.execute("ALTER TABLE scrum_item_linked_file ADD COLUMN auth_user TEXT NOT NULL DEFAULT ''")
    if "auth_pass" not in cols:
        conn.execute("ALTER TABLE scrum_item_linked_file ADD COLUMN auth_pass TEXT NOT NULL DEFAULT ''")
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_linked_file_auth_v1')")


def _migrate_scrum_item_checklist_v1(conn: sqlite3.Connection) -> None:
    """Task checklist rows per sprint item (Items To Finish / status / LE / done till date)."""
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_item_checklist_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_item_checklist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            sort_order INTEGER NOT NULL DEFAULT 0,
            items_to_finish TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT '',
            le_to_complete TEXT NOT NULL DEFAULT '',
            done_till_date TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT '',
            updated_by TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_item_checklist_item ON scrum_item_checklist(item_id, sort_order)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_item_checklist_v1')")


def _hash_manager_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 260000)
    return f"pbkdf2:sha256:260000${salt}${digest.hex()}"


def _verify_manager_password(stored_hash: str, candidate: str) -> bool:
    try:
        parts = stored_hash.split("$")
        if len(parts) != 3:
            return False
        salt = parts[1]
        expected_hex = parts[2]
        digest = hashlib.pbkdf2_hmac("sha256", candidate.encode("utf-8"), salt.encode("utf-8"), 260000)
        return secrets.compare_digest(digest.hex(), expected_hex)
    except Exception:
        return False


def _seed_primary_owner_manager_from_master_pin(app: Flask, conn: sqlite3.Connection) -> None:
    """If PRIMARY_OWNER_MANAGER_EMAIL and MANAGER_DASHBOARD_PASSWORD are set, create that managers row once.

    Use this to mirror the legacy master-pin flow: one email + password_hash from MANAGER_DASHBOARD_PASSWORD.
    Leave PRIMARY_OWNER_MANAGER_EMAIL unset to skip automatic seeding (register via /register or insert manually).
    """
    email = (app.config.get("PRIMARY_OWNER_MANAGER_EMAIL") or "").strip()
    if not email or "@" not in email:
        return
    if not conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name='managers' LIMIT 1"
    ).fetchone():
        return
    if conn.execute("SELECT 1 FROM managers WHERE email = ? COLLATE NOCASE", (email,)).fetchone():
        return
    plain = (app.config.get("MANAGER_DASHBOARD_PASSWORD") or "").strip()
    if not plain:
        return
    team_row = conn.execute(
        "SELECT name FROM teams WHERE lower(trim(name)) = 'default' LIMIT 1"
    ).fetchone()
    if not team_row:
        team_row = conn.execute("SELECT name FROM teams ORDER BY id ASC LIMIT 1").fetchone()
    team_name = str(team_row["name"]).strip() if team_row and team_row["name"] else "Default"
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn.execute(
        "INSERT INTO managers (email, password_hash, team_name, created_at) VALUES (?, ?, ?, ?)",
        (email, _hash_manager_password(plain), team_name, ts),
    )


def init_db(app: Flask) -> None:
    conn = get_db(app)
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS leave_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_name TEXT NOT NULL,
            reason TEXT NOT NULL,
            description TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            duration_type TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL,
            submitted_ip TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_name TEXT NOT NULL,
            work_date TEXT NOT NULL,
            status TEXT NOT NULL,
            note TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            UNIQUE(employee_name, work_date)
        );

        CREATE TABLE IF NOT EXISTS meet_attendance (
            employee_name TEXT NOT NULL,
            work_date TEXT NOT NULL,
            pl_verified TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            PRIMARY KEY (employee_name, work_date)
        );

        CREATE TABLE IF NOT EXISTS meet_leave_day (
            leave_id INTEGER NOT NULL,
            work_date TEXT NOT NULL,
            decision TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (leave_id, work_date),
            CHECK (decision IN ('approved', 'removed'))
        );

        CREATE TABLE IF NOT EXISTS teams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL DEFAULT '',
            hub_mode TEXT NOT NULL DEFAULT 'leave',
            owner_email TEXT NOT NULL DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS team_roster (
            team_id INTEGER NOT NULL REFERENCES teams(id) ON DELETE CASCADE,
            employee_name TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (team_id, employee_name)
        );

        CREATE INDEX IF NOT EXISTS idx_team_roster_team ON team_roster(team_id);

        CREATE TABLE IF NOT EXISTS leave_tracker_eleaves (
            team_id INTEGER NOT NULL REFERENCES teams(id) ON DELETE CASCADE,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            employee_name TEXT NOT NULL,
            days REAL NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (team_id, year, month, employee_name)
        );

        CREATE INDEX IF NOT EXISTS idx_leave_tracker_eleaves_tm ON leave_tracker_eleaves(team_id, year, month);

        CREATE TABLE IF NOT EXISTS scrum_sprint (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER NOT NULL REFERENCES teams(id) ON DELETE CASCADE,
            name TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            goal TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_scrum_sprint_team ON scrum_sprint(team_id, start_date);

        CREATE TABLE IF NOT EXISTS scrum_sprint_item (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sprint_id INTEGER NOT NULL REFERENCES scrum_sprint(id) ON DELETE CASCADE,
            assignee TEXT NOT NULL DEFAULT '',
            title TEXT NOT NULL,
            estimate_hours REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'open',
            notes TEXT NOT NULL DEFAULT '',
            dod TEXT NOT NULL DEFAULT '',
            done_artifacts TEXT NOT NULL DEFAULT '[]',
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            kanban_column TEXT NOT NULL DEFAULT 'backlog',
            task_kind TEXT NOT NULL DEFAULT 'task',
            area TEXT NOT NULL DEFAULT ''
        );

        CREATE INDEX IF NOT EXISTS idx_scrum_item_sprint ON scrum_sprint_item(sprint_id);

        CREATE TABLE IF NOT EXISTS scrum_daily_task (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER NOT NULL REFERENCES teams(id) ON DELETE CASCADE,
            work_date TEXT NOT NULL,
            assignee TEXT NOT NULL DEFAULT '',
            title TEXT NOT NULL,
            planned_hours REAL NOT NULL DEFAULT 0,
            committed_hours REAL NOT NULL DEFAULT 0,
            actual_hours REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'open',
            notes TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            sprint_id INTEGER REFERENCES scrum_sprint(id) ON DELETE SET NULL,
            sprint_item_id INTEGER REFERENCES scrum_sprint_item(id) ON DELETE SET NULL
        );

        CREATE INDEX IF NOT EXISTS idx_scrum_daily_team_date ON scrum_daily_task(team_id, work_date);

        CREATE TABLE IF NOT EXISTS scrum_sprint_member_goal (
            sprint_id INTEGER NOT NULL REFERENCES scrum_sprint(id) ON DELETE CASCADE,
            employee_name TEXT NOT NULL,
            goal TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            PRIMARY KEY (sprint_id, employee_name)
        );

        CREATE INDEX IF NOT EXISTS idx_sprint_member_goal_sprint ON scrum_sprint_member_goal(sprint_id);

        CREATE TABLE IF NOT EXISTS scrum_item_activity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            body TEXT NOT NULL DEFAULT '',
            committed_hours REAL,
            from_column TEXT NOT NULL DEFAULT '',
            to_column TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_scrum_activity_item ON scrum_item_activity(item_id, created_at DESC);

        CREATE TABLE IF NOT EXISTS app_migrations (
            id TEXT PRIMARY KEY
        );
        """
    )
    cols_teams = {row[1] for row in conn.execute("PRAGMA table_info(teams)")}
    if "hub_mode" not in cols_teams:
        conn.execute("ALTER TABLE teams ADD COLUMN hub_mode TEXT NOT NULL DEFAULT 'leave'")
    cols = {row[1] for row in conn.execute("PRAGMA table_info(leave_requests)")}
    if "submitted_ip" not in cols:
        conn.execute("ALTER TABLE leave_requests ADD COLUMN submitted_ip TEXT NOT NULL DEFAULT ''")
    for old, new in (
        ("vacation", "pl"),
        ("personal", "pl"),
        ("family", "ul"),
        ("bereavement", "ul"),
        ("sick", "sl"),
        ("other", "ul"),
    ):
        conn.execute("UPDATE leave_requests SET reason = ? WHERE reason = ?", (new, old))
    conn.execute("UPDATE leave_requests SET reason = 'compoff' WHERE lower(trim(reason)) = 'wfh'")
    conn.execute(
        "UPDATE leave_requests SET reason = 'ul' WHERE reason NOT IN ('pl', 'ul', 'sl', 'll', 'compoff')"
    )
    cols_scrum = {row[1] for row in conn.execute("PRAGMA table_info(scrum_daily_task)")}
    if "sprint_id" not in cols_scrum:
        conn.execute(
            "ALTER TABLE scrum_daily_task ADD COLUMN sprint_id INTEGER REFERENCES scrum_sprint(id) ON DELETE SET NULL"
        )
    if "sprint_item_id" not in cols_scrum:
        conn.execute(
            "ALTER TABLE scrum_daily_task ADD COLUMN sprint_item_id INTEGER REFERENCES scrum_sprint_item(id) ON DELETE SET NULL"
        )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_scrum_daily_sprint ON scrum_daily_task(sprint_id, work_date)")
    cols_item = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint_item)")}
    if "kanban_column" not in cols_item:
        conn.execute("ALTER TABLE scrum_sprint_item ADD COLUMN kanban_column TEXT NOT NULL DEFAULT 'backlog'")
    if "task_kind" not in cols_item:
        conn.execute("ALTER TABLE scrum_sprint_item ADD COLUMN task_kind TEXT NOT NULL DEFAULT 'task'")
    cols_item2 = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint_item)")}
    if "sticky_color_hex" not in cols_item2:
        conn.execute("ALTER TABLE scrum_sprint_item ADD COLUMN sticky_color_hex TEXT")
    cols_item3 = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint_item)")}
    if "dod" not in cols_item3:
        conn.execute("ALTER TABLE scrum_sprint_item ADD COLUMN dod TEXT NOT NULL DEFAULT ''")
    cols_item4 = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint_item)")}
    if "done_artifacts" not in cols_item4:
        conn.execute("ALTER TABLE scrum_sprint_item ADD COLUMN done_artifacts TEXT NOT NULL DEFAULT '[]'")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_item_activity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            body TEXT NOT NULL DEFAULT '',
            committed_hours REAL,
            from_column TEXT NOT NULL DEFAULT '',
            to_column TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_activity_item ON scrum_item_activity(item_id, created_at DESC)"
    )
    conn.execute("CREATE TABLE IF NOT EXISTS app_migrations (id TEXT PRIMARY KEY)")
    if not conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_kanban_v1",)).fetchone():
        conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_kanban_v1')")
        conn.execute(
            """
            UPDATE scrum_sprint_item SET kanban_column = CASE lower(trim(status))
                WHEN 'done' THEN 'done'
                WHEN 'doing' THEN 'doing'
                WHEN 'blocked' THEN 'doing'
                ELSE 'backlog'
            END
            """
        )
    _migrate_employee_portal_v1(conn)
    _migrate_scrum_sprint_item_area_v1(conn)
    _migrate_scrum_sprint_item_attachment_v1(conn)
    _migrate_roster_email_v1(conn)
    _migrate_portal_employee_access_code_v1(conn)
    _migrate_portal_employee_access_code_email_v1(conn)
    _seed_default_team_if_empty(conn)
    _seed_portal_demo_scrum_items_if_requested(conn)
    _migrate_scrum_team_task_kinds(conn)
    _migrate_scrum_task_kind_short_labels(conn)
    _migrate_scrum_task_kind_fixed_three_v1(conn)
    _migrate_scrum_task_kind_builtin_five_v1(conn)
    _migrate_scrum_sprint_team_capacity_hours_v1(conn)
    _migrate_scrum_sprint_is_closed_v1(conn)
    _migrate_scrum_item_appreciation_v1(conn)
    _migrate_portal_email_otp_v1(conn)
    _migrate_scrum_portal_proposal_v1(conn)
    _migrate_scrum_item_activity_committed_hours_nullable_v1(conn)
    _migrate_managers_table_v1(conn)
    _migrate_teams_owner_email_v1(conn)
    _migrate_lpo_manager_emails_v1(conn)
    _migrate_lpo_manager_team_access_v1(conn)
    _migrate_strip_approved_employee_change_prefix_v1(conn)
    _migrate_scrum_item_linked_files_v1(conn)
    _migrate_scrum_item_linked_file_auth_v1(conn)
    _migrate_scrum_item_checklist_v1(conn)
    _seed_primary_owner_manager_from_master_pin(app, conn)
    conn.commit()
    conn.close()


def get_union_roster_for_app(app: Flask) -> tuple[str, ...]:
    """Distinct members across all teams (for public leave / typeahead)."""
    conn = get_db(app)
    rows = conn.execute(
        "SELECT DISTINCT employee_name FROM team_roster ORDER BY employee_name COLLATE NOCASE"
    ).fetchall()
    conn.close()
    return tuple(r[0] for r in rows) if rows else EMPLOYEES


def build_team_roster_export_rows(conn: sqlite3.Connection) -> list[tuple[str, str, str]]:
    """
    Rows for roster Excel/CSV: (TeamName, EmployeeName, EmployeeEmail) for each team_roster row,
    then Default rows for seeded EMPLOYEES not on any team (case-insensitive).
    """
    rows = conn.execute(
        """
        SELECT t.name AS team_name, tr.employee_name AS employee_name,
               tr.employee_email AS employee_email
        FROM team_roster tr
        JOIN teams t ON t.id = tr.team_id
        ORDER BY t.name COLLATE NOCASE, tr.sort_order, tr.employee_name COLLATE NOCASE
        """
    ).fetchall()
    triplets: list[tuple[str, str, str]] = []
    for r in rows:
        tn = str(r["team_name"] or "").strip() or "Default"
        en = str(r["employee_name"] or "").strip()
        em = normalize_email(str(r["employee_email"] or ""))
        if en:
            triplets.append((tn, en, em))
    mapped_ci = {n.casefold() for _, n, _ in triplets}
    drow = conn.execute("SELECT name FROM teams WHERE lower(trim(name)) = 'default' LIMIT 1").fetchone()
    default_label = "Default"
    if drow and (drow["name"] or "").strip():
        default_label = str(drow["name"]).strip()
    for emp in EMPLOYEES:
        e = (emp or "").strip()
        if not e or e.casefold() in mapped_ci:
            continue
        triplets.append((default_label, e, ""))
        mapped_ci.add(e.casefold())
    return triplets


def _find_csv_col(header: list[str], candidates: tuple[str, ...]) -> int | None:
    hlow = [(i, (c or "").strip().lower()) for i, c in enumerate(header)]
    for cand in candidates:
        c = cand.lower()
        for i, h in hlow:
            if h == c or h.replace(" ", "") == c.replace(" ", ""):
                return i
    for cand in candidates:
        c = cand.lower()
        for i, h in hlow:
            if c in h:
                return i
    return None


def _roster_table_rows_to_entries(header: list[str], data_rows: list[list[str]]) -> list[tuple[str, str, str]]:
    """Map tabular rows to (team_name, employee_name, employee_email) using flexible header matching."""
    header = [(c or "").strip() for c in header]
    ti = _find_csv_col(header, ("teamname", "team", "team name", "squad", "group", "pod"))
    ni = _find_csv_col(
        header, ("employeename", "employee name", "name", "employee", "member", "display name", "full name", "person")
    )
    ei = _find_csv_col(
        header, ("employeeemail", "employee email", "email", "e-mail", "mail", "nokia email", "work email")
    )
    if ni is None and len(header) >= 2 and ti == 0:
        ni = 1
    if ni is None and len(header) >= 2 and ti is None:
        ti, ni = 0, 1
    if ei is None and len(header) >= 3:
        used = {x for x in (ti, ni) if x is not None}
        for idx in range(len(header)):
            if idx not in used:
                ei = idx
                break
    if ni is None:
        raise ValueError(
            "Could not detect employee column (try headers: Team Name, Employee Name, Employee Email)."
        )

    entries: list[tuple[str, str, str]] = []
    for r in data_rows:

        def cell(idx: int | None) -> str:
            if idx is None or idx >= len(r):
                return ""
            return (r[idx] or "").strip()

        emp = cell(ni)
        if not emp:
            continue
        team = cell(ti) if ti is not None else "Default"
        if not team:
            team = "Default"
        email = normalize_email(cell(ei)) if ei is not None else ""
        entries.append((team, emp, email))
    return entries


def _roster_table_rows_to_pairs(header: list[str], data_rows: list[list[str]]) -> list[tuple[str, str]]:
    """Legacy helper: team + employee only (email ignored)."""
    return [(t, n) for t, n, _e in _roster_table_rows_to_entries(header, data_rows)]


def _sqlite_db_snapshot_copy(db_path: Path, label: str) -> Path | None:
    """Copy the SQLite file beside the live DB (e.g. before roster import). Returns path or None."""
    db_path = db_path.resolve()
    if not db_path.is_file():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = db_path.parent / f"{db_path.stem}.{label}_{ts}{db_path.suffix}"
    try:
        shutil.copy2(db_path, dest)
        return dest
    except OSError:
        logging.getLogger(__name__).warning("DB snapshot failed: %s → %s", db_path, dest, exc_info=True)
        return None


def ingest_team_roster_entries(
    app: Flask, entries: list[tuple[str, str, str]]
) -> tuple[int, int, list[str]]:
    """Replace roster rows for each team present in entries (teams are created if missing)."""
    warnings: list[str] = []
    if not entries:
        raise ValueError("No member rows found below the header.")

    _sqlite_db_snapshot_copy(Path(app.config["DB_PATH"]), "pre_roster_upload")

    by_team: dict[str, list[tuple[str, str]]] = {}
    for team, emp, email in entries:
        by_team.setdefault(team, []).append((emp, normalize_email(email)))

    conn = get_db(app)
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    teams_updated = 0
    people_rows = 0
    for team_name, members in by_team.items():
        conn.execute("INSERT OR IGNORE INTO teams (name, created_at) VALUES (?, ?)", (team_name, ts))
        row = conn.execute("SELECT id FROM teams WHERE name = ?", (team_name,)).fetchone()
        if not row:
            warnings.append(f"Could not resolve team id for {team_name!r}.")
            continue
        tid = int(row["id"])
        teams_updated += 1
        conn.execute("DELETE FROM team_roster WHERE team_id = ?", (tid,))
        deduped: list[tuple[str, str]] = []
        seen: set[str] = set()
        for n, em in members:
            if n not in seen:
                seen.add(n)
                deduped.append((n, em))
        for i, (emp, email) in enumerate(deduped):
            conn.execute(
                """
                INSERT INTO team_roster (team_id, employee_name, employee_email, sort_order)
                VALUES (?, ?, ?, ?)
                """,
                (tid, emp, email, i),
            )
            people_rows += 1
    conn.commit()
    conn.close()
    return teams_updated, people_rows, warnings


def ingest_team_roster_pairs(app: Flask, pairs: list[tuple[str, str]]) -> tuple[int, int, list[str]]:
    """Legacy: team + employee only (empty email)."""
    entries = [(team, emp, "") for team, emp in pairs]
    return ingest_team_roster_entries(app, entries)


def ingest_team_roster_xlsx(app: Flask, data: bytes) -> tuple[int, int, list[str]]:
    """Parse first worksheet of an .xlsx workbook; columns TeamName + EmployeeName (or legacy headers)."""
    from openpyxl import load_workbook

    if not data:
        raise ValueError("Empty file.")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as ex:  # noqa: BLE001
        raise ValueError("Could not read Excel file (.xlsx).") from ex
    rows: list[list[str]] = []
    try:
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 5000:
                break
            cells = [("" if c is None else str(c)).strip() for c in row]
            if not any(cells):
                continue
            rows.append(cells)
    finally:
        wb.close()
    if not rows:
        raise ValueError("No rows in workbook.")
    header = rows[0]
    entries = _roster_table_rows_to_entries(header, rows[1:])
    return ingest_team_roster_entries(app, entries)


def ingest_team_roster_csv(app: Flask, text: str) -> tuple[int, int, list[str]]:
    """
    Replace roster rows for each team present in the CSV (teams are created if missing).
    Expected columns include team + member name (TeamName / EmployeeName, or Team / Name, etc.).
    Returns (teams_updated, people_rows, warnings).
    """
    raw = (text or "").strip()
    if not raw:
        raise ValueError("Empty file.")
    sample = raw[:8192] if len(raw) > 8192 else raw
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(raw), dialect)
    rows = [r for r in reader if r and any((c or "").strip() for c in r)]
    if not rows:
        raise ValueError("No rows in CSV.")
    header = [(c or "").strip() for c in rows[0]]
    entries = _roster_table_rows_to_entries(header, rows[1:])
    if not entries:
        raise ValueError("No member rows found below the header.")
    return ingest_team_roster_entries(app, entries)


def ingest_team_roster_json(app: Flask, data: bytes) -> tuple[int, int, list[str]]:
    """Parse roster JSON: {rows: [{team_name, employee_name, employee_email}, ...]} or a bare array."""
    if not data:
        raise ValueError("Empty file.")
    try:
        body = json.loads(data.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as ex:
        raise ValueError("Could not parse roster JSON.") from ex
    raw_rows = body.get("rows") if isinstance(body, dict) else body
    if not isinstance(raw_rows, list):
        raise ValueError("Roster JSON must be an array or an object with a rows array.")
    entries: list[tuple[str, str, str]] = []
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        team = str(item.get("team_name") or item.get("team") or item.get("Team Name") or "Default").strip()
        name = str(item.get("employee_name") or item.get("name") or item.get("Employee Name") or "").strip()
        email = normalize_email(
            str(item.get("employee_email") or item.get("email") or item.get("Employee Email") or "")
        )
        if not name:
            continue
        if not team:
            team = "Default"
        entries.append((team, name, email))
    if not entries:
        raise ValueError("No roster rows found in JSON.")
    return ingest_team_roster_entries(app, entries)


def _safe_next_path(nxt: str | None) -> str | None:
    if not nxt:
        return None
    u = urlparse(nxt)
    if u.scheme or u.netloc:
        return None
    path = u.path or "/"
    if not path.startswith("/") or path.startswith("//"):
        return None
    return path + (f"?{u.query}" if u.query else "")


def _portal_session() -> dict | None:
    raw = session.get("portal_user")
    if not isinstance(raw, dict):
        return None
    return raw


def _portal_dod_default_json() -> str:
    return json.dumps([False] * len(PORTAL_DOD_CHECKLIST_LABELS))


def _portal_dod_parse(raw: str | None) -> list[bool]:
    n = len(PORTAL_DOD_CHECKLIST_LABELS)
    try:
        data = json.loads(raw or "[]")
        if not isinstance(data, list):
            return [False] * n
        return [bool(data[i]) if i < len(data) else False for i in range(n)]
    except Exception:
        return [False] * n


def _portal_dod_all_done(states: list[bool]) -> bool:
    n = len(PORTAL_DOD_CHECKLIST_LABELS)
    return len(states) == n and all(states)


def _working_weekdays_count(sd: date, ed: date) -> int:
    return sum(1 for d in _daterange_inclusive(sd, ed) if d.weekday() < 5)


def _portal_status_to_kanban(ps: str) -> str:
    x = (ps or "").strip().lower()
    if x == "progress":
        return "doing"
    if x == "done":
        return "done"
    return "backlog"


def _kanban_to_portal_status(col: str | None) -> str:
    c = _normalize_kanban_column(col)
    if c == "done":
        return "done"
    if c in ("doing", "do"):
        return "progress"
    return "todo"


def _migrate_scrum_sprint_item_area_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_sprint_item_area_v1",)).fetchone():
        return
    cols = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint_item)")}
    if "area" not in cols:
        conn.execute("ALTER TABLE scrum_sprint_item ADD COLUMN area TEXT NOT NULL DEFAULT ''")
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_sprint_item_area_v1')")


def _migrate_scrum_sprint_item_attachment_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("scrum_sprint_item_attachment_v1",)).fetchone():
        return
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS scrum_sprint_item_attachment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL REFERENCES scrum_sprint_item(id) ON DELETE CASCADE,
            rel_path TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            content_type TEXT NOT NULL DEFAULT 'application/octet-stream',
            size_bytes INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_scrum_item_att_item ON scrum_sprint_item_attachment(item_id, id)"
    )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('scrum_sprint_item_attachment_v1')")


def _migrate_employee_portal_v1(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("employee_portal_v1",)).fetchone():
        return
    cols_leave = {row[1] for row in conn.execute("PRAGMA table_info(leave_requests)")}
    if "submitted_email" not in cols_leave:
        conn.execute(
            "ALTER TABLE leave_requests ADD COLUMN submitted_email TEXT NOT NULL DEFAULT ''"
        )
    if "portal_leave_label" not in cols_leave:
        conn.execute(
            "ALTER TABLE leave_requests ADD COLUMN portal_leave_label TEXT NOT NULL DEFAULT ''"
        )
    cols_item = {row[1] for row in conn.execute("PRAGMA table_info(scrum_sprint_item)")}
    if "portal_dod_json" not in cols_item:
        conn.execute(
            "ALTER TABLE scrum_sprint_item ADD COLUMN portal_dod_json TEXT NOT NULL DEFAULT '[]'"
        )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('employee_portal_v1')")


def _seed_portal_demo_scrum_items_if_requested(conn: sqlite3.Connection) -> None:
    """Optional demo stickies (3 per directory member) when PORTAL_SEED_DEMO_ITEMS=1."""
    flag = os.environ.get("PORTAL_SEED_DEMO_ITEMS", "").strip().lower()
    if flag not in ("1", "true", "yes"):
        return
    if conn.execute("SELECT 1 FROM app_migrations WHERE id = ?", ("portal_seed_demo_v1",)).fetchone():
        return
    team = conn.execute("SELECT id FROM teams ORDER BY id LIMIT 1").fetchone()
    if not team:
        return
    team_id = int(team["id"])
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    sp = conn.execute(
        "SELECT id FROM scrum_sprint WHERE team_id = ? AND name = ?",
        (team_id, "Portal demo seed"),
    ).fetchone()
    if sp:
        sid = int(sp["id"])
    else:
        cur = conn.execute(
            """
            INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                team_id,
                "Portal demo seed",
                date.today().isoformat(),
                (date.today() + timedelta(days=14)).isoformat(),
                "Auto-seeded for Microsoft portal demo",
                ts,
                ts,
            ),
        )
        sid = int(cur.lastrowid)
    titles = ("Sample task A", "Sample task B", "Sample task C")
    for _em, _disp, roster in NOKIA_PORTAL_DIRECTORY:
        for idx, title in enumerate(titles):
            conn.execute(
                """
                INSERT INTO scrum_sprint_item
                (sprint_id, assignee, title, estimate_hours, status, notes, dod, done_artifacts,
                 sort_order, created_at, updated_at, kanban_column, task_kind, portal_dod_json)
                VALUES (?, ?, ?, 4.0, 'open', '', '', '[]', ?, ?, ?, 'backlog', 'ndy', ?)
                """,
                (sid, roster, title, idx, ts, ts, _portal_dod_default_json()),
            )
    conn.execute("INSERT INTO app_migrations (id) VALUES ('portal_seed_demo_v1')")


def _manager_password_configured(app: Flask) -> bool:
    return bool((app.config.get("MANAGER_DASHBOARD_PASSWORD") or "").strip())


def _manager_logged_in() -> bool:
    return bool(session.get("manager"))


def _check_manager_password(app: Flask, candidate: str) -> bool:
    expected = (app.config.get("MANAGER_DASHBOARD_PASSWORD") or "").strip()
    if not expected:
        return False
    try:
        return secrets.compare_digest(candidate.encode("utf-8"), expected.encode("utf-8"))
    except Exception:
        return False


def _lpo_sm_password_configured(app: Flask) -> bool:
    return bool((app.config.get("LPO_SM_DASHBOARD_PASSWORD") or "").strip())


def _check_lpo_sm_password(app: Flask, candidate: str) -> bool:
    expected = (app.config.get("LPO_SM_DASHBOARD_PASSWORD") or "").strip()
    if not expected:
        return False
    try:
        return secrets.compare_digest(candidate.encode("utf-8"), expected.encode("utf-8"))
    except Exception:
        return False


def _lpo_assigned_team_ids(conn: sqlite3.Connection, email: str) -> list[int]:
    """Return team IDs explicitly assigned to this LPO email (empty = none assigned yet)."""
    rows = conn.execute(
        "SELECT team_id FROM lpo_manager_team_access WHERE lpo_email = ? COLLATE NOCASE",
        (normalize_email(email),),
    ).fetchall()
    return [int(r["team_id"]) for r in rows]


def _fetch_lpo_manager_emails(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        "SELECT email FROM lpo_manager_emails ORDER BY email COLLATE NOCASE"
    ).fetchall()
    return [normalize_email(str(r[0])) for r in rows if str(r[0] or "").strip()]


def _is_lpo_manager_email(app: Flask, email: str) -> bool:
    em = normalize_email(email)
    if not em:
        return False
    conn = get_db(app)
    row = conn.execute(
        "SELECT 1 FROM lpo_manager_emails WHERE email = ? COLLATE NOCASE",
        (em,),
    ).fetchone()
    conn.close()
    return row is not None


def _resolve_portal_signin_identity(app: Flask, email: str) -> dict | None:
    """Roster directory entry, or a synthetic identity for configured LPO manager emails."""
    em = normalize_email(email)
    if not em:
        return None
    hit = lookup_portal_directory(em)
    if hit:
        return hit
    if _is_lpo_manager_email(app, em):
        local = em.split("@")[0].replace(".", " ").replace("_", " ").title()
        return {"email": em, "display_name": local, "roster_name": local}
    return None


def _establish_lpo_manager_session(email: str) -> None:
    em = normalize_email(email)
    session["manager"] = True
    session["manager_role"] = "lpo_sm"
    session["manager_user_email"] = em
    session.pop("my_employee_name", None)
    session.pop("portal_user", None)
    session.pop("portal_otp_email", None)
    session.pop("portal_otp_fails", None)


def _team_rows_for_roster_name_conn(
    conn: sqlite3.Connection, roster_name: str
) -> list[dict[str, Any]]:
    name = (roster_name or "").strip()
    if not name:
        return []
    rows = conn.execute(
        """
        SELECT DISTINCT t.id AS team_id, t.name AS team_name
        FROM team_roster tr
        JOIN teams t ON t.id = tr.team_id
        WHERE trim(tr.employee_name) COLLATE NOCASE = trim(?)
        ORDER BY t.name COLLATE NOCASE
        """,
        (name,),
    ).fetchall()
    return [
        {"team_id": int(r["team_id"]), "team_name": str(r["team_name"] or "").strip()}
        for r in rows
    ]


def _fetch_portal_employee_team_rows_conn(
    conn: sqlite3.Connection, roster_name: str
) -> list[dict[str, Any]]:
    name = (roster_name or "").strip()
    if not name:
        return []
    rows = _team_rows_for_roster_name_conn(conn, name)
    if rows:
        return rows
    all_names = [
        str(r[0]).strip()
        for r in conn.execute(
            "SELECT DISTINCT employee_name FROM team_roster WHERE trim(employee_name) != ''"
        ).fetchall()
        if str(r[0] or "").strip()
    ]
    canonical = _match_roster_employee_name(name, all_names)
    if canonical:
        return _team_rows_for_roster_name_conn(conn, canonical)
    return []


def _portal_name_match_candidates(
    email: str, display_name: str, roster_name: str
) -> list[str]:
    """Name strings to try when matching a signed-in employee to team_roster."""
    out: list[str] = []

    def add(value: str | None) -> None:
        v = (value or "").strip()
        if v and v not in out:
            out.append(v)

    add(roster_name)
    add(display_name)
    dir_hit = lookup_portal_directory(email) if email else None
    if dir_hit:
        add(str(dir_hit.get("roster_name") or ""))
        add(str(dir_hit.get("display_name") or ""))
    return out


def _match_roster_employee_name(candidate: str, roster_names: Sequence[str]) -> str | None:
    """Return the exact team_roster spelling for this person, if recognized."""
    cand = (candidate or "").strip()
    if not cand or not roster_names:
        return None
    cand_n = _norm(cand)
    for rn in roster_names:
        if _norm(rn) == cand_n:
            return str(rn)
    resolved = resolve_employee_name(cand, roster=roster_names)
    if resolved:
        return resolved
    fuzzy = fuzzy_employee_matches(cand, 1, roster=roster_names)
    return fuzzy[0] if fuzzy else None


def _resolve_portal_employee_roster_name(
    app: Flask,
    *,
    email: str = "",
    display_name: str = "",
    roster_name: str = "",
) -> str | None:
    """Map portal sign-in identity to the canonical name on team_roster."""
    union = get_union_roster_for_app(app)
    if not union:
        name = (roster_name or display_name or "").strip()
        return name or None
    for cand in _portal_name_match_candidates(email, display_name, roster_name):
        hit = _match_roster_employee_name(cand, union)
        if hit:
            return hit
    return None


def _portal_effective_roster_name(app: Flask, pu: dict | None) -> str:
    """Canonical roster name for leave, sprint, and team mapping."""
    if not pu:
        return ""
    email = str(pu.get("email") or "")
    display = str(pu.get("name") or "")
    roster = str(pu.get("roster_name") or "")
    canonical = _resolve_portal_employee_roster_name(
        app, email=email, display_name=display, roster_name=roster
    )
    if canonical and session.get("portal_user") and canonical != roster:
        session["portal_user"]["roster_name"] = canonical
    return canonical or roster


def _portal_employee_team_rows(app: Flask, roster_name: str) -> list[dict[str, Any]]:
    """Teams this portal employee belongs to (from roster upload mapping)."""
    pu = _portal_session()
    candidates: list[str] = []
    if pu:
        roster_name = _portal_effective_roster_name(app, pu) or roster_name
        candidates.extend(
            _portal_name_match_candidates(
                str(pu.get("email") or ""),
                str(pu.get("name") or ""),
                str(pu.get("roster_name") or ""),
            )
        )
    name = (roster_name or "").strip()
    if name and name not in candidates:
        candidates.insert(0, name)
    elif name:
        candidates = [name]
    if not candidates:
        return []
    conn = get_db(app)
    rows: list[dict[str, Any]] = []
    seen_team_ids: set[int] = set()
    for cand in candidates:
        for row in _fetch_portal_employee_team_rows_conn(conn, cand):
            tid = int(row["team_id"])
            if tid in seen_team_ids:
                continue
            seen_team_ids.add(tid)
            rows.append(row)
    conn.close()
    rows.sort(key=lambda r: str(r.get("team_name") or "").casefold())
    return rows


def _portal_employee_team_ids(app: Flask, roster_name: str) -> set[int]:
    return {int(t["team_id"]) for t in _portal_employee_team_rows(app, roster_name)}


def _portal_employee_can_access_team(app: Flask, roster_name: str, team_id: int) -> bool:
    return int(team_id) in _portal_employee_team_ids(app, roster_name)


def _resolve_portal_selected_team_id(
    app: Flask,
    roster_name: str,
    team_id_hint: int | None = None,
    *,
    persist: bool = True,
) -> int | None:
    """Active portal team: from ?team_id=, session, or sole roster mapping."""
    teams = _portal_employee_team_rows(app, roster_name)
    allowed = {int(t["team_id"]) for t in teams}
    if not allowed:
        if persist:
            session.pop("portal_team_id", None)
        return None
    chosen: int | None = None
    if team_id_hint is not None and int(team_id_hint) in allowed:
        chosen = int(team_id_hint)
    else:
        raw_sess = session.get("portal_team_id")
        if raw_sess is not None:
            try:
                sid = int(raw_sess)
                if sid in allowed:
                    chosen = sid
            except (TypeError, ValueError):
                pass
    if chosen is None and len(allowed) == 1:
        chosen = next(iter(allowed))
    if persist:
        if chosen is not None:
            session["portal_team_id"] = chosen
        elif len(allowed) > 1:
            session.pop("portal_team_id", None)
    return chosen


def _portal_team_name_for_id(teams: list[dict[str, Any]], team_id: int | None) -> str | None:
    if team_id is None:
        return None
    for t in teams:
        if int(t["team_id"]) == int(team_id):
            nm = str(t.get("team_name") or "").strip()
            return nm or None
    return None


def _portal_employee_primary_team_id(app: Flask, roster_name: str) -> int | None:
    """Team used for the employee's current sprint board (roster mapping)."""
    teams = _portal_employee_team_rows(app, roster_name)
    if not teams:
        return None
    allowed = {int(t["team_id"]) for t in teams}

    def _pick_from_hint(raw: object | None) -> int | None:
        if raw is None:
            return None
        try:
            tid = int(raw)
        except (TypeError, ValueError):
            return None
        return tid if tid in allowed else None

    pu = _portal_session()
    if pu:
        chosen = _pick_from_hint(pu.get("team_id"))
        if chosen is not None:
            return chosen
    chosen = _pick_from_hint(session.get("portal_team_id"))
    if chosen is not None:
        return chosen
    if len(teams) == 1:
        return int(teams[0]["team_id"])
    conn = get_db(app)
    today = date.today().isoformat()
    in_window: int | None = None
    with_sprint: int | None = None
    for t in teams:
        tid = int(t["team_id"])
        row = conn.execute(
            """
            SELECT id FROM scrum_sprint
            WHERE team_id = ? AND start_date <= ? AND end_date >= ?
            ORDER BY start_date DESC
            LIMIT 1
            """,
            (tid, today, today),
        ).fetchone()
        if row:
            in_window = tid
            break
        if with_sprint is None and _pick_default_sprint_id(conn, tid, None) is not None:
            with_sprint = tid
    conn.close()
    if in_window is not None:
        return in_window
    if with_sprint is not None:
        return with_sprint
    return int(teams[0]["team_id"])


def _portal_current_sprint_id_for_employee(app: Flask, roster_name: str) -> int | None:
    """Current sprint for this employee's mapped team (in-window, else latest)."""
    team_id = _portal_employee_primary_team_id(app, roster_name)
    if team_id is None:
        return None
    conn = get_db(app)
    sid = _pick_default_sprint_id(conn, int(team_id), None)
    conn.close()
    return sid


def _portal_employee_landing_response(app: Flask):
    """Default employee portal home: current sprint Kanban (Sprint work tab)."""
    pu = _portal_session()
    if not pu:
        return redirect(url_for("home"))
    roster_name = _portal_effective_roster_name(app, pu)
    if not roster_name:
        flash("Your profile is missing a roster name.", "error")
        return redirect(url_for("home"))
    team_id = _portal_employee_primary_team_id(app, roster_name)
    if team_id is None:
        flash(
            "You are not on a team roster yet. Ask your manager or LPO to add you in Settings.",
            "info",
        )
        return render_template(
            "portal_dashboard.html",
            hide_nav=True,
        )
    session["portal_team_id"] = int(team_id)
    sprint_id = _portal_current_sprint_id_for_employee(app, roster_name)
    if sprint_id is None:
        flash("No sprint has been created for your team yet.", "info")
        return render_template(
            "portal_dashboard.html",
            hide_nav=True,
        )
    return redirect(url_for("portal_scrum_kanban_board", sprint_id=int(sprint_id)))


def _list_portal_employee_sprints(
    app: Flask, roster_name: str, team_id: int | None = None
) -> list[dict[str, Any]]:
    """Manager/LPO sprints for the employee's selected roster-mapped team."""
    team_rows = _portal_employee_team_rows(app, roster_name)
    allowed = {int(t["team_id"]) for t in team_rows}
    if not allowed:
        return []
    if team_id is not None:
        if int(team_id) not in allowed:
            return []
        team_ids = [int(team_id)]
    elif len(allowed) == 1:
        team_ids = [next(iter(allowed))]
    else:
        return []
    conn = get_db(app)
    placeholders = ",".join("?" * len(team_ids))
    rows = conn.execute(
        f"""
        SELECT id, name, start_date, end_date, COALESCE(is_closed, 0) AS is_closed
        FROM scrum_sprint
        WHERE team_id IN ({placeholders})
        ORDER BY start_date DESC, name COLLATE NOCASE
        """,
        team_ids,
    ).fetchall()
    for r in rows:
        _maybe_auto_close_scrum_sprint(conn, int(r["id"]))
    if rows:
        conn.commit()
        sprint_ids = [int(r["id"]) for r in rows]
        placeholders = ",".join("?" * len(sprint_ids))
        rows = conn.execute(
            f"""
            SELECT id, name, start_date, end_date, COALESCE(is_closed, 0) AS is_closed
            FROM scrum_sprint
            WHERE id IN ({placeholders})
            ORDER BY start_date DESC, name COLLATE NOCASE
            """,
            sprint_ids,
        ).fetchall()
    conn.close()
    out: list[dict[str, Any]] = []
    for r in rows:
        freeze = _sprint_board_frozen_reason_for_row(r)
        closed = _sprint_row_is_closed(r)
        past = _sprint_row_past_end(r)
        if closed:
            label = "CLOSED"
        elif past:
            label = "ENDED"
        else:
            label = "OPEN"
        out.append(
            {
                "id": int(r["id"]),
                "name": str(r["name"] or f"Sprint {r['id']}"),
                "start_date": str(r["start_date"] or ""),
                "end_date": str(r["end_date"] or ""),
                "is_editable": freeze is None,
                "status_label": label,
            }
        )
    return out


def _complete_employee_signin(app: Flask, hit: dict):
    """After Microsoft or email OTP auth: LPO emails become managers; others use the employee portal."""
    em = normalize_email(str(hit.get("email") or ""))
    name = str(hit.get("display_name") or em.split("@")[0])
    if _is_lpo_manager_email(app, em):
        _establish_lpo_manager_session(em)
        flash(f"Welcome, {name}. Signed in with LPO manager access.", "success")
        return redirect(url_for("dashboard"))
    session["portal_user"] = {
        "email": em,
        "name": name,
        "roster_name": hit["roster_name"],
        "role": "employee",
        "auth": hit.get("auth") or "portal",
    }
    canonical = _resolve_portal_employee_roster_name(
        app,
        email=em,
        display_name=name,
        roster_name=str(hit.get("roster_name") or ""),
    )
    if canonical:
        session["portal_user"]["roster_name"] = canonical
    roster_for_teams = canonical or str(hit.get("roster_name") or "")
    emp_teams = _portal_employee_team_rows(app, roster_for_teams)
    team_id = _portal_employee_primary_team_id(app, roster_for_teams)
    if team_id is not None:
        session["portal_team_id"] = int(team_id)
        if len(emp_teams) == 1:
            session["portal_user"]["team_id"] = int(team_id)
            session["portal_user"]["team_name"] = str(emp_teams[0]["team_name"])
    elif emp_teams:
        session["portal_user"]["team_name"] = ", ".join(
            str(t["team_name"]) for t in emp_teams if t.get("team_name")
        )
    session.permanent = True
    flash(f"Welcome, {name}.", "success")
    return _portal_employee_landing_response(app)

def _daterange_inclusive(a: date, b: date):
    d = a
    while d <= b:
        yield d
        d += timedelta(days=1)


def _weekdays_in_range_inclusive(a: date, b: date) -> list[date]:
    """Monday–Friday dates in [a, b] inclusive (Saturday/Sunday omitted)."""
    out: list[date] = []
    for d in _daterange_inclusive(a, b):
        if d.weekday() < 5:
            out.append(d)
    return out


def _load_meet_leave_day_map(app: Flask, start_iso: str, end_iso: str) -> dict[tuple[int, str], str]:
    conn = get_db(app)
    rows = conn.execute(
        """
        SELECT leave_id, work_date, decision
        FROM meet_leave_day
        WHERE work_date >= ? AND work_date <= ?
        """,
        (start_iso, end_iso),
    ).fetchall()
    conn.close()
    return {(int(r["leave_id"]), r["work_date"]): str(r["decision"]) for r in rows}


def _overlapping_leaves_for_day(
    rows: list[sqlite3.Row],
    d: date,
    day_dec: dict[tuple[int, str], str],
) -> list[sqlite3.Row]:
    d_iso = d.isoformat()
    overlapping: list[sqlite3.Row] = []
    for row in rows:
        try:
            sd = date.fromisoformat(row["start_date"])
            ed = date.fromisoformat(row["end_date"])
        except ValueError:
            continue
        if not (sd <= d <= ed):
            continue
        lid = int(row["id"])
        if day_dec.get((lid, d_iso)) == "removed":
            continue
        overlapping.append(row)
    return overlapping


def _effective_leave_status(row: sqlite3.Row, d_iso: str, day_dec: dict[tuple[int, str], str]) -> str:
    if row["status"] == "pending" and day_dec.get((int(row["id"]), d_iso)) == "approved":
        return "approved"
    return row["status"]


def _is_effectively_pending(row: sqlite3.Row, d_iso: str, day_dec: dict[tuple[int, str], str]) -> bool:
    return row["status"] == "pending" and day_dec.get((int(row["id"]), d_iso)) != "approved"


def _maybe_promote_leave_after_meet_days(app: Flask, leave_id: int) -> None:
    """If every calendar day in the request span has meet_leave_day=approved, mark the row approved."""
    conn = get_db(app)
    row = conn.execute("SELECT * FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()
    if not row or row["status"] != "pending":
        conn.close()
        return
    try:
        sd = date.fromisoformat(row["start_date"])
        ed = date.fromisoformat(row["end_date"])
    except ValueError:
        conn.close()
        return
    for d in _daterange_inclusive(sd, ed):
        iso = d.isoformat()
        r = conn.execute(
            "SELECT decision FROM meet_leave_day WHERE leave_id = ? AND work_date = ?",
            (leave_id, iso),
        ).fetchone()
        if not r or r["decision"] != "approved":
            conn.close()
            return
    conn.execute("UPDATE leave_requests SET status = 'approved' WHERE id = ?", (leave_id,))
    conn.commit()
    conn.close()


def _purge_meet_leave_day_rows_for_leave(conn: sqlite3.Connection, leave_id: int) -> None:
    """Drop DSM per-day rows for this request so rejected/withdrawn leave cannot affect any calendar or capacity logic."""
    conn.execute("DELETE FROM meet_leave_day WHERE leave_id = ?", (int(leave_id),))


def _portal_withdraw_leave_request(app: Flask, roster_name: str, leave_id: int) -> str | None:
    """Delete one non-Nokia leave owned by the portal employee. Returns an error key or None."""
    roster_name = (roster_name or "").strip()
    if not roster_name or leave_id <= 0:
        return "missing"
    conn = get_db(app)
    row = conn.execute(
        """
        SELECT id, description FROM leave_requests
        WHERE id = ? AND employee_name = ? COLLATE NOCASE
        """,
        (leave_id, roster_name),
    ).fetchone()
    if not row:
        conn.close()
        return "not_found"
    if NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in str(row["description"] or ""):
        conn.close()
        return "nokia_locked"
    _purge_meet_leave_day_rows_for_leave(conn, int(row["id"]))
    conn.execute("DELETE FROM leave_requests WHERE id = ?", (int(row["id"]),))
    conn.commit()
    conn.close()
    return None


def _leave_row_on_roster(row: sqlite3.Row, roster: Sequence[str]) -> bool:
    """True when the stored employee_name maps to this roster (exact / normalized only)."""
    return _exact_roster_name_match(str(row["employee_name"] or ""), roster) is not None


def _delete_single_day_leave_rows_for_employee(
    conn: sqlite3.Connection, employee_name: str, work_date: str
) -> int:
    """Delete all single-day pending/approved leave rows for one employee on one calendar day."""
    rows = conn.execute(
        """
        SELECT id FROM leave_requests
        WHERE employee_name = ?
          AND start_date = ? AND end_date = ?
          AND status IN ('pending', 'approved')
        """,
        (employee_name, work_date, work_date),
    ).fetchall()
    for row in rows:
        lid = int(row["id"])
        _purge_meet_leave_day_rows_for_leave(conn, lid)
        conn.execute("DELETE FROM leave_requests WHERE id = ?", (lid,))
    return len(rows)


def _apply_leave_tracker_day_removal(app: Flask, leave_id: int, work_date: str) -> bool:
    """Remove one day from the tracker. Single-day rows are deleted from leave_requests; multi-day rows use meet_leave_day."""
    work_date = work_date.strip()[:10]
    conn = get_db(app)
    row = conn.execute("SELECT * FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()
    if not row:
        conn.close()
        return False
    sd = str(row["start_date"] or "")[:10]
    ed = str(row["end_date"] or "")[:10]
    emp = str(row["employee_name"] or "").strip()
    if sd == ed == work_date:
        _delete_single_day_leave_rows_for_employee(conn, emp, work_date)
    else:
        ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        conn.execute(
            """
            INSERT INTO meet_leave_day (leave_id, work_date, decision, updated_at)
            VALUES (?, ?, 'removed', ?)
            ON CONFLICT(leave_id, work_date) DO UPDATE SET
                decision = excluded.decision,
                updated_at = excluded.updated_at
            """,
            (leave_id, work_date, ts),
        )
    conn.commit()
    conn.close()
    return True


def _meet_validate_leave_in_window(
    app: Flask,
    leave_id_raw: str | None,
    work_date: str | None,
    anchor_raw: str | None,
    roster: Sequence[str],
) -> tuple[sqlite3.Row | None, date, str | None]:
    """Returns (leave_row, anchor, error_key)."""
    roster_t = tuple(roster)
    if not leave_id_raw or not work_date:
        return None, date.today(), "bad_input"
    try:
        leave_id = int(str(leave_id_raw).strip())
    except ValueError:
        return None, date.today(), "bad_input"
    anchor = _parse_meet_anchor(anchor_raw)
    try:
        wd = date.fromisoformat(work_date.strip()[:10])
    except ValueError:
        return None, anchor, "bad_date"
    d0, d4 = _meet_window(anchor)
    if not (d0 <= wd <= d4):
        return None, anchor, "out_of_window"
    conn = get_db(app)
    row = conn.execute("SELECT * FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()
    conn.close()
    if not row:
        return None, anchor, "not_found"
    if not _leave_row_on_roster(row, roster_t):
        return None, anchor, "bad_emp"
    if row["status"] not in ("pending", "approved"):
        return None, anchor, "bad_status"
    try:
        sd = date.fromisoformat(row["start_date"])
        ed = date.fromisoformat(row["end_date"])
    except ValueError:
        return None, anchor, "bad_leave_dates"
    if not (sd <= wd <= ed):
        return None, anchor, "day_not_in_leave"
    return row, anchor, None


def _dashboard_validate_leave_for_month(
    app: Flask,
    leave_id_raw: str | None,
    work_date: str | None,
    year: int,
    month: int,
    roster: Sequence[str],
) -> tuple[sqlite3.Row | None, str | None]:
    """Validate leave_id + work_date for dashboard month grid (same row checks as meet, date must fall in that month)."""
    roster_t = tuple(roster)
    if not leave_id_raw or not work_date:
        return None, "bad_input"
    try:
        leave_id = int(str(leave_id_raw).strip())
    except ValueError:
        return None, "bad_input"
    try:
        wd = date.fromisoformat(work_date.strip()[:10])
    except ValueError:
        return None, "bad_date"
    if not (1 <= month <= 12):
        return None, "bad_month"
    first = date(year, month, 1)
    _, last_day = monthrange(year, month)
    last = date(year, month, last_day)
    if not (first <= wd <= last):
        return None, "out_of_month"
    conn = get_db(app)
    row = conn.execute("SELECT * FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()
    conn.close()
    if not row:
        return None, "not_found"
    if not _leave_row_on_roster(row, roster_t):
        return None, "bad_emp"
    if row["status"] not in ("pending", "approved"):
        return None, "bad_status"
    try:
        sd = date.fromisoformat(row["start_date"])
        ed = date.fromisoformat(row["end_date"])
    except ValueError:
        return None, "bad_leave_dates"
    if not (sd <= wd <= ed):
        return None, "day_not_in_leave"
    return row, None


def build_month_context(
    app: Flask,
    year: int,
    month: int,
    roster: Sequence[str] | None = None,
) -> dict:
    """Month leave grid: per-row ``eleave_days`` = Nokia-approved day-units (same rules as green A cells)."""
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    first = date(year, month, 1)
    _, last_day = monthrange(year, month)
    last = date(year, month, last_day)
    month_start_iso = first.isoformat()
    month_end_iso = last.isoformat()

    conn = get_db(app)
    leaves = conn.execute(
        """
        SELECT * FROM leave_requests
        WHERE start_date <= ? AND end_date >= ?
          AND status IN ('pending', 'approved')
        ORDER BY employee_name ASC, start_date ASC
        """,
        (month_end_iso, month_start_iso),
    ).fetchall()
    conn.close()

    day_dec = _load_meet_leave_day_map(app, month_start_iso, month_end_iso)

    by_emp = _bucket_leave_requests_by_roster(leaves, roster_t)

    reason_l = dict(LEAVE_REASONS)
    dur_l = dict(DURATION_CHOICES)

    days_meta: list[dict] = []
    for d in _daterange_inclusive(first, last):
        wd = d.weekday()  # Mon=0
        is_weekend = wd >= 5
        days_meta.append(
            {
                "date": d,
                "iso": d.isoformat(),
                "weekday": calendar.day_abbr[wd],
                "day": d.day,
                "is_weekend": is_weekend,
            }
        )

    grid_rows: list[dict] = []
    for emp in roster_t:
        cells: list[dict | None] = []
        leave_days_total = 0.0
        nokia_a_day_units = 0.0
        for dm in days_meta:
            d = dm["date"]
            d_iso = d.isoformat()
            overlapping = _overlapping_leaves_for_day(by_emp[emp], d, day_dec)
            cell: dict | None = None
            if overlapping:
                pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
                pool = pending_only or overlapping
                row = max(pool, key=_leave_row_display_tiebreak)
                eff = _effective_leave_status(row, d_iso, day_dec)
                desc = str(row["description"] or "")
                if eff in ("pending", "approved"):
                    dur = (row["duration_type"] or "full").strip()
                    day_unit = 0.5 if dur in ("half_am", "half_pm") else 1.0
                    if _leave_reason_counts_toward_day_units(row["reason"]):
                        leave_days_total += day_unit
                    if NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in desc:
                        nokia_a_day_units += day_unit
                code = leave_cell_code(row["reason"], row["duration_type"], eff, desc)
                rlab = reason_l.get(row["reason"], row["reason"])
                dlab = dur_l.get(row["duration_type"], row["duration_type"])
                plab = (
                    str(row["portal_leave_label"] or "").strip()
                    if "portal_leave_label" in row.keys()
                    else ""
                )
                leave_type = plab if plab else rlab
                title = f"{rlab} · {dlab} · {eff} · single-click: approve day · double-click: remove day"
                css = cell_css_class(row["reason"], eff, desc)
                cell = {
                    "code": code,
                    "title": title,
                    "css": css,
                    "status": eff,
                    "leave_id": int(row["id"]),
                    "work_date": d_iso,
                    "leave_type": leave_type,
                    "reason": str(row["reason"] or ""),
                    "duration_type": str(row["duration_type"] or ""),
                    "start_date": str(row["start_date"] or "")[:10],
                    "end_date": str(row["end_date"] or "")[:10],
                    "description": desc,
                    "detail_summary": f"{leave_type} · {dlab} · {eff}",
                    "detail_range": f"{str(row['start_date'])[:10]} → {str(row['end_date'])[:10]}",
                }
            cells.append(cell)
        el_days = round(nokia_a_day_units, 2)
        tot_r = round(leave_days_total, 2)
        gap_days = round(el_days - tot_r, 2)
        grid_rows.append(
            {
                "employee": emp,
                "leave_days_total": tot_r,
                "eleave_days": round(el_days, 2),
                "gap_days": gap_days,
                "cells": cells,
            }
        )

    return {
        "year": year,
        "month": month,
        "month_label": first.strftime("%B %Y"),
        "month_name": first.strftime("%B"),
        "days": days_meta,
        "grid_rows": grid_rows,
        "month_start": month_start_iso,
        "month_end": month_end_iso,
    }


def _employee_leave_tracker_months_in_year(app: Flask, employee_name: str, year: int) -> set[int]:
    """Months (1–12) where the leave tracker grid shows at least one day cell for this employee."""
    name = (employee_name or "").strip()
    if not name:
        return set()
    roster_t = (name,)
    start_iso = f"{year}-01-01"
    end_iso = f"{year}-12-31"
    conn = get_db(app)
    leaves = conn.execute(
        """
        SELECT * FROM leave_requests
        WHERE start_date <= ? AND end_date >= ?
          AND status IN ('pending', 'approved')
        ORDER BY employee_name ASC, start_date ASC
        """,
        (end_iso, start_iso),
    ).fetchall()
    conn.close()

    by_emp = _bucket_leave_requests_by_roster(leaves, roster_t)

    months: set[int] = set()
    for m in range(1, 13):
        first = date(year, m, 1)
        _, last_day = monthrange(year, m)
        last = date(year, m, last_day)
        day_dec = _load_meet_leave_day_map(app, first.isoformat(), last.isoformat())
        for d in _daterange_inclusive(first, last):
            d_iso = d.isoformat()
            overlapping = _overlapping_leaves_for_day(by_emp[name], d, day_dec)
            if not overlapping:
                continue
            pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
            pool = pending_only or overlapping
            row = max(pool, key=_leave_row_display_tiebreak)
            eff = _effective_leave_status(row, d_iso, day_dec)
            if eff in ("pending", "approved"):
                months.add(m)
                break
    return months


def _employee_leave_tracker_row(month_ctx: dict) -> dict | None:
    """First grid row from a leave-tracker month context (single-employee roster)."""
    rows = month_ctx.get("grid_rows") or []
    return rows[0] if rows else None


def _portal_employee_team_roster(app: Flask, roster_name: str) -> tuple[str, ...]:
    """Team roster for the portal employee's team (same names as manager leave tracker)."""
    name = (roster_name or "").strip()
    team_id = _resolve_portal_selected_team_id(app, name, persist=False) if name else None
    if team_id is None and name:
        team_id = _portal_employee_primary_team_id(app, name)
    conn = get_db(app)
    try:
        if team_id is not None:
            trow = conn.execute("SELECT name FROM teams WHERE id = ?", (int(team_id),)).fetchone()
            tname = str(trow["name"] or "").strip() if trow else ""
            if tname.casefold() == "default":
                urows = conn.execute(
                    "SELECT DISTINCT employee_name FROM team_roster ORDER BY employee_name COLLATE NOCASE"
                ).fetchall()
            else:
                urows = conn.execute(
                    """
                    SELECT employee_name FROM team_roster
                    WHERE team_id = ?
                    ORDER BY sort_order, employee_name COLLATE NOCASE
                    """,
                    (int(team_id),),
                ).fetchall()
            if urows:
                return tuple(str(r[0]).strip() for r in urows if str(r[0] or "").strip())
        union = get_union_roster_for_app(app)
        if union:
            return union
        return EMPLOYEES
    finally:
        conn.close()


def _team_leave_tracker_months_in_year(
    app: Flask, roster: Sequence[str], year: int
) -> set[int]:
    """Months where the leave tracker shows at least one filled day for any roster employee."""
    roster_t = tuple(str(r).strip() for r in roster if str(r or "").strip())
    if not roster_t:
        return set()
    months: set[int] = set()
    for m in range(1, 13):
        ctx = build_month_context(app, year, m, roster=roster_t)
        for row in ctx.get("grid_rows") or []:
            if any(c is not None for c in row.get("cells") or []):
                months.add(m)
                break
    return months


def build_leave_tracker_month_xlsx_bytes(month_ctx: dict[str, Any]) -> tuple[bytes | None, str | None]:
    """
    Build .xlsx for the manager month leave grid (month name column, Total, eLeaveCount, GAP, then day codes).
    Returns (bytes, None) on success or (None, error_message) if openpyxl is missing.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "Excel export requires openpyxl (install dependencies)."

    days: list[dict[str, Any]] = list(month_ctx.get("days") or [])
    grid_rows: list[dict[str, Any]] = list(month_ctx.get("grid_rows") or [])
    month_label = str(month_ctx.get("month_label") or "Leave tracker").strip()
    month_name = str(month_ctx.get("month_name") or month_label.split()[0] or "Month").strip()
    sheet_title = month_label[:31] if month_label else "Leave tracker"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_fill_sat = PatternFill("solid", fgColor="4338CA")
    hdr_fill_sun = PatternFill("solid", fgColor="6D28D9")
    hdr_fill_wknd = PatternFill("solid", fgColor="3730A3")
    body_fill_sat = PatternFill("solid", fgColor="E0E7FF")
    body_fill_sun = PatternFill("solid", fgColor="EDE9FE")
    hdr_font = Font(color="F8FAFC", bold=True, size=10)
    thin = Side(style="thin", color="94A3B8")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:A2")
    a1 = ws.cell(row=1, column=1, value="#")
    a1.fill = hdr_fill
    a1.font = hdr_font
    a1.alignment = center
    a1.border = grid_border

    ws.merge_cells("B1:B2")
    b1 = ws.cell(row=1, column=2, value=month_name)
    b1.fill = hdr_fill
    b1.font = hdr_font
    b1.alignment = center
    b1.border = grid_border

    ws.merge_cells("C1:C2")
    c1 = ws.cell(row=1, column=3, value="Total")
    c1.fill = hdr_fill
    c1.font = hdr_font
    c1.alignment = center
    c1.border = grid_border

    ws.merge_cells("D1:D2")
    d1 = ws.cell(row=1, column=4, value="eLeaveCount")
    d1.fill = hdr_fill
    d1.font = hdr_font
    d1.alignment = center
    d1.border = grid_border

    ws.merge_cells("E1:E2")
    e1 = ws.cell(row=1, column=5, value="GAP")
    e1.fill = hdr_fill
    e1.font = hdr_font
    e1.alignment = center
    e1.border = grid_border

    day0 = 6
    for i, dm in enumerate(days):
        col = day0 + i
        wd_label = str(dm.get("weekday") or "").strip()
        wd_l = wd_label.casefold()
        wknd = bool(dm.get("is_weekend"))
        if wd_l == "sat":
            hdr_fill_day = hdr_fill_sat
        elif wd_l == "sun":
            hdr_fill_day = hdr_fill_sun
        else:
            hdr_fill_day = hdr_fill_wknd if wknd else hdr_fill
        h1 = ws.cell(row=1, column=col, value=wd_label)
        h1.fill = hdr_fill_day
        h1.font = hdr_font
        h1.alignment = center
        h1.border = grid_border
        h2 = ws.cell(row=2, column=col, value=dm.get("day"))
        h2.fill = hdr_fill_day
        h2.font = hdr_font
        h2.alignment = center
        h2.border = grid_border

    body_font = Font(size=10)
    r = 3
    for idx, grow in enumerate(grid_rows, start=1):
        n0 = ws.cell(row=r, column=1, value=idx)
        n0.font = body_font
        n0.alignment = center
        n0.border = grid_border
        n1 = ws.cell(row=r, column=2, value=str(grow.get("employee") or ""))
        n1.font = body_font
        n1.border = grid_border
        n1.alignment = Alignment(vertical="center")
        tot = grow.get("leave_days_total", 0)
        try:
            tot_v: int | float = float(tot) if tot is not None else 0.0
        except (TypeError, ValueError):
            tot_v = 0.0
        n2 = ws.cell(row=r, column=3, value=tot_v)
        n2.font = body_font
        n2.alignment = center
        n2.border = grid_border
        try:
            el_v = float(grow.get("eleave_days") or 0)
        except (TypeError, ValueError):
            el_v = 0.0
        try:
            gap_v = float(grow.get("gap_days") or 0)
        except (TypeError, ValueError):
            gap_v = 0.0
        n3 = ws.cell(row=r, column=4, value=el_v)
        n3.font = body_font
        n3.alignment = center
        n3.border = grid_border
        n4 = ws.cell(row=r, column=5, value=gap_v)
        n4.font = Font(size=10, color="FCA5A5") if gap_v < -0.0001 else body_font
        n4.alignment = center
        n4.border = grid_border
        cells = list(grow.get("cells") or [])
        for j, cell in enumerate(cells):
            col = day0 + j
            v = str(cell.get("code") or "") if cell else ""
            css = str(cell.get("css") or "") if cell else ""
            c = ws.cell(row=r, column=col, value=v if v else "")
            c.alignment = center
            c.border = grid_border
            code_l = v.lower()
            is_nokia = "cell-nokia-approved" in css or (v and v.upper().startswith("A"))
            if is_nokia:
                bg, fg = _LEAVE_NOKIA_APPROVED_XLSX_STYLE
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(size=10, bold=True, color=fg)
            elif code_l and code_l in _LEAVE_CODE_XLSX_STYLE:
                bg, fg = _LEAVE_CODE_XLSX_STYLE[code_l]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(size=10, bold=True, color=fg)
            elif j < len(days):
                dmj = days[j]
                wdj = str(dmj.get("weekday") or "").strip().casefold()
                if wdj == "sat":
                    c.fill = body_fill_sat
                    c.font = Font(size=10, color="4338CA")
                elif wdj == "sun":
                    c.fill = body_fill_sun
                    c.font = Font(size=10, color="6D28D9")
                else:
                    c.font = body_font
            else:
                c.font = body_font
        r += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    for i in range(len(days)):
        col_letter = get_column_letter(day0 + i)
        ws.column_dimensions[col_letter].width = 5.5

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


_LEAVE_CODE_XLSX_STYLE: dict[str, tuple[str, str]] = {
    # code → (bg_hex, font_hex)  — matches worksheet.css leave cell colours
    "pl":      ("DBEAFE", "1D4ED8"),
    "ul":      ("FEF3C7", "92400E"),
    "sl":      ("FEE2E2", "991B1B"),
    "ll":      ("EDE9FE", "5B21B6"),
    "compoff": ("E2E8F0", "475569"),
    "full":    ("DBEAFE", "1D4ED8"),   # treat plain 'full' as PL colour
    "half_am": ("BFDBFE", "1E40AF"),   # slightly deeper blue for half-days
    "half_pm": ("BFDBFE", "1E40AF"),
}
# Nokia e-tool approved cells (code starts with "A") — green matching cell-nokia-approved
_LEAVE_NOKIA_APPROVED_XLSX_STYLE: tuple[str, str] = ("DCFCE7", "166534")


def build_leave_tracker_year_xlsx_bytes(
    app: "Flask",
    year: int,
    current_month: int,
    roster: "Sequence[str]",
) -> "tuple[bytes | None, str | None]":
    """
    Build a multi-sheet .xlsx with one sheet per month from January to current_month
    (inclusive).  Each sheet mirrors the single-month export layout so it can be
    re-uploaded to import/overwrite leave data.

    A hidden row 1 carries machine-readable metadata:
        LEAVE_IMPORT_META | <year> | <month_number>
    Row 2 is the human column header, data starts at row 3.
    Columns:  A=#  B=Employee  C=Total  D=eLeaveCount  E=GAP  F…=day cells (leave codes)

    On round-trip upload the importer reads row1 for year/month, row2 for day-numbers,
    then rows 3+ for employee name (col B) and day codes (col F onward).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "Excel export requires openpyxl (install dependencies)."

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_fill_sat = PatternFill("solid", fgColor="4338CA")
    hdr_fill_sun = PatternFill("solid", fgColor="6D28D9")
    hdr_fill_wknd = PatternFill("solid", fgColor="3730A3")
    body_fill_sat = PatternFill("solid", fgColor="E0E7FF")
    body_fill_sun = PatternFill("solid", fgColor="EDE9FE")
    meta_fill = PatternFill("solid", fgColor="0F172A")
    hdr_font = Font(color="F8FAFC", bold=True, size=10)
    meta_font = Font(color="0F172A", size=8)  # near-invisible: same colour as fill
    body_font = Font(size=10)
    thin = Side(style="thin", color="94A3B8")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    month_names = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    day0 = 6  # column F = first day column

    for mo in range(1, current_month + 1):
        ctx = build_month_context(app, year, mo, roster=roster)
        days: list = list(ctx.get("days") or [])
        grid_rows: list = list(ctx.get("grid_rows") or [])
        month_name = month_names[mo - 1]

        ws = wb.create_sheet(title=f"{month_name} {year}")

        # ── hidden meta row (row 1) ──────────────────────────────────────────
        meta_cell = ws.cell(row=1, column=1, value="LEAVE_IMPORT_META")
        meta_cell.font = meta_font
        meta_cell.fill = meta_fill
        ws.cell(row=1, column=2, value=year).font = meta_font
        ws.cell(row=1, column=2).fill = meta_fill
        ws.cell(row=1, column=3, value=mo).font = meta_font
        ws.cell(row=1, column=3).fill = meta_fill
        # store day-numbers in meta row F onwards for round-trip parsing
        for i, dm in enumerate(days):
            mc = ws.cell(row=1, column=day0 + i, value=int(dm.get("day") or 0))
            mc.font = meta_font
            mc.fill = meta_fill
        ws.row_dimensions[1].hidden = True

        # ── header row (row 2) ───────────────────────────────────────────────
        for col, val in [(1, "#"), (2, month_name), (3, "Total"),
                         (4, "eLeaveCount"), (5, "GAP")]:
            c = ws.cell(row=2, column=col, value=val)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = center
            c.border = grid_border

        for i, dm in enumerate(days):
            col = day0 + i
            wd_label = str(dm.get("weekday") or "").strip()
            wd_l = wd_label.casefold()
            wknd = bool(dm.get("is_weekend"))
            if wd_l == "sat":
                fill_d = hdr_fill_sat
            elif wd_l == "sun":
                fill_d = hdr_fill_sun
            else:
                fill_d = hdr_fill_wknd if wknd else hdr_fill
            hd = ws.cell(row=2, column=col, value=f"{wd_label}\n{dm.get('day')}")
            hd.fill = fill_d
            hd.font = hdr_font
            hd.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            hd.border = grid_border

        # ── data rows (row 3+) ───────────────────────────────────────────────
        for idx, grow in enumerate(grid_rows, start=1):
            r = idx + 2
            ws.cell(row=r, column=1, value=idx).font = body_font
            ws.cell(row=r, column=1).alignment = center
            ws.cell(row=r, column=1).border = grid_border
            ws.cell(row=r, column=2, value=str(grow.get("employee") or "")).font = body_font
            ws.cell(row=r, column=2).border = grid_border
            try:
                tot_v = float(grow.get("leave_days_total") or 0)
            except (TypeError, ValueError):
                tot_v = 0.0
            try:
                el_v = float(grow.get("eleave_days") or 0)
            except (TypeError, ValueError):
                el_v = 0.0
            try:
                gap_v = float(grow.get("gap_days") or 0)
            except (TypeError, ValueError):
                gap_v = 0.0
            ws.cell(row=r, column=3, value=tot_v).font = body_font
            ws.cell(row=r, column=3).alignment = center
            ws.cell(row=r, column=3).border = grid_border
            ws.cell(row=r, column=4, value=el_v).font = body_font
            ws.cell(row=r, column=4).alignment = center
            ws.cell(row=r, column=4).border = grid_border
            gap_cell = ws.cell(row=r, column=5, value=gap_v)
            gap_cell.font = Font(size=10, color="FCA5A5") if gap_v < -0.0001 else body_font
            gap_cell.alignment = center
            gap_cell.border = grid_border

            cells = list(grow.get("cells") or [])
            for j, cell in enumerate(cells):
                col = day0 + j
                v = str(cell.get("code") or "") if cell else ""
                css = str(cell.get("css") or "") if cell else ""
                dc = ws.cell(row=r, column=col, value=v if v else "")
                dc.alignment = center
                dc.border = grid_border
                code_l = v.lower()
                is_nokia = "cell-nokia-approved" in css or (v and v.upper().startswith("A"))
                if is_nokia:
                    bg, fg = _LEAVE_NOKIA_APPROVED_XLSX_STYLE
                    dc.fill = PatternFill("solid", fgColor=bg)
                    dc.font = Font(size=10, bold=True, color=fg)
                elif code_l and code_l in _LEAVE_CODE_XLSX_STYLE:
                    bg, fg = _LEAVE_CODE_XLSX_STYLE[code_l]
                    dc.fill = PatternFill("solid", fgColor=bg)
                    dc.font = Font(size=10, bold=True, color=fg)
                elif j < len(days):
                    wd_l = str(days[j].get("weekday") or "").strip().casefold()
                    if wd_l == "sat":
                        dc.fill = body_fill_sat
                        dc.font = Font(size=10, color="4338CA")
                    elif wd_l == "sun":
                        dc.fill = body_fill_sun
                        dc.font = Font(size=10, color="6D28D9")
                    else:
                        dc.font = body_font
                else:
                    dc.font = body_font

        # ── column widths ───────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 7
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 6
        ws.row_dimensions[2].height = 28
        for i in range(len(days)):
            ws.column_dimensions[get_column_letter(day0 + i)].width = 5.5

        ws.freeze_panes = "C3"

    buf = __import__("io").BytesIO()
    wb.save(buf)
    return buf.getvalue(), None


def build_sprint_leave_tracker_context(
    conn: sqlite3.Connection,
    app: Flask,
    sd: date,
    ed: date,
    roster: Sequence[str],
) -> tuple[list[dict], list[dict]]:
    """
    Leave-tracker style grid for a sprint date window: each roster member × each day,
    same leave / meet-day rules as the monthly worksheet and sticky-board capacity strip.
    Returns (days_meta, grid_rows) where grid_rows items are {employee, leave_days_total, cells} and each
    cell is None or {code, title, status, css} (code = PL/UL/SL/LL/CO… for Excel / HTML).
    """
    roster_t = tuple(roster)
    if not roster_t:
        return [], []
    start_iso = sd.isoformat()
    end_iso = ed.isoformat()
    leaves = list(
        conn.execute(
            """
            SELECT * FROM leave_requests
            WHERE start_date <= ? AND end_date >= ?
              AND status IN ('pending', 'approved')
            ORDER BY employee_name ASC, start_date ASC
            """,
            (end_iso, start_iso),
        )
    )
    day_dec = _load_meet_leave_day_map(app, start_iso, end_iso)

    by_emp: dict[str, list[sqlite3.Row]] = {e: [] for e in roster_t}
    for row in leaves:
        name = row["employee_name"]
        if name in by_emp:
            by_emp[name].append(row)

    reason_l = dict(LEAVE_REASONS)
    dur_l = dict(DURATION_CHOICES)

    days_meta: list[dict] = []
    for d in _daterange_inclusive(sd, ed):
        wd = d.weekday()
        is_weekend = wd >= 5
        days_meta.append(
            {
                "date": d,
                "iso": d.isoformat(),
                "weekday": calendar.day_abbr[wd],
                "day": d.day,
                "is_weekend": is_weekend,
            }
        )

    grid_rows: list[dict] = []
    for emp in roster_t:
        cells: list[dict | None] = []
        leave_days_total = 0.0
        for dm in days_meta:
            d = dm["date"]
            d_iso = d.isoformat()
            overlapping = _overlapping_leaves_for_day(by_emp[emp], d, day_dec)
            cell: dict | None = None
            if overlapping:
                pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
                pool = pending_only or overlapping
                row = max(pool, key=_leave_row_display_tiebreak)
                eff = _effective_leave_status(row, d_iso, day_dec)
                if eff in ("pending", "approved") and _leave_reason_counts_toward_day_units(row["reason"]):
                    dur = (row["duration_type"] or "full").strip()
                    if dur in ("half_am", "half_pm"):
                        leave_days_total += 0.5
                    else:
                        leave_days_total += 1.0
                desc = str(row["description"] or "")
                code = leave_cell_code(row["reason"], row["duration_type"], eff, desc)
                rlab = reason_l.get(row["reason"], row["reason"])
                dlab = dur_l.get(row["duration_type"], row["duration_type"])
                title = f"{rlab} · {dlab} · {eff}"
                css = cell_css_class(row["reason"], eff, desc)
                cell = {"code": code, "title": title, "status": eff, "css": css}
            cells.append(cell)
        grid_rows.append(
            {"employee": emp, "leave_days_total": round(leave_days_total, 2), "cells": cells}
        )
    return days_meta, grid_rows


def build_kanban_leave_worksheet_context(
    app: Flask,
    sprint_id: int,
    sprint_start_iso: str,
    sprint_end_iso: str,
    assignee: str,
) -> dict:
    """Leave strip + capacity for one assignee over sprint dates (same cell rules as manager worksheet)."""
    sd = date.fromisoformat(str(sprint_start_iso).strip()[:10])
    ed = date.fromisoformat(str(sprint_end_iso).strip()[:10])
    start_iso = sd.isoformat()
    end_iso = ed.isoformat()

    conn = get_db(app)
    assigned_row = conn.execute(
        """
        SELECT COALESCE(SUM(estimate_hours), 0) AS s
        FROM scrum_sprint_item
        WHERE sprint_id = ? AND assignee = ?
        """,
        (int(sprint_id), assignee),
    ).fetchone()
    assigned = float(assigned_row["s"] or 0)
    leaves = list(
        conn.execute(
            """
            SELECT * FROM leave_requests
            WHERE employee_name = ?
              AND start_date <= ? AND end_date >= ?
              AND status IN ('pending', 'approved')
            ORDER BY id ASC
            """,
            (assignee, end_iso, start_iso),
        )
    )
    hours_rows = list(
        conn.execute(
            """
            SELECT date(a.created_at) AS d, COALESCE(SUM(a.committed_hours), 0) AS h
            FROM scrum_item_activity a
            INNER JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ? AND i.assignee = ?
              AND date(a.created_at) >= date(?)
              AND date(a.created_at) <= date(?)
            GROUP BY date(a.created_at)
            """,
            (int(sprint_id), assignee, start_iso, end_iso),
        )
    )
    burnt_total_row = conn.execute(
        """
        SELECT COALESCE(SUM(a.committed_hours), 0) AS s
        FROM scrum_item_activity a
        INNER JOIN scrum_sprint_item i ON i.id = a.item_id
        WHERE i.sprint_id = ? AND i.assignee = ?
        """,
        (int(sprint_id), assignee),
    ).fetchone()
    total_burnt_hours = float(burnt_total_row["s"] or 0.0)
    sprint_meta = conn.execute(
        "SELECT team_id FROM scrum_sprint WHERE id = ?", (int(sprint_id),)
    ).fetchone()
    team_id = int(sprint_meta["team_id"]) if sprint_meta else 0
    sticky_rows = list(
        conn.execute(
            """
            SELECT i.id, i.title, i.task_kind, i.kanban_column, i.sort_order,
                   k.label AS kind_label
            FROM scrum_sprint_item i
            LEFT JOIN scrum_team_task_kind k ON k.team_id = ? AND k.code = i.task_kind
            WHERE i.sprint_id = ? AND i.assignee = ?
            ORDER BY i.sort_order ASC, i.id ASC
            """,
            (team_id, int(sprint_id), assignee),
        )
    )
    item_hours_rows = list(
        conn.execute(
            """
            SELECT i.id AS item_id, date(a.created_at) AS d, COALESCE(SUM(a.committed_hours), 0) AS h
            FROM scrum_item_activity a
            INNER JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ? AND i.assignee = ?
              AND date(a.created_at) >= date(?)
              AND date(a.created_at) <= date(?)
            GROUP BY i.id, date(a.created_at)
            """,
            (int(sprint_id), assignee, start_iso, end_iso),
        )
    )
    conn.close()
    hours_by_iso: dict[str, float] = {}
    for r in hours_rows:
        raw_d = r["d"]
        d_key = str(raw_d)[:10] if raw_d is not None else ""
        if d_key:
            hours_by_iso[d_key] = float(r["h"] or 0.0)

    day_dec = _load_meet_leave_day_map(app, start_iso, end_iso)
    reason_l = dict(LEAVE_REASONS)
    dur_l = dict(DURATION_CHOICES)

    days_meta: list[dict] = []
    cells: list[dict | None] = []
    hours_logged: list[str] = []
    gross = 0.0
    leave_debit_total = 0.0
    available_sum = 0.0

    for d in _daterange_inclusive(sd, ed):
        wd = d.weekday()
        is_weekend = wd >= 5
        d_iso = d.isoformat()
        days_meta.append(
            {
                "date": d,
                "iso": d_iso,
                "weekday": calendar.day_abbr[wd],
                "day": d.day,
                "is_weekend": is_weekend,
            }
        )
        base = SCRUM_KANBAN_WEEKDAY_HOURS if wd < 5 else 0.0
        gross += base
        overlapping = _overlapping_leaves_for_day(leaves, d, day_dec)
        debit = 0.0
        cell: dict | None = None
        if overlapping:
            pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
            pool = pending_only or overlapping
            row = max(pool, key=_leave_row_display_tiebreak)
            eff = _effective_leave_status(row, d_iso, day_dec)
            desc = str(row["description"] or "")
            code = leave_cell_code(row["reason"], row["duration_type"], eff, desc)
            rlab = reason_l.get(row["reason"], row["reason"])
            dlab = dur_l.get(row["duration_type"], row["duration_type"])
            title = f"{rlab} · {dlab} · {eff}"
            css = cell_css_class(row["reason"], eff, desc)
            cell = {"code": code, "title": title, "css": css, "status": eff}
            if (
                base > 0
                and eff in ("pending", "approved")
                and _leave_reason_counts_toward_day_units(row["reason"])
            ):
                dur = (row["duration_type"] or "full").strip()
                if dur in ("half_am", "half_pm"):
                    debit = min(base, SCRUM_KANBAN_WEEKDAY_HOURS / 2.0)
                else:
                    debit = min(base, SCRUM_KANBAN_WEEKDAY_HOURS)
        leave_debit_total += debit
        available_sum += max(0.0, base - debit)
        cells.append(cell)
        h_day = float(hours_by_iso.get(d_iso, 0.0))
        hours_logged.append(f"{h_day:.1f}h")

    stretch = assigned > available_sum + SCRUM_HOUR_EPS
    scale_max = max(assigned, available_sum, total_burnt_hours, SCRUM_HOUR_EPS)
    pct_available_bar = round(100.0 * available_sum / scale_max, 1) if scale_max > 0 else 0.0
    pct_assigned_within = round(100.0 * min(assigned, available_sum) / scale_max, 1) if scale_max > 0 else 0.0
    pct_assigned_stretch = (
        round(100.0 * max(0.0, assigned - available_sum) / scale_max, 1) if scale_max > 0 else 0.0
    )
    pct_burnt_bar = round(100.0 * total_burnt_hours / scale_max, 1) if scale_max > 0 else 0.0
    if assigned > SCRUM_HOUR_EPS:
        burn_pct_of_estimate = int(round(100.0 * total_burnt_hours / assigned))
    else:
        burn_pct_of_estimate = None

    item_hours_map: dict[int, dict[str, float]] = {}
    for r in item_hours_rows:
        iid = int(r["item_id"])
        d_key = str(r["d"])[:10] if r["d"] is not None else ""
        if d_key:
            item_hours_map.setdefault(iid, {})[d_key] = float(r["h"] or 0.0)

    task_detail_rows: list[dict[str, Any]] = []
    for it in sticky_rows:
        iid = int(it["id"])
        title = (str(it["title"] or "").strip()) or f"Sticky #{iid}"
        title_short = title if len(title) <= 36 else f"{title[:33]}…"
        col = _normalize_kanban_column(it["kanban_column"] if "kanban_column" in it.keys() else None)
        kind_label = (str(it["kind_label"] or "").strip() if it["kind_label"] else "") or str(
            it["task_kind"] or ""
        ).upper()
        day_hours = [
            f"{float(item_hours_map.get(iid, {}).get(dm['iso'], 0.0)):.1f}h" for dm in days_meta
        ]
        task_detail_rows.append(
            {
                "id": iid,
                "title": title,
                "title_short": title_short,
                "kind_label": kind_label,
                "kanban_column_label": _kanban_column_public_label(col),
                "day_hours": day_hours,
            }
        )

    return {
        "kb_leave_days": days_meta,
        "kb_leave_cells": cells,
        "kb_leave_hours_logged": hours_logged,
        "kb_assignee": assignee,
        "kb_leave_range_label": f"{start_iso} → {end_iso}",
        "kb_capacity_weekday_gross_hours": round(gross, 1),
        "kb_capacity_leave_hours": round(leave_debit_total, 1),
        "kb_capacity_available_hours": round(available_sum, 1),
        "kb_capacity_assigned_hours": round(assigned, 1),
        "kb_capacity_burnt_hours": round(total_burnt_hours, 1),
        "kb_capacity_burn_pct_of_estimate": burn_pct_of_estimate,
        "kb_capacity_stretch": stretch,
        "kb_capacity_scale_max": round(scale_max, 1),
        "kb_capacity_pct_available_bar": pct_available_bar,
        "kb_capacity_pct_assigned_within_bar": pct_assigned_within,
        "kb_capacity_pct_assigned_stretch_bar": pct_assigned_stretch,
        "kb_capacity_pct_burnt_bar": pct_burnt_bar,
        "kb_task_detail_rows": task_detail_rows,
    }


def _available_hours_for_assignee_sprint_window(app: Flask, assignee: str, sd: date, ed: date) -> float:
    """Mon–Fri gross SCRUM_KANBAN_WEEKDAY_HOURS per day minus pending/approved leave (same rules as sticky-board capacity strip)."""
    start_iso = sd.isoformat()
    end_iso = ed.isoformat()
    conn = get_db(app)
    leaves = list(
        conn.execute(
            """
            SELECT * FROM leave_requests
            WHERE employee_name = ?
              AND start_date <= ? AND end_date >= ?
              AND status IN ('pending', 'approved')
            ORDER BY id ASC
            """,
            (assignee, end_iso, start_iso),
        )
    )
    conn.close()
    day_dec = _load_meet_leave_day_map(app, start_iso, end_iso)
    available_sum = 0.0
    for d in _daterange_inclusive(sd, ed):
        wd = d.weekday()
        base = SCRUM_KANBAN_WEEKDAY_HOURS if wd < 5 else 0.0
        d_iso = d.isoformat()
        overlapping = _overlapping_leaves_for_day(leaves, d, day_dec)
        debit = 0.0
        if overlapping:
            pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
            pool = pending_only or overlapping
            row = max(pool, key=_leave_row_display_tiebreak)
            eff = _effective_leave_status(row, d_iso, day_dec)
            if (
                base > 0
                and eff in ("pending", "approved")
                and _leave_reason_counts_toward_day_units(row["reason"])
            ):
                dur = (row["duration_type"] or "full").strip()
                if dur in ("half_am", "half_pm"):
                    debit = min(base, SCRUM_KANBAN_WEEKDAY_HOURS / 2.0)
                else:
                    debit = min(base, SCRUM_KANBAN_WEEKDAY_HOURS)
        available_sum += max(0.0, base - debit)
    return available_sum


def compute_team_sprint_capacity_leave_hours(app: Flask, roster: Sequence[str], sd: date, ed: date) -> float:
    """Total leave-adjusted capacity (h) for all roster names across sprint dates (weekends 0; leave days 0 or half)."""
    return sum(_available_hours_for_assignee_sprint_window(app, emp, sd, ed) for emp in roster)


def compute_team_sprint_gross_weekday_capacity_hours(roster: Sequence[str], sd: date, ed: date) -> float:
    """Mon–Fri gross hours (8h × roster × sprint weekdays) before leave debits; basis for HPPM % columns when absences are included."""
    roster_t = tuple((r or "").strip() for r in roster if (r or "").strip())
    if not roster_t:
        return 0.0
    gross = 0.0
    for _e in roster_t:
        for d in _daterange_inclusive(sd, ed):
            if d.weekday() < 5:
                gross += float(SCRUM_KANBAN_WEEKDAY_HOURS)
    return gross


def compute_team_sprint_leave_debit_hours(app: Flask, roster: Sequence[str], sd: date, ed: date) -> float:
    """Weekday hours lost to pending/approved leave for the roster in the sprint window (gross Mon–Fri − capacity)."""
    roster_t = tuple((r or "").strip() for r in roster if (r or "").strip())
    if not roster_t:
        return 0.0
    gross = compute_team_sprint_gross_weekday_capacity_hours(roster_t, sd, ed)
    cap = compute_team_sprint_capacity_leave_hours(app, roster_t, sd, ed)
    return max(0.0, gross - cap)


def compute_team_sprint_leave_absence_hours_hppm(app: Flask, roster: Sequence[str], sd: date, ed: date) -> float:
    """
    HPPM “Absences” row: total leave across roster in the sprint window as weekday day-units × 8h
    (full day = 1 → 8h, half day = 0.5 → 4h). Uses the same leave rows / meet-day rules as the leave tracker;
    weekend dates are excluded so totals align with Mon–Fri capacity.
    """
    detail, _totals = _leave_tracker_day_rows_for_range(app, roster, sd, ed)
    units = 0.0
    for r in detail:
        try:
            d = date.fromisoformat(str(r["leave_date"])[:10])
        except ValueError:
            continue
        if d.weekday() >= 5:
            continue
        units += float(r.get("day_units") or 0.0)
    return round(units * float(SCRUM_KANBAN_WEEKDAY_HOURS), 4)


def _parse_meet_anchor(raw: str | None) -> date:
    if not raw:
        return date.today()
    try:
        return date.fromisoformat(raw.strip()[:10])
    except ValueError:
        return date.today()


def _meet_window(anchor: date) -> tuple[date, date]:
    return anchor + timedelta(days=-1), anchor + timedelta(days=1)


def build_meet_context(app: Flask, anchor: date, roster: Sequence[str] | None = None) -> dict:
    """Three-day window: anchor−1 … anchor+1 (center column is the anchor day, highlighted)."""
    roster_t = tuple(roster) if roster is not None else EMPLOYEES
    dates_list = [anchor + timedelta(days=k) for k in (-1, 0, 1)]
    start_iso = dates_list[0].isoformat()
    end_iso = dates_list[-1].isoformat()
    conn = get_db(app)
    leaves = list(
        conn.execute(
            """
            SELECT * FROM leave_requests
            WHERE start_date <= ? AND end_date >= ?
              AND status IN ('pending', 'approved')
            ORDER BY id ASC
            """,
            (end_iso, start_iso),
        )
    )
    conn.close()

    day_dec = _load_meet_leave_day_map(app, start_iso, end_iso)

    by_emp: dict[str, list[sqlite3.Row]] = {e: [] for e in roster_t}
    for row in leaves:
        if row["employee_name"] in by_emp:
            by_emp[row["employee_name"]].append(row)

    reason_l = dict(LEAVE_REASONS)
    dur_l = dict(DURATION_CHOICES)
    today = date.today()
    days_out: list[dict] = []
    for d in dates_list:
        wd = d.weekday()
        days_out.append(
            {
                "date": d,
                "iso": d.isoformat(),
                "weekday": calendar.day_abbr[wd],
                "label": d.strftime("%a %d %b"),
                "is_anchor": d == anchor,
                "is_today": d == today,
                "is_weekend": wd >= 5,
            }
        )

    grid_rows: list[dict] = []
    for emp in roster_t:
        cells: list[dict] = []
        for d in dates_list:
            d_iso = d.isoformat()
            overlapping = _overlapping_leaves_for_day(by_emp[emp], d, day_dec)
            cell: dict = {
                "iso": d_iso,
                "has_leave": False,
                "code": "",
                "css": "",
                "title": "",
                "status": "",
                "leave_id": None,
                "is_anchor": d == anchor,
                "is_today": d == today,
                "is_weekend": d.weekday() >= 5,
            }
            if overlapping:
                pending_only = [r for r in overlapping if _is_effectively_pending(r, d_iso, day_dec)]
                pool = pending_only or overlapping
                row = max(pool, key=_leave_row_display_tiebreak)
                eff = _effective_leave_status(row, d_iso, day_dec)
                desc = str(row["description"] or "")
                code = leave_cell_code(row["reason"], row["duration_type"], eff, desc)
                rlab = reason_l.get(row["reason"], row["reason"])
                dlab = dur_l.get(row["duration_type"], row["duration_type"])
                title = f"{rlab} · {dlab} · {eff}"
                css = cell_css_class(row["reason"], eff, desc)
                cell.update(
                    {
                        "has_leave": True,
                        "code": code,
                        "css": css,
                        "title": title,
                        "status": eff,
                        "leave_id": int(row["id"]),
                    }
                )
            cells.append(cell)
        grid_rows.append({"employee": emp, "cells": cells})

    return {
        "anchor": anchor,
        "anchor_label": anchor.strftime("%A %d %B %Y"),
        "prev_anchor": (anchor - timedelta(days=1)).isoformat(),
        "next_anchor": (anchor + timedelta(days=1)).isoformat(),
        "days": days_out,
        "grid_rows": grid_rows,
        "window_start": start_iso,
        "window_end": end_iso,
    }


def _normalize_hub_mode(raw: str | None) -> str:
    m = (raw or "leave").strip().lower()
    return m if m in ("leave", "scrum") else "leave"


def _normalize_scrum_status(raw: str | None) -> str:
    s = (raw or "open").strip().lower()
    return s if s in {c[0] for c in SCRUM_STATUS_CHOICES} else "open"


def _parse_hours_field(raw: str | None) -> float:
    if raw is None:
        return 0.0
    t = str(raw).strip().replace(",", ".")
    if not t:
        return 0.0
    try:
        return max(0.0, min(9999.0, float(t)))
    except ValueError:
        return 0.0


def _estimate_hours_for_backlog_to_do_move(
    data: dict, *, from_col: str, to_col: str
) -> tuple[float | None, str | None]:
    """Backlog → To DO requires a positive estimate_hours in the request body."""
    if from_col != "backlog" or to_col != "do":
        return None, None
    raw = data.get("estimate_hours")
    if isinstance(raw, bool):
        return None, "estimate_required"
    if isinstance(raw, (int, float)):
        est = max(0.0, min(9999.0, float(raw)))
    else:
        est = _parse_hours_field(str(raw if raw is not None else ""))
    if est <= SCRUM_HOUR_EPS:
        return None, "estimate_required"
    return est, None


def _parse_committed_hours_for_kanban_move(
    raw: object,
    *,
    from_col: str,
    to_col: str,
) -> float | None:
    """Hours logged on a column move. Do → In progress: blank / JSON null → NULL in DB (unset), not zero."""
    if from_col != "do" or to_col != "doing":
        return _parse_hours_field(str(raw if raw is not None else ""))
    if raw is None:
        return None
    if isinstance(raw, bool):
        return _parse_hours_field(str(raw))
    if isinstance(raw, (int, float)):
        return max(0.0, min(9999.0, float(raw)))
    t = str(raw).strip().replace(",", ".")
    if not t:
        return None
    try:
        return max(0.0, min(9999.0, float(t)))
    except ValueError:
        return 0.0


def _hours_close(a: float, b: float, eps: float = SCRUM_HOUR_EPS) -> bool:
    return abs(float(a) - float(b)) <= eps


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _working_weekdays(start: date, end: date) -> list[date]:
    """Mon–Fri dates inclusive between start and end (for ideal burndown)."""
    out: list[date] = []
    d = start
    while d <= end:
        if d.weekday() < 5:
            out.append(d)
        d += timedelta(days=1)
    return out


def _parse_optional_int(raw: str | int | None) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw if raw > 0 else None
    try:
        v = int(str(raw).strip())
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


def _sprint_row_for_team(conn: sqlite3.Connection, sprint_id: int, team_id: int) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), int(team_id))
    ).fetchone()


def _sprint_row_is_closed(row: sqlite3.Row | None) -> bool:
    if row is None:
        return False
    if "is_closed" not in row.keys():
        return False
    try:
        return int(row["is_closed"] or 0) != 0
    except (TypeError, ValueError):
        return False


def _sprint_is_closed_by_id(conn: sqlite3.Connection, sprint_id: int) -> bool:
    r = conn.execute("SELECT is_closed FROM scrum_sprint WHERE id = ?", (int(sprint_id),)).fetchone()
    return _sprint_row_is_closed(r)


def _sprint_row_past_end(row: sqlite3.Row | None) -> bool:
    """True after the sprint's inclusive ``end_date`` has fully elapsed in the configured sprint-close timezone."""
    if row is None or "end_date" not in row.keys():
        return False
    return _sprint_inclusive_calendar_window_ended(str(row["end_date"] or ""))


def _maybe_auto_close_scrum_sprint(conn: sqlite3.Connection, sprint_id: int) -> None:
    """Set ``is_closed`` when the sprint window has ended so the sprint is stored as a closed snapshot."""
    r = conn.execute(
        "SELECT COALESCE(is_closed, 0) AS c, end_date FROM scrum_sprint WHERE id = ?",
        (int(sprint_id),),
    ).fetchone()
    if not r or int(r["c"] or 0) != 0:
        return
    if not _sprint_inclusive_calendar_window_ended(str(r["end_date"] or "")):
        return
    ts = _utc_stamp()
    conn.execute(
        "UPDATE scrum_sprint SET is_closed = 1, updated_at = ? WHERE id = ? AND COALESCE(is_closed, 0) = 0",
        (ts, int(sprint_id)),
    )


def _sprint_board_frozen_reason_for_row(row: sqlite3.Row | None) -> str | None:
    """``sprint_closed`` (manual) takes precedence over ``sprint_ended`` (past ``end_date``)."""
    if row is None:
        return None
    if _sprint_row_is_closed(row):
        return "sprint_closed"
    if _sprint_row_past_end(row):
        return "sprint_ended"
    return None


def _sprint_board_frozen_reason(conn: sqlite3.Connection, sprint_id: int) -> str | None:
    _maybe_auto_close_scrum_sprint(conn, int(sprint_id))
    r = conn.execute(
        "SELECT is_closed, end_date FROM scrum_sprint WHERE id = ?", (int(sprint_id),)
    ).fetchone()
    return _sprint_board_frozen_reason_for_row(r)


def _sprint_board_frozen_flash_for_reason(reason: str | None) -> str | None:
    if reason == "sprint_closed":
        return SCRUM_SPRINT_READONLY_FLASH
    if reason == "sprint_ended":
        return SCRUM_SPRINT_AFTER_END_FLASH
    return None


def _sprint_team_page_template_flags(conn: sqlite3.Connection, team_id: int, sprint_id: int) -> dict[str, Any]:
    """UI flags for sprint team / kanban / HPPM (board read-only when closed or past end)."""
    freeze = _sprint_board_frozen_reason(conn, int(sprint_id))
    sprint = conn.execute(
        "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), int(team_id))
    ).fetchone()
    if sprint is None:
        return {
            "sprint_board_readonly": False,
            "sprint_manually_closed": False,
            "sprint_status_label": "OPEN",
            "sprint_freeze_mode": None,
        }
    freeze = _sprint_board_frozen_reason_for_row(sprint)
    closed = _sprint_row_is_closed(sprint)
    past = _sprint_row_past_end(sprint)
    if closed:
        label = "CLOSED"
    elif past:
        label = "ENDED"
    else:
        label = "OPEN"
    return {
        "sprint_board_readonly": freeze is not None,
        "sprint_manually_closed": closed,
        "sprint_status_label": label,
        "sprint_freeze_mode": freeze,
    }


def _sprint_name_exists_for_team(conn: sqlite3.Connection, team_id: int, name: str) -> bool:
    """True if this team already has a sprint with the same name (case-insensitive, trimmed)."""
    n = (name or "").strip()
    if not n:
        return False
    return (
        conn.execute(
            """
            SELECT 1 FROM scrum_sprint
            WHERE team_id = ? AND lower(trim(name)) = lower(trim(?))
            LIMIT 1
            """,
            (int(team_id), n),
        ).fetchone()
        is not None
    )


def _sprint_name_taken_by_other(
    conn: sqlite3.Connection, team_id: int, name: str, exclude_sprint_id: int
) -> bool:
    """True if another sprint on this team (not exclude_sprint_id) already uses this name."""
    n = (name or "").strip()
    if not n:
        return False
    return (
        conn.execute(
            """
            SELECT 1 FROM scrum_sprint
            WHERE team_id = ? AND id != ? AND lower(trim(name)) = lower(trim(?))
            LIMIT 1
            """,
            (int(team_id), int(exclude_sprint_id), n),
        ).fetchone()
        is not None
    )


def _pick_default_sprint_id(conn: sqlite3.Connection, team_id: int, preferred: int | None) -> int | None:
    if preferred is not None:
        row = conn.execute(
            "SELECT id FROM scrum_sprint WHERE id = ? AND team_id = ?",
            (int(preferred), int(team_id)),
        ).fetchone()
        if row:
            return int(row["id"])
    today = date.today().isoformat()
    row = conn.execute(
        """
        SELECT id FROM scrum_sprint
        WHERE team_id = ? AND start_date <= ? AND end_date >= ?
        ORDER BY start_date DESC
        LIMIT 1
        """,
        (int(team_id), today, today),
    ).fetchone()
    if row:
        return int(row["id"])
    row = conn.execute(
        """
        SELECT id FROM scrum_sprint
        WHERE team_id = ?
        ORDER BY end_date DESC, id DESC
        LIMIT 1
        """,
        (int(team_id),),
    ).fetchone()
    return int(row["id"]) if row else None


def _normalize_kanban_column(raw: str | None) -> str:
    c = (raw or "backlog").strip().lower()
    return c if c in SCRUM_KANBAN_COLUMNS else "backlog"


def _normalize_hex_color(raw: str | None, fallback: str = "#64748b") -> str:
    s = (raw or "").strip()
    if _SCRUM_HEX_COLOR.match(s):
        return s.lower()
    return fallback


def _slug_task_kind_code(label: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "-", (label or "").lower()).strip("-")[:20]
    if not base:
        base = "kind"
    return f"{base}-{secrets.token_hex(3)}"


def _ensure_team_task_kinds(conn: sqlite3.Connection, team_id: int) -> None:
    for code, label, color, so in SCRUM_BUILTIN_TASK_KIND_ROWS:
        conn.execute(
            """
            INSERT OR IGNORE INTO scrum_team_task_kind (team_id, code, label, color_hex, sort_order)
            VALUES (?, ?, ?, ?, ?)
            """,
            (int(team_id), code, label, color, so),
        )


def _list_team_task_kinds(conn: sqlite3.Connection, team_id: int) -> list[dict]:
    _ensure_team_task_kinds(conn, team_id)
    codes = tuple(sorted(SCRUM_TASK_KIND_CODES))
    ph = ",".join("?" * len(codes))
    return [
        {"code": r["code"], "label": r["label"], "color_hex": r["color_hex"]}
        for r in conn.execute(
            f"""
            SELECT code, label, color_hex FROM scrum_team_task_kind
            WHERE team_id = ? AND code IN ({ph})
            ORDER BY sort_order ASC, code ASC
            """,
            (int(team_id), *codes),
        )
    ]


def _resolve_task_kind_code(conn: sqlite3.Connection, team_id: int, raw: str | None) -> str:
    """Return a task kind code, or '' when the type is cleared (no kind)."""
    if raw is None:
        return ""
    k = raw.strip().lower()
    if not k or k in ("none", "-", "__none__"):
        return ""
    if k in SCRUM_LEGACY_TASK_KIND_MAP:
        k = SCRUM_LEGACY_TASK_KIND_MAP[k]
    row = conn.execute(
        "SELECT code FROM scrum_team_task_kind WHERE team_id = ? AND code = ?",
        (int(team_id), k),
    ).fetchone()
    if row:
        return str(row["code"])
    return "ndy"


def _coerce_sprint_item_task_kind(conn: sqlite3.Connection, team_id: int, raw: str | None) -> str:
    """Normalize to a built-in task kind code (for persistence and dropdowns)."""
    c = _resolve_task_kind_code(conn, team_id, raw)
    if c in SCRUM_TASK_KIND_CODES:
        return c
    return "ndy"


def _insert_team_task_kind(conn: sqlite3.Connection, team_id: int, label: str, color_hex: str) -> str:
    _ensure_team_task_kinds(conn, team_id)
    safe_label = (label or "").strip()[:120] or "Custom kind"
    hx = _normalize_hex_color(color_hex)
    code = _slug_task_kind_code(safe_label)
    mx = conn.execute(
        "SELECT COALESCE(MAX(sort_order), -1) + 1 AS n FROM scrum_team_task_kind WHERE team_id = ?",
        (int(team_id),),
    ).fetchone()
    sort_order = int(mx["n"] if mx is not None else 0)
    conn.execute(
        """
        INSERT INTO scrum_team_task_kind (team_id, code, label, color_hex, sort_order)
        VALUES (?, ?, ?, ?, ?)
        """,
        (int(team_id), code, safe_label, hx, sort_order),
    )
    return code


def _task_kind_color_for_item(
    conn: sqlite3.Connection, team_id: int, task_kind_code: str, custom_color: str | None
) -> str:
    if not (task_kind_code or "").strip():
        return "#64748b"
    code = _resolve_task_kind_code(conn, team_id, task_kind_code)
    if not code:
        return "#64748b"
    row = conn.execute(
        "SELECT color_hex FROM scrum_team_task_kind WHERE team_id = ? AND code = ?",
        (int(team_id), code),
    ).fetchone()
    if row and (row["color_hex"] or "").strip():
        return _normalize_hex_color(str(row["color_hex"]))
    return "#64748b"


def _status_for_kanban_column(col: str) -> str:
    c = _normalize_kanban_column(col)
    if c == "done":
        return "done"
    if c == "doing":
        return "doing"
    return "open"


def _carry_forward_do_doing_to_new_sprint(
    conn: sqlite3.Connection,
    team_id: int,
    new_sprint_id: int,
    new_start: date,
    ts: str,
) -> int:
    """Copy Do and Doing stickies from the latest prior sprint (ends before new_start) into this sprint's backlog.

    Each copy is a new sticky row: **estimate_hours** matches the source sticky; **Burnt** starts at 0 because
    ``scrum_item_activity`` rows are not copied (only the prior sprint item had the old burn history).
    """
    prev = conn.execute(
        """
        SELECT id FROM scrum_sprint
        WHERE team_id = ? AND id != ? AND end_date < ?
        ORDER BY end_date DESC, id DESC
        LIMIT 1
        """,
        (team_id, new_sprint_id, new_start.isoformat()),
    ).fetchone()
    if not prev:
        return 0
    prev_id = int(prev["id"])
    rows = list(
        conn.execute(
            """
            SELECT id, assignee, title, estimate_hours, notes, dod, task_kind, sticky_color_hex, done_artifacts, sort_order, area
            FROM scrum_sprint_item
            WHERE sprint_id = ? AND lower(trim(kanban_column)) IN ('do', 'doing')
            ORDER BY assignee ASC, sort_order ASC, id ASC
            """,
            (prev_id,),
        )
    )
    if not rows:
        return 0
    st = _status_for_kanban_column("backlog")
    last_so: dict[str, int] = {}
    copied = 0
    for r in rows:
        emp = (r["assignee"] or "").strip()
        last_so[emp] = last_so.get(emp, -1) + 1
        so = last_so[emp]
        title = (r["title"] or "").strip()
        raw_est = float(r["estimate_hours"] or 0)
        est = max(0.0, round(raw_est, 2))
        notes = (r["notes"] or "").strip()[:2000]
        dod = (r["dod"] or "").strip()[:4000]
        tkind = (r["task_kind"] or "").strip() or "task"
        sticky_raw = (r["sticky_color_hex"] or "").strip() if "sticky_color_hex" in r.keys() else ""
        sticky_hex = sticky_raw if sticky_raw and _SCRUM_HEX_COLOR.match(sticky_raw) else None
        da_raw = (r["done_artifacts"] if "done_artifacts" in r.keys() else None) or "[]"
        da_str = str(da_raw).strip()[:8000] or "[]"
        area_s = str(r["area"] or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        conn.execute(
            """
            INSERT INTO scrum_sprint_item
            (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind, sticky_color_hex, done_artifacts, area)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'backlog', ?, ?, ?, ?)
            """,
            (new_sprint_id, emp, title, est, st, notes, dod, so, ts, ts, tkind, sticky_hex, da_str, area_s),
        )
        copied += 1
    return copied


def _activity_calendar_day(raw_ts: str | None) -> date | None:
    if not raw_ts:
        return None
    s = str(raw_ts).strip()[:10]
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def build_sprint_burndown_chart_context(
    conn: sqlite3.Connection, sprint_id: int, start_iso: str, end_iso: str
) -> dict:
    """
    Ideal vs remaining-effort burndown for the whole sprint (all members).
    Remaining = sum per open sticky of max(0, estimate - cumulative Burnt hours) until Done.
    """
    empty = {
        "burndown_has_chart": False,
        "burndown_message": "Add stickies with estimates to see a burndown.",
        "burndown_ideal_d": "",
        "burndown_actual_d": "",
        "burndown_area_d": "",
        "burndown_y_ticks": [],
        "burndown_x_labels": [],
        "burndown_svg_w": 520,
        "burndown_svg_h": 228,
        "burndown_total_hours": 0.0,
        "burndown_remaining_hours": 0.0,
        "burndown_axis_label": "",
        "burndown_y_max": 0.0,
        "burndown_last_day_label": "—",
        "burndown_actual_dots": [],
    }
    try:
        sd = date.fromisoformat(str(start_iso)[:10])
        ed = date.fromisoformat(str(end_iso)[:10])
    except ValueError:
        return empty
    if ed < sd:
        sd, ed = ed, sd
    days: list[date] = list(_daterange_inclusive(sd, ed))
    n = len(days)
    if n < 1:
        return empty

    items = list(
        conn.execute(
            """
            SELECT id, estimate_hours, kanban_column, updated_at
            FROM scrum_sprint_item
            WHERE sprint_id = ?
            """,
            (int(sprint_id),),
        )
    )
    if not items:
        return empty

    ids = [int(r["id"]) for r in items]
    ph = ",".join("?" * len(ids))

    first_done: dict[int, date] = {}
    for r in conn.execute(
        f"""
        SELECT item_id, MIN(created_at) AS ts
        FROM scrum_item_activity
        WHERE item_id IN ({ph}) AND lower(trim(to_column)) = 'done'
        GROUP BY item_id
        """,
        ids,
    ):
        d = _activity_calendar_day(str(r["ts"] or ""))
        if d and d <= ed:
            first_done[int(r["item_id"])] = d

    acts: dict[int, list[tuple[date, float]]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT item_id, committed_hours, created_at
        FROM scrum_item_activity
        WHERE item_id IN ({ph})
        """,
        ids,
    ):
        iid = int(r["item_id"])
        d = _activity_calendar_day(str(r["created_at"] or ""))
        if d is None:
            continue
        if d > ed:
            continue
        h = float(r["committed_hours"] or 0)
        if abs(h) <= SCRUM_HOUR_EPS:
            continue
        acts[iid].append((d, h))
    for iid in acts:
        acts[iid].sort(key=lambda t: t[0])

    def cum_burnt_through(iid: int, d: date) -> float:
        t = 0.0
        for ad, h in acts.get(iid, []):
            if ad <= d:
                t += h
        return t

    def remaining_on_day(d: date) -> float:
        rem = 0.0
        for it in items:
            iid = int(it["id"])
            est = float(it["estimate_hours"] or 0)
            col = _normalize_kanban_column(it["kanban_column"] if it["kanban_column"] else None)
            fd = first_done.get(iid)
            if fd is not None and fd > ed:
                fd = ed
            if fd is None and col == "done":
                upd = _activity_calendar_day(str(it["updated_at"] or ""))
                if upd is not None:
                    fd = min(upd, ed)
                else:
                    fd = ed
            if fd is not None and d >= fd:
                continue
            burnt = cum_burnt_through(iid, d)
            rem += max(0.0, est - burnt)
        return rem

    day0 = days[0]
    remaining_start = remaining_on_day(day0)
    if remaining_start <= SCRUM_HOUR_EPS:
        remaining_start = sum(
            float(it["estimate_hours"] or 0)
            for it in items
            if _normalize_kanban_column(it["kanban_column"] if it["kanban_column"] else None) != "done"
        )
    if remaining_start <= SCRUM_HOUR_EPS:
        remaining_start = sum(float(it["estimate_hours"] or 0) for it in items)
    if remaining_start <= SCRUM_HOUR_EPS:
        return empty

    total_scope = float(remaining_start)
    today = date.today()
    last_actual_day = min(ed, today) if today >= sd else None

    ideal_vals: list[float] = []
    for i in range(n):
        if n <= 1:
            ideal_vals.append(0.0)
        else:
            ideal_vals.append(total_scope * (n - 1 - i) / (n - 1))

    actual_vals: list[float | None] = []
    for d in days:
        if last_actual_day is None or d > last_actual_day:
            actual_vals.append(None)
        else:
            actual_vals.append(remaining_on_day(d))

    actual_numeric = [float(x) for x in actual_vals if x is not None]
    y_max = max(max(ideal_vals), max(actual_numeric) if actual_numeric else 0.0, total_scope, 1.0) * 1.08
    y_max = max(y_max, 1.0)

    W, H = 520, 228
    pad_l, pad_r, pad_t, pad_b = 50, 14, 26, 34
    iw = W - pad_l - pad_r
    ih = H - pad_t - pad_b
    base_y = pad_t + ih

    def x_at(i: int) -> float:
        if n <= 1:
            return pad_l + iw / 2
        return pad_l + i * iw / (n - 1)

    def y_at(val: float) -> float:
        v = max(0.0, min(val, y_max))
        return pad_t + (y_max - v) / y_max * ih

    ideal_pts: list[tuple[float, float]] = [(x_at(i), y_at(ideal_vals[i])) for i in range(n)]
    ideal_d = f"M {ideal_pts[0][0]:.1f},{ideal_pts[0][1]:.1f}"
    if len(ideal_pts) > 1:
        ideal_d += "".join(f" L {x:.1f},{y:.1f}" for x, y in ideal_pts[1:])

    actual_pts: list[tuple[float, float]] = []
    for i, d in enumerate(days):
        av = actual_vals[i]
        if av is None:
            break
        actual_pts.append((x_at(i), y_at(av)))
    actual_d = ""
    if len(actual_pts) >= 2:
        actual_d = f"M {actual_pts[0][0]:.1f},{actual_pts[0][1]:.1f}"
        actual_d += "".join(f" L {x:.1f},{y:.1f}" for x, y in actual_pts[1:])
    elif len(actual_pts) == 1:
        actual_d = f"M {actual_pts[0][0]:.1f},{actual_pts[0][1]:.1f}"

    area_d = ""
    if len(actual_pts) >= 2:
        ax0, _y0 = actual_pts[0]
        ax1, _y1 = actual_pts[-1]
        area_d = f"M {ax0:.1f},{base_y:.1f}"
        for x, y in actual_pts:
            area_d += f" L {x:.1f},{y:.1f}"
        area_d += f" L {ax1:.1f},{base_y:.1f} Z"
    elif len(actual_pts) == 1:
        ax, ay = actual_pts[0]
        area_d = f"M {ax:.1f},{base_y:.1f} L {ax:.1f},{ay:.1f} L {ax + 0.5:.1f},{base_y:.1f} Z"

    y_tick_vals = [0.0, y_max * 0.25, y_max * 0.5, y_max * 0.75, y_max]
    y_ticks_out: list[dict[str, float | str]] = []
    for tv in y_tick_vals:
        y_ticks_out.append({"y": y_at(tv), "label": f"{tv:.0f}h" if tv >= 10 else f"{tv:.1f}h"})

    x_labels: list[dict[str, float | str]] = []
    for idx in sorted({0, n // 2, n - 1}):
        if idx < 0 or idx >= n:
            continue
        d = days[idx]
        x_labels.append({"x": x_at(idx), "label": f"{calendar.month_abbr[d.month]} {d.day}"})

    current_remaining = float(actual_numeric[-1]) if actual_numeric else remaining_on_day(min(ed, max(today, sd)))

    actual_dots = [{"x": round(x, 1), "y": round(y, 1)} for x, y in actual_pts]

    return {
        "burndown_has_chart": True,
        "burndown_message": "",
        "burndown_total_hours": round(total_scope, 1),
        "burndown_remaining_hours": round(current_remaining, 1),
        "burndown_axis_label": "Hours (remaining work)",
        "burndown_svg_w": W,
        "burndown_svg_h": H,
        "burndown_ideal_d": ideal_d,
        "burndown_actual_d": actual_d,
        "burndown_area_d": area_d,
        "burndown_actual_dots": actual_dots,
        "burndown_y_ticks": y_ticks_out,
        "burndown_x_labels": x_labels,
        "burndown_y_max": round(y_max, 1),
        "burndown_last_day_label": f"{calendar.month_abbr[last_actual_day.month]} {last_actual_day.day}"
        if last_actual_day
        else "—",
    }


def build_sprint_team_meter_context(
    *,
    sprint_team_capacity_hours: float | None,
    kind_stack_total_est: float,
    kind_stack_total_burnt: float,
) -> dict[str, float | None]:
    """Sprint hero gauge: team capacity vs total estimate vs total burnt hours."""
    cap = float(sprint_team_capacity_hours or 0)
    est = float(kind_stack_total_est or 0)
    burnt = float(kind_stack_total_burnt or 0)
    stretch = max(0.0, burnt - est)
    mx = max(cap, est, burnt, 0.01)
    pct: float | None
    if est > SCRUM_HOUR_EPS:
        pct = round(100.0 * burnt / est, 1)
    else:
        pct = None
    return {
        "sprint_meter_cap_hours": round(cap, 1),
        "sprint_meter_est_hours": round(est, 1),
        "sprint_meter_burnt_hours": round(burnt, 1),
        "sprint_meter_stretch_hours": round(stretch, 1),
        "sprint_meter_scale_max": round(mx, 1),
        "sprint_meter_burnt_pct": pct,
    }


def build_sprint_task_kind_stack_chart_context(
    conn: sqlite3.Connection,
    team_id: int,
    sprint_id: int,
    *,
    horizontal: bool = False,
    chart_palette: str = "default",
    absence_burnt_hours: float | None = None,
    match_svg_height: int | None = None,
) -> dict:
    """
    Stacked bars per task type: burnt within estimate, burnt beyond estimate, remaining estimate.
    Default is vertical columns (sprint team hero). With ``horizontal=True``, rows are types and
    the hours scale runs left-to-right (HPPM view only).
    ``chart_palette="hppm"`` uses a distinct color set (violet / orange / teal).
    When ``horizontal`` and ``absence_burnt_hours`` is set, a final **Absence** row is drawn using
    the same sprint leave hours as the HPPM summary / leave tracker (est = burnt so it reads as
    burnt ≤ est.). ``match_svg_height`` (e.g. area-stack SVG height) stretches the horizontal chart
    to align with the adjacent HPPM graph.
    """
    pal = (chart_palette or "default").strip().lower()
    if horizontal:
        W, H = (640, 252) if pal == "hppm" else (720, 248)
    else:
        W, H = (520, 228)
    empty: dict = {
        "kind_stack_has_chart": False,
        "kind_stack_message": "No sticky estimates by type yet.",
        "kind_stack_horizontal": horizontal,
        "kind_stack_svg_w": W,
        "kind_stack_svg_h": H,
        "kind_stack_y_ticks": [],
        "kind_stack_x_ticks": [],
        "kind_stack_y_max": 0.0,
        "kind_stack_x_max": 0.0,
        "kind_stack_chart_x0": 0.0,
        "kind_stack_chart_y0": 0.0,
        "kind_stack_chart_y1": 0.0,
        "kind_stack_x_axis_y": 0.0,
        "kind_stack_bars": [],
        "kind_stack_axis_label": "Hours (by task type)",
        "kind_stack_total_est": 0.0,
        "kind_stack_total_burnt": 0.0,
    }
    _ensure_team_task_kinds(conn, team_id)
    kind_meta: dict[str, dict[str, str]] = {}
    for r in conn.execute(
        "SELECT code, label, color_hex FROM scrum_team_task_kind WHERE team_id = ?",
        (int(team_id),),
    ):
        kind_meta[str(r["code"])] = {
            "label": str(r["label"] or r["code"]),
            "color_hex": str(r["color_hex"] or "#94a3b8"),
        }

    kind_est = {code: 0.0 for code in SCRUM_TASK_KIND_CODES}
    kind_com = {code: 0.0 for code in SCRUM_TASK_KIND_CODES}
    for r in conn.execute(
        """
        SELECT task_kind, SUM(estimate_hours) AS h
        FROM scrum_sprint_item
        WHERE sprint_id = ?
        GROUP BY task_kind
        """,
        (int(sprint_id),),
    ).fetchall():
        code = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
        if code in kind_est:
            kind_est[code] += float(r["h"] or 0)
    for r in conn.execute(
        """
        SELECT i.task_kind, SUM(a.committed_hours) AS h
        FROM scrum_item_activity a
        JOIN scrum_sprint_item i ON i.id = a.item_id
        WHERE i.sprint_id = ?
        GROUP BY i.task_kind
        """,
        (int(sprint_id),),
    ).fetchall():
        code = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
        if code in kind_com:
            kind_com[code] += float(r["h"] or 0)

    ordered = SPRINT_TEAM_KIND_STACK_ORDER
    any_tasks = any(
        kind_est[c] > SCRUM_HOUR_EPS or kind_com[c] > SCRUM_HOUR_EPS for c in ordered
    )
    absence_h_raw = (
        max(0.0, float(absence_burnt_hours))
        if horizontal and absence_burnt_hours is not None
        else 0.0
    )
    add_absence_row = bool(horizontal and absence_burnt_hours is not None and absence_h_raw > SCRUM_HOUR_EPS)
    if not any_tasks and not add_absence_row:
        return empty

    tot_est = sum(kind_est[c] for c in ordered)
    tot_burnt = sum(kind_com[c] for c in ordered)
    h_max = max(max(kind_est[c], kind_com[c]) for c in ordered)
    if add_absence_row:
        h_max = max(h_max, absence_h_raw)
    h_max = max(h_max, 1.0) * 1.08
    f_burnt, f_over, f_rem = _stack_segment_fills(pal)

    if horizontal:
        pad_lbl = 56.0
        pad_r, pad_t, pad_b = 12.0, 18.0, 36.0
        nbar = len(ordered) + (1 if add_absence_row else 0)
        if match_svg_height and pal == "hppm":
            try:
                mh = int(match_svg_height)
            except (TypeError, ValueError):
                mh = 0
            if mh >= 200:
                H = mh
        plot_w = W - pad_lbl - pad_r
        row_gap = max(5.0, (H - pad_t - pad_b) * 0.03)
        plot_h = H - pad_t - pad_b
        bar_h = (plot_h - row_gap * (nbar - 1)) / nbar if nbar else plot_h
        chart_x0 = pad_lbl
        chart_y0 = pad_t
        chart_y1 = pad_t + nbar * bar_h + (nbar - 1) * row_gap

        def width_for(hrs: float) -> float:
            return max(0.0, hrs / h_max) * plot_w

        x_tick_vals = [0.0, h_max * 0.25, h_max * 0.5, h_max * 0.75, h_max]
        x_ticks_out: list[dict[str, float | str]] = []
        for tv in x_tick_vals:
            v = max(0.0, min(tv, h_max))
            xx = chart_x0 + (v / h_max) * plot_w if h_max > 0 else chart_x0
            x_ticks_out.append({"x": round(xx, 1), "label": f"{v:.0f}h" if v >= 10 else f"{v:.1f}h"})

        bars_h: list[dict] = []
        for i, code in enumerate(ordered):
            est = float(kind_est.get(code, 0.0))
            burnt = float(kind_com.get(code, 0.0))
            bi = min(est, burnt)
            over = max(0.0, burnt - est)
            rem = max(0.0, est - burnt)
            y_row = pad_t + i * (bar_h + row_gap)
            meta = kind_meta.get(code, {"label": code.upper(), "color_hex": "#94a3b8"})
            label = str(meta["label"])
            short = label
            if code == "process_tools":
                short = "P&T"
            elif len(label) > 10:
                short = label[:9] + "…"

            segs: list[dict[str, float | str]] = []
            xc = chart_x0
            w1 = width_for(bi)
            if w1 > 0.08:
                segs.append(
                    {
                        "x": round(xc, 1),
                        "y": round(y_row, 1),
                        "w": round(w1, 1),
                        "h": round(bar_h, 1),
                        "fill": f_burnt,
                        "kind": "burnt",
                    }
                )
                xc += w1
            w2 = width_for(over)
            if w2 > 0.08:
                segs.append(
                    {
                        "x": round(xc, 1),
                        "y": round(y_row, 1),
                        "w": round(w2, 1),
                        "h": round(bar_h, 1),
                        "fill": f_over,
                        "kind": "over",
                    }
                )
                xc += w2
            w3 = width_for(rem)
            if w3 > 0.08:
                segs.append(
                    {
                        "x": round(xc, 1),
                        "y": round(y_row, 1),
                        "w": round(w3, 1),
                        "h": round(bar_h, 1),
                        "fill": f_rem,
                        "kind": "remaining",
                    }
                )

            bars_h.append(
                {
                    "code": code,
                    "label": label,
                    "label_short": short,
                    "row_label_x": 4.0,
                    "row_label_y": round(y_row + bar_h / 2.0 + 3.5, 1),
                    "est": round(est, 1),
                    "burnt": round(burnt, 1),
                    "title": f"{label}: {est:.1f}h estimated, {burnt:.1f}h burnt (this sprint)",
                    "segments": segs,
                }
            )

        if add_absence_row:
            i = len(ordered)
            est = absence_h_raw
            burnt = absence_h_raw
            bi = min(est, burnt)
            over = max(0.0, burnt - est)
            rem = max(0.0, est - burnt)
            y_row = pad_t + i * (bar_h + row_gap)
            label = "Absence (sprint leave)"
            short = "Absence"
            segs = []
            xc = chart_x0
            w1 = width_for(bi)
            if w1 > 0.08:
                segs.append(
                    {
                        "x": round(xc, 1),
                        "y": round(y_row, 1),
                        "w": round(w1, 1),
                        "h": round(bar_h, 1),
                        "fill": f_burnt,
                        "kind": "burnt",
                    }
                )
                xc += w1
            w2 = width_for(over)
            if w2 > 0.08:
                segs.append(
                    {
                        "x": round(xc, 1),
                        "y": round(y_row, 1),
                        "w": round(w2, 1),
                        "h": round(bar_h, 1),
                        "fill": f_over,
                        "kind": "over",
                    }
                )
                xc += w2
            w3 = width_for(rem)
            if w3 > 0.08:
                segs.append(
                    {
                        "x": round(xc, 1),
                        "y": round(y_row, 1),
                        "w": round(w3, 1),
                        "h": round(bar_h, 1),
                        "fill": f_rem,
                        "kind": "remaining",
                    }
                )
            bars_h.append(
                {
                    "code": "__absence__",
                    "label": label,
                    "label_short": short,
                    "row_label_x": 4.0,
                    "row_label_y": round(y_row + bar_h / 2.0 + 3.5, 1),
                    "est": round(est, 1),
                    "burnt": round(burnt, 1),
                    "title": f"{label}: {est:.1f}h (same total as HPPM Absences / sprint leave tracker)",
                    "segments": segs,
                }
            )

        tot_burnt_out = tot_burnt + (absence_h_raw if add_absence_row else 0.0)
        x_axis_y = round(chart_y1 + 14.0, 1)
        return {
            "kind_stack_has_chart": True,
            "kind_stack_message": "",
            "kind_stack_horizontal": True,
            "kind_stack_svg_w": W,
            "kind_stack_svg_h": H,
            "kind_stack_y_ticks": [],
            "kind_stack_x_ticks": x_ticks_out,
            "kind_stack_y_max": 0.0,
            "kind_stack_x_max": round(h_max, 1),
            "kind_stack_chart_x0": round(chart_x0, 1),
            "kind_stack_chart_y0": round(chart_y0, 1),
            "kind_stack_chart_y1": round(chart_y1, 1),
            "kind_stack_x_axis_y": x_axis_y,
            "kind_stack_bars": bars_h,
            "kind_stack_axis_label": "Hours (est. vs burnt by type)",
            "kind_stack_total_est": round(tot_est, 1),
            "kind_stack_total_burnt": round(tot_burnt_out, 1),
        }

    y_max = h_max
    pad_l, pad_r, pad_t, pad_b = 46, 12, 22, 34
    iw = W - pad_l - pad_r
    ih = H - pad_t - pad_b
    base_y = pad_t + ih
    nbar = len(ordered)
    gap = max(6.0, iw * 0.018)
    bar_w = (iw - gap * (nbar - 1)) / nbar

    def height_for(hrs: float) -> float:
        return max(0.0, hrs / y_max) * ih

    y_tick_vals = [0.0, y_max * 0.25, y_max * 0.5, y_max * 0.75, y_max]
    y_ticks_out: list[dict[str, float | str]] = []
    for tv in y_tick_vals:
        v = max(0.0, min(tv, y_max))
        yy = pad_t + (y_max - v) / y_max * ih
        y_ticks_out.append({"y": yy, "label": f"{tv:.0f}h" if tv >= 10 else f"{tv:.1f}h"})

    bars: list[dict] = []
    for i, code in enumerate(ordered):
        est = float(kind_est.get(code, 0.0))
        burnt = float(kind_com.get(code, 0.0))
        bi = min(est, burnt)
        over = max(0.0, burnt - est)
        rem = max(0.0, est - burnt)
        x = pad_l + i * (bar_w + gap)
        meta = kind_meta.get(code, {"label": code.upper(), "color_hex": "#94a3b8"})
        label = str(meta["label"])
        short = label
        if code == "process_tools":
            short = "P&T"
        elif len(label) > 8:
            short = label[:7] + "…"

        segs: list[dict[str, float | str]] = []
        yc = base_y
        h1 = height_for(bi)
        if h1 > 0.08:
            segs.append(
                {
                    "x": round(x, 1),
                    "y": round(yc - h1, 1),
                    "w": round(bar_w, 1),
                    "h": round(h1, 1),
                    "fill": f_burnt,
                    "kind": "burnt",
                }
            )
            yc -= h1
        h2 = height_for(over)
        if h2 > 0.08:
            segs.append(
                {
                    "x": round(x, 1),
                    "y": round(yc - h2, 1),
                    "w": round(bar_w, 1),
                    "h": round(h2, 1),
                    "fill": f_over,
                    "kind": "over",
                }
            )
            yc -= h2
        h3 = height_for(rem)
        if h3 > 0.08:
            segs.append(
                {
                    "x": round(x, 1),
                    "y": round(yc - h3, 1),
                    "w": round(bar_w, 1),
                    "h": round(h3, 1),
                    "fill": f_rem,
                    "kind": "remaining",
                }
            )

        bars.append(
            {
                "code": code,
                "label": label,
                "label_short": short,
                "x": round(x, 1),
                "bar_w": round(bar_w, 1),
                "cx": round(x + bar_w / 2, 1),
                "est": round(est, 1),
                "burnt": round(burnt, 1),
                "title": f"{label}: {est:.1f}h estimated, {burnt:.1f}h burnt (this sprint)",
                "segments": segs,
            }
        )

    return {
        "kind_stack_has_chart": True,
        "kind_stack_message": "",
        "kind_stack_horizontal": False,
        "kind_stack_svg_w": W,
        "kind_stack_svg_h": H,
        "kind_stack_y_ticks": y_ticks_out,
        "kind_stack_x_ticks": [],
        "kind_stack_y_max": round(y_max, 1),
        "kind_stack_x_max": 0.0,
        "kind_stack_chart_x0": 0.0,
        "kind_stack_chart_y0": 0.0,
        "kind_stack_chart_y1": 0.0,
        "kind_stack_x_axis_y": 0.0,
        "kind_stack_bars": bars,
        "kind_stack_axis_label": "Hours (est. vs burnt by type)",
        "kind_stack_total_est": round(tot_est, 1),
        "kind_stack_total_burnt": round(tot_burnt, 1),
    }


def build_sprint_task_area_stack_chart_context(
    conn: sqlite3.Connection,
    team_id: int,
    sprint_id: int,
    *,
    assignee: str | None = None,
    chart_palette: str = "default",
) -> dict:
    """
    Horizontal stacked bars per task Area: each row is one area; segments show burnt ≤ est,
    burnt over est, and remaining estimate left-to-right on a shared hours scale.
    When ``assignee`` is set, only that person's stickies are included (for per-member boards).
    Areas that differ only by letter case are merged into one bucket (display label picks longer text,
    then a stable tie-break).
    ``chart_palette="hppm"`` uses the same alternate segment colors as the HPPM type stack.
    """
    _default_area_stack_w, _default_area_stack_h = 920, 320
    pal = (chart_palette or "default").strip().lower()
    assignee_s = (assignee or "").strip()
    empty_msg = (
        "No area estimates for this assignee's stickies yet."
        if assignee_s
        else "No sticky estimates by area yet."
    )
    empty: dict = {
        "area_stack_has_chart": False,
        "area_stack_message": empty_msg,
        "area_stack_svg_w": _default_area_stack_w,
        "area_stack_svg_h": _default_area_stack_h,
        "area_stack_y_ticks": [],
        "area_stack_x_ticks": [],
        "area_stack_chart_x0": 120.0,
        "area_stack_chart_y0": 10.0,
        "area_stack_chart_y1": 200.0,
        "area_stack_x_tick_y": 298.0,
        "area_stack_axis_title_y": 312.0,
        "area_stack_pad_r": 16.0,
        "area_stack_band_x": 116.0,
        "area_stack_band_w": 800.0,
        "area_stack_y_max": 0.0,
        "area_stack_bars": [],
        "area_stack_axis_label": "Hours (by task area)",
        "area_stack_total_est": 0.0,
        "area_stack_total_burnt": 0.0,
    }

    def _area_merge_key(raw: str | None) -> str:
        t = (raw or "").strip()
        if not t:
            return SCRUM_AREA_STACK_NO_AREA_LABEL.casefold()
        return t.casefold()

    def _pick_area_display(a: str, b: str) -> str:
        x, y = a.strip(), b.strip()
        if len(x) > len(y):
            return x
        if len(y) > len(x):
            return y
        return min(x, y, key=lambda s: (s.casefold(), s))

    est_sql = """
        SELECT TRIM(COALESCE(i.area, '')) AS ar, SUM(i.estimate_hours) AS h
        FROM scrum_sprint_item i
        INNER JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE i.sprint_id = ? AND s.team_id = ?
    """
    est_params: list[Any] = [int(sprint_id), int(team_id)]
    if assignee_s:
        est_sql += " AND i.assignee = ?"
        est_params.append(assignee_s)
    est_sql += "\n        GROUP BY TRIM(COALESCE(i.area, ''))"

    est_by: dict[str, float] = {}
    display_for: dict[str, str] = {}
    for r in conn.execute(est_sql, est_params).fetchall():
        raw = str(r["ar"]) if r["ar"] is not None else ""
        mk = _area_merge_key(raw)
        h = float(r["h"] or 0)
        disp = raw.strip() or SCRUM_AREA_STACK_NO_AREA_LABEL
        est_by[mk] = est_by.get(mk, 0.0) + h
        if mk in display_for:
            display_for[mk] = _pick_area_display(display_for[mk], disp)
        else:
            display_for[mk] = disp

    burnt_sql = """
        SELECT TRIM(COALESCE(i.area, '')) AS ar, SUM(a.committed_hours) AS h
        FROM scrum_item_activity a
        JOIN scrum_sprint_item i ON i.id = a.item_id
        INNER JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE i.sprint_id = ? AND s.team_id = ?
    """
    burnt_params: list[Any] = [int(sprint_id), int(team_id)]
    if assignee_s:
        burnt_sql += " AND i.assignee = ?"
        burnt_params.append(assignee_s)
    burnt_sql += "\n        GROUP BY TRIM(COALESCE(i.area, ''))"

    burnt_by: dict[str, float] = {}
    for r in conn.execute(burnt_sql, burnt_params).fetchall():
        raw = str(r["ar"]) if r["ar"] is not None else ""
        mk = _area_merge_key(raw)
        h = float(r["h"] or 0)
        disp = raw.strip() or SCRUM_AREA_STACK_NO_AREA_LABEL
        burnt_by[mk] = burnt_by.get(mk, 0.0) + h
        if mk in display_for:
            display_for[mk] = _pick_area_display(display_for[mk], disp)
        else:
            display_for[mk] = disp

    all_keys = sorted(set(est_by) | set(burnt_by), key=lambda k: k.casefold())
    if not all_keys:
        return empty

    def score(k: str) -> float:
        return max(est_by.get(k, 0.0), burnt_by.get(k, 0.0))

    ranked = sorted(all_keys, key=lambda k: (-score(k), k.casefold()))
    max_b = int(SCRUM_AREA_STACK_MAX_BUCKETS)
    if len(ranked) > max_b:
        head = ranked[: max_b - 1]
        tail = ranked[max_b - 1 :]
        other_key = "__other__"
        est_o = sum(est_by.get(k, 0.0) for k in tail)
        burnt_o = sum(burnt_by.get(k, 0.0) for k in tail)
        ordered = head + [other_key]
        est_by[other_key] = est_o
        burnt_by[other_key] = burnt_o
        label_for: dict[str, str] = {k: display_for.get(k, k) for k in head}
        label_for[other_key] = SCRUM_AREA_STACK_OTHER_LABEL
    else:
        ordered = ranked
        label_for = {k: display_for.get(k, k) for k in ordered}

    any_h = any(
        est_by.get(k, 0.0) > SCRUM_HOUR_EPS or burnt_by.get(k, 0.0) > SCRUM_HOUR_EPS for k in ordered
    )
    if not any_h:
        return empty

    tot_est = sum(est_by.get(k, 0.0) for k in ordered)
    tot_burnt = sum(burnt_by.get(k, 0.0) for k in ordered)
    h_max = max(max(est_by.get(k, 0.0), burnt_by.get(k, 0.0)) for k in ordered)
    h_max = max(h_max, 1.0) * 1.08
    f_burnt, f_over, f_rem = _stack_segment_fills(pal)

    nbar = len(ordered)
    pad_r = 16.0
    pad_side = 6.0
    pad_t = 10.0
    pad_b = 36.0
    row_gap = max(3.0, min(7.0, 52.0 / max(nbar, 1)))
    row_h_min = 20.0
    plot_h_needed = nbar * row_h_min + max(0, nbar - 1) * row_gap
    H = int(max(_default_area_stack_h, math.ceil(pad_t + pad_b + plot_h_needed)))
    H = min(H, 720)

    max_raw_label = max(len(label_for.get(k, k)) for k in ordered)
    label_col_w = min(240.0, max(104.0, min(max_raw_label * 6.8, 280.0)))
    plot_x0 = pad_side + label_col_w
    min_plot_w = 400.0
    W = int(max(_default_area_stack_w, math.ceil(plot_x0 + min_plot_w + pad_r)))
    W = min(W, 5200)
    plot_w = W - plot_x0 - pad_r
    plot_y0 = pad_t
    plot_y1 = H - pad_b
    plot_h_total = plot_y1 - plot_y0
    row_h = (plot_h_total - (nbar - 1) * row_gap) / nbar if nbar > 0 else plot_h_total
    bar_h = max(7.0, min(16.0, row_h * 0.58))

    def width_for(hrs: float) -> float:
        return max(0.0, hrs / h_max) * plot_w

    x_tick_vals = [0.0, h_max * 0.25, h_max * 0.5, h_max * 0.75, h_max]
    x_ticks_out: list[dict[str, float | str]] = []
    for tv in x_tick_vals:
        v = max(0.0, min(tv, h_max))
        xv = plot_x0 + v / h_max * plot_w
        x_ticks_out.append({"x": round(xv, 1), "label": f"{v:.0f}h" if v >= 10 else f"{v:.1f}h"})

    max_label_chars = max(10, min(48, int((label_col_w - 10) / 6.2)))

    x_tick_y = H - 22.0
    axis_title_y = H - 6.0

    bars: list[dict] = []
    for i, key in enumerate(ordered):
        est = float(est_by.get(key, 0.0))
        burnt = float(burnt_by.get(key, 0.0))
        bi = min(est, burnt)
        over = max(0.0, burnt - est)
        rem = max(0.0, est - burnt)
        row_top = pad_t + i * (row_h + row_gap)
        y_mid = row_top + row_h / 2.0
        bar_y = y_mid - bar_h / 2.0
        full_label = label_for.get(key, key)
        label = full_label
        if len(label) > max_label_chars:
            label = label[: max_label_chars - 1] + "…"

        segs: list[dict[str, float | str]] = []
        xc = plot_x0
        w1 = width_for(bi)
        if w1 > 0.08:
            segs.append(
                {
                    "x": round(xc, 1),
                    "y": round(bar_y, 1),
                    "w": round(w1, 1),
                    "h": round(bar_h, 1),
                    "fill": f_burnt,
                    "kind": "burnt",
                }
            )
            xc += w1
        w2 = width_for(over)
        if w2 > 0.08:
            segs.append(
                {
                    "x": round(xc, 1),
                    "y": round(bar_y, 1),
                    "w": round(w2, 1),
                    "h": round(bar_h, 1),
                    "fill": f_over,
                    "kind": "over",
                }
            )
            xc += w2
        w3 = width_for(rem)
        if w3 > 0.08:
            segs.append(
                {
                    "x": round(xc, 1),
                    "y": round(bar_y, 1),
                    "w": round(w3, 1),
                    "h": round(bar_h, 1),
                    "fill": f_rem,
                    "kind": "remaining",
                }
            )

        bars.append(
            {
                "code": key,
                "label": label,
                "label_short": label,
                "label_x": round(plot_x0 - 8.0, 1),
                "label_y": round(y_mid + 4.0, 1),
                "row_y0": round(row_top, 1),
                "row_y1": round(row_top + row_h, 1),
                "row_idx": i,
                "est": round(est, 1),
                "burnt": round(burnt, 1),
                "title": f"{full_label}: {est:.1f}h estimated, {burnt:.1f}h burnt (this sprint)",
                "segments": segs,
            }
        )

    return {
        "area_stack_has_chart": True,
        "area_stack_message": "",
        "area_stack_svg_w": W,
        "area_stack_svg_h": H,
        "area_stack_pad_r": pad_r,
        "area_stack_band_x": round(plot_x0 - 4.0, 1),
        "area_stack_band_w": round(W - pad_r - (plot_x0 - 4.0), 1),
        "area_stack_chart_x0": round(plot_x0, 1),
        "area_stack_chart_y0": round(plot_y0, 1),
        "area_stack_chart_y1": round(plot_y1, 1),
        "area_stack_x_tick_y": round(x_tick_y, 1),
        "area_stack_axis_title_y": round(axis_title_y, 1),
        "area_stack_x_ticks": x_ticks_out,
        "area_stack_y_max": round(h_max, 1),
        "area_stack_bars": bars,
        "area_stack_axis_label": "Hours (est. vs burnt by area)",
        "area_stack_total_est": round(tot_est, 1),
        "area_stack_total_burnt": round(tot_burnt, 1),
    }


def build_hppm_sprint_page_extra_context(
    conn: sqlite3.Connection,
    app: Flask,
    *,
    team_id: int,
    sprint_id: int,
    roster: Sequence[str],
    sprint_start: str,
    sprint_end: str,
) -> dict[str, Any]:
    """HPPM summary by work type (estimates, burnt, absences; % columns = each row's share of table total for that measure), and per-sticky rows grouped by area."""
    team_row = conn.execute("SELECT name FROM teams WHERE id = ?", (int(team_id),)).fetchone()
    pool_name = str(team_row["name"] or "") if team_row else ""
    try:
        sd = date.fromisoformat(str(sprint_start)[:10])
        ed = date.fromisoformat(str(sprint_end)[:10])
    except ValueError:
        sd = ed = date.today()
    if sd > ed:
        sd, ed = ed, sd
    team_cap_net = float(compute_team_sprint_capacity_leave_hours(app, roster, sd, ed))
    team_cap_gross = float(compute_team_sprint_gross_weekday_capacity_hours(roster, sd, ed))
    absence_h = round(compute_team_sprint_leave_absence_hours_hppm(app, roster, sd, ed), 4)

    def pct_of_column_total(part: float, whole: float) -> float | None:
        if whole <= SCRUM_HOUR_EPS:
            return None
        return round(100.0 * float(part) / whole, 2)

    def pct_burnt_vs_estimate(burnt: float, estimate: float) -> float | None:
        """100 × burnt ÷ estimate; None if estimate ~0 but burnt > 0."""
        est = float(estimate)
        br = float(burnt)
        if est <= SCRUM_HOUR_EPS:
            if br <= SCRUM_HOUR_EPS:
                return 0.0
            return None
        return round(100.0 * br / est, 2)

    est_by: dict[str, float] = {c: 0.0 for c in SCRUM_TASK_KIND_CODES}
    burnt_by: dict[str, float] = {c: 0.0 for c in SCRUM_TASK_KIND_CODES}
    _ensure_team_task_kinds(conn, team_id)
    for r in conn.execute(
        """
        SELECT task_kind, SUM(estimate_hours) AS h
        FROM scrum_sprint_item
        WHERE sprint_id = ?
        GROUP BY task_kind
        """,
        (int(sprint_id),),
    ):
        code = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
        if code in est_by:
            est_by[code] += float(r["h"] or 0)
    for r in conn.execute(
        """
        SELECT i.task_kind, SUM(a.committed_hours) AS h
        FROM scrum_item_activity a
        JOIN scrum_sprint_item i ON i.id = a.item_id
        WHERE i.sprint_id = ?
        GROUP BY i.task_kind
        """,
        (int(sprint_id),),
    ):
        code = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
        if code in burnt_by:
            burnt_by[code] += float(r["h"] or 0)

    summary_rows: list[dict[str, Any]] = []
    total_est_summary = 0.0
    total_burnt_summary = 0.0
    for code, label in SCRUM_HPPM_SUMMARY_ROWS:
        if code is None:
            ah = float(absence_h)
            total_burnt_summary += ah
            total_est_summary += ah
            summary_rows.append(
                {
                    "work_type": label,
                    "estimate_hours": ah,
                    "burnt_hours": ah,
                    "is_absence": True,
                }
            )
        else:
            eh = float(est_by.get(code, 0.0))
            bh = float(burnt_by.get(code, 0.0))
            total_est_summary += eh
            total_burnt_summary += bh
            summary_rows.append(
                {
                    "work_type": label,
                    "estimate_hours": eh,
                    "burnt_hours": bh,
                    "is_absence": False,
                }
            )

    te = total_est_summary
    tb = total_burnt_summary
    for row in summary_rows:
        eh = float(row["estimate_hours"])
        bh = float(row["burnt_hours"])
        row["pct_capacity_estimate"] = pct_of_column_total(eh, te)
        row["pct_capacity_burnt"] = pct_of_column_total(bh, tb)

    summary_totals: dict[str, Any] = {
        "estimate_hours": round(te, 4),
        "burnt_hours": round(tb, 4),
        "pct_capacity_estimate": 100.0 if te > SCRUM_HOUR_EPS else None,
        "pct_capacity_burnt": 100.0 if tb > SCRUM_HOUR_EPS else None,
    }

    task_rows: list[dict[str, Any]] = []
    for r in conn.execute(
        """
        SELECT i.title, i.assignee, i.task_kind, i.estimate_hours,
               TRIM(COALESCE(i.area, '')) AS area,
               COALESCE(k.label, i.task_kind) AS kind_label,
               COALESCE(
                 (SELECT SUM(committed_hours) FROM scrum_item_activity WHERE item_id = i.id),
                 0
               ) AS total_burnt_hours
        FROM scrum_sprint_item i
        LEFT JOIN scrum_team_task_kind k ON k.team_id = ? AND k.code = i.task_kind
        WHERE i.sprint_id = ?
        ORDER BY i.assignee COLLATE NOCASE, i.sort_order, i.id
        """,
        (int(team_id), int(sprint_id)),
    ):
        rk = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
        ar_raw = str(r["area"] if r["area"] is not None else "").strip()
        eh = float(r["estimate_hours"] or 0)
        bh = float(r["total_burnt_hours"] or 0)
        task_rows.append(
            {
                "work_type": SCRUM_HPPM_LABEL_BY_CODE.get(rk, str(r["kind_label"] or rk)),
                "area_type": ar_raw if ar_raw else SCRUM_AREA_STACK_NO_AREA_LABEL,
                "title": str(r["title"] or ""),
                "assignee": str(r["assignee"] or ""),
                "estimate_hours": eh,
                "burnt_hours": bh,
                "burnt_pct": pct_burnt_vs_estimate(bh, eh),
            }
        )

    groups: dict[str, list[dict[str, Any]]] = {}
    for row in task_rows:
        ak = str(row.get("area_type") or "")
        groups.setdefault(ak, []).append(row)
    ordered_task_rows: list[dict[str, Any]] = []
    for area_key in sorted(
        groups.keys(),
        key=lambda ak: (
            -sum(float(r["burnt_hours"]) for r in groups[ak]),
            (ak or "").lower(),
        ),
    ):
        chunk = sorted(
            groups[area_key],
            key=lambda r: (
                -float(r.get("burnt_hours") or 0),
                str(r.get("assignee") or "").lower(),
                str(r.get("title") or "").lower(),
            ),
        )
        ordered_task_rows.extend(chunk)
    task_rows = ordered_task_rows

    i = 0
    while i < len(task_rows):
        j = i + 1
        area_i = str(task_rows[i].get("area_type") or "")
        while j < len(task_rows) and str(task_rows[j].get("area_type") or "") == area_i:
            j += 1
        blk = task_rows[i:j]
        total_burnt = sum(float(r["burnt_hours"]) for r in blk)
        for k, r in enumerate(blk):
            r["area_group_burnt_total"] = float(total_burnt)
            r["area_group_first"] = k == 0
            r["area_total_rowspan"] = len(blk) if k == 0 else None
        i = j

    _hppm_area_band_n = 8
    _unique_areas_order: list[str] = []
    _seen_area: set[str] = set()
    for _row in task_rows:
        _ak = str(_row.get("area_type") or "")
        if _ak not in _seen_area:
            _seen_area.add(_ak)
            _unique_areas_order.append(_ak)
    _area_to_band = {a: i % _hppm_area_band_n for i, a in enumerate(_unique_areas_order)}
    for _row in task_rows:
        _row["area_band_idx"] = int(_area_to_band.get(str(_row.get("area_type") or ""), 0))

    return {
        "hppm_pool_name": pool_name,
        "hppm_summary_rows": summary_rows,
        "hppm_summary_totals": summary_totals,
        "hppm_task_rows": task_rows,
        "hppm_absence_hours": absence_h,
        "hppm_team_capacity_gross_hours": round(team_cap_gross, 2) if team_cap_gross > SCRUM_HOUR_EPS else None,
        "hppm_team_capacity_net_hours": round(team_cap_net, 2) if team_cap_net > SCRUM_HOUR_EPS else None,
    }


def _sanitize_xlsx_base_filename(name: str) -> str:
    s = (name or "").strip() or "sprint"
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", s)
    s = re.sub(r"\s+", "_", s).strip("._")
    return (s[:96] if s else "sprint")


def build_sprint_export_xlsx_bytes(
    conn: sqlite3.Connection,
    team_id: int,
    sprint_id: int,
    *,
    app: Flask | None = None,
    manager_roster: Sequence[str] | None = None,
) -> tuple[bytes | None, str]:
    """
    Build a multi-sheet .xlsx: Summary (metrics A:B; leave tracker column D; planned capacity % =
    (sum estimates + Sprint leaves) ÷ sprint capacity; free capacity hours; member goals; task kinds),
    PNG snapshot tab, SprintStatus (stickies +
    day-wise hours like FB2610), HPPM (summary by work type + all stickies, same as HPPM web view),
    Daily tasks Details (per-member metrics with team + day columns; compact sticky rows; daily task rows),
    and Appriciation (one row per appreciation; columns C, D, G, H only — author, comment, title, assignee).
    On failure returns (None, flash_message). On success returns (bytes, download_filename).
    """
    try:
        from collections import defaultdict

        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "Excel export requires openpyxl (install dependencies)."

    sprint = conn.execute(
        "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), int(team_id))
    ).fetchone()
    if not sprint:
        return None, "Sprint not found."

    team_row = conn.execute("SELECT name FROM teams WHERE id = ?", (int(team_id),)).fetchone()
    team_name = str(team_row["name"]) if team_row else ""

    items = list(
        conn.execute(
            """
            SELECT
              i.id,
              i.assignee,
              i.title,
              i.area,
              i.estimate_hours,
              i.status,
              i.kanban_column,
              i.task_kind,
              COALESCE(k.label, i.task_kind) AS task_kind_label,
              i.sticky_color_hex,
              i.notes,
              i.dod,
              i.done_artifacts,
              i.sort_order,
              i.created_at,
              i.updated_at,
              COALESCE(
                (SELECT SUM(committed_hours) FROM scrum_item_activity WHERE item_id = i.id),
                0
              ) AS total_burnt_hours
            FROM scrum_sprint_item i
            LEFT JOIN scrum_team_task_kind k ON k.team_id = ? AND k.code = i.task_kind
            WHERE i.sprint_id = ?
            ORDER BY i.assignee COLLATE NOCASE, i.sort_order, i.id
            """,
            (int(team_id), int(sprint_id)),
        )
    )

    activity = list(
        conn.execute(
            """
            SELECT
              a.id AS activity_id,
              a.item_id,
              i.title AS sticky_title,
              i.assignee,
              a.created_at,
              a.committed_hours,
              a.from_column,
              a.to_column,
              a.body AS note
            FROM scrum_item_activity a
            JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ?
            ORDER BY a.created_at ASC, a.id ASC
            """,
            (int(sprint_id),),
        )
    )

    appreciation_export_rows = list(
        conn.execute(
            """
            SELECT
              a.author AS appreciation_author,
              a.comment AS appreciation_comment,
              i.title,
              i.assignee
            FROM scrum_item_appreciation a
            JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ?
            ORDER BY a.created_at ASC, a.id ASC
            """,
            (int(sprint_id),),
        )
    )

    # Fetch checklist rows for all sprint items
    all_item_ids_export = [int(r["id"]) for r in items]
    checklist_export_map = _fetch_item_checklist_map(conn, all_item_ids_export) if all_item_ids_export else {}

    try:
        sd_date_export = date.fromisoformat(str(sprint["start_date"])[:10])
        ed_date_export = date.fromisoformat(str(sprint["end_date"])[:10])
    except ValueError:
        sd_date_export = ed_date_export = date.today()
    sprint_calendar_dates: list[date] = []
    _walk = sd_date_export
    while _walk <= ed_date_export:
        sprint_calendar_dates.append(_walk)
        _walk += timedelta(days=1)
    sprint_day_set = frozenset(sprint_calendar_dates)

    def _activity_day(ts: object) -> date | None:
        if ts is None:
            return None
        s = str(ts).strip()
        if len(s) < 10 or s[4] != "-" or s[7] != "-":
            return None
        try:
            d = date.fromisoformat(s[:10])
        except ValueError:
            return None
        return d if d in sprint_day_set else None

    item_day_hours: dict[int, dict[date, float]] = defaultdict(lambda: defaultdict(float))
    item_day_snips: dict[int, dict[date, list[str]]] = defaultdict(lambda: defaultdict(list))
    assignee_day_hours: dict[str, dict[date, float]] = defaultdict(lambda: defaultdict(float))
    for ar in activity:
        ad = _activity_day(ar["created_at"])
        if ad is None:
            continue
        iid = int(ar["item_id"])
        h = float(ar["committed_hours"] or 0)
        item_day_hours[iid][ad] += h
        em = str(ar["assignee"] or "").strip()
        if em:
            assignee_day_hours[em][ad] += h
        fc = str(ar["from_column"] or "").strip()
        tc = str(ar["to_column"] or "").strip()
        body = str(ar["note"] or "").strip()
        # Standup note: show the note body as the primary text; include status transition as context
        if body:
            # Truncate long notes to keep cell readable
            display_body = body if len(body) <= 200 else body[:197] + "…"
            snip = f"[{fc}→{tc}] {display_body}"
        else:
            snip = f"[{fc}→{tc}] +{h:.1f}h"
        item_day_snips[iid][ad].append(snip)

    member_goals = list(
        conn.execute(
            """
            SELECT employee_name, goal, updated_at
            FROM scrum_sprint_member_goal
            WHERE sprint_id = ?
            ORDER BY employee_name COLLATE NOCASE
            """,
            (int(sprint_id),),
        )
    )

    daily_rows = list(
        conn.execute(
            """
            SELECT
              d.id,
              d.work_date,
              d.assignee,
              d.title,
              d.planned_hours,
              d.committed_hours,
              d.actual_hours,
              d.status,
              d.notes,
              d.sort_order,
              d.created_at,
              d.updated_at,
              d.sprint_item_id,
              COALESCE(i.title, '') AS linked_sticky_title
            FROM scrum_daily_task d
            LEFT JOIN scrum_sprint_item i
              ON i.id = d.sprint_item_id AND i.sprint_id = d.sprint_id
            WHERE d.sprint_id = ?
            ORDER BY d.work_date, d.assignee COLLATE NOCASE, d.sort_order, d.id
            """,
            (int(sprint_id),),
        )
    )

    roster_xlsx: list[str]
    if manager_roster:
        roster_xlsx = [str(x).strip() for x in manager_roster if str(x).strip()]
    else:
        roster_xlsx = [
            str(row["employee_name"])
            for row in conn.execute(
                """
                SELECT employee_name FROM team_roster
                WHERE team_id = ?
                ORDER BY sort_order, employee_name COLLATE NOCASE
                """,
                (int(team_id),),
            )
        ]
    if not roster_xlsx:
        roster_xlsx = sorted(
            {str(r["assignee"] or "").strip() for r in items if (r["assignee"] or "").strip()},
            key=str.casefold,
        )
    member_overview = _sprint_team_overview_rows(app, conn, team_id, sprint_id, roster_xlsx)

    task_kinds = list(
        conn.execute(
            """
            SELECT code, label, color_hex, sort_order
            FROM scrum_team_task_kind
            WHERE team_id = ?
            ORDER BY sort_order, id ASC
            """,
            (int(team_id),),
        )
    )

    tot_est = sum(float(r["estimate_hours"] or 0) for r in items)
    tot_burnt = sum(float(r["total_burnt_hours"] or 0) for r in items)
    done_n = sum(
        1
        for r in items
        if _normalize_kanban_column(r["kanban_column"] if r["kanban_column"] else None) == "done"
    )
    open_n = len(items) - done_n

    col_counts: dict[str, int] = {}
    for r in items:
        kc = _normalize_kanban_column(r["kanban_column"] if r["kanban_column"] else None) or "unknown"
        col_counts[kc] = col_counts.get(kc, 0) + 1

    raw_cap = sprint["team_capacity_hours"] if "team_capacity_hours" in sprint.keys() else None
    total_sprint_capacity_h: float | None = float(raw_cap) if raw_cap is not None else None
    if total_sprint_capacity_h is None and app is not None and roster_xlsx:
        try:
            sd_cap = date.fromisoformat(str(sprint["start_date"])[:10])
            ed_cap = date.fromisoformat(str(sprint["end_date"])[:10])
            total_sprint_capacity_h = round(
                compute_team_sprint_capacity_leave_hours(app, roster_xlsx, sd_cap, ed_cap), 2
            )
        except ValueError:
            total_sprint_capacity_h = None

    sd_iso_sprint = str(sprint["start_date"])[:10]
    ed_iso_sprint = str(sprint["end_date"])[:10]
    bd_ctx_export = build_sprint_burndown_chart_context(conn, sprint_id, sd_iso_sprint, ed_iso_sprint)
    png_sprint_snapshot: bytes | None = None
    try:
        from sprint_hub_snapshot_png import build_sprint_hub_snapshot_png_bytes

        png_sprint_snapshot = build_sprint_hub_snapshot_png_bytes(
            team_name=team_name,
            sprint_name=str(sprint["name"]),
            sprint_start=sd_iso_sprint,
            sprint_end=ed_iso_sprint,
            capacity_h=total_sprint_capacity_h,
            bd_ctx=bd_ctx_export,
            member_rows=member_overview,
        )
    except Exception:
        png_sprint_snapshot = None

    sprint_leaves_h = 0.0
    if app is not None and roster_xlsx:
        try:
            sprint_leaves_h = float(
                compute_team_sprint_leave_absence_hours_hppm(app, roster_xlsx, sd_date_export, ed_date_export)
            )
        except Exception:
            sprint_leaves_h = 0.0
    sprint_leaves_h = max(0.0, round(sprint_leaves_h, 4))

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="0F172A")
    hdr_font = Font(color="F8FAFC", bold=True, size=11)
    sub_font = Font(bold=True, size=10, color="1E293B")
    title_font = Font(color="F8FAFC", bold=True, size=14)
    title_fill = PatternFill("solid", fgColor="1E3A5F")
    subtitle_font = Font(size=11, color="334155", italic=True)
    thin = Side(style="thin", color="94A3B8")
    hdr_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    pct_fill = PatternFill("solid", fgColor="FFF59D")   # light yellow for % cells
    pct_font = Font(color="5D4037", bold=False, size=10) # dark-brown text on yellow

    def style_header_row(ws, row_idx: int, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = hdr_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autofit(ws: object, max_w: float = 52.0) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            maxlen = 10
            for cell in col:
                if cell.value is None:
                    continue
                maxlen = max(maxlen, min(240, len(str(cell.value))))
            ws.column_dimensions[letter].width = min(max_w, max(10.0, maxlen * 1.05 + 1.0))

    def apply_data_grid(ws: object, min_row: int, max_row: int, ncols: int) -> None:
        if max_row < min_row:
            return
        for r in range(min_row, max_row + 1):
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).border = grid_border

    def apply_data_grid_region(ws: object, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
        if max_row < min_row or max_col < min_col:
            return
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = grid_border

    # --- Summary ---
    ws0 = wb.active
    ws0.title = "Summary"
    exported_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    sd_s = str(sprint["start_date"])[:10]
    ed_s = str(sprint["end_date"])[:10]
    cap_display = f"{total_sprint_capacity_h:.2f}" if total_sprint_capacity_h is not None else "—"

    pairs: list[tuple[str, str]] = [
        ("Sprint name", str(sprint["name"])),
        ("Total Sprint Capacity (h)", cap_display),
        ("Start date", sd_s),
        ("End date", ed_s),
        ("Team", team_name),
        ("Sprint ID", str(int(sprint["id"]))),
        ("Exported at", exported_at),
        ("", ""),
        ("Stickies (count)", str(len(items))),
        ("Done stickies", str(done_n)),
        ("Open / not-done stickies", str(open_n)),
        ("Sum of estimates (h)", f"{tot_est:.2f}"),
        ("Sum of Burnt hours (h)", f"{tot_burnt:.2f}"),
        ("Sprint leaves (h)", f"{sprint_leaves_h:.2f}"),
    ]
    if total_sprint_capacity_h is not None and float(total_sprint_capacity_h) > SCRUM_HOUR_EPS:
        capv = float(total_sprint_capacity_h)
        planned_use_h = tot_est + sprint_leaves_h
        pairs.append(
            (
                "Planned capacity ((sum estimates + Sprint leaves) ÷ sprint capacity, %)",
                f"{100.0 * planned_use_h / capv:.1f}%",
            )
        )
        pairs.append(
            (
                "Free sprint capacity (sprint capacity − sum estimates − Sprint leaves, h)",
                f"{capv - planned_use_h:.2f}",
            )
        )
    else:
        pairs.append(("Planned capacity ((sum estimates + Sprint leaves) ÷ sprint capacity, %)", "—"))
        pairs.append(("Free sprint capacity (sprint capacity − sum estimates − Sprint leaves, h)", "—"))
    for col_name, n in sorted(col_counts.items()):
        pairs.append((f"Stickies in column: {col_name}", str(n)))
    while 4 + len(pairs) - 1 > 22 and len(pairs) > 8:
        tail = pairs[-1][0] if pairs else ""
        if isinstance(tail, str) and tail.startswith("Stickies in column:"):
            pairs.pop()
        else:
            break

    ws0.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws0.row_dimensions[1].height = 30
    ban = ws0.cell(row=1, column=1, value=f"Sprint export — {str(sprint['name'])}")
    ban.font = title_font
    ban.fill = title_fill
    ban.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws0.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    sub = ws0.cell(row=2, column=1, value=f"{team_name} · {sd_s} → {ed_s}")
    sub.font = subtitle_font
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    summary_start_row = 4
    for i, (k, v) in enumerate(pairs, start=summary_start_row):
        ws0.cell(row=i, column=1, value=k).font = sub_font if k else Font()
        c2 = ws0.cell(row=i, column=2, value=v)
        c2.alignment = wrap
    ws0.column_dimensions["A"].width = 42
    ws0.column_dimensions["B"].width = 72
    last_summary = summary_start_row + len(pairs) - 1
    apply_data_grid(ws0, summary_start_row, last_summary, 2)

    # Leave tracker grid in columns D onward, row 1+ (alongside A:B sprint metrics).
    weekend_fill_lv = PatternFill("solid", fgColor="312E81")
    last_leave_row = last_summary
    try:
        sdd = date.fromisoformat(sd_s)
        edd = date.fromisoformat(ed_s)
    except ValueError:
        sdd = edd = date.today()
    leave_first_col = 4  # Column D
    weekday_row_lv = 1
    date_row_lv = 2
    member_start_lv = 3
    if app is not None and roster_xlsx:
        days_lv, grid_lv = build_sprint_leave_tracker_context(conn, app, sdd, edd, roster_xlsx)
        if days_lv and grid_lv:
            ncols_lv = len(days_lv) + 1
            last_col_lv = leave_first_col + ncols_lv - 1
            corner = ws0.cell(row=weekday_row_lv, column=leave_first_col, value="")
            corner.border = grid_border
            corner.fill = hdr_fill
            for idx, dm in enumerate(days_lv):
                col = leave_first_col + 1 + idx
                ccell = ws0.cell(row=weekday_row_lv, column=col, value=str(dm["weekday"]).upper())
                ccell.font = hdr_font
                ccell.fill = hdr_fill if not dm["is_weekend"] else weekend_fill_lv
                ccell.border = hdr_border
                ccell.alignment = Alignment(horizontal="center", vertical="center")
            name_hdr = ws0.cell(row=date_row_lv, column=leave_first_col, value="NAME")
            name_hdr.font = hdr_font
            name_hdr.fill = hdr_fill
            name_hdr.border = hdr_border
            name_hdr.alignment = Alignment(horizontal="center", vertical="center")
            for idx, dm in enumerate(days_lv):
                col = leave_first_col + 1 + idx
                ccell = ws0.cell(row=date_row_lv, column=col, value=dm["date"].strftime("%m/%d"))
                ccell.font = hdr_font
                ccell.fill = hdr_fill if not dm["is_weekend"] else weekend_fill_lv
                ccell.border = hdr_border
                ccell.alignment = Alignment(horizontal="center", vertical="center")
            mr = member_start_lv
            max_lv_rows = max(0, 22 - member_start_lv + 1)
            for grow in grid_lv[:max_lv_rows]:
                ws0.cell(row=mr, column=leave_first_col, value=grow["employee"]).alignment = wrap
                ws0.cell(row=mr, column=leave_first_col).border = grid_border
                for idx, cel in enumerate(grow["cells"]):
                    dm = days_lv[idx]
                    col = leave_first_col + 1 + idx
                    ccell = ws0.cell(row=mr, column=col)
                    ccell.border = grid_border
                    if dm["is_weekend"]:
                        ccell.fill = weekend_fill_lv
                    if cel:
                        ccell.value = cel["code"]
                        ccell.alignment = Alignment(horizontal="center", vertical="center")
                        if cel.get("title"):
                            ccell.comment = Comment(str(cel["title"])[:1000], "Leave")
                    else:
                        ccell.value = ""
                mr += 1
            last_leave_row = mr - 1
            apply_data_grid_region(ws0, weekday_row_lv, last_leave_row, leave_first_col, last_col_lv)
            ws0.column_dimensions["D"].width = max(float(ws0.column_dimensions["D"].width or 10), 28)
            for idx in range(len(days_lv)):
                letter = get_column_letter(leave_first_col + 1 + idx)
                ws0.column_dimensions[letter].width = 5.5

    # Member goals & task kinds (only if they fit on or before row 22; charts start at row 23 on Summary).
    last_mg = max(last_summary, last_leave_row)
    r = last_mg + 2
    if r <= 20 and member_goals:
        ws0.cell(row=r, column=1, value="Member sprint goals").font = sub_font
        r += 1
        mg_hdr_row = r
        mg_h = ["Member", "Sprint goal / notes", "Last updated (timestamp)"]
        for j, h in enumerate(mg_h, start=1):
            ws0.cell(row=r, column=j, value=h)
        style_header_row(ws0, mg_hdr_row, len(mg_h))
        r = mg_hdr_row + 1
        max_goal_rows = max(0, 22 - r)
        for mg in member_goals[:max_goal_rows]:
            ws0.cell(row=r, column=1, value=str(mg["employee_name"] or "")).alignment = top
            ws0.cell(row=r, column=2, value=str(mg["goal"] or "")).alignment = wrap
            ws0.cell(row=r, column=3, value=str(mg["updated_at"] or "")).alignment = top
            r += 1
        last_mg = r - 1
        apply_data_grid(ws0, mg_hdr_row, max(mg_hdr_row, last_mg), len(mg_h))

    last_tk = last_mg
    r = last_mg + 2
    if r <= 19 and task_kinds:
        ws0.cell(row=r, column=1, value="Task kinds (team reference)").font = sub_font
        r += 1
        tk_hdr_row = r
        tk_h = ["Code", "Label", "Color (hex)", "Sort order"]
        for j, h in enumerate(tk_h, start=1):
            ws0.cell(row=r, column=j, value=h)
        style_header_row(ws0, tk_hdr_row, len(tk_h))
        r = tk_hdr_row + 1
        max_tk = max(0, 22 - r)
        for tk in task_kinds[:max_tk]:
            ws0.cell(row=r, column=1, value=str(tk["code"] or "")).alignment = top
            ws0.cell(row=r, column=2, value=str(tk["label"] or "")).alignment = wrap
            ws0.cell(row=r, column=3, value=str(tk["color_hex"] or "")).alignment = top
            ws0.cell(row=r, column=4, value=int(tk["sort_order"] or 0)).alignment = top
            r += 1
        last_tk = r - 1
        apply_data_grid(ws0, tk_hdr_row, max(tk_hdr_row, last_tk), len(tk_h))

    ws0.column_dimensions["C"].width = max(float(ws0.column_dimensions["C"].width or 12), 22)
    ws0.column_dimensions["D"].width = max(float(ws0.column_dimensions["D"].width or 10), 14)

    if ws0.max_row > 22:
        ws0.delete_rows(23, ws0.max_row - 22)

    # --- Hidden chart source + single merged chart on Summary (row 23+) ---
    try:
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList

        ws_cd = wb.create_sheet("_chart_data")
        ws_cd.sheet_state = "hidden"

        def _lbl(code: str) -> str:
            return {
                "ndy": "NDY",
                "fsy": "FSY",
                "code": "CODE",
                "improvement": "Improvement",
                "process_tools": "P&T",
            }.get(code, code)

        cr = 1
        ws_cd.cell(row=cr, column=1, value="Sprint export — chart backing (hidden)")
        cr += 1
        ws_cd.cell(row=cr, column=1, value="Merged: Total sprint + members — est/burnt by kind (h)")
        cr += 1

        k_est, k_brn = _export_roll_kind_est_burnt(items)
        hrow = cr
        ws_cd.cell(row=hrow, column=1, value="Name")
        ncol = 2
        for code in _SUMMARY_EXPORT_KIND_CHART_CODES:
            lab = _lbl(code)
            ws_cd.cell(row=hrow, column=ncol, value=f"{lab} burnt (h)")
            ncol += 1
            ws_cd.cell(row=hrow, column=ncol, value=f"{lab} est (h)")
            ncol += 1
        last_col = ncol - 1
        cr = hrow + 1
        d0 = cr

        def _write_kind_burnt_est_row(
            name: str,
            est_d: dict[str, float],
            brn_d: dict[str, float],
            hide_zero_est_kinds: bool = False,
        ) -> None:
            nonlocal cr
            ws_cd.cell(row=cr, column=1, value=name)
            cc = 2
            for code in _SUMMARY_EXPORT_KIND_CHART_CODES:
                est_v = est_d.get(code, 0.0)
                brn_v = brn_d.get(code, 0.0)
                # When hide_zero_est_kinds is True, write None so Excel renders no bar for that kind.
                if hide_zero_est_kinds and est_v <= SCRUM_HOUR_EPS:
                    ws_cd.cell(row=cr, column=cc, value=None)
                    cc += 1
                    ws_cd.cell(row=cr, column=cc, value=None)
                else:
                    ws_cd.cell(row=cr, column=cc, value=round(brn_v, 2))
                    cc += 1
                    ws_cd.cell(row=cr, column=cc, value=round(est_v, 2))
                cc += 1
            cr += 1

        _write_kind_burnt_est_row("Total sprint", k_est, k_brn)

        max_member_rows = 12
        shown = 0
        for emp in roster_xlsx:
            if shown >= max_member_rows:
                break
            if not any(str(it["assignee"] or "").strip() == emp.strip() for it in items):
                continue
            me_est, me_brn = _export_roll_kind_est_burnt_assignee(items, emp)
            # Skip members whose total estimate across all kinds is zero — they produce empty bars.
            if sum(me_est.values()) <= SCRUM_HOUR_EPS:
                continue
            # Per-kind: hide est+burnt bars where that kind has zero estimate for this member.
            _write_kind_burnt_est_row(str(emp).strip(), me_est, me_brn, hide_zero_est_kinds=True)
            shown += 1

        d1 = cr - 1
        if d1 >= d0:
            ch = BarChart()
            ch.type = "bar"
            ch.grouping = "clustered"
            ch.title = "Sprint + members — burnt vs est by task kind (h); includes Total sprint"
            ch.x_axis.title = "Hours"
            ch.add_data(
                Reference(ws_cd, min_col=2, min_row=hrow, max_row=d1, max_col=last_col),
                titles_from_data=True,
            )
            ch.set_categories(Reference(ws_cd, min_col=1, min_row=d0, max_row=d1, max_col=1))
            ch.legend.position = "b"
            # Size in cm — generous plot so category names, 10 clustered series, and labels stay readable.
            ncat = d1 - d0 + 1
            ch.height = min(64.0, max(20.0, 16.0 + 1.35 * float(ncat)))
            ch.width = 52
            ch.gapWidth = 90
            try:
                ch.dataLabels = DataLabelList()
                ch.dataLabels.showVal = True
            except Exception:
                pass
            ws0.add_chart(ch, "A23")
            # Patch chart font to 7pt by injecting DrawingML font attributes directly into the XML.
            # openpyxl doesn't expose chart font size through a clean API, so we edit the element tree.
            try:
                from lxml import etree as _et
                _NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
                _NS_C = "http://schemas.openxmlformats.org/drawingml/2006/chart"
                _SZ = "700"  # 700 hundredths-of-a-point = 7 pt

                def _inject_font(parent_el: object) -> None:
                    """Recursively set sz on every <a:defRPr> and <a:rPr> found."""
                    for el in parent_el.iter():  # type: ignore[union-attr]
                        tag = el.tag if isinstance(el.tag, str) else ""
                        local = tag.split("}")[-1] if "}" in tag else tag
                        if local in ("defRPr", "rPr"):
                            el.set("sz", _SZ)
                            el.set("b", "0")

                root = ch._element  # type: ignore[attr-defined]
                # Apply to axis tick labels (catAx / valAx / txPr)
                for axis_tag in (f"{{{_NS_C}}}catAx", f"{{{_NS_C}}}valAx"):
                    for ax in root.iter(axis_tag):
                        # Ensure txPr/bodyPr/lstStyle/p/pPr/defRPr exist
                        txPr = ax.find(f"{{{_NS_C}}}txPr")
                        if txPr is None:
                            txPr = _et.SubElement(ax, f"{{{_NS_C}}}txPr")
                        if txPr.find(f"{{{_NS_A}}}bodyPr") is None:
                            _et.SubElement(txPr, f"{{{_NS_A}}}bodyPr")
                        if txPr.find(f"{{{_NS_A}}}lstStyle") is None:
                            _et.SubElement(txPr, f"{{{_NS_A}}}lstStyle")
                        p = txPr.find(f"{{{_NS_A}}}p")
                        if p is None:
                            p = _et.SubElement(txPr, f"{{{_NS_A}}}p")
                        pPr = p.find(f"{{{_NS_A}}}pPr")
                        if pPr is None:
                            pPr = _et.SubElement(p, f"{{{_NS_A}}}pPr")
                        defRPr = pPr.find(f"{{{_NS_A}}}defRPr")
                        if defRPr is None:
                            defRPr = _et.SubElement(pPr, f"{{{_NS_A}}}defRPr")
                        defRPr.set("sz", _SZ)
                        defRPr.set("b", "0")
                # Apply to data labels
                for dlbls in root.iter(f"{{{_NS_C}}}dLbls"):
                    _inject_font(dlbls)
                # Apply to legend
                for leg in root.iter(f"{{{_NS_C}}}legend"):
                    _inject_font(leg)
                # Apply to chart title
                for title_el in root.iter(f"{{{_NS_C}}}title"):
                    _inject_font(title_el)
            except Exception:
                pass
    except Exception:
        pass

    # Sprint hub snapshot as PNG on a tab named after the sprint (second sheet).
    if png_sprint_snapshot:
        try:
            from openpyxl.drawing.image import Image as XLImage
            from sprint_hub_snapshot_png import sanitize_excel_sheet_title

            sheet_title = sanitize_excel_sheet_title(str(sprint["name"]))
            reserved = {
                "Summary",
                "SprintStatus",
                "HPPM",
                "Daily tasks Details",
                "Appriciation",
                "_chart_data",
            }
            if sheet_title in reserved:
                sheet_title = (sheet_title[:18] + " hub")[:31]
            if sheet_title in wb.sheetnames:
                sheet_title = (sheet_title[:20] + "_")[:31]
            ws_snap = wb.create_sheet(sheet_title, 1)
            xl_img = XLImage(io.BytesIO(png_sprint_snapshot))
            tw = 920
            if getattr(xl_img, "width", 0) and xl_img.width > 0:
                sc = tw / float(xl_img.width)
                xl_img.width = int(xl_img.width * sc)
                xl_img.height = int(xl_img.height * sc)
            ws_snap.add_image(xl_img, "A1")
        except Exception:
            pass

    # --- Sprint status (FB2610-style: sticky metadata + day-wise logged hours) ---
    ws1 = wb.create_sheet("SprintStatus")
    fb_base_headers = [
        "TEAM",
        "Sticky ID",
        "Title",
        "Task type",
        "Kanban",
        "Owner",
        "Estimate (h)",
        "Burnt total (h)",
        "Burn %",
        "DoD (summary)",
        "Task Checklist",
    ]
    sticky_headers = fb_base_headers + [d.isoformat() for d in sprint_calendar_dates]
    nstatcols = len(sticky_headers)
    for j, h in enumerate(sticky_headers, start=1):
        ws1.cell(row=1, column=j, value=h)
    style_header_row(ws1, 1, nstatcols)
    for ri, r in enumerate(items, start=2):
        est = float(r["estimate_hours"] or 0)
        burnt = float(r["total_burnt_hours"] or 0)
        bpct = round(burnt / est, 6) if est > SCRUM_HOUR_EPS else None  # 0–1 for Excel % format
        dod = str(r["dod"] or "")
        if len(dod) > 800:
            dod = dod[:797] + "…"
        iid = int(r["id"])
        # Build checklist text: each row on its own line as "• <item> | <status> | LE:<le> | <comments>"
        checklist_rows = checklist_export_map.get(iid, [])
        checklist_parts = []
        for cl in checklist_rows:
            parts = [str(cl.get("items_to_finish") or "").strip()]
            if (cl.get("status") or "").strip():
                parts.append(str(cl["status"]).strip())
            if (cl.get("le_to_complete") or "").strip():
                parts.append("LE:" + str(cl["le_to_complete"]).strip())
            if (cl.get("done_till_date") or "").strip():
                parts.append(str(cl["done_till_date"]).strip())
            checklist_parts.append("• " + " | ".join(p for p in parts if p))
        checklist_text = "\n".join(checklist_parts)
        if len(checklist_text) > 1500:
            checklist_text = checklist_text[:1497] + "…"
        base_vals: list[object] = [
            team_name,
            iid,
            str(r["title"] or ""),
            str(r["task_kind_label"] or r["task_kind"] or ""),
            str(r["kanban_column"] or ""),
            str(r["assignee"] or ""),
            est,
            burnt,
            bpct,
            dod,
            checklist_text,
        ]
        for j, val in enumerate(base_vals, start=1):
            write_val = "—" if (j == 9 and val is None) else val
            cell = ws1.cell(row=ri, column=j, value=write_val)
            cell.alignment = wrap
            if j in (7, 8):
                cell.number_format = "0.00"
            if j == 9 and val is not None:
                cell.number_format = "0.0%"
                cell.fill = pct_fill
                cell.font = pct_font
        for j, d in enumerate(sprint_calendar_dates, start=len(base_vals) + 1):
            hv = float(item_day_hours.get(iid, {}).get(d, 0.0))
            snips = list(item_day_snips.get(iid, {}).get(d, []))
            # Cell value: standup notes (text); hours go into the comment
            if snips:
                note_text = "\n".join(snips)
                ccell = ws1.cell(row=ri, column=j, value=note_text[:900])
                ccell.alignment = Alignment(vertical="top", wrap_text=True)
                if hv > SCRUM_HOUR_EPS:
                    ccell.comment = Comment(f"{round(hv, 2):.2f} h logged", "scrum")
            elif hv > SCRUM_HOUR_EPS:
                # No note text but hours exist — show hours so the cell isn't blank
                ccell = ws1.cell(row=ri, column=j, value=round(hv, 2))
                ccell.number_format = "0.00"
                ccell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                ws1.cell(row=ri, column=j, value=None)
    ws1.freeze_panes = "A2"
    autofit(ws1)
    apply_data_grid(ws1, 1, max(1, len(items) + 1), nstatcols)

    # --- HPPM (same summary + sticky table as /scrum/sprint/<id>/hppm) ---
    ws_hppm = wb.create_sheet("HPPM")
    rh = 1
    ws_hppm.merge_cells(start_row=rh, start_column=1, end_row=rh, end_column=8)
    h_title = ws_hppm.cell(row=rh, column=1, value=f"HPPM view — {str(sprint['name'])} · {sd_s} → {ed_s}")
    h_title.font = title_font
    h_title.fill = title_fill
    h_title.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rh += 2
    if app is not None and roster_xlsx:
        try:
            hppm_ctx = build_hppm_sprint_page_extra_context(
                conn,
                app,
                team_id=int(team_id),
                sprint_id=int(sprint_id),
                roster=roster_xlsx,
                sprint_start=sd_s,
                sprint_end=ed_s,
            )
            gh = hppm_ctx.get("hppm_team_capacity_gross_hours")
            nh = hppm_ctx.get("hppm_team_capacity_net_hours")
            cap_bits: list[str] = []
            if gh is not None:
                cap_bits.append(f"Gross team capacity (before leave): {float(gh):.1f} h")
            if nh is not None:
                cap_bits.append(f"Net team capacity (after leave): {float(nh):.1f} h")
            if cap_bits:
                ccap = ws_hppm.cell(row=rh, column=1, value=" · ".join(cap_bits))
                ccap.font = subtitle_font
                ccap.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                rh += 1
            cnote = ws_hppm.cell(
                row=rh,
                column=1,
                value="% Team Sprint Estimate = row estimate ÷ total estimate in this table (Total = 100%). "
                "% Team Burnt = row burnt hours ÷ total burnt hours in this table (Total = 100%).",
            )
            cnote.font = subtitle_font
            cnote.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            rh += 2

            sum_hdr = (
                "Pool name",
                "Work type",
                "Estimate sprint (h)",
                "% Team Sprint Estimate",
                f"Burnt (h) — {str(sprint['name'])}",
                "% Team Burnt",
            )
            sum_hdr_row = rh
            for jc, lab in enumerate(sum_hdr, start=1):
                ws_hppm.cell(row=sum_hdr_row, column=jc, value=lab)
            style_header_row(ws_hppm, sum_hdr_row, len(sum_hdr))
            rh = sum_hdr_row + 1
            pool_h = str(hppm_ctx.get("hppm_pool_name") or team_name or "")
            for srow in hppm_ctx.get("hppm_summary_rows") or []:
                pe = srow.get("pct_capacity_estimate")
                pb = srow.get("pct_capacity_burnt")
                ws_hppm.cell(row=rh, column=1, value=pool_h).alignment = wrap
                ws_hppm.cell(row=rh, column=2, value=str(srow.get("work_type") or "")).alignment = wrap
                c_est = ws_hppm.cell(row=rh, column=3, value=float(srow.get("estimate_hours") or 0.0))
                c_est.number_format = "0.0000"
                c_est.alignment = top
                c_pe = ws_hppm.cell(row=rh, column=4, value=round(float(pe) / 100.0, 6) if pe is not None else "—")
                c_pe.alignment = top
                if pe is not None:
                    c_pe.number_format = "0.00%"
                    c_pe.fill = pct_fill
                    c_pe.font = pct_font
                c_br = ws_hppm.cell(row=rh, column=5, value=float(srow.get("burnt_hours") or 0.0))
                c_br.number_format = "0.0000"
                c_br.alignment = top
                c_pb = ws_hppm.cell(row=rh, column=6, value=round(float(pb) / 100.0, 6) if pb is not None else "—")
                c_pb.alignment = top
                if pb is not None:
                    c_pb.number_format = "0.00%"
                    c_pb.fill = pct_fill
                    c_pb.font = pct_font
                rh += 1
            tot = hppm_ctx.get("hppm_summary_totals") or {}
            pte = tot.get("pct_capacity_estimate")
            ptb = tot.get("pct_capacity_burnt")
            ws_hppm.cell(row=rh, column=1, value="Total").font = sub_font
            ws_hppm.cell(row=rh, column=2, value="").font = sub_font
            c_te = ws_hppm.cell(row=rh, column=3, value=float(tot.get("estimate_hours") or 0.0))
            c_te.number_format = "0.0000"
            c_te.font = sub_font
            ws_hppm.cell(row=rh, column=4, value=f"{float(pte):.2f}%" if pte is not None else "—").font = sub_font
            c_tb = ws_hppm.cell(row=rh, column=5, value=float(tot.get("burnt_hours") or 0.0))
            c_tb.number_format = "0.0000"
            c_tb.font = sub_font
            ws_hppm.cell(row=rh, column=6, value=f"{float(ptb):.2f}%" if ptb is not None else "—").font = sub_font
            last_sum = rh
            rh += 1
            apply_data_grid(ws_hppm, sum_hdr_row, last_sum, len(sum_hdr))

            rh += 2
            ws_hppm.cell(row=rh, column=1, value="All stickies (HPPM work type & burnt)").font = sub_font
            rh += 1
            stk_hdr = (
                "Area type",
                "Work type",
                "Title",
                "Assignee",
                "Est. (h)",
                "Burnt (h)",
                "Burnt %",
                "Total burnt (h)",
            )
            stk_hdr_row = rh
            for jc, lab in enumerate(stk_hdr, start=1):
                ws_hppm.cell(row=stk_hdr_row, column=jc, value=lab)
            style_header_row(ws_hppm, stk_hdr_row, len(stk_hdr))
            rh = stk_hdr_row + 1
            task_rows_h = hppm_ctx.get("hppm_task_rows") or []
            if not task_rows_h:
                ws_hppm.cell(row=rh, column=1, value="(No stickies in this sprint.)").alignment = wrap
                last_stk = rh
                rh += 1
                stk_total_merges: list[tuple[int, int]] = []
            else:
                stk_total_merges = []
                for tr in task_rows_h:
                    bp = tr.get("burnt_pct")
                    ws_hppm.cell(row=rh, column=1, value=str(tr.get("area_type") or "")).alignment = wrap
                    ws_hppm.cell(row=rh, column=2, value=str(tr.get("work_type") or "")).alignment = wrap
                    ws_hppm.cell(row=rh, column=3, value=str(tr.get("title") or "")).alignment = wrap
                    ws_hppm.cell(row=rh, column=4, value=str(tr.get("assignee") or "")).alignment = wrap
                    ce = ws_hppm.cell(row=rh, column=5, value=float(tr.get("estimate_hours") or 0.0))
                    ce.number_format = "0.00"
                    cb = ws_hppm.cell(row=rh, column=6, value=float(tr.get("burnt_hours") or 0.0))
                    cb.number_format = "0.0000"
                    _bp_cell = ws_hppm.cell(
                        row=rh,
                        column=7,
                        value=round(float(bp) / 100.0, 6) if bp is not None else "—",
                    )
                    _bp_cell.alignment = top
                    if bp is not None:
                        _bp_cell.number_format = "0.0%"
                        _bp_cell.fill = pct_fill
                        _bp_cell.font = pct_font
                    if tr.get("area_group_first"):
                        rs = int(tr.get("area_total_rowspan") or 1)
                        c_tot = ws_hppm.cell(row=rh, column=8, value=float(tr.get("area_group_burnt_total") or 0.0))
                        c_tot.number_format = "0.0000"
                        c_tot.alignment = Alignment(horizontal="right", vertical="center")
                        if rs > 1:
                            stk_total_merges.append((rh, rs))
                    rh += 1
                last_stk = rh - 1
            apply_data_grid(ws_hppm, stk_hdr_row, max(stk_hdr_row, last_stk), len(stk_hdr))
            for sr, rsp in stk_total_merges:
                if rsp > 1:
                    ws_hppm.merge_cells(start_row=sr, start_column=8, end_row=sr + rsp - 1, end_column=8)
                    ws_hppm.cell(row=sr, column=8).alignment = Alignment(horizontal="right", vertical="center")
            ws_hppm.freeze_panes = f"A{sum_hdr_row + 1}"
            autofit(ws_hppm)
        except Exception:
            ws_hppm.cell(row=5, column=1, value="HPPM sheet could not be generated (export error).")
    else:
        ws_hppm.cell(row=3, column=1, value="HPPM sheet skipped (app or roster not available in this export).")

    # --- Daily tasks Details: sprint team view per member + all stickies + scrum_daily_task rows ---
    # Sentinel: returns the raw float (÷100 for Excel %-format) or None for "—"
    def _xlsx_pct_disp(v: float | None) -> float | None:
        if v is None:
            return None
        return round(float(v) / 100.0, 6)  # Excel stores 0–1; "0.0%" format shows "39.5%"

    ws3 = wb.create_sheet("Daily tasks Details")
    rz = 1
    tcell = ws3.cell(row=rz, column=1, value="Daily tasks Details — Sprint team view, stickies, and daily task records")
    tcell.font = Font(bold=True, size=12, color="1E293B")
    rz += 2

    ws3.cell(row=rz, column=1, value="Per-member sprint overview (same metrics as Sprint team page)").font = sub_font
    rz += 1
    memb_headers = (
        ["Team", "Member"]
        + [
            "Sprint burnt %",
            "NDY % burnt (by type)",
            "FSY % burnt (by type)",
            "CODE % burnt (by type)",
            "NDY est (h)",
            "NDY burnt (h)",
            "FSY est (h)",
            "FSY burnt (h)",
            "CODE est (h)",
            "CODE burnt (h)",
            "Improvement % burnt (by type)",
            "Process&Tools % burnt (by type)",
            "Improvement est (h)",
            "Improvement burnt (h)",
            "Process&Tools est (h)",
            "Process&Tools burnt (h)",
            "Queue backlog (B)",
            "Queue do (D)",
            "Queue in progress (Dg)",
            "Queue done (Dn)",
            "Total est all stickies (h)",
            "Total logged (h)",
            "In progress titles (preview)",
        ]
        + [d.strftime("%m/%d") for d in sprint_calendar_dates]
    )
    mem_hdr_row = rz
    for jc, h in enumerate(memb_headers, start=1):
        ws3.cell(row=rz, column=jc, value=h)
    style_header_row(ws3, mem_hdr_row, len(memb_headers))
    rz += 1
    freeze_member_row = rz
    for m in member_overview:
        cnt = m["counts"]
        emp = str(m["name"] or "").strip()
        day_vals = [round(float(assignee_day_hours.get(emp, {}).get(d, 0.0)), 2) for d in sprint_calendar_dates]
        row_mv = [
            team_name,
            emp,
            _xlsx_pct_disp(m.get("progress_pct")),
            _xlsx_pct_disp(m.get("burn_kind_ndy_pct")),
            _xlsx_pct_disp(m.get("burn_kind_fsy_pct")),
            _xlsx_pct_disp(m.get("burn_kind_code_pct")),
            float(m["burn_kind_ndy_est"]),
            float(m["burn_kind_ndy_com"]),
            float(m["burn_kind_fsy_est"]),
            float(m["burn_kind_fsy_com"]),
            float(m["burn_kind_code_est"]),
            float(m["burn_kind_code_com"]),
            _xlsx_pct_disp(m.get("burn_kind_improvement_pct")),
            _xlsx_pct_disp(m.get("burn_kind_process_tools_pct")),
            float(m["burn_kind_improvement_est"]),
            float(m["burn_kind_improvement_com"]),
            float(m["burn_kind_process_tools_est"]),
            float(m["burn_kind_process_tools_com"]),
            int(cnt.get("backlog", 0)),
            int(cnt.get("do", 0)),
            int(cnt.get("doing", 0)),
            int(cnt.get("done", 0)),
            float(m["est_total_hours"]),
            float(m["committed_total_hours"]),
            "; ".join(str(x) for x in (m.get("doing_preview") or [])),
        ] + day_vals
        # Columns (1-based) that hold percentage values from _xlsx_pct_disp (stored as 0–1 float)
        # col 3 = Sprint burnt %, 4 = NDY%, 5 = FSY%, 6 = CODE%, 13 = Improvement%, 14 = P&T%
        pct_cols = {3, 4, 5, 6, 13, 14}
        for jc, val in enumerate(row_mv, start=1):
            if jc in pct_cols:
                if val is None:
                    cell = ws3.cell(row=rz, column=jc, value="—")
                else:
                    cell = ws3.cell(row=rz, column=jc, value=val)
                    cell.number_format = "0.0%"
                cell.fill = pct_fill
                cell.font = pct_font
            else:
                cell = ws3.cell(row=rz, column=jc, value=val)
                if isinstance(val, float):
                    cell.number_format = "0.00"
            cell.alignment = wrap
        rz += 1
    last_mem = rz - 1
    apply_data_grid(ws3, mem_hdr_row, max(mem_hdr_row, last_mem), len(memb_headers))

    rz += 1
    ws3.cell(row=rz, column=1, value="Every sticky — compact (no duplicate columns vs SprintStatus)").font = sub_font
    rz += 1
    sticky_detail_headers = [
        "Team",
        "Sticky ID",
        "Title",
        "Kanban",
        "Est (h)",
        "Burnt (h)",
        "Burn %",
    ]
    st_hdr_row = rz
    for jc, h in enumerate(sticky_detail_headers, start=1):
        ws3.cell(row=rz, column=jc, value=h)
    style_header_row(ws3, st_hdr_row, len(sticky_detail_headers))
    rz += 1
    for it in items:
        est = float(it["estimate_hours"] or 0)
        burnt = float(it["total_burnt_hours"] or 0)
        sburn = round(burnt / est, 6) if est > SCRUM_HOUR_EPS else None  # 0–1 for Excel % format
        row_st = [
            team_name,
            int(it["id"]),
            str(it["title"] or ""),
            str(it["kanban_column"] or ""),
            est,
            burnt,
            sburn,
        ]
        for jc, val in enumerate(row_st, start=1):
            cell = ws3.cell(row=rz, column=jc, value=val if val is not None else "—")
            cell.alignment = wrap
            if jc in (5, 6):
                cell.number_format = "0.00"
            if jc == 7 and val is not None:
                cell.number_format = "0.0%"
                cell.fill = pct_fill
                cell.font = pct_font
        rz += 1
    last_st = rz - 1
    apply_data_grid(ws3, st_hdr_row, max(st_hdr_row, last_st), len(sticky_detail_headers))

    rz += 1
    ws3.cell(row=rz, column=1, value="Daily task rows (scrum_daily_task table)").font = sub_font
    rz += 1
    daily_headers = [
        "Team",
        "Row ID",
        "Work date",
        "Assignee",
        "Title",
        "Planned (h)",
        "Committed (h)",
        "Actual (h)",
        "Status",
        "Notes",
        "Sort order",
        "Linked sticky ID",
        "Linked sticky title",
        "Created at (timestamp)",
        "Last updated (timestamp)",
    ]
    daily_hdr_row = rz
    for jc, h in enumerate(daily_headers, start=1):
        ws3.cell(row=rz, column=jc, value=h)
    style_header_row(ws3, daily_hdr_row, len(daily_headers))
    rz += 1
    for dr in daily_rows:
        row_d = [
            team_name,
            int(dr["id"]),
            str(dr["work_date"] or ""),
            str(dr["assignee"] or ""),
            str(dr["title"] or ""),
            float(dr["planned_hours"] or 0),
            float(dr["committed_hours"] or 0),
            float(dr["actual_hours"] or 0),
            str(dr["status"] or ""),
            str(dr["notes"] or ""),
            int(dr["sort_order"] or 0),
            int(dr["sprint_item_id"]) if dr["sprint_item_id"] is not None else "",
            str(dr["linked_sticky_title"] or ""),
            str(dr["created_at"] or ""),
            str(dr["updated_at"] or ""),
        ]
        for jc, val in enumerate(row_d, start=1):
            cell = ws3.cell(row=rz, column=jc, value=val)
            cell.alignment = wrap
            if jc in (6, 7, 8):
                cell.number_format = "0.00"
        rz += 1
    last_daily = rz - 1
    apply_data_grid(ws3, daily_hdr_row, max(daily_hdr_row, last_daily), len(daily_headers))

    ws3.freeze_panes = f"A{freeze_member_row}"
    autofit(ws3)

    # Daily tasks Details: allow many rows (day columns + roster); trim only extreme overflow.
    if ws3.max_row > 1200:
        n_del3 = ws3.max_row - 1200
        if n_del3 > 0:
            ws3.delete_rows(1201, n_del3)

    # --- Appriciation: author, comment, title, assignee only (columns C, D, G, H) ---
    ws4 = wb.create_sheet("Appriciation")
    r4 = 1
    ws4.merge_cells(start_row=r4, start_column=3, end_row=r4, end_column=8)
    t4 = ws4.cell(
        row=r4,
        column=3,
        value="Appriciation — one row per manager comment (author, comment, title, assignee)",
    )
    t4.font = Font(bold=True, size=12, color="1E293B")
    t4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    r4 += 2
    ap_hdr_row = r4
    ap_cols = (
        (3, "Appriciation author"),
        (4, "Appriciation comment"),
        (7, "Title"),
        (8, "Assignee"),
    )
    for col_idx, label in ap_cols:
        c = ws4.cell(row=r4, column=col_idx, value=label)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = hdr_border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r4 += 1
    first_data_row = r4

    def _ap_short_text(val: object, cap: int = 12000) -> str:
        s = str(val if val is not None else "")
        if len(s) > cap:
            return s[: cap - 1] + "…"
        return s

    if not appreciation_export_rows:
        c0 = ws4.cell(row=r4, column=3, value="(No appreciation comments for this sprint.)")
        c0.alignment = wrap
        last_ap = r4
    else:
        for ar in appreciation_export_rows:
            ws4.cell(row=r4, column=3, value=str(ar["appreciation_author"] or "")).alignment = wrap
            ws4.cell(row=r4, column=4, value=_ap_short_text(ar["appreciation_comment"])).alignment = wrap
            ws4.cell(row=r4, column=7, value=str(ar["title"] or "")).alignment = wrap
            ws4.cell(row=r4, column=8, value=str(ar["assignee"] or "")).alignment = wrap
            r4 += 1
        last_ap = r4 - 1

    apply_data_grid_region(ws4, ap_hdr_row, max(ap_hdr_row, last_ap), 3, 4)
    apply_data_grid_region(ws4, ap_hdr_row, max(ap_hdr_row, last_ap), 7, 8)
    ws4.freeze_panes = f"C{first_data_row}"
    for letter, w in (("C", 18.0), ("D", 48.0), ("G", 36.0), ("H", 28.0)):
        ws4.column_dimensions[letter].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    base = _sanitize_xlsx_base_filename(str(sprint["name"]))
    fname = f"{base}_sprint_{int(sprint_id)}.xlsx"
    return bio.getvalue(), fname


def _sprint_item_completion(conn: sqlite3.Connection, sprint_id: int) -> dict:
    rows = list(
        conn.execute(
            """
            SELECT estimate_hours, status, kanban_column
            FROM scrum_sprint_item WHERE sprint_id = ?
            """,
            (sprint_id,),
        )
    )
    if not rows:
        return {"item_count": 0, "done_count": 0, "pct_done": 0.0, "open_estimate": 0.0}
    n = len(rows)
    done = 0
    open_est = 0.0
    for r in rows:
        col = _normalize_kanban_column(r["kanban_column"] if "kanban_column" in r.keys() else None)
        st = _normalize_scrum_status(r["status"])
        is_done = col == "done" or st == "done"
        if is_done:
            done += 1
        else:
            open_est += float(r["estimate_hours"] or 0)
    return {
        "item_count": n,
        "done_count": done,
        "pct_done": round(100.0 * done / n, 1) if n else 0.0,
        "open_estimate": round(open_est, 2),
    }


def _dashboard_sprint_summaries(conn: sqlite3.Connection, team_id: int, limit: int = 8) -> list[dict]:
    rows = list(
        conn.execute(
            """
            SELECT id, name, start_date, end_date, goal
            FROM scrum_sprint
            WHERE team_id = ?
            ORDER BY end_date DESC, id DESC
            LIMIT ?
            """,
            (int(team_id), int(limit)),
        )
    )
    today = date.today().isoformat()
    out: list[dict] = []
    for r in rows:
        sid = int(r["id"])
        comp = _sprint_item_completion(conn, sid)
        logged = conn.execute(
            """
            SELECT COALESCE(SUM(a.committed_hours), 0) AS h
            FROM scrum_item_activity a
            JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ?
            """,
            (sid,),
        ).fetchone()
        h = float(logged["h"] or 0) if logged else 0.0
        s0, s1 = str(r["start_date"])[:10], str(r["end_date"])[:10]
        in_window = s0 <= today <= s1
        out.append(
            {
                "id": sid,
                "name": r["name"],
                "start": s0,
                "end": s1,
                "goal": (r["goal"] or "").strip()[:200],
                "pct_done": comp["pct_done"],
                "done_count": comp["done_count"],
                "item_count": comp["item_count"],
                "logged_hours": round(h, 1),
                "in_sprint_window": in_window,
            }
        )
    return out


def _parse_activity_created_at_utc(raw: object) -> datetime | None:
    """Parse scrum_item_activity.created_at (ISO, usually …Z) to aware UTC datetime."""
    s = str(raw or "").strip()
    if not s:
        return None
    try:
        if s.endswith("Z"):
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        else:
            dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except ValueError:
        return None


def _last_notes_for_assignee(conn: sqlite3.Connection, sprint_id: int, assignee: str, limit: int = 3) -> list[dict]:
    rows = conn.execute(
        """
        SELECT a.body, a.committed_hours, a.from_column, a.to_column, a.created_at, i.title
        FROM scrum_item_activity a
        JOIN scrum_sprint_item i ON i.id = a.item_id
        WHERE i.sprint_id = ? AND i.assignee = ?
          AND (
            trim(a.body) != ''
            OR IFNULL(ABS(a.committed_hours), 0) > 0.00001
            OR lower(trim(COALESCE(a.from_column, ''))) != lower(trim(COALESCE(a.to_column, '')))
          )
        ORDER BY a.created_at DESC
        LIMIT ?
        """,
        (int(sprint_id), assignee, int(limit)),
    ).fetchall()
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=24)
    out: list[dict] = []
    for r in rows:
        created = _parse_activity_created_at_utc(r["created_at"])
        recent_24h = bool(created is not None and created >= cutoff)
        out.append(
            {
                "body": (r["body"] or "").strip(),
                "committed_hours": float(r["committed_hours"] or 0),
                "from_column": r["from_column"],
                "to_column": r["to_column"],
                "created_at": r["created_at"],
                "title": (r["title"] or "").strip(),
                "recent_24h": recent_24h,
            }
        )
    return out


TEAM_TRACKER_TEAM_BUNDLE_FORMAT = "team_tracker_team_bundle_v1"

_BUNDLE_INSERT_TABLES = frozenset(
    {
        "scrum_team_task_kind",
        "scrum_sprint",
        "scrum_sprint_item",
        "scrum_item_activity",
        "scrum_sprint_member_goal",
        "scrum_daily_task",
        "scrum_item_appreciation",
        "scrum_portal_proposal",
    }
)


def _sqlite_table_exists(conn: sqlite3.Connection, name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ? LIMIT 1",
        (name,),
    ).fetchone()
    return row is not None


def _row_to_bundle_dict(r: sqlite3.Row) -> dict[str, object]:
    out: dict[str, object] = {}
    for k in r.keys():
        v = r[k]
        if isinstance(v, bytes):
            out[k] = v.decode("utf-8", errors="replace")
        else:
            out[k] = v
    return out


def _pragma_insert_columns(conn: sqlite3.Connection, table: str) -> list[str]:
    if table not in _BUNDLE_INSERT_TABLES:
        raise ValueError(f"Unsupported bundle table {table!r}")
    return [str(r[1]) for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]


def _insert_bundle_row(
    conn: sqlite3.Connection, table: str, row: dict[str, object], overrides: dict[str, object]
) -> int:
    merged: dict[str, object] = {k: v for k, v in row.items() if k != "id"}
    merged.update(overrides)
    pragma_order = _pragma_insert_columns(conn, table)
    cols = [c for c in pragma_order if c != "id" and c in merged]
    if not cols:
        raise ValueError(f"No insertable columns for {table}")
    vals = [merged[c] for c in cols]
    ph = ",".join("?" * len(cols))
    cur = conn.execute(f"INSERT INTO {table} ({','.join(cols)}) VALUES ({ph})", vals)
    return int(cur.lastrowid)


def build_manager_team_bundle_dict(conn: sqlite3.Connection, team_id: int) -> dict[str, object]:
    """Full Scrum-related snapshot for one team (JSON-serializable)."""
    team_row = conn.execute("SELECT * FROM teams WHERE id = ?", (int(team_id),)).fetchone()
    if not team_row:
        raise ValueError("Team not found")
    tid = int(team_id)
    sprints = list(
        conn.execute("SELECT * FROM scrum_sprint WHERE team_id = ? ORDER BY id", (tid,)).fetchall()
    )
    sid_list = [int(s["id"]) for s in sprints]
    items: list[sqlite3.Row] = []
    if sid_list:
        ph = ",".join("?" * len(sid_list))
        items = list(
            conn.execute(
                f"SELECT * FROM scrum_sprint_item WHERE sprint_id IN ({ph}) ORDER BY id",
                sid_list,
            ).fetchall()
        )
    iid_list = [int(x["id"]) for x in items]
    activity: list[sqlite3.Row] = []
    appreciation: list[sqlite3.Row] = []
    if iid_list:
        ph = ",".join("?" * len(iid_list))
        activity = list(
            conn.execute(
                f"SELECT * FROM scrum_item_activity WHERE item_id IN ({ph}) ORDER BY id",
                iid_list,
            ).fetchall()
        )
        if _sqlite_table_exists(conn, "scrum_item_appreciation"):
            appreciation = list(
                conn.execute(
                    f"SELECT * FROM scrum_item_appreciation WHERE item_id IN ({ph}) ORDER BY id",
                    iid_list,
                ).fetchall()
            )
    member_goals: list[sqlite3.Row] = []
    if sid_list:
        ph = ",".join("?" * len(sid_list))
        member_goals = list(
            conn.execute(
                f"SELECT * FROM scrum_sprint_member_goal WHERE sprint_id IN ({ph})",
                sid_list,
            ).fetchall()
        )
    daily = list(conn.execute("SELECT * FROM scrum_daily_task WHERE team_id = ? ORDER BY id", (tid,)).fetchall())
    roster = list(
        conn.execute(
            "SELECT * FROM team_roster WHERE team_id = ? ORDER BY sort_order, employee_name COLLATE NOCASE",
            (tid,),
        ).fetchall()
    )
    kinds: list[sqlite3.Row] = []
    if _sqlite_table_exists(conn, "scrum_team_task_kind"):
        kinds = list(
            conn.execute(
                "SELECT * FROM scrum_team_task_kind WHERE team_id = ? ORDER BY sort_order, id",
                (tid,),
            ).fetchall()
        )
    proposals: list[sqlite3.Row] = []
    if _sqlite_table_exists(conn, "scrum_portal_proposal"):
        proposals = list(
            conn.execute("SELECT * FROM scrum_portal_proposal WHERE team_id = ? ORDER BY id", (tid,)).fetchall()
        )
    return {
        "format": TEAM_TRACKER_TEAM_BUNDLE_FORMAT,
        "exported_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "team": _row_to_bundle_dict(team_row),
        "team_roster": [_row_to_bundle_dict(r) for r in roster],
        "scrum_team_task_kind": [_row_to_bundle_dict(r) for r in kinds],
        "scrum_sprint": [_row_to_bundle_dict(r) for r in sprints],
        "scrum_sprint_item": [_row_to_bundle_dict(r) for r in items],
        "scrum_item_activity": [_row_to_bundle_dict(r) for r in activity],
        "scrum_sprint_member_goal": [_row_to_bundle_dict(r) for r in member_goals],
        "scrum_daily_task": [_row_to_bundle_dict(r) for r in daily],
        "scrum_item_appreciation": [_row_to_bundle_dict(r) for r in appreciation],
        "scrum_portal_proposal": [_row_to_bundle_dict(r) for r in proposals],
    }


def import_manager_team_bundle(conn: sqlite3.Connection, team_id: int, bundle: dict[str, object]) -> None:
    """Replace this team's Scrum roster + data with the contents of a bundle (same format as export)."""
    if bundle.get("format") != TEAM_TRACKER_TEAM_BUNDLE_FORMAT:
        raise ValueError("Unrecognized bundle format (expected team_tracker_team_bundle_v1).")
    tid = int(team_id)
    if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (tid,)).fetchone():
        raise ValueError("Team not found.")

    if _sqlite_table_exists(conn, "scrum_portal_proposal"):
        conn.execute("DELETE FROM scrum_portal_proposal WHERE team_id = ?", (tid,))
    conn.execute("DELETE FROM scrum_daily_task WHERE team_id = ?", (tid,))
    conn.execute("DELETE FROM scrum_sprint WHERE team_id = ?", (tid,))
    if _sqlite_table_exists(conn, "scrum_team_task_kind"):
        conn.execute("DELETE FROM scrum_team_task_kind WHERE team_id = ?", (tid,))
    conn.execute("DELETE FROM team_roster WHERE team_id = ?", (tid,))

    team_meta = bundle.get("team") or {}
    hub = str(team_meta.get("hub_mode") or "leave").strip()
    if hub not in ("leave", "scrum"):
        hub = "leave"
    conn.execute("UPDATE teams SET hub_mode = ? WHERE id = ?", (hub, tid))

    for row in bundle.get("team_roster") or []:
        if not isinstance(row, dict):
            continue
        name = str(row.get("employee_name") or "").strip()
        if not name:
            continue
        so = int(row.get("sort_order") or 0)
        email = normalize_email(str(row.get("employee_email") or ""))
        conn.execute(
            """
            INSERT INTO team_roster (team_id, employee_name, employee_email, sort_order)
            VALUES (?, ?, ?, ?)
            """,
            (tid, name, email, so),
        )

    if _sqlite_table_exists(conn, "scrum_team_task_kind"):
        for row in bundle.get("scrum_team_task_kind") or []:
            if not isinstance(row, dict):
                continue
            _insert_bundle_row(conn, "scrum_team_task_kind", row, {"team_id": tid})

    sprint_map: dict[int, int] = {}
    for sp in sorted(bundle.get("scrum_sprint") or [], key=lambda x: int(x["id"])):
        if not isinstance(sp, dict):
            continue
        old_sid = int(sp["id"])
        new_sid = _insert_bundle_row(conn, "scrum_sprint", sp, {"team_id": tid})
        sprint_map[old_sid] = new_sid

    item_map: dict[int, int] = {}
    for it in sorted(bundle.get("scrum_sprint_item") or [], key=lambda x: int(x["id"])):
        if not isinstance(it, dict):
            continue
        old_iid = int(it["id"])
        old_sp = int(it["sprint_id"])
        new_sp = sprint_map.get(old_sp)
        if new_sp is None:
            continue
        new_iid = _insert_bundle_row(conn, "scrum_sprint_item", it, {"sprint_id": new_sp})
        item_map[old_iid] = new_iid

    for row in bundle.get("scrum_item_activity") or []:
        if not isinstance(row, dict):
            continue
        old_item = int(row["item_id"])
        new_item = item_map.get(old_item)
        if new_item is None:
            continue
        _insert_bundle_row(conn, "scrum_item_activity", row, {"item_id": new_item})

    for row in bundle.get("scrum_sprint_member_goal") or []:
        if not isinstance(row, dict):
            continue
        old_sp = int(row["sprint_id"])
        new_sp = sprint_map.get(old_sp)
        if new_sp is None:
            continue
        _insert_bundle_row(conn, "scrum_sprint_member_goal", row, {"sprint_id": new_sp})

    if _sqlite_table_exists(conn, "scrum_item_appreciation"):
        for row in bundle.get("scrum_item_appreciation") or []:
            if not isinstance(row, dict):
                continue
            old_item = int(row["item_id"])
            new_item = item_map.get(old_item)
            if new_item is None:
                continue
            _insert_bundle_row(conn, "scrum_item_appreciation", row, {"item_id": new_item})

    if _sqlite_table_exists(conn, "scrum_portal_proposal"):
        for row in bundle.get("scrum_portal_proposal") or []:
            if not isinstance(row, dict):
                continue
            old_sp = int(row["sprint_id"])
            new_sp = sprint_map.get(old_sp)
            if new_sp is None:
                continue
            ov: dict[str, object] = {"team_id": tid, "sprint_id": new_sp}
            raw_item = row.get("item_id")
            if raw_item is not None:
                old_item = int(raw_item)
                ov["item_id"] = item_map.get(old_item)
            _insert_bundle_row(conn, "scrum_portal_proposal", row, ov)

    for row in bundle.get("scrum_daily_task") or []:
        if not isinstance(row, dict):
            continue
        ov: dict[str, object] = {"team_id": tid}
        raw_sp = row.get("sprint_id")
        if raw_sp is not None:
            ov["sprint_id"] = sprint_map.get(int(raw_sp))
        raw_it = row.get("sprint_item_id")
        if raw_it is not None:
            ov["sprint_item_id"] = item_map.get(int(raw_it))
        _insert_bundle_row(conn, "scrum_daily_task", row, ov)


def _sprint_team_overview_rows(
    app: Flask | None, conn: sqlite3.Connection, team_id: int, sprint_id: int, roster: Sequence[str]
) -> list[dict]:
    _ensure_team_task_kinds(conn, team_id)
    srow = conn.execute(
        "SELECT start_date, end_date FROM scrum_sprint WHERE id = ? AND team_id = ?",
        (int(sprint_id), int(team_id)),
    ).fetchone()
    if not srow:
        return []
    try:
        sd = date.fromisoformat(str(srow["start_date"])[:10])
        ed = date.fromisoformat(str(srow["end_date"])[:10])
    except ValueError:
        sd = ed = date.today()
    out: list[dict] = []
    for emp in roster:
        counts = {c: 0 for c in SCRUM_KANBAN_COLUMNS}
        rows = conn.execute(
            """
            SELECT kanban_column, status FROM scrum_sprint_item
            WHERE sprint_id = ? AND assignee = ?
            """,
            (int(sprint_id), emp),
        ).fetchall()
        for r in rows:
            col = _normalize_kanban_column(r["kanban_column"] if "kanban_column" in r.keys() else None)
            if _normalize_scrum_status(r["status"]) == "done" and col != "done":
                col = "done"
            counts[col] = counts.get(col, 0) + 1
        done_n = counts["done"]
        doing_preview = [
            str(t[0] or "").strip()
            for t in conn.execute(
                """
                SELECT title FROM scrum_sprint_item
                WHERE sprint_id = ? AND assignee = ? AND lower(trim(kanban_column)) = 'doing'
                ORDER BY sort_order, id LIMIT 4
                """,
                (int(sprint_id), emp),
            ).fetchall()
        ]
        notes = _last_notes_for_assignee(conn, sprint_id, emp, 4)
        est_row = conn.execute(
            """
            SELECT COALESCE(SUM(estimate_hours), 0) AS h
            FROM scrum_sprint_item
            WHERE sprint_id = ? AND assignee = ?
            """,
            (int(sprint_id), emp),
        ).fetchone()
        com_row = conn.execute(
            """
            SELECT COALESCE(SUM(a.committed_hours), 0) AS h
            FROM scrum_item_activity a
            JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ? AND i.assignee = ?
            """,
            (int(sprint_id), emp),
        ).fetchone()
        est_total = float(est_row["h"] or 0) if est_row else 0.0
        committed_total = float(com_row["h"] or 0) if com_row else 0.0
        if app is not None:
            capacity_available = round(_available_hours_for_assignee_sprint_window(app, emp, sd, ed), 2)
        else:
            capacity_available = round(
                sum(
                    SCRUM_KANBAN_WEEKDAY_HOURS
                    for d in _daterange_inclusive(sd, ed)
                    if d.weekday() < 5
                ),
                2,
            )
        mx = max(est_total, committed_total, capacity_available, 0.01)
        stretch = max(0.0, committed_total - est_total)
        pct_capacity = round(100.0 * capacity_available / mx, 2)
        pct_est = round(100.0 * est_total / mx, 2)
        reg = min(committed_total, est_total)
        pct_committed_regular = round(100.0 * reg / mx, 2)
        pct_stretch = round(100.0 * stretch / mx, 2)
        task_count = sum(int(counts[c]) for c in SCRUM_KANBAN_COLUMNS)
        progress_pct: float | None
        if task_count > 0 and est_total > SCRUM_HOUR_EPS:
            progress_pct = round(100.0 * committed_total / est_total, 1)
        else:
            progress_pct = None
        kind_est = {code: 0.0 for code in SCRUM_TASK_KIND_CODES}
        kind_com = {code: 0.0 for code in SCRUM_TASK_KIND_CODES}
        for r in conn.execute(
            """
            SELECT task_kind, SUM(estimate_hours) AS h
            FROM scrum_sprint_item
            WHERE sprint_id = ? AND assignee = ?
            GROUP BY task_kind
            """,
            (int(sprint_id), emp),
        ).fetchall():
            code = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
            if code in kind_est:
                kind_est[code] += float(r["h"] or 0)
        for r in conn.execute(
            """
            SELECT i.task_kind, SUM(a.committed_hours) AS h
            FROM scrum_item_activity a
            JOIN scrum_sprint_item i ON i.id = a.item_id
            WHERE i.sprint_id = ? AND i.assignee = ?
            GROUP BY i.task_kind
            """,
            (int(sprint_id), emp),
        ).fetchall():
            code = _resolve_task_kind_code(conn, team_id, str(r["task_kind"] or ""))
            if code in kind_com:
                kind_com[code] += float(r["h"] or 0)

        def _kind_burn_pct(est: float, com: float) -> float | None:
            if est > SCRUM_HOUR_EPS:
                return round(100.0 * com / est, 1)
            return None

        burn_ndy_pct = _kind_burn_pct(kind_est["ndy"], kind_com["ndy"])
        burn_fsy_pct = _kind_burn_pct(kind_est["fsy"], kind_com["fsy"])
        burn_code_pct = _kind_burn_pct(kind_est["code"], kind_com["code"])
        burn_improvement_pct = _kind_burn_pct(kind_est["improvement"], kind_com["improvement"])
        burn_process_tools_pct = _kind_burn_pct(kind_est["process_tools"], kind_com["process_tools"])
        any_kind_est = any(kind_est[c] > SCRUM_HOUR_EPS for c in SCRUM_TASK_KIND_CODES)
        show_kind_burn_breakdown = bool(
            progress_pct is not None
            and progress_pct < SCRUM_TEAM_KIND_BURN_BREAKDOWN_BELOW_PCT
            and any_kind_est
        )
        collapsed_kind_lines: list[dict[str, object]] = []
        _kind_lbl = {
            "ndy": "NDY",
            "fsy": "FSY",
            "code": "CODE",
            "process_tools": "Process&Tools",
            "improvement": "Improvement",
        }
        for _code in SPRINT_TEAM_KIND_STACK_ORDER:
            est_k = float(kind_est[_code])
            if est_k <= SCRUM_HOUR_EPS:
                continue
            com_k = float(kind_com[_code])
            collapsed_kind_lines.append(
                {
                    "label": _kind_lbl.get(_code, _code.upper()),
                    "pct": _kind_burn_pct(est_k, com_k),
                    "com": round(com_k, 2),
                    "est": round(est_k, 2),
                }
            )
            if len(collapsed_kind_lines) >= 3:
                break
        out.append(
            {
                "name": emp,
                "counts": counts,
                "done_total": done_n,
                "doing_preview": doing_preview,
                "last_notes": notes,
                "est_total_hours": round(est_total, 2),
                "committed_total_hours": round(committed_total, 2),
                "capacity_available_hours": capacity_available,
                "stretch_load_hours": round(stretch, 2),
                "effort_scale_max": round(mx, 2),
                "pct_capacity_bar": pct_capacity,
                "pct_est_bar": pct_est,
                "pct_committed_regular_bar": pct_committed_regular,
                "pct_stretch_bar": pct_stretch,
                "task_count": task_count,
                "progress_pct": progress_pct,
                "burn_kind_ndy_pct": burn_ndy_pct,
                "burn_kind_fsy_pct": burn_fsy_pct,
                "burn_kind_code_pct": burn_code_pct,
                "burn_kind_improvement_pct": burn_improvement_pct,
                "burn_kind_process_tools_pct": burn_process_tools_pct,
                "burn_kind_ndy_est": round(kind_est["ndy"], 2),
                "burn_kind_ndy_com": round(kind_com["ndy"], 2),
                "burn_kind_fsy_est": round(kind_est["fsy"], 2),
                "burn_kind_fsy_com": round(kind_com["fsy"], 2),
                "burn_kind_code_est": round(kind_est["code"], 2),
                "burn_kind_code_com": round(kind_com["code"], 2),
                "burn_kind_improvement_est": round(kind_est["improvement"], 2),
                "burn_kind_improvement_com": round(kind_com["improvement"], 2),
                "burn_kind_process_tools_est": round(kind_est["process_tools"], 2),
                "burn_kind_process_tools_com": round(kind_com["process_tools"], 2),
                "show_kind_burn_breakdown": show_kind_burn_breakdown,
                "collapsed_kind_lines": collapsed_kind_lines,
            }
        )
    return out


def _sprint_team_detail_one_liner(
    members: list[dict],
    sprint_name: str,
    sprint_team_capacity_hours: float | None,
) -> str:
    """Single-line sprint summary for the team overview detailed page."""
    n_stick = sum(int(m.get("task_count") or 0) for m in members)
    tot_est = sum(float(m.get("est_total_hours") or 0) for m in members)
    tot_log = sum(float(m.get("committed_total_hours") or 0) for m in members)
    pend = sum(len(m.get("pending_portal") or []) for m in members)
    parts = [
        f"«{sprint_name}»",
        f"{n_stick} stickies",
        f"team est {tot_est:.1f}h",
        f"logged {tot_log:.1f}h",
    ]
    if sprint_team_capacity_hours is not None:
        parts.append(f"Sprint cap {float(sprint_team_capacity_hours):.1f}h (net weekdays after leave)")
    if pend:
        parts.append(f"{pend} portal change(s) pending")
    return " · ".join(parts)


def _sprint_team_overview_detailed_rows(
    app: Flask, conn: sqlite3.Connection, team_id: int, sprint_id: int, roster: Sequence[str]
) -> list[dict]:
    """Like ``_sprint_team_overview_rows`` plus per-sticky full activity log and burn metrics."""
    base = _sprint_team_overview_rows(app, conn, team_id, sprint_id, roster)
    for m in base:
        buckets = _load_kanban_cards(conn, sprint_id, m["name"], team_id)
        for col in SCRUM_KANBAN_COLUMNS:
            _enrich_kanban_burn_cards(conn, buckets[col])
        all_ids = [int(c["id"]) for col in SCRUM_KANBAN_COLUMNS for c in buckets[col]]
        act_map = _fetch_kanban_item_activity_log_map(conn, all_ids) if all_ids else {}
        checklist_map = _fetch_item_checklist_map(conn, all_ids) if all_ids else {}
        linked_files_map = _fetch_item_linked_files_map(conn, all_ids) if all_ids else {}
        stickies: list[dict] = []
        for col in SCRUM_KANBAN_COLUMNS:
            for c in buckets[col]:
                iid = int(c["id"])
                row = dict(c)
                row["kanban_column"] = col
                row["kanban_column_label"] = _kanban_column_public_label(col)
                acts = list(act_map.get(iid, []))
                # Stand-ups are listed under "In progress updates"; omit duplicate doing→doing from full log.
                if row.get("standup_updates"):
                    acts = [a for a in acts if not (a.get("from_column") == "doing" and a.get("to_column") == "doing")]
                row["activity_log"] = acts
                row["checklist"] = checklist_map.get(iid, [])
                row["linked_files"] = linked_files_map.get(iid, [])
                stickies.append(row)
        m["stickies"] = stickies
    return base


def _build_scrum_item_detail_payload(
    conn: sqlite3.Connection,
    team_id: int,
    sprint_id: int,
    item_id: int,
) -> dict[str, Any] | None:
    """Single sticky detail for Kanban task-detail modal (checklist + activity)."""
    item_row = conn.execute(
        """
        SELECT i.id FROM scrum_sprint_item i
        JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE i.id = ? AND i.sprint_id = ? AND s.team_id = ?
        """,
        (int(item_id), int(sprint_id), int(team_id)),
    ).fetchone()
    if not item_row:
        return None
    assignee_row = conn.execute(
        "SELECT assignee FROM scrum_sprint_item WHERE id = ?", (int(item_id),)
    ).fetchone()
    assignee = str(assignee_row["assignee"] or "").strip() if assignee_row else ""
    if not assignee:
        return None
    buckets = _load_kanban_cards(conn, int(sprint_id), assignee, int(team_id))
    card: dict[str, Any] | None = None
    col_key = ""
    for col in SCRUM_KANBAN_COLUMNS:
        for c in buckets[col]:
            if int(c["id"]) == int(item_id):
                card = dict(c)
                col_key = col
                break
        if card:
            break
    if not card:
        return None
    card["kanban_column"] = col_key
    card["kanban_column_label"] = _kanban_column_public_label(col_key)
    iid = int(item_id)
    checklist = _fetch_item_checklist_map(conn, [iid]).get(iid, [])
    act_map = _fetch_kanban_item_activity_log_map(conn, [iid])
    acts = list(act_map.get(iid, []))
    standups = list(card.get("standup_updates") or [])
    if standups:
        acts = [
            a
            for a in acts
            if not (a.get("from_column") == "doing" and a.get("to_column") == "doing")
        ]
    return {
        "item": card,
        "checklist": checklist,
        "standup_updates": standups,
        "activity_log": acts,
    }


def _scrum_checklist_item_on_team(conn: sqlite3.Connection, item_id: int, team_id: int) -> bool:
    row = conn.execute(
        """
        SELECT i.id FROM scrum_sprint_item i
        JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE i.id = ? AND s.team_id = ?
        """,
        (int(item_id), int(team_id)),
    ).fetchone()
    return row is not None


def _scrum_portal_checklist_updated_by(pu: dict[str, Any], roster_name: str) -> str:
    email = str(pu.get("email") or "").strip()
    return email or roster_name or "portal"


def _scrum_attachment_root_dir(app: Flask) -> Path:
    root = Path(app.config["DB_PATH"]).resolve().parent / "scrum_item_attachments"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _attachment_allowed_suffix(original_name: str) -> str | None:
    base = (original_name or "").strip().replace("\\", "/")
    if not base or ".." in base:
        return None
    low = base.lower()
    for suf in sorted(SCRUM_ITEM_ATTACHMENT_SUFFIXES, key=len, reverse=True):
        if low.endswith(suf):
            return suf
    return None


def _fetch_item_attachments_map(conn: sqlite3.Connection, item_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not item_ids:
        return {}
    ids = sorted({int(i) for i in item_ids})
    ph = ",".join("?" * len(ids))
    out: dict[int, list[dict[str, Any]]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT id, item_id, original_filename, size_bytes, created_at, rel_path
        FROM scrum_sprint_item_attachment
        WHERE item_id IN ({ph})
        ORDER BY id ASC
        """,
        ids,
    ):
        iid = int(r["item_id"])
        if iid not in out:
            continue
        out[iid].append(
            {
                "id": int(r["id"]),
                "original_filename": (str(r["original_filename"] or "")).strip()[:200],
                "size_bytes": int(r["size_bytes"] or 0),
                "created_at": str(r["created_at"] or ""),
                "rel_path": (str(r["rel_path"] or "")).strip(),
            }
        )
    return out


def _fetch_item_linked_files_map(conn: sqlite3.Connection, item_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not item_ids:
        return {}
    ids = sorted({int(i) for i in item_ids})
    ph = ",".join("?" * len(ids))
    out: dict[int, list[dict[str, Any]]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT id, item_id, sort_order, display_name, url, auth_user, auth_pass, updated_at, updated_by
        FROM scrum_item_linked_file
        WHERE item_id IN ({ph})
        ORDER BY item_id ASC, sort_order ASC, id ASC
        """,
        ids,
    ):
        iid = int(r["item_id"])
        if iid not in out:
            continue
        out[iid].append(
            {
                "id": int(r["id"]),
                "sort_order": int(r["sort_order"] or 0),
                "display_name": str(r["display_name"] or ""),
                "url": str(r["url"] or ""),
                "auth_user": str(r["auth_user"] or ""),
                "has_auth": bool((r["auth_user"] or "").strip()),
                "updated_at": str(r["updated_at"] or ""),
                "updated_by": str(r["updated_by"] or ""),
            }
        )
    return out


def _fetch_item_checklist_map(conn: sqlite3.Connection, item_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not item_ids:
        return {}
    ids = sorted({int(i) for i in item_ids})
    ph = ",".join("?" * len(ids))
    out: dict[int, list[dict[str, Any]]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT id, item_id, sort_order, items_to_finish, status, le_to_complete, done_till_date, updated_at, updated_by
        FROM scrum_item_checklist
        WHERE item_id IN ({ph})
        ORDER BY item_id ASC, sort_order ASC, id ASC
        """,
        ids,
    ):
        iid = int(r["item_id"])
        if iid not in out:
            continue
        out[iid].append(
            {
                "id": int(r["id"]),
                "sort_order": int(r["sort_order"] or 0),
                "items_to_finish": str(r["items_to_finish"] or ""),
                "status": str(r["status"] or ""),
                "le_to_complete": str(r["le_to_complete"] or ""),
                "done_till_date": str(r["done_till_date"] or ""),
                "updated_at": str(r["updated_at"] or ""),
                "updated_by": str(r["updated_by"] or ""),
            }
        )
    return out


def _scrum_attachment_auth_row(conn: sqlite3.Connection, attachment_id: int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT a.id AS aid, a.item_id, a.rel_path, a.original_filename, a.content_type, a.size_bytes, a.created_at,
               i.assignee, i.sprint_id, s.team_id AS team_id
        FROM scrum_sprint_item_attachment a
        JOIN scrum_sprint_item i ON i.id = a.item_id
        JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE a.id = ?
        """,
        (int(attachment_id),),
    ).fetchone()


def _scrum_attachment_build_preview_html(path: Path, original_fn: str) -> str | None:
    """
    Build a small self-contained HTML document for iframe preview (first worksheet of Excel
    workbooks, or CSV as a simple table). Returns None if the format is unsupported or read fails.
    """
    import csv
    import html as html_mod

    low = (original_fn or "").lower()
    title = html_mod.escape((Path(original_fn).name or "preview")[:160])
    css = (
        "body{margin:0;padding:0.5rem;font:12px/1.35 system-ui,Segoe UI,sans-serif;background:#0f172a;color:#e2e8f0;}"
        ".meta{margin:0 0 0.5rem;font-size:0.75rem;color:#94a3b8;}"
        ".wrap{overflow:auto;max-height:calc(100vh - 2rem);}"
        "table{border-collapse:collapse;width:100%;}"
        "td{border:1px solid #334155;padding:4px 6px;vertical-align:top;white-space:nowrap;max-width:18rem;"
        "overflow:hidden;text-overflow:ellipsis;}"
    )
    head = (
        "<!DOCTYPE html><html><head><meta charset=\"utf-8\"/><title>"
        + title
        + "</title><style>"
        + css
        + "</style></head><body>"
    )
    tail = "</body></html>"

    def _td(val: object) -> str:
        s = "" if val is None else str(val)
        return "<td>" + html_mod.escape(s[:8000]) + "</td>"

    if low.endswith(".csv"):
        try:
            raw = path.read_bytes()
        except OSError:
            return None
        text = raw.decode("utf-8-sig", errors="replace")
        rows = []
        try:
            for i, row in enumerate(csv.reader(text.splitlines())):
                if i >= 500:
                    break
                rows.append(row[:45])
        except csv.Error:
            return None
        if not rows:
            return head + "<p>(empty file)</p>" + tail
        out = [head, "<div class=\"wrap\"><table>"]
        for row in rows:
            out.append("<tr>")
            for j in range(max(len(row), 1)):
                out.append(_td(row[j] if j < len(row) else ""))
            out.append("</tr>")
        out.append("</table></div>")
        return "".join(out) + tail

    if low.endswith((".xlsx", ".xlsm", ".xltx")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None
        wb = None
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            if not wb.sheetnames:
                return head + "<p>(empty workbook)</p>" + tail
            ws = wb[wb.sheetnames[0]]
            meta = html_mod.escape(str(ws.title)[:120])
            parts: list[str] = [head, '<p class="meta">Sheet: ' + meta + '</p><div class="wrap"><table>']
            for row in ws.iter_rows(min_row=1, max_row=220, max_col=40, values_only=True):
                parts.append("<tr>")
                for v in row:
                    parts.append(_td(v))
                parts.append("</tr>")
            parts.append("</table></div>")
            return "".join(parts) + tail
        except Exception:
            return None
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    return None


def _scrum_attachment_upload_core(
    app: Flask,
    conn: sqlite3.Connection,
    *,
    team_id: int,
    sprint_id: int,
    item_id: int,
    roster_gate: str | None,
    fobj,
) -> tuple[bool, str, str]:
    """
    Validate and store one uploaded attachment. ``roster_gate`` when set must equal the sticky assignee (portal).
    Returns ``(success, flash_message, assignee_for_redirect)``. Commits on success; rolls back on failure after disk cleanup attempt.
    """
    if not _sprint_row_for_team(conn, sprint_id, team_id):
        return False, "Invalid sprint.", ""
    fr = _sprint_board_frozen_reason(conn, sprint_id)
    if fr:
        msg = _sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only."
        return False, msg, ""
    row = conn.execute(
        "SELECT assignee FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
        (item_id, sprint_id),
    ).fetchone()
    if not row:
        return False, "Item not found.", ""
    assignee = (str(row["assignee"] or "")).strip()
    if roster_gate is not None and assignee != roster_gate.strip():
        return False, "You can only add files to stickies assigned to you.", assignee
    if not fobj or not (fobj.filename or "").strip():
        return False, "Choose a file to upload.", assignee
    orig = (fobj.filename or "").strip()
    suf = _attachment_allowed_suffix(orig)
    if not suf:
        return (
            False,
            "File type not allowed. Use a common document or image extension "
            f"({', '.join(sorted(SCRUM_ITEM_ATTACHMENT_SUFFIXES)[:8])}…).",
            assignee,
        )
    n_att = int(
        conn.execute(
            "SELECT COUNT(*) AS c FROM scrum_sprint_item_attachment WHERE item_id = ?",
            (item_id,),
        ).fetchone()["c"]
    )
    if n_att >= SCRUM_ITEM_ATTACHMENTS_MAX_PER_STICKY:
        return False, f"Maximum {SCRUM_ITEM_ATTACHMENTS_MAX_PER_STICKY} attachments per sticky.", assignee
    blob = fobj.read(SCRUM_ITEM_ATTACHMENT_MAX_BYTES + 1)
    if len(blob) > SCRUM_ITEM_ATTACHMENT_MAX_BYTES:
        return False, f"File too large (max {SCRUM_ITEM_ATTACHMENT_MAX_BYTES // (1024 * 1024)} MB).", assignee
    ct = (fobj.mimetype or "application/octet-stream").strip()[:200] or "application/octet-stream"
    safe_tail = secure_filename(orig) or "file"
    if len(safe_tail) > 120:
        safe_tail = safe_tail[:120]
    disk_name = f"{secrets.token_hex(16)}{suf}"
    rel_path = f"{team_id}/{item_id}/{disk_name}"
    dest_dir = _scrum_attachment_root_dir(app) / str(team_id) / str(item_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / disk_name
    try:
        dest_path.write_bytes(blob)
    except OSError as ex:
        return False, f"Could not save file: {ex}", assignee
    ts = _utc_stamp()
    disp_name = orig[:200]
    try:
        conn.execute(
            """
            INSERT INTO scrum_sprint_item_attachment (item_id, rel_path, original_filename, content_type, size_bytes, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (item_id, rel_path, disp_name, ct, len(blob), ts),
        )
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ? AND sprint_id = ?",
            (ts, item_id, sprint_id),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        try:
            dest_path.unlink(missing_ok=True)
        except OSError:
            pass
        return False, "Could not record attachment.", assignee
    return True, "Attachment added.", assignee


def _scrum_attachment_upload_file_objects_from_request() -> list:
    """Non-empty uploaded file objects from ``files`` (multi-select) or legacy single ``file``."""
    found: list = []
    for f in request.files.getlist("files"):
        if f and (getattr(f, "filename", None) or "").strip():
            found.append(f)
    if not found:
        f = request.files.get("file")
        if f and (f.filename or "").strip():
            found.append(f)
    return found


def _scrum_attachment_upload_run_batch(
    app: Flask,
    conn: sqlite3.Connection,
    *,
    team_id: int,
    sprint_id: int,
    item_id: int,
    roster_gate: str | None,
) -> tuple[str, str, str]:
    """
    Upload every file in the current request for one sticky.
    Returns ``(flash_category, flash_message, assignee_for_redirect)``.
    """
    files = _scrum_attachment_upload_file_objects_from_request()
    if not files:
        return "error", "Choose at least one file to upload.", ""
    n_ok = 0
    assignee_out = ""
    first_err: str | None = None
    for fobj in files:
        ok, msg, asg = _scrum_attachment_upload_core(
            app, conn, team_id=team_id, sprint_id=sprint_id, item_id=item_id, roster_gate=roster_gate, fobj=fobj
        )
        if asg:
            assignee_out = asg
        if ok:
            n_ok += 1
        elif first_err is None:
            first_err = msg
    n = len(files)
    if n_ok == n:
        if n == 1:
            return "success", "Attachment added.", assignee_out
        return "success", f"Added {n_ok} attachments.", assignee_out
    if n_ok:
        tail = (" " + first_err) if first_err else ""
        return "success", f"Added {n_ok} of {n} files.{tail}", assignee_out
    return "error", first_err or "Upload failed.", assignee_out


def _scrum_attachment_delete_core(
    app: Flask,
    conn: sqlite3.Connection,
    *,
    team_id: int,
    sprint_id: int,
    attachment_id: int,
    roster_gate: str | None,
) -> tuple[bool, str, str]:
    """``roster_gate`` when set must match the sticky assignee. Returns (success, message, assignee). Commits on success."""
    if not _sprint_row_for_team(conn, sprint_id, team_id):
        return False, "Invalid sprint.", ""
    fr = _sprint_board_frozen_reason(conn, sprint_id)
    if fr:
        msg = _sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only."
        return False, msg, ""
    ar = _scrum_attachment_auth_row(conn, int(attachment_id))
    if not ar or int(ar["team_id"]) != team_id or int(ar["sprint_id"]) != int(sprint_id):
        return False, "Attachment not found.", ""
    assignee = (str(ar["assignee"] or "")).strip()
    if roster_gate is not None and assignee != roster_gate.strip():
        return False, "You can only remove files from your own stickies.", assignee
    rel = (str(ar["rel_path"]) or "").strip().replace("\\", "/")
    if ".." in rel or rel.startswith("/"):
        return False, "Invalid attachment path.", assignee
    root = _scrum_attachment_root_dir(app)
    path = root / rel
    try:
        conn.execute("DELETE FROM scrum_sprint_item_attachment WHERE id = ?", (int(attachment_id),))
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ? AND sprint_id = ?",
            (ts, int(ar["item_id"]), sprint_id),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        return False, "Could not remove attachment.", assignee
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass
    return True, "Attachment removed.", assignee


def _load_kanban_cards(conn: sqlite3.Connection, sprint_id: int, assignee: str, team_id: int) -> dict[str, list[dict]]:
    buckets: dict[str, list[dict]] = {c: [] for c in SCRUM_KANBAN_COLUMNS}
    _ensure_team_task_kinds(conn, team_id)
    for r in conn.execute(
        """
        SELECT i.id, i.title, i.estimate_hours, i.status, i.notes, i.dod, i.done_artifacts, i.sort_order, i.kanban_column, i.task_kind,
               i.sticky_color_hex, i.area, k.label AS kind_label, k.color_hex AS kind_color_hex
        FROM scrum_sprint_item i
        LEFT JOIN scrum_team_task_kind k ON k.team_id = ? AND k.code = i.task_kind
        WHERE i.sprint_id = ? AND i.assignee = ?
        ORDER BY i.sort_order ASC, i.id ASC
        """,
        (int(team_id), int(sprint_id), assignee),
    ):
        col = _normalize_kanban_column(r["kanban_column"] if "kanban_column" in r.keys() else None)
        if _normalize_scrum_status(r["status"]) == "done" and col != "done":
            col = "done"
        raw_kind = (r["task_kind"] or "").strip()
        code = _coerce_sprint_item_task_kind(conn, team_id, raw_kind or None)
        sticky_hex = (r["sticky_color_hex"] or "").strip() if "sticky_color_hex" in r.keys() else ""
        kind_color = (r["kind_color_hex"] or "").strip() if "kind_color_hex" in r.keys() else ""
        if sticky_hex and _SCRUM_HEX_COLOR.match(sticky_hex):
            kind_color = _normalize_hex_color(sticky_hex)
        elif not kind_color:
            kind_color = _task_kind_color_for_item(conn, team_id, code, None)
        kind_label = (r["kind_label"] or "").strip() if "kind_label" in r.keys() else ""
        if not kind_label:
            kind_label = code.upper()
        da_list = _parse_done_artifacts_db(
            str(r["done_artifacts"] or "[]") if "done_artifacts" in r.keys() else "[]"
        )
        buckets[col].append(
            {
                "id": int(r["id"]),
                "title": (r["title"] or "").strip(),
                "estimate_hours": float(r["estimate_hours"] or 0),
                "notes": (r["notes"] or "").strip(),
                "workflow_status": _normalize_scrum_status(str(r["status"] or "")),
                "dod": (r["dod"] or "").strip() if "dod" in r.keys() else "",
                "done_artifacts": da_list,
                "done_artifacts_lines": _done_artifacts_to_lines(da_list),
                "task_kind": code,
                "kind_label": kind_label,
                "kind_color_hex": kind_color,
                "area": (str(r["area"] or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]),
                "sticky_color_hex_raw": _normalize_hex_color(sticky_hex)
                if sticky_hex and _SCRUM_HEX_COLOR.match(sticky_hex)
                else "",
                "column": col,
            }
        )
    all_item_ids = [int(c["id"]) for col in SCRUM_KANBAN_COLUMNS for c in buckets[col]]
    stand_map = _fetch_kanban_standup_updates_map(conn, all_item_ids)
    for col in SCRUM_KANBAN_COLUMNS:
        for c in buckets[col]:
            c["standup_updates"] = stand_map.get(int(c["id"]), [])
    if buckets["done"]:
        done_ids = [int(c["id"]) for c in buckets["done"]]
        act_map = _fetch_kanban_item_activity_log_map(conn, done_ids)
        for c in buckets["done"]:
            c["activity_log"] = act_map.get(int(c["id"]), [])
    _enrich_kanban_burn_cards(conn, buckets["doing"])
    _enrich_kanban_burn_cards(conn, buckets["done"])
    _enrich_kanban_appreciation(conn, buckets)
    att_map = _fetch_item_attachments_map(conn, all_item_ids)
    for col in SCRUM_KANBAN_COLUMNS:
        for c in buckets[col]:
            c["attachments"] = att_map.get(int(c["id"]), [])
    return buckets


def _enrich_kanban_appreciation(conn: sqlite3.Connection, buckets: dict[str, list[dict]]) -> None:
    """Attach appreciation_count and appreciation_preview (latest) for Well Done comments."""
    all_cards: list[dict] = [c for col in buckets.values() for c in col]
    for c in all_cards:
        c["appreciation_count"] = 0
        c["appreciation_preview"] = ""
        c["appreciation_messages"] = []
    if not all_cards:
        return
    ids = sorted({int(c["id"]) for c in all_cards})
    if not ids:
        return
    ph = ",".join("?" * len(ids))
    counts = {
        int(r["item_id"]): int(r["c"] or 0)
        for r in conn.execute(
            f"SELECT item_id, COUNT(*) AS c FROM scrum_item_appreciation WHERE item_id IN ({ph}) GROUP BY item_id",
            ids,
        )
    }
    previews: dict[int, str] = {}
    for r in conn.execute(
        f"""
        SELECT a.item_id, a.comment, a.author
        FROM scrum_item_appreciation a
        INNER JOIN (
            SELECT item_id, MAX(id) AS mid FROM scrum_item_appreciation WHERE item_id IN ({ph}) GROUP BY item_id
        ) t ON a.id = t.mid
        """,
        ids,
    ):
        iid = int(r["item_id"])
        body = (str(r["comment"] or "")).strip()
        auth = (str(r["author"] or "")).strip()
        line = f"{auth}: {body}" if auth else body
        previews[iid] = line[:500] if line else "Well done"
    # Fetch all appreciation messages for portal display
    all_messages: dict[int, list[dict]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT item_id, author, comment, created_at
        FROM scrum_item_appreciation
        WHERE item_id IN ({ph})
        ORDER BY created_at ASC, id ASC
        """,
        ids,
    ):
        iid2 = int(r["item_id"])
        if iid2 in all_messages:
            all_messages[iid2].append({
                "author": (str(r["author"] or "")).strip(),
                "comment": (str(r["comment"] or "")).strip(),
                "created_at_display": _format_activity_ts(str(r["created_at"] or "")),
            })
    for c in all_cards:
        iid = int(c["id"])
        n = counts.get(iid, 0)
        c["appreciation_count"] = n
        if n:
            c["appreciation_preview"] = previews.get(iid, "Appreciation")
        c["appreciation_messages"] = all_messages.get(iid, [])


def _format_activity_ts(raw: str) -> str:
    s = (raw or "").strip().replace("+00:00", "Z")
    if not s:
        return ""
    if "T" in s:
        date_part, rest = s.split("T", 1)
        time_part = rest.replace("Z", "").split(".")[0][:8]
        if len(time_part) >= 5:
            time_part = time_part[:5]
        return f"{date_part} {time_part} UTC"
    return s[:19]


def _strip_appended_standup_lines_from_notes(notes: str) -> str:
    """Remove legacy lines appended by older stand-up saves (started with [YYYY-MM-DD])."""
    lines = (notes or "").split("\n")
    kept: list[str] = []
    for line in lines:
        s = line.strip()
        if len(s) >= 12 and s.startswith("[") and s[1:11].count("-") == 2 and s[11] == "]":
            continue
        kept.append(line)
    return "\n".join(kept).strip()


def _fetch_kanban_standup_updates_map(conn: sqlite3.Connection, item_ids: list[int]) -> dict[int, list[dict]]:
    """In-progress stand-up rows (doing→doing) keyed by item id."""
    if not item_ids:
        return {}
    ids = sorted({int(i) for i in item_ids})
    ph = ",".join("?" * len(ids))
    standups: dict[int, list[dict]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT id, item_id, body, committed_hours, created_at
        FROM scrum_item_activity
        WHERE item_id IN ({ph}) AND lower(trim(from_column)) = 'doing' AND lower(trim(to_column)) = 'doing'
        ORDER BY created_at ASC, id ASC
        """,
        ids,
    ):
        iid = int(r["item_id"])
        if iid not in standups:
            continue
        standups[iid].append(
            {
                "id": r["id"],
                "body": (r["body"] or "").strip(),
                "committed_hours": float(r["committed_hours"] or 0),
                "created_at": str(r["created_at"] or ""),
                "created_at_display": _format_activity_ts(str(r["created_at"] or "")),
            }
        )
    return standups


def _fetch_kanban_item_activity_log_map(conn: sqlite3.Connection, item_ids: list[int]) -> dict[int, list[dict]]:
    """Chronological board activity (column moves + stand-ups) for Done history."""
    if not item_ids:
        return {}
    ids = sorted({int(i) for i in item_ids})
    ph = ",".join("?" * len(ids))
    by_item: dict[int, list[dict]] = {i: [] for i in ids}
    for r in conn.execute(
        f"""
        SELECT id, item_id, body, committed_hours, from_column, to_column, created_at
        FROM scrum_item_activity
        WHERE item_id IN ({ph})
        ORDER BY created_at ASC, id ASC
        """,
        ids,
    ):
        iid = int(r["item_id"])
        if iid not in by_item:
            continue
        fc = _normalize_kanban_column(str(r["from_column"] or ""))
        tc = _normalize_kanban_column(str(r["to_column"] or ""))
        if fc == "doing" and tc == "doing":
            kind_label = "Stand-up"
        else:
            kind_label = "Move"
        ch_raw = r["committed_hours"]
        try:
            chf = float(ch_raw) if ch_raw is not None else None
        except (TypeError, ValueError):
            chf = None
        by_item[iid].append(
            {
                "id": int(r["id"]),
                "body": (str(r["body"] or "")).strip(),
                "committed_hours": chf,
                "from_column": fc,
                "to_column": tc,
                "from_label": _kanban_column_public_label(fc),
                "to_label": _kanban_column_public_label(tc),
                "created_at": str(r["created_at"] or ""),
                "created_at_display": _format_activity_ts(str(r["created_at"] or "")),
                "kind_label": kind_label,
            }
        )
    return by_item


def _kanban_burn_metrics(estimate_hours: float, committed_sum: float) -> dict[str, Any]:
    """Shared est vs logged burn math for sticky cards and activity-update API JSON."""
    est = float(estimate_hours or 0)
    committed = float(committed_sum or 0)
    mx = max(est, committed, 0.01)
    stretch = max(0.0, committed - est)
    pct_est = round(100.0 * est / mx, 2)
    reg = min(committed, est)
    pct_committed_regular = round(100.0 * reg / mx, 2)
    pct_stretch = round(100.0 * stretch / mx, 2)
    burn_pct: float | None
    if est > SCRUM_HOUR_EPS:
        burn_pct = round(100.0 * committed / est, 1)
    else:
        burn_pct = None
    return {
        "estimate_hours": round(est, 2),
        "committed_logged_hours": round(committed, 2),
        "burn_pct": burn_pct,
        "effort_scale_max": round(mx, 2),
        "pct_est_bar": pct_est,
        "pct_committed_regular_bar": pct_committed_regular,
        "pct_stretch_bar": pct_stretch,
        "stretch_load_hours": round(stretch, 2),
    }


def _kanban_single_item_burn_payload(conn: sqlite3.Connection, item_id: int) -> dict[str, Any] | None:
    """Recompute task burn totals from DB (sum of scrum_item_activity.committed_hours for this sticky)."""
    row = conn.execute(
        "SELECT estimate_hours FROM scrum_sprint_item WHERE id = ?",
        (int(item_id),),
    ).fetchone()
    if not row:
        return None
    est = float(row["estimate_hours"] or 0)
    srow = conn.execute(
        "SELECT COALESCE(SUM(committed_hours), 0) AS h FROM scrum_item_activity WHERE item_id = ?",
        (int(item_id),),
    ).fetchone()
    committed = float(srow["h"] or 0)
    return _kanban_burn_metrics(est, committed)


def _enrich_kanban_burn_cards(conn: sqlite3.Connection, kanban_cards: list[dict]) -> None:
    """Task burn %, logged vs estimate, and bar widths (Doing + Done sticky cards)."""
    if not kanban_cards:
        return
    ids = [int(c["id"]) for c in kanban_cards]
    ph = ",".join("?" * len(ids))
    sums = {
        int(r["item_id"]): float(r["h"] or 0)
        for r in conn.execute(
            f"SELECT item_id, COALESCE(SUM(committed_hours), 0) AS h FROM scrum_item_activity WHERE item_id IN ({ph}) GROUP BY item_id",
            ids,
        )
    }
    for card in kanban_cards:
        iid = int(card["id"])
        est = float(card.get("estimate_hours") or 0)
        committed = float(sums.get(iid, 0.0))
        m = _kanban_burn_metrics(est, committed)
        for k, v in m.items():
            card[k] = v
        card["notes_display"] = _strip_appended_standup_lines_from_notes(str(card.get("notes") or ""))


def _scrum_distinct_area_suggestions(conn: sqlite3.Connection, team_id: int, q: str, limit: int = 25) -> list[str]:
    """Distinct non-empty Area strings for this team, filtered by substring (case-insensitive)."""
    raw = (q or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
    if not raw:
        return []
    needle = raw.casefold()
    rows = conn.execute(
        """
        SELECT DISTINCT TRIM(i.area) AS a
        FROM scrum_sprint_item i
        INNER JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE s.team_id = ?
          AND LENGTH(TRIM(COALESCE(i.area, ''))) > 0
          AND instr(lower(trim(i.area)), ?) > 0
        ORDER BY TRIM(i.area) COLLATE NOCASE ASC
        LIMIT ?
        """,
        (int(team_id), needle, int(limit)),
    ).fetchall()
    out: list[str] = []
    seen: set[str] = set()
    for r in rows:
        t = (str(r["a"]) if r["a"] is not None else "").strip()
        if not t:
            continue
        k = t.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(t)
    return out


def _item_belongs_to_sprint_team(
    conn: sqlite3.Connection, item_id: int, sprint_id: int, team_id: int, assignee: str
) -> bool:
    row = conn.execute(
        """
        SELECT i.id FROM scrum_sprint_item i
        JOIN scrum_sprint s ON s.id = i.sprint_id
        WHERE i.id = ? AND i.sprint_id = ? AND s.team_id = ? AND i.assignee = ?
        """,
        (int(item_id), int(sprint_id), int(team_id), assignee),
    ).fetchone()
    return row is not None


def _sprint_team_id(conn: sqlite3.Connection, sprint_id: int) -> int | None:
    row = conn.execute("SELECT team_id FROM scrum_sprint WHERE id = ?", (int(sprint_id),)).fetchone()
    if not row:
        return None
    return int(row["team_id"])


SCRUM_PORTAL_PROPOSAL_ACTIONS: frozenset[str] = frozenset(
    {
        "item_move",
        "item_note",
        "item_activity_update",
        "item_add",
        "item_update_do",
        "item_update_done",
        "item_update_doing_estimate",
        "item_delete",
        "portal_checklist",
    }
)


def _approved_portal_activity_note(note: str) -> str:
    b = (note or "").strip()
    prefix = "[approved employee change] "
    if b.startswith(prefix):
        b = b[len(prefix):]
    return b[:2000]


def _insert_scrum_portal_proposal(
    conn: sqlite3.Connection,
    team_id: int,
    sprint_id: int,
    item_id: int | None,
    action: str,
    proposer_roster_name: str,
    payload: dict,
) -> None:
    if action not in SCRUM_PORTAL_PROPOSAL_ACTIONS:
        raise ValueError("bad proposal action")
    ts = _utc_stamp()
    cursor = conn.execute(
        """
        INSERT INTO scrum_portal_proposal
        (team_id, sprint_id, item_id, proposer_roster_name, action, payload_json, status, created_at, resolved_at, resolution_note)
        VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, NULL, '')
        """,
        (
            int(team_id),
            int(sprint_id),
            int(item_id) if item_id is not None else None,
            proposer_roster_name,
            action,
            json.dumps(payload),
            ts,
        ),
    )
    # Always apply employee changes immediately — no manager approval step.
    proposal_id = cursor.lastrowid
    prop = conn.execute("SELECT * FROM scrum_portal_proposal WHERE id = ?", (proposal_id,)).fetchone()
    if prop:
        ok, _err = _apply_scrum_portal_proposal_core(conn, prop, int(team_id))
        if ok:
            conn.execute(
                "UPDATE scrum_portal_proposal SET status = 'approved', resolved_at = ?, resolution_note = 'Auto-saved' WHERE id = ?",
                (_utc_stamp(), proposal_id),
            )


def _count_pending_scrum_portal_proposals(conn: sqlite3.Connection, team_id: int, sprint_id: int | None = None) -> int:
    if sprint_id is None:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM scrum_portal_proposal WHERE team_id = ? AND status = 'pending'",
            (int(team_id),),
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM scrum_portal_proposal WHERE team_id = ? AND sprint_id = ? AND status = 'pending'",
            (int(team_id), int(sprint_id)),
        ).fetchone()
    return int(row["c"]) if row else 0


def _count_pending_scrum_portal_proposals_for_proposer(conn: sqlite3.Connection, roster_name: str) -> int:
    row = conn.execute(
        "SELECT COUNT(*) AS c FROM scrum_portal_proposal WHERE proposer_roster_name = ? AND status = 'pending'",
        (roster_name,),
    ).fetchone()
    return int(row["c"]) if row else 0


def _portal_proposal_summary_line(action: str, payload: dict, item_title: str | None) -> str:
    t = (item_title or "").strip() or "(item)"
    if action == "item_move":
        col = _normalize_kanban_column(payload.get("to_column"))
        return f"Move «{t}» → {col}"
    if action == "item_note":
        return f"Stand-up / hours on «{t}»"
    if action == "item_add":
        title = (payload.get("title") or "").strip() or "New sticky"
        return f"Add sticky in To DO: «{title}»"
    if action == "item_update_do":
        return f"Edit plan (To DO) for «{t}»"
    if action == "item_update_done":
        return f"Edit Done details for «{t}»"
    if action == "item_update_doing_estimate":
        return f"Adjust In progress estimate for «{t}»"
    if action == "item_delete":
        return f"Delete «{t}»"
    if action == "portal_checklist":
        return f"My sprint checklist update for «{t}»"
    return action


def _kanban_column_public_label(col: str | None) -> str:
    c = _normalize_kanban_column(col)
    return {"backlog": "To do", "do": "To DO", "doing": "In progress", "done": "Done"}.get(c, c or "—")


def _pending_portal_proposals_grouped_for_sprint(
    conn: sqlite3.Connection, team_id: int, sprint_id: int
) -> dict[str, list[dict]]:
    """Pending portal proposals for this sprint, grouped by employee roster name (for team overview)."""
    action_labels = {
        "item_move": "Column move",
        "item_note": "Stand-up / hours",
        "item_add": "New sticky",
        "item_update_do": "Edit plan (To DO)",
        "item_update_done": "Edit Done",
        "item_update_doing_estimate": "Adjust estimate (In progress)",
        "item_delete": "Delete sticky",
        "portal_checklist": "Status & DoD",
    }
    rows = list(
        conn.execute(
            """
            SELECT p.id, p.item_id, p.proposer_roster_name, p.action, p.payload_json, p.created_at,
              (SELECT title FROM scrum_sprint_item WHERE id = p.item_id) AS item_title
            FROM scrum_portal_proposal p
            WHERE p.team_id = ? AND p.sprint_id = ? AND p.status = 'pending'
            ORDER BY p.created_at ASC
            """,
            (int(team_id), int(sprint_id)),
        )
    )
    by: dict[str, list[dict]] = {}
    for r in rows:
        emp = (r["proposer_roster_name"] or "").strip()
        if not emp:
            continue
        try:
            payload = json.loads(r["payload_json"] or "{}")
        except json.JSONDecodeError:
            payload = {}
        it = None
        if "item_title" in r.keys() and r["item_title"] is not None:
            it = (str(r["item_title"]) or "").strip() or None
        action = str(r["action"] or "")
        summary = _portal_proposal_summary_line(action, payload, it)
        card: dict = {
            "id": int(r["id"]),
            "created_at": str(r["created_at"] or ""),
            "created_at_short": (str(r["created_at"] or "")[:16].replace("T", " ")),
            "action": action,
            "action_label": action_labels.get(action, action),
            "summary": summary,
            "item_title": it or ("(new sticky)" if action == "item_add" else "(item)"),
            "dod_rows": None,
            "dod_all_met": None,
            "dod_checked_count": None,
            "note_preview": None,
            "target_column_label": None,
            "hours_hint": None,
            "estimate_hint": None,
            "from_column_label": None,
            "last_sync_body": None,
            "last_sync_at_short": None,
            "last_sync_hours": None,
            "artifact_items": [],
            "verification_urls": [],
            "proposed_standup_note": None,
        }
        da_raw = payload.get("done_artifacts_json")
        if isinstance(da_raw, str) and da_raw.strip():
            card["artifact_items"] = _parse_done_artifacts_db(da_raw)
        card["verification_urls"] = _normalize_verification_urls_mixed(payload.get("verification_urls"))
        if action == "portal_checklist":
            dod_json = (payload.get("portal_dod_json") or "[]").strip()
            states = _portal_dod_parse(dod_json)
            card["dod_rows"] = [
                {
                    "label": PORTAL_DOD_CHECKLIST_LABELS[i],
                    "checked": bool(states[i]) if i < len(states) else False,
                }
                for i in range(len(PORTAL_DOD_CHECKLIST_LABELS))
            ]
            card["dod_all_met"] = _portal_dod_all_done(states)
            card["dod_checked_count"] = sum(1 for x in card["dod_rows"] if x.get("checked"))
            card["target_column_label"] = _kanban_column_public_label(str(payload.get("kanban_column")))
        if action == "item_move":
            card["target_column_label"] = _kanban_column_public_label(str(payload.get("to_column")))
            card["note_preview"] = (payload.get("note") or "").strip()[:160]
            raw_ch = payload.get("committed_hours")
            ch: float | None
            try:
                ch = float(raw_ch) if raw_ch is not None and raw_ch != "" else None
            except (TypeError, ValueError):
                ch = 0.0
            if ch is not None and abs(ch) > SCRUM_HOUR_EPS:
                card["hours_hint"] = f"+{ch:g}h on this move"
            try:
                move_est = float(payload.get("estimate_hours", 0) or 0)
            except (TypeError, ValueError):
                move_est = 0.0
            if move_est > SCRUM_HOUR_EPS and str(payload.get("to_column") or "").strip().lower() == "do":
                card["estimate_hint"] = f"{move_est:g}h plan (To DO)"
        if action == "item_note":
            note_full = (payload.get("note") or "").strip()
            card["proposed_standup_note"] = note_full
            card["note_preview"] = note_full[:160]
            try:
                ch = float(payload.get("committed_hours", 0) or 0)
            except (TypeError, ValueError):
                ch = 0.0
            if abs(ch) > SCRUM_HOUR_EPS:
                card["hours_hint"] = f"+{ch:g}h"
        if action == "item_add":
            card["note_preview"] = (payload.get("title") or "").strip()[:160]
            try:
                est = float(payload.get("estimate_hours", 0) or 0)
            except (TypeError, ValueError):
                est = 0.0
            if est > SCRUM_HOUR_EPS:
                card["estimate_hint"] = f"{est:g}h estimate"
        if action in ("item_update_do", "item_update_done"):
            card["note_preview"] = (payload.get("notes") or "").strip()[:120]
        if action == "item_update_doing_estimate":
            try:
                est = float(payload.get("estimate_hours", 0) or 0)
            except (TypeError, ValueError):
                est = 0.0
            if est > SCRUM_HOUR_EPS:
                card["estimate_hint"] = f"{est:g}h plan estimate"
        raw_item = r["item_id"]
        item_id_int = int(raw_item) if raw_item is not None else None
        if item_id_int is not None:
            crow = conn.execute(
                "SELECT kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
                (item_id_int, int(sprint_id)),
            ).fetchone()
            if crow:
                card["from_column_label"] = _kanban_column_public_label(str(crow["kanban_column"]))
            act = conn.execute(
                """
                SELECT body, created_at, committed_hours
                FROM scrum_item_activity
                WHERE item_id = ? AND from_column = 'doing' AND to_column = 'doing'
                ORDER BY id DESC LIMIT 1
                """,
                (item_id_int,),
            ).fetchone()
            if act:
                card["last_sync_body"] = (str(act["body"] or "")).strip()
                card["last_sync_at_short"] = (str(act["created_at"] or "")[:16].replace("T", " "))
                try:
                    lh = float(act["committed_hours"] or 0)
                except (TypeError, ValueError):
                    lh = 0.0
                if abs(lh) > SCRUM_HOUR_EPS:
                    card["last_sync_hours"] = lh
        by.setdefault(emp, []).append(card)
    return by


def _redirect_after_portal_proposal(next_raw: str | None):
    n = (next_raw or "").strip()
    if n.startswith("/") and not n.startswith("//") and ".." not in n and "\n" not in n and "\r" not in n:
        return redirect(n)
    return redirect(url_for("portal_my_sprint_kanban"))


def _apply_scrum_portal_proposal_core(
    conn: sqlite3.Connection, proposal: sqlite3.Row, manager_team_id: int
) -> tuple[bool, str | None]:
    if int(proposal["team_id"]) != int(manager_team_id):
        return False, "team"
    if (proposal["status"] or "").strip().lower() != "pending":
        return False, "not_pending"
    action = (proposal["action"] or "").strip()
    if action not in SCRUM_PORTAL_PROPOSAL_ACTIONS:
        return False, "bad_action"
    sprint_id = int(proposal["sprint_id"])
    team_id = int(proposal["team_id"])
    fr = _sprint_board_frozen_reason(conn, sprint_id)
    if fr:
        return False, fr
    proposer = (proposal["proposer_roster_name"] or "").strip()
    try:
        payload = json.loads(proposal["payload_json"] or "{}")
    except json.JSONDecodeError:
        return False, "bad_payload"
    raw_item_id = proposal["item_id"]
    item_id: int | None = int(raw_item_id) if raw_item_id is not None else None
    ts = _utc_stamp()

    if action == "item_move":
        if item_id is None:
            return False, "missing_item"
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, proposer):
            return False, "forbidden"
        to_col = _normalize_kanban_column(payload.get("to_column"))
        note_raw = (payload.get("note") or "").strip()[:2000]
        body = _approved_portal_activity_note(note_raw)
        row = conn.execute(
            "SELECT kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row:
            return False, "item_gone"
        from_col = _normalize_kanban_column(row["kanban_column"] if row else None)
        if from_col == to_col:
            return True, None
        est_new, est_err = _estimate_hours_for_backlog_to_do_move(payload, from_col=from_col, to_col=to_col)
        if est_err:
            return False, est_err
        if est_new is not None:
            note_raw = ""
            body = _approved_portal_activity_note(note_raw)
        st = _status_for_kanban_column(to_col)
        if from_col == "backlog" and to_col == "do":
            ch = 0.0
        elif from_col == "do" and to_col == "doing":
            conn.execute("DELETE FROM scrum_item_activity WHERE item_id = ?", (item_id,))
            ch = _parse_committed_hours_for_kanban_move(
                payload.get("committed_hours"), from_col=from_col, to_col=to_col
            )
        elif from_col == "doing" and to_col == "do":
            conn.execute("DELETE FROM scrum_item_activity WHERE item_id = ?", (item_id,))
            ch = 0.0
        else:
            ch = _parse_committed_hours_for_kanban_move(
                payload.get("committed_hours"), from_col=from_col, to_col=to_col
            )
        artifacts_json: str | None = None
        if to_col == "done":
            raw_art = payload.get("done_artifacts_json")
            if raw_art is not None and isinstance(raw_art, str):
                artifacts_json = raw_art
            else:
                artifacts_json = "[]"
        if to_col == "done" and artifacts_json is not None:
            conn.execute(
                """
                UPDATE scrum_sprint_item
                SET kanban_column = ?, status = ?, updated_at = ?, done_artifacts = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (to_col, st, ts, artifacts_json, item_id, sprint_id),
            )
        elif est_new is not None:
            conn.execute(
                """
                UPDATE scrum_sprint_item
                SET kanban_column = ?, status = ?, updated_at = ?, estimate_hours = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (to_col, st, ts, est_new, item_id, sprint_id),
            )
        else:
            conn.execute(
                """
                UPDATE scrum_sprint_item
                SET kanban_column = ?, status = ?, updated_at = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (to_col, st, ts, item_id, sprint_id),
            )
        conn.execute(
            """
            INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (item_id, body, ch, from_col, to_col, ts),
        )
        return True, None

    if action == "item_note":
        if item_id is None:
            return False, "missing_item"
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, proposer):
            return False, "forbidden"
        row = conn.execute(
            "SELECT kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row or _normalize_kanban_column(row["kanban_column"]) != "doing":
            return False, "only_doing"
        note_raw = (payload.get("note") or "").strip()[:2000]
        ch = _parse_hours_field(str(payload.get("committed_hours", "0")))
        if not note_raw and ch <= SCRUM_HOUR_EPS:
            return False, "missing"
        body = _approved_portal_activity_note(note_raw)
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ?",
            (ts, item_id),
        )
        conn.execute(
            """
            INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
            VALUES (?, ?, ?, 'doing', 'doing', ?)
            """,
            (item_id, body, ch, ts),
        )
        return True, None

    if action == "item_activity_update":
        if item_id is None:
            return False, "missing_item"
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, proposer):
            return False, "forbidden"
        activity_id = payload.get("activity_id")
        if not activity_id:
            return False, "missing_activity_id"
        note_raw = (payload.get("note") or "").strip()[:2000]
        ch = _parse_hours_field(str(payload.get("committed_hours", "0")))
        body = _approved_portal_activity_note(note_raw)
        conn.execute(
            "UPDATE scrum_item_activity SET body = ?, committed_hours = ? WHERE id = ? AND item_id = ?",
            (body, ch, activity_id, item_id),
        )
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ?",
            (ts, item_id),
        )
        return True, None

    if action == "item_add":
        if not _sprint_row_for_team(conn, sprint_id, team_id):
            return False, "sprint_gone"
        title = (payload.get("title") or "").strip()[:500]
        notes = (payload.get("notes") or "").strip()[:2000]
        dod = (payload.get("dod") or "").strip()[:4000]
        est = _parse_hours_field(str(payload.get("estimate_hours", "0")))
        raw_tc = (payload.get("task_kind_code") or "ndy").strip().lower()[:64]
        if raw_tc in SCRUM_LEGACY_TASK_KIND_MAP:
            raw_tc = SCRUM_LEGACY_TASK_KIND_MAP[raw_tc]
        tcode = raw_tc if raw_tc in SCRUM_TASK_KIND_CODES else "ndy"
        if not title or est <= SCRUM_HOUR_EPS:
            return False, "missing"
        area = (payload.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        _ensure_team_task_kinds(conn, team_id)
        st = _status_for_kanban_column("do")
        mx_row = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) + 1 AS n FROM scrum_sprint_item WHERE sprint_id = ?",
            (sprint_id,),
        ).fetchone()
        mx = int(mx_row["n"] if mx_row is not None else 0)
        conn.execute(
            """
            INSERT INTO scrum_sprint_item
            (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind, sticky_color_hex, area)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'do', ?, ?, ?)
            """,
            (sprint_id, proposer, title, est, st, notes, dod, mx, ts, ts, tcode, None, area),
        )
        return True, None

    if action == "item_update_do":
        if item_id is None:
            return False, "missing_item"
        row = conn.execute(
            "SELECT assignee, kanban_column, sticky_color_hex FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row or (row["assignee"] or "").strip() != proposer:
            return False, "forbidden"
        cur_col = _normalize_kanban_column(row["kanban_column"] if "kanban_column" in row.keys() else None)
        if cur_col != "do":
            return False, "wrong_column"
        title = (payload.get("title") or "").strip()[:500]
        if not title:
            return False, "missing"
        est = _parse_hours_field(str(payload.get("estimate_hours", "0")))
        if est <= SCRUM_HOUR_EPS:
            return False, "estimate_required"
        tkind = _coerce_sprint_item_task_kind(conn, team_id, (payload.get("task_kind") or "ndy").strip()[:64])
        notes = (payload.get("notes") or "").strip()[:2000]
        dod = (payload.get("dod") or "").strip()[:4000]
        area = (payload.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        sticky_hex = None
        st = _status_for_kanban_column(cur_col)
        conn.execute(
            """
            UPDATE scrum_sprint_item SET
                title = ?, estimate_hours = ?, task_kind = ?, notes = ?, dod = ?, status = ?, updated_at = ?, sticky_color_hex = ?, area = ?
            WHERE id = ? AND sprint_id = ?
            """,
            (title, est, tkind, notes, dod, st, ts, sticky_hex, area, item_id, sprint_id),
        )
        return True, None

    if action == "item_update_done":
        if item_id is None:
            return False, "missing_item"
        row = conn.execute(
            "SELECT assignee, kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row or (row["assignee"] or "").strip() != proposer:
            return False, "forbidden"
        if _normalize_kanban_column(row["kanban_column"] if "kanban_column" in row.keys() else None) != "done":
            return False, "wrong_column"
        notes = (payload.get("notes") or "").strip()[:2000]
        da_json = (payload.get("done_artifacts_json") or "[]").strip()
        if not da_json.startswith("["):
            da_json = "[]"
        conn.execute(
            """
            UPDATE scrum_sprint_item SET notes = ?, done_artifacts = ?, updated_at = ?
            WHERE id = ? AND sprint_id = ?
            """,
            (notes, da_json, ts, item_id, sprint_id),
        )
        return True, None

    if action == "item_update_doing_estimate":
        if item_id is None:
            return False, "missing_item"
        row = conn.execute(
            "SELECT assignee, kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row or (row["assignee"] or "").strip() != proposer:
            return False, "forbidden"
        if _normalize_kanban_column(row["kanban_column"] if "kanban_column" in row.keys() else None) != "doing":
            return False, "wrong_column"
        est = _parse_hours_field(str(payload.get("estimate_hours", "0")))
        if est <= SCRUM_HOUR_EPS:
            return False, "estimate_required"
        conn.execute(
            """
            UPDATE scrum_sprint_item SET estimate_hours = ?, updated_at = ?
            WHERE id = ? AND sprint_id = ?
            """,
            (est, ts, item_id, sprint_id),
        )
        return True, None

    if action == "item_delete":
        if item_id is None:
            return False, "missing_item"
        ar = conn.execute(
            "SELECT assignee, kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not ar or (ar["assignee"] or "").strip() != proposer:
            return False, "forbidden"
        if _normalize_kanban_column(ar["kanban_column"] if "kanban_column" in ar.keys() else None) != "do":
            return False, "remove only while the sticky is in To DO"
        conn.execute("DELETE FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?", (item_id, sprint_id))
        return True, None

    if action == "portal_checklist":
        if item_id is None:
            return False, "missing_item"
        chk = conn.execute(
            "SELECT id FROM scrum_sprint_item WHERE id = ? AND assignee = ?",
            (item_id, proposer),
        ).fetchone()
        if not chk:
            return False, "forbidden"
        col = _normalize_kanban_column(payload.get("kanban_column"))
        st = (payload.get("item_status") or "").strip()[:32]
        dod_json = (payload.get("portal_dod_json") or "[]").strip()[:8000]
        conn.execute(
            """
            UPDATE scrum_sprint_item
            SET kanban_column = ?, status = ?, portal_dod_json = ?, updated_at = ?
            WHERE id = ? AND assignee = ?
            """,
            (col, st, dod_json, ts, item_id, proposer),
        )
        return True, None

    return False, "bad_action"


def _csrf_api_ok(app: Flask) -> bool:
    if not app.config.get("WTF_CSRF_ENABLED", True):
        return True
    try:
        token = request.form.get("csrf_token")
        if not token and request.is_json:
            token = (request.get_json(silent=True) or {}).get("csrf_token")
        if not token:
            token = request.headers.get("X-CSRFToken")
        validate_csrf(token)
        return True
    except Exception:
        return False


def create_app() -> Flask:
    load_application_environment()

    app = Flask(__name__)
    apply_flask_config_from_environ(app)
    ms_id = (app.config.get("MICROSOFT_OAUTH_CLIENT_ID") or "").strip()

    CSRFProtect(app)
    register_blueprints(app)

    oauth_ms = None
    if ms_id:
        from authlib.integrations.flask_client import OAuth

        oauth_ms = OAuth(app)
        tenant = app.config["MICROSOFT_OAUTH_TENANT_ID"]
        meta = f"https://login.microsoftonline.com/{tenant}/v2.0/.well-known/openid-configuration"
        oauth_ms.register(
            name="microsoft",
            client_id=ms_id,
            client_secret=app.config["MICROSOFT_OAUTH_CLIENT_SECRET"],
            server_metadata_url=meta,
            client_kwargs={"scope": "openid profile email offline_access User.Read"},
        )

    @app.route("/auth/microsoft/start")
    def portal_ms_start():
        if not oauth_ms:
            if _portal_otp_mail_configured(app):
                flash("Microsoft OAuth is not configured. Use email sign-in below, or set MICROSOFT_OAUTH_*.", "error")
            else:
                flash(
                    "Employee sign-in is not configured. Set MICROSOFT_OAUTH_CLIENT_ID and "
                    "MICROSOFT_OAUTH_CLIENT_SECRET for Microsoft, or set PORTAL_OTP_SMTP_HOST and PORTAL_OTP_FROM "
                    "for email codes. For local development, leave TEAM_TRACKER_PRODUCTION unset (dev OTP auto-enables) "
                    "or set PORTAL_OTP_DEV_CONSOLE=1. In production set TEAM_TRACKER_PRODUCTION=1 and real auth. See README.",
                    "error",
                )
            return redirect(url_for("home"))
        redirect_uri = url_for("portal_ms_callback", _external=True)
        return oauth_ms.microsoft.authorize_redirect(redirect_uri)

    @app.route("/auth/microsoft/callback")
    def portal_ms_callback():
        if not oauth_ms:
            return redirect(url_for("home"))
        try:
            token = oauth_ms.microsoft.authorize_access_token()
        except Exception:
            flash("Microsoft sign-in did not complete. Try again.", "error")
            return redirect(url_for("home"))
        access = (token or {}).get("access_token")
        if not access:
            flash("Microsoft sign-in failed (no access token).", "error")
            return redirect(url_for("home"))
        try:
            gr = requests.get(
                "https://graph.microsoft.com/v1.0/me",
                headers={"Authorization": f"Bearer {access}"},
                timeout=20,
            )
        except requests.RequestException:
            flash("Could not reach Microsoft to read your profile.", "error")
            return redirect(url_for("home"))
        if gr.status_code != 200:
            flash("Could not read your Microsoft profile.", "error")
            return redirect(url_for("home"))
        me = gr.json()
        mail = (me.get("mail") or me.get("userPrincipalName") or "").strip()
        if not mail:
            flash("No email address returned from Microsoft.", "error")
            return redirect(url_for("home"))
        low = normalize_email(mail)
        if not low.endswith("@nokia.com"):
            flash("Access restricted to Nokia corporate accounts.", "error")
            return redirect(url_for("home"))
        hit = _resolve_portal_signin_identity(app, low)
        if not hit:
            flash("Your account is not registered in this team. Contact your admin.", "error")
            return redirect(url_for("home"))
        hit["auth"] = "microsoft"
        return _complete_employee_signin(app, hit)

    @app.post("/auth/logout")
    def portal_logout():
        if not _csrf_form_ok():
            flash("Invalid security token.", "error")
            return redirect(url_for("home"))
        session.pop("portal_user", None)
        session.pop("portal_team_id", None)
        session.pop("portal_otp_email", None)
        session.pop("portal_otp_fails", None)
        session.pop("portal_access_code_fails", None)
        flash("Signed out.", "success")
        return redirect(url_for("home"))

    @app.post("/auth/employee-access-code")
    def portal_employee_access_code_signin():
        if not _csrf_form_ok():
            flash("Invalid security token.", "error")
            return redirect(url_for("login_page"))
        identifier = (request.form.get("identifier") or request.form.get("email") or "").strip()
        code = (request.form.get("access_code") or request.form.get("code") or "").strip()
        if not identifier:
            flash("Enter your work email or roster name.", "error")
            return redirect(url_for("login_page"))
        if not code:
            flash("Enter your 6-digit access code.", "error")
            return redirect(url_for("login_page"))
        hit = _resolve_portal_access_code_identity(app, identifier)
        if not hit:
            flash("We could not match that name or email to the team roster.", "error")
            return redirect(url_for("login_page"))
        roster_name = str(hit["roster_name"] or "").strip()
        if not _verify_portal_employee_access_code(
            app,
            roster_name,
            code,
            login_email=normalize_email(identifier) if "@" in identifier else "",
        ):
            fails = int(session.get("portal_access_code_fails") or 0) + 1
            session["portal_access_code_fails"] = fails
            if fails >= 8:
                session.pop("portal_access_code_fails", None)
                flash("Too many failed attempts. Try again later or contact your manager.", "error")
            else:
                flash("Invalid access code. Check with your manager if you need a reset.", "error")
            return redirect(url_for("login_page"))
        session.pop("portal_access_code_fails", None)
        hit["auth"] = "access_code"
        return _complete_employee_signin(app, hit)

    @app.post("/auth/email-otp/send")
    def portal_email_otp_send():
        if not _portal_otp_mail_configured(app):
            flash("Email sign-in is not configured on this server.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Invalid security token.", "error")
            return redirect(url_for("home"))
        sess_em = session.get("portal_otp_email")
        if sess_em and isinstance(sess_em, str):
            em = normalize_email(sess_em)
        else:
            raw_email = (request.form.get("email") or "").strip()
            em = normalize_email(raw_email)
        if not em:
            flash("Enter your Nokia email address.", "error")
            return redirect(url_for("home"))
        hit = _resolve_portal_signin_identity(app, em)
        if not hit:
            flash(
                "If that address is registered for this team, a sign-in code will arrive shortly. "
                "Otherwise no email was sent.",
                "success",
            )
            return redirect(url_for("home"))

        now = datetime.now(timezone.utc)
        since = (now - timedelta(hours=1)).isoformat(timespec="seconds").replace("+00:00", "Z")
        ip = client_ip()
        conn = get_db(app)
        n_email = int(
            conn.execute(
                "SELECT COUNT(*) AS c FROM portal_email_otp WHERE email = ? AND created_at >= ?",
                (em, since),
            ).fetchone()["c"]
        )
        nip = int(
            conn.execute(
                "SELECT COUNT(*) AS c FROM portal_email_otp WHERE client_ip = ? AND created_at >= ?",
                (ip, since),
            ).fetchone()["c"]
        )
        if n_email >= 5:
            conn.close()
            flash("Too many code requests for this address in the last hour. Try again later.", "error")
            return redirect(url_for("home"))
        if nip >= 40:
            conn.close()
            flash("Too many code requests from this network. Try again later.", "error")
            return redirect(url_for("home"))

        code = f"{secrets.randbelow(900000) + 100000:06d}"
        ttl = max(5, min(int(app.config.get("PORTAL_OTP_TTL_MINUTES") or 15), 120))
        exp = now + timedelta(minutes=ttl)
        exp_iso = exp.isoformat(timespec="seconds").replace("+00:00", "Z")
        cre_iso = now.isoformat(timespec="seconds").replace("+00:00", "Z")
        digest = _portal_otp_code_hmac(app, code)
        conn.execute("DELETE FROM portal_email_otp WHERE email = ?", (em,))
        conn.execute(
            """
            INSERT INTO portal_email_otp (email, code_hmac, expires_at, created_at, client_ip)
            VALUES (?, ?, ?, ?, ?)
            """,
            (em, digest, exp_iso, cre_iso, ip),
        )
        conn.commit()
        conn.close()
        if _portal_otp_smtp_ready(app):
            try:
                send_portal_otp_smtp(app, em, code)
            except Exception as exc:
                _log.warning("portal OTP email failed: %s", exc)
                conn2 = get_db(app)
                conn2.execute("DELETE FROM portal_email_otp WHERE email = ?", (em,))
                conn2.commit()
                conn2.close()
                flash("Could not send the email. Check SMTP settings or try again later.", "error")
                return redirect(url_for("home"))
        elif app.config.get("PORTAL_OTP_DEV_CONSOLE"):
            _log.warning(
                "PORTAL_OTP_DEV_CONSOLE: one-time sign-in for %s — code=%s (local dev only; never in production)",
                em,
                code,
            )
            flash(
                f"Development mode: your code is {code}. It is also in the server log. "
                "Unset PORTAL_OTP_DEV_CONSOLE for production.",
                "success",
            )
        else:
            conn2 = get_db(app)
            conn2.execute("DELETE FROM portal_email_otp WHERE email = ?", (em,))
            conn2.commit()
            conn2.close()
            flash("Email delivery is not configured.", "error")
            return redirect(url_for("home"))

        session["portal_otp_email"] = em
        session.pop("portal_otp_fails", None)
        flash("Check your inbox for a 6-digit sign-in code.", "success")
        return redirect(url_for("home"))

    @app.post("/auth/email-otp/verify")
    def portal_email_otp_verify():
        if not _portal_otp_mail_configured(app):
            flash("Email sign-in is not configured on this server.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Invalid security token.", "error")
            return redirect(url_for("home"))
        em = session.get("portal_otp_email")
        if not em or not isinstance(em, str):
            flash("Request a sign-in code first.", "error")
            return redirect(url_for("home"))
        code = (request.form.get("code") or "").replace(" ", "").strip()
        if len(code) != 6 or not code.isdigit():
            flash("Enter the 6-digit code from your email.", "error")
            return redirect(url_for("home"))

        conn = get_db(app)
        row = conn.execute(
            "SELECT id, code_hmac, expires_at FROM portal_email_otp WHERE email = ? ORDER BY id DESC LIMIT 1",
            (em,),
        ).fetchone()
        if not row:
            conn.close()
            flash("No active code. Request a new one.", "error")
            return redirect(url_for("home"))

        exp_raw = str(row["expires_at"] or "")
        try:
            exp_dt = datetime.fromisoformat(exp_raw.replace("Z", "+00:00"))
        except ValueError:
            conn.execute("DELETE FROM portal_email_otp WHERE id = ?", (int(row["id"]),))
            conn.commit()
            conn.close()
            flash("Invalid challenge state. Request a new code.", "error")
            return redirect(url_for("home"))

        if datetime.now(timezone.utc) > exp_dt:
            conn.execute("DELETE FROM portal_email_otp WHERE id = ?", (int(row["id"]),))
            conn.commit()
            conn.close()
            flash("That code has expired. Request a new one.", "error")
            return redirect(url_for("home"))

        digest = _portal_otp_code_hmac(app, code)
        if not hmac.compare_digest(digest, str(row["code_hmac"] or "")):
            fails = int(session.get("portal_otp_fails") or 0) + 1
            session["portal_otp_fails"] = fails
            conn.close()
            if fails >= 8:
                conn2 = get_db(app)
                conn2.execute("DELETE FROM portal_email_otp WHERE email = ?", (em,))
                conn2.commit()
                conn2.close()
                session.pop("portal_otp_fails", None)
                flash("Too many wrong attempts. Request a new code.", "error")
            else:
                flash("Incorrect code. Try again.", "error")
            return redirect(url_for("home"))

        hit = _resolve_portal_signin_identity(app, em)
        if not hit:
            conn.execute("DELETE FROM portal_email_otp WHERE id = ?", (int(row["id"]),))
            conn.commit()
            conn.close()
            session.pop("portal_otp_email", None)
            flash("Your email is no longer in the team directory.", "error")
            return redirect(url_for("home"))

        conn.execute("DELETE FROM portal_email_otp WHERE id = ?", (int(row["id"]),))
        conn.commit()
        conn.close()
        session.pop("portal_otp_email", None)
        session.pop("portal_otp_fails", None)
        hit["auth"] = "email_otp"
        return _complete_employee_signin(app, hit)

    @app.route("/auth/email-otp/cancel")
    def portal_email_otp_cancel():
        session.pop("portal_otp_email", None)
        session.pop("portal_otp_fails", None)
        flash("Cancelled email sign-in.", "info")
        return redirect(url_for("login_page"))

    @app.route("/portal")
    def portal_dashboard():
        """Legacy URL: open current sprint Kanban for this employee."""
        return _portal_employee_landing_response(app)

    @app.route("/portal/leave/apply", methods=["GET", "POST"])
    def portal_leave_apply():
        pu = _portal_session()
        if not pu:
            return redirect(url_for("home"))
        today_iso = date.today().isoformat()
        if request.method == "POST":
            if not _csrf_form_ok():
                flash("Invalid security token.", "error")
                return redirect(url_for("portal_leave_apply"))
            kind = (request.form.get("leave_kind") or "").strip()
            allowed = {c for c, _ in PORTAL_LEAVE_FORM_CHOICES}
            if kind not in allowed:
                flash("Select a valid leave type.", "error")
                return render_template(
                    "portal_leave_apply.html",
                    hide_nav=True,
                    leave_choices=PORTAL_LEAVE_FORM_CHOICES,
                    form=request.form,
                    default_start=today_iso,
                    default_end=today_iso,
                )
            start_date = (request.form.get("start_date") or "").strip()
            end_date = (request.form.get("end_date") or "").strip()
            reason = (request.form.get("reason") or "").strip()[:300]
            errors: list[str] = []
            if not start_date:
                errors.append("Start date is required.")
            if not end_date:
                errors.append("End date is required.")
            if not reason:
                errors.append("Reason is required.")
            sd: date | None = None
            ed: date | None = None
            if not errors:
                try:
                    sd = date.fromisoformat(start_date[:10])
                    ed = date.fromisoformat(end_date[:10])
                except ValueError:
                    errors.append("Invalid date format.")
            if not errors and sd and ed:
                if ed < sd:
                    errors.append("End date cannot be before start date.")
                elif sd.weekday() >= 5 or ed.weekday() >= 5:
                    errors.append("Leave cannot start or end on a weekend — choose Monday through Friday.")
            if errors:
                for e in errors:
                    flash(e, "error")
                return render_template(
                    "portal_leave_apply.html",
                    hide_nav=True,
                    leave_choices=PORTAL_LEAVE_FORM_CHOICES,
                    form=request.form,
                    default_start=today_iso,
                    default_end=today_iso,
                )
            assert sd is not None and ed is not None
            duration_type = "multi" if ed > sd else "full"
            label = dict(PORTAL_LEAVE_FORM_CHOICES).get(kind, kind)
            created = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
            ip = client_ip()
            conn = get_db(app)
            conn.execute(
                """
                INSERT INTO leave_requests
                (employee_name, reason, description, start_date, end_date, duration_type, status,
                 created_at, submitted_ip, submitted_email, portal_leave_label)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?, ?)
                """,
                (
                    _portal_effective_roster_name(app, pu),
                    kind,
                    reason,
                    start_date[:10],
                    end_date[:10],
                    duration_type,
                    created,
                    ip,
                    pu["email"],
                    label,
                ),
            )
            conn.commit()
            conn.close()
            wd = _working_weekdays_count(sd, ed)
            if any(d.weekday() >= 5 for d in _daterange_inclusive(sd, ed)):
                flash(
                    f"Leave request submitted successfully. Your range includes weekend days; "
                    f"working weekdays in this span: {wd}.",
                    "success",
                )
            else:
                flash("Leave request submitted successfully.", "success")
            return redirect(url_for("portal_my_sprint_kanban"))

        return render_template(
            "portal_leave_apply.html",
            hide_nav=True,
            leave_choices=PORTAL_LEAVE_FORM_CHOICES,
            form={},
            default_start=today_iso,
            default_end=today_iso,
        )

    @app.route("/portal/leave/history")
    def portal_leave_history():
        pu = _portal_session()
        if not pu:
            return redirect(url_for("home"))
        conn = get_db(app)
        rows = list(
            conn.execute(
                """
                SELECT * FROM leave_requests
                WHERE employee_name = ?
                ORDER BY created_at DESC
                LIMIT 200
                """,
                (_portal_effective_roster_name(app, pu),),
            )
        )
        conn.close()
        base_reasons = dict(LEAVE_REASONS)
        display_rows: list[dict] = []
        for r in rows:
            try:
                sd = date.fromisoformat(str(r["start_date"])[:10])
                ed = date.fromisoformat(str(r["end_date"])[:10])
                wd = _working_weekdays_count(sd, ed)
            except ValueError:
                wd = 0
            plab = (r["portal_leave_label"] if "portal_leave_label" in r.keys() else "") or ""
            leave_type = plab.strip() if plab.strip() else base_reasons.get(r["reason"], r["reason"])
            display_rows.append({"raw": r, "working_days": wd, "leave_type": leave_type})
        return render_template(
            "portal_leave_history.html",
            hide_nav=True,
            display_rows=display_rows,
        )

    @app.route("/portal/my-sprint/board", methods=["GET"])
    def portal_my_sprint_kanban():
        """Employee Sprint work: Kanban board for the current team sprint."""
        return _portal_employee_landing_response(app)

    @app.route("/portal/sprint", methods=["GET", "POST"])
    def portal_sprint():
        pu = _portal_session()
        if not pu:
            return redirect(url_for("home"))
        if request.method == "GET":
            return redirect(url_for("portal_my_sprint_kanban"))
        if not _csrf_form_ok():
            flash("Invalid security token.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        item_id = _parse_optional_int(request.form.get("item_id"))
        if not item_id:
            flash("Missing item.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        status = (request.form.get("portal_status") or "").strip().lower()
        if status not in ("todo", "progress", "done"):
            flash("Invalid status.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        dod_flags: list[bool] = []
        for i in range(len(PORTAL_DOD_CHECKLIST_LABELS)):
            dod_flags.append((request.form.get(f"dod_{i}") or "").strip() == "1")
        dod_json = json.dumps(dod_flags)
        col = _portal_status_to_kanban(status)
        st = _status_for_kanban_column(col)
        roster_name = _portal_effective_roster_name(app, pu)
        conn = get_db(app)
        row = conn.execute(
            """
            SELECT i.id, i.sprint_id, s.team_id
            FROM scrum_sprint_item i
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE i.id = ? AND i.assignee = ?
            """,
            (int(item_id), roster_name),
        ).fetchone()
        if not row:
            conn.close()
            flash("That work item was not found or is not assigned to you.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        sprint_id = int(row["sprint_id"])
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        pl = {"kanban_column": col, "item_status": st, "portal_dod_json": dod_json}
        _insert_scrum_portal_proposal(
            conn,
            int(row["team_id"]),
            sprint_id,
            int(item_id),
            "portal_checklist",
            roster_name,
            pl,
        )
        conn.commit()
        sid = int(row["sprint_id"])
        conn.close()
        flash("Change saved.", "success")
        return redirect(url_for("portal_scrum_kanban_board", sprint_id=sid))

    @app.route("/portal/sprint/<int:sprint_id>/board", methods=["GET"])
    def portal_scrum_kanban_board(sprint_id: int):
        pu = _portal_session()
        if not pu:
            return redirect(url_for("home"))
        roster_name = _portal_effective_roster_name(app, pu)
        if not roster_name:
            flash("Your profile is missing a roster name.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        if not _portal_employee_can_access_team(app, roster_name, int(team_id)):
            conn.close()
            flash("That sprint is not on your team roster.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        session["portal_team_id"] = int(team_id)
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        _maybe_auto_close_scrum_sprint(conn, int(sprint_id))
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        cards_by_col = _load_kanban_cards(conn, int(sprint_id), roster_name, team_id)
        task_kinds_rows = _list_team_task_kinds(conn, team_id)
        tf = _sprint_team_page_template_flags(conn, team_id, int(sprint_id))
        conn.commit()
        conn.close()
        kb_leave_ctx = build_kanban_leave_worksheet_context(
            app,
            int(sprint_id),
            str(sprint["start_date"])[:10],
            str(sprint["end_date"])[:10],
            roster_name,
        )
        return render_template(
            "scrum_kanban.html",
            hide_nav=True,
            sprint_id=int(sprint_id),
            sprint_name=sprint["name"],
            assignee=roster_name,
            columns=SCRUM_KANBAN_COLUMNS,
            cards_by_col=cards_by_col,
            task_kinds_rows=task_kinds_rows,
            kb_back_url=url_for("portal_my_sprint_kanban"),
            kb_api_urls={
                "item_move": url_for("portal_scrum_api_item_move"),
                "item_note": url_for("portal_scrum_api_item_note"),
                "item_activity_update": url_for("portal_scrum_api_activity_update"),
                "item_add": url_for("portal_scrum_api_item_add"),
                "area_suggest": url_for("portal_scrum_api_areas", sprint_id=int(sprint_id)),
            },
            kb_form_urls={
                "item_update": url_for("portal_scrum_sprint_item_update"),
                "item_delete": url_for("portal_scrum_sprint_item_delete"),
                "attachment_upload": url_for("portal_scrum_item_attachment_upload"),
                "attachment_delete": url_for("portal_scrum_item_attachment_delete"),
            },
            kb_checklist_urls={
                "add": url_for("portal_scrum_api_item_checklist_add"),
                "update": url_for("portal_scrum_api_item_checklist_update"),
                "delete": url_for("portal_scrum_api_item_checklist_delete"),
            },
            kb_item_detail_url=url_for("portal_scrum_api_item_detail"),
            portal_kanban=True,
            sprint_readonly=tf["sprint_board_readonly"],
            sprint_freeze_mode=tf["sprint_freeze_mode"],
            **kb_leave_ctx,
        )

    @app.get("/portal/scrum/sprint/item/attachment/<int:attachment_id>/file")
    def portal_scrum_item_attachment_download(attachment_id: int):
        pu = _portal_session()
        if not pu:
            return redirect(url_for("home"))
        roster_name = _portal_effective_roster_name(app, pu)
        if not roster_name:
            return Response("Forbidden.", status=403, mimetype="text/plain")
        conn = get_db(app)
        ar = _scrum_attachment_auth_row(conn, int(attachment_id))
        conn.close()
        if not ar or (str(ar["assignee"] or "").strip() != roster_name):
            return Response("Not found.", status=404, mimetype="text/plain")
        rel = (str(ar["rel_path"]) or "").strip().replace("\\", "/")
        if ".." in rel or rel.startswith("/"):
            return Response("Bad path.", status=400, mimetype="text/plain")
        path = _scrum_attachment_root_dir(app) / rel
        if not path.is_file():
            return Response("File missing on server.", status=404, mimetype="text/plain")
        return send_file(
            path,
            as_attachment=True,
            download_name=str(ar["original_filename"] or "download").strip()[:200] or "download",
            mimetype=str(ar["content_type"] or "application/octet-stream").strip()[:200]
            or "application/octet-stream",
        )

    @app.get("/portal/scrum/sprint/item/attachment/<int:attachment_id>/preview-html")
    def portal_scrum_item_attachment_preview_html(attachment_id: int):
        pu = _portal_session()
        if not pu:
            return Response("Forbidden.", status=403, mimetype="text/plain")
        roster_name = _portal_effective_roster_name(app, pu)
        if not roster_name:
            return Response("Forbidden.", status=403, mimetype="text/plain")
        conn = get_db(app)
        ar = _scrum_attachment_auth_row(conn, int(attachment_id))
        conn.close()
        if not ar or (str(ar["assignee"] or "").strip() != roster_name):
            return Response("Not found.", status=404, mimetype="text/plain")
        rel = (str(ar["rel_path"]) or "").strip().replace("\\", "/")
        if ".." in rel or rel.startswith("/"):
            return Response("Bad path.", status=400, mimetype="text/plain")
        path = _scrum_attachment_root_dir(app) / rel
        if not path.is_file():
            return Response("File missing on server.", status=404, mimetype="text/plain")
        low = (str(ar["original_filename"]) or "").lower()
        if not (low.endswith((".xlsx", ".xlsm", ".xltx", ".csv"))):
            return Response("Preview not available for this file type.", status=415, mimetype="text/plain")
        doc = _scrum_attachment_build_preview_html(path, str(ar["original_filename"] or ""))
        if not doc:
            return Response("Could not build preview (install openpyxl for Excel files).", status=500, mimetype="text/plain")
        return Response(doc, mimetype="text/html; charset=utf-8")

    @app.post("/portal/scrum/sprint/item/attachment")
    def portal_scrum_item_attachment_upload():
        pu = _portal_session()
        if not pu:
            flash("Sign in required.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        roster_name = _portal_effective_roster_name(app, pu)
        item_id = _parse_optional_int(request.form.get("item_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        if not item_id or not sprint_id or not roster_name:
            flash("Missing backlog item.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        cat, msg, _assignee = _scrum_attachment_upload_run_batch(
            app,
            conn,
            team_id=int(team_id),
            sprint_id=sprint_id,
            item_id=item_id,
            roster_gate=roster_name,
        )
        conn.close()
        flash(msg, cat)
        return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))

    @app.post("/portal/scrum/sprint/item/attachment/delete")
    def portal_scrum_item_attachment_delete():
        pu = _portal_session()
        if not pu:
            flash("Sign in required.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        roster_name = _portal_effective_roster_name(app, pu)
        aid = _parse_optional_int(request.form.get("attachment_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        if not aid or not sprint_id or not roster_name:
            flash("Missing attachment.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        ok, msg, _assignee = _scrum_attachment_delete_core(
            app,
            conn,
            team_id=int(team_id),
            sprint_id=sprint_id,
            attachment_id=int(aid),
            roster_gate=roster_name,
        )
        conn.close()
        flash(msg, "success" if ok else "error")
        return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))

    @app.get("/portal/sprint/<int:sprint_id>/api/areas")
    def portal_scrum_api_areas(sprint_id: int):
        pu = _portal_session()
        if not pu:
            return jsonify({"matches": []}), 403
        q = (request.args.get("q") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        if len(q) < 1:
            return jsonify({"matches": []})
        conn = get_db(app)
        try:
            team_id = _sprint_team_id(conn, sprint_id)
            if team_id is None:
                return jsonify({"matches": []}), 404
            sprint = conn.execute(
                "SELECT id FROM scrum_sprint WHERE id = ? AND team_id = ?",
                (int(sprint_id), team_id),
            ).fetchone()
            if not sprint:
                return jsonify({"matches": []}), 404
            out = _scrum_distinct_area_suggestions(conn, team_id, q)
        finally:
            conn.close()
        return jsonify({"matches": out})

    @app.post("/portal/scrum/api/item/move")
    def portal_scrum_api_item_move():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        to_col = _normalize_kanban_column(data.get("to_column"))
        note_raw = (data.get("note") or "").strip()[:2000]
        if not item_id or not sprint_id or not roster_name:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, roster_name):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        row = conn.execute(
            "SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (item_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "missing"}), 400
        from_col = _normalize_kanban_column(row["kanban_column"])
        est_bl2do, est_err = _estimate_hours_for_backlog_to_do_move(data, from_col=from_col, to_col=to_col)
        if est_err:
            conn.close()
            return jsonify({"ok": False, "error": est_err}), 400
        if from_col == "backlog" and to_col == "do":
            ch = 0.0
            note_raw = ""
        else:
            ch = _parse_committed_hours_for_kanban_move(
                data.get("committed_hours"), from_col=from_col, to_col=to_col
            )
        artifacts_json: str | None = None
        if to_col == "done":
            artifacts_json, aerr = _normalize_done_artifacts_from_api(data.get("artifacts"))
            if aerr:
                conn.close()
                return jsonify({"ok": False, "error": aerr}), 400
        pl: dict = {"to_column": to_col, "committed_hours": ch, "note": note_raw}
        if est_bl2do is not None:
            pl["estimate_hours"] = est_bl2do
        if to_col == "done" and artifacts_json is not None:
            pl["done_artifacts_json"] = artifacts_json
        vurls = _normalize_verification_urls_mixed(data.get("verification_urls"))
        if vurls and est_bl2do is None:
            pl["verification_urls"] = vurls
        _insert_scrum_portal_proposal(conn, team_id, sprint_id, item_id, "item_move", roster_name, pl)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": True})

    @app.post("/portal/scrum/api/item/note")
    def portal_scrum_api_item_note():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        note_raw = (data.get("note") or "").strip()[:2000]
        ch = _parse_hours_field(str(data.get("committed_hours")))
        if not item_id or not sprint_id or not roster_name:
            return jsonify({"ok": False, "error": "missing"}), 400
        if not note_raw and ch < 0:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, roster_name):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        row = conn.execute(
            "SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (item_id,)
        ).fetchone()
        if not row or _normalize_kanban_column(row["kanban_column"]) != "doing":
            conn.close()
            return jsonify({"ok": False, "error": "only_doing"}), 400
        pl = {"note": note_raw, "committed_hours": ch}
        vurls = _normalize_verification_urls_mixed(data.get("verification_urls"))
        if vurls:
            pl["verification_urls"] = vurls
        _insert_scrum_portal_proposal(conn, team_id, sprint_id, item_id, "item_note", roster_name, pl)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": True})

    @app.post("/portal/scrum/api/item/activity_update")
    def portal_scrum_api_activity_update():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        activity_id = _parse_optional_int(data.get("activity_id"))
        note_raw = (data.get("note") or "").strip()[:2000]
        ch = _parse_hours_field(str(data.get("committed_hours")))
        if not item_id or not sprint_id or not roster_name or not activity_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, roster_name):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        
        pl = {"activity_id": activity_id, "note": note_raw, "committed_hours": ch}
        _insert_scrum_portal_proposal(conn, team_id, sprint_id, item_id, "item_activity_update", roster_name, pl)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": True})

    @app.post("/portal/scrum/api/item/add")
    def portal_scrum_api_item_add():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        title = (data.get("title") or "").strip()[:500]
        notes = (data.get("notes") or "").strip()[:2000]
        dod = (data.get("dod") or "").strip()[:4000]
        est = _parse_hours_field(data.get("estimate_hours"))
        if not sprint_id or not title or not roster_name:
            return jsonify({"ok": False, "error": "missing"}), 400
        if est < 0:
            return jsonify({"ok": False, "error": "estimate_required"}), 400
        new_l = (data.get("new_kind_label") or "").strip()
        if new_l:
            return jsonify({"ok": False, "error": "no_new_kind"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        _ensure_team_task_kinds(conn, team_id)
        specified = "task_kind_code" in data
        v = data.get("task_kind_code") if specified else "ndy"
        raw = str(v or "ndy").strip().lower()
        if raw in SCRUM_LEGACY_TASK_KIND_MAP:
            raw = SCRUM_LEGACY_TASK_KIND_MAP[raw]
        tcode = raw if raw in SCRUM_TASK_KIND_CODES else "ndy"
        pl: dict = {
            "title": title,
            "notes": notes,
            "dod": dod,
            "estimate_hours": est,
            "task_kind_code": tcode,
            "area": (data.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN],
        }
        _insert_scrum_portal_proposal(conn, team_id, sprint_id, None, "item_add", roster_name, pl)
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": True})

    @app.get("/portal/scrum/api/item/detail")
    def portal_scrum_api_item_detail():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        roster_name = _portal_effective_roster_name(app, pu)
        item_id = _parse_optional_int(request.args.get("item_id"))
        sprint_id = _parse_optional_int(request.args.get("sprint_id"))
        if not item_id or not sprint_id or not roster_name:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(
            conn, item_id, sprint_id, team_id, roster_name
        ):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        payload = _build_scrum_item_detail_payload(conn, team_id, sprint_id, item_id)
        conn.close()
        if not payload:
            return jsonify({"ok": False, "error": "not_found"}), 404
        return jsonify({"ok": True, **payload})

    @app.post("/portal/scrum/api/item/checklist/add")
    def portal_scrum_api_item_checklist_add():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        if not item_id or not sprint_id or not roster_name:
            return jsonify({"ok": False, "error": "missing"}), 400
        initial_item_name = (data.get("items_to_finish") or "").strip()[:500]
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(
            conn, item_id, sprint_id, team_id, roster_name
        ):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM scrum_item_checklist WHERE item_id = ?",
            (item_id,),
        ).fetchone()[0]
        ts = _utc_stamp()
        updated_by = _scrum_portal_checklist_updated_by(pu, roster_name)
        cur = conn.execute(
            """
            INSERT INTO scrum_item_checklist (item_id, sort_order, items_to_finish, status, le_to_complete, done_till_date, updated_at, updated_by)
            VALUES (?, ?, ?, '', '', '', ?, ?)
            """,
            (item_id, int(max_order) + 1, initial_item_name, ts, updated_by),
        )
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id, "sort_order": int(max_order) + 1})

    @app.post("/portal/scrum/api/item/checklist/update")
    def portal_scrum_api_item_checklist_update():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        row_id = _parse_optional_int(data.get("id"))
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        field = (data.get("field") or "").strip()
        value = (data.get("value") or "")[:2000]
        allowed_fields = {"items_to_finish", "status", "le_to_complete", "done_till_date"}
        if not row_id or not item_id or not sprint_id or field not in allowed_fields:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(
            conn, item_id, sprint_id, team_id, roster_name
        ):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        row = conn.execute(
            """
            SELECT c.id FROM scrum_item_checklist c
            JOIN scrum_sprint_item i ON i.id = c.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE c.id = ? AND c.item_id = ? AND s.team_id = ? AND i.assignee = ?
            """,
            (row_id, item_id, team_id, roster_name),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        ts = _utc_stamp()
        updated_by = _scrum_portal_checklist_updated_by(pu, roster_name)
        conn.execute(
            f"UPDATE scrum_item_checklist SET {field} = ?, updated_at = ?, updated_by = ? WHERE id = ?",
            (value, ts, updated_by, row_id),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/portal/scrum/api/item/checklist/delete")
    def portal_scrum_api_item_checklist_delete():
        pu = _portal_session()
        if not pu:
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        roster_name = _portal_effective_roster_name(app, pu)
        data = request.get_json(silent=True) or {}
        row_id = _parse_optional_int(data.get("id"))
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        if not row_id or not item_id or not sprint_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _item_belongs_to_sprint_team(
            conn, item_id, sprint_id, team_id, roster_name
        ):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        row = conn.execute(
            """
            SELECT c.id FROM scrum_item_checklist c
            JOIN scrum_sprint_item i ON i.id = c.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE c.id = ? AND c.item_id = ? AND s.team_id = ? AND i.assignee = ?
            """,
            (row_id, item_id, team_id, roster_name),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        conn.execute("DELETE FROM scrum_item_checklist WHERE id = ?", (row_id,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/portal/scrum/sprint/item/update")
    def portal_scrum_sprint_item_update():
        pu = _portal_session()
        if not pu:
            flash("Sign in required.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        roster_name = _portal_effective_roster_name(app, pu)
        item_id = _parse_optional_int(request.form.get("item_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        if not item_id or not sprint_id or not roster_name:
            flash("Missing backlog item.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        row = conn.execute(
            "SELECT assignee, kanban_column, sticky_color_hex FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row or (row["assignee"] or "").strip() != roster_name:
            conn.close()
            flash("Item not found.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        cur_col = _normalize_kanban_column(row["kanban_column"] if "kanban_column" in row.keys() else None)

        if cur_col == "done":
            notes = (request.form.get("notes") or "").strip()[:2000]
            da_lines = request.form.get("done_artifacts_lines") or ""
            da_json, da_err = _normalize_done_artifacts_from_lines(da_lines)
            if da_err:
                conn.close()
                if da_err == "artifacts_limit":
                    msg = "Too many artifact links (maximum 20)."
                else:
                    msg = "Artifact links must be valid http(s) URLs (one per line, optional label before |)."
                flash(msg, "error")
                return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
            pl = {"notes": notes, "done_artifacts_json": da_json}
            vurls = _normalize_verification_urls_mixed(request.form.get("verification_urls_lines") or "")
            if vurls:
                pl["verification_urls"] = vurls
            _insert_scrum_portal_proposal(conn, team_id, sprint_id, item_id, "item_update_done", roster_name, pl)
            conn.commit()
            conn.close()
            flash("Changes saved.", "success")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))

        if cur_col == "doing":
            est = _parse_hours_field(request.form.get("estimate_hours"))
            if est <= SCRUM_HOUR_EPS:
                conn.close()
                flash("Estimated hours must be greater than zero while this sticky is In progress.", "error")
                return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
            pl = {"estimate_hours": est}
            _insert_scrum_portal_proposal(
                conn, team_id, sprint_id, item_id, "item_update_doing_estimate", roster_name, pl
            )
            conn.commit()
            conn.close()
            flash("Plan estimate updated.", "success")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))

        if cur_col != "do":
            conn.close()
            flash("Planning edits are only in To DO; details and artifacts can be edited in Done.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        title = (request.form.get("title") or "").strip()
        if not title:
            conn.close()
            flash("Title is required.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        est = _parse_hours_field(request.form.get("estimate_hours"))
        if est < 0:
            conn.close()
            flash("Estimate cannot be negative.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        tkind = _coerce_sprint_item_task_kind(conn, team_id, request.form.get("task_kind"))
        notes = (request.form.get("notes") or "").strip()[:2000]
        dod = (request.form.get("dod") or "").strip()[:4000]
        pl: dict = {
            "title": title,
            "estimate_hours": est,
            "task_kind": tkind,
            "notes": notes,
            "dod": dod,
            "area": (request.form.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN],
        }
        _insert_scrum_portal_proposal(conn, team_id, sprint_id, item_id, "item_update_do", roster_name, pl)
        conn.commit()
        conn.close()
        flash("Changes saved.", "success")
        return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))

    @app.post("/portal/scrum/sprint/item/delete")
    def portal_scrum_sprint_item_delete():
        pu = _portal_session()
        if not pu:
            flash("Sign in required.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        roster_name = _portal_effective_roster_name(app, pu)
        item_id = _parse_optional_int(request.form.get("item_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        if not item_id or not sprint_id or not roster_name:
            flash("Missing item.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        conn = get_db(app)
        team_id = _sprint_team_id(conn, sprint_id)
        if team_id is None or not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("portal_my_sprint_kanban"))
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        ar = conn.execute(
            "SELECT assignee, kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not ar or (ar["assignee"] or "").strip() != roster_name:
            conn.close()
            flash("Item not found.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        if _normalize_kanban_column(ar["kanban_column"] if "kanban_column" in ar.keys() else None) != "do":
            conn.close()
            flash("Stickies can only be removed while they are in To DO.", "error")
            return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))
        _insert_scrum_portal_proposal(conn, team_id, sprint_id, item_id, "item_delete", roster_name, {})
        conn.commit()
        conn.close()
        flash("Changes saved.", "success")
        return redirect(url_for("portal_scrum_kanban_board", sprint_id=sprint_id))

    @app.before_request
    def _manager_team_roster_g() -> None:
        g.manager_teams = []
        g.manager_team_id = None
        g.manager_team_name = None
        g.manager_roster = EMPLOYEES
        g.team_hub_mode = "leave"
        if not _manager_logged_in():
            return
        conn = get_db(app)
        manager_email = session.get("manager_user_email") or ""
        manager_role = (session.get("manager_role") or "").strip()
        is_lpo = manager_role == "lpo_sm"

        all_teams = list(conn.execute("SELECT id, name, hub_mode, COALESCE(owner_email,'') AS owner_email FROM teams ORDER BY name COLLATE NOCASE"))

        if is_lpo and manager_email:
            registered = conn.execute(
                "SELECT 1 FROM lpo_manager_emails WHERE email = ? COLLATE NOCASE",
                (normalize_email(manager_email),),
            ).fetchone()
            if not registered:
                # Unknown email somehow got LPO session → blank Default sandbox
                conn.close()
                g.manager_teams = [{"id": -1, "name": "Default", "hub_mode": "leave"}]
                g.manager_team_id = None
                g.manager_team_name = "Default"
                g.manager_roster = ()
                return
            # Registered LPO — restrict to assigned teams only
            assigned_ids = set(_lpo_assigned_team_ids(conn, manager_email))
            teams = [t for t in all_teams if int(t["id"]) in assigned_ids]
            if not teams:
                # Registered but no teams assigned yet → show empty placeholder
                conn.close()
                g.manager_teams = []
                g.manager_team_id = None
                g.manager_team_name = None
                g.manager_roster = ()
                return
        elif manager_email and not is_lpo:
            teams = [t for t in all_teams if (t["owner_email"] or "").strip() == "" or
                     normalize_email(str(t["owner_email"] or "")) == normalize_email(manager_email)]
        else:
            teams = all_teams

        if not teams:
            conn.close()
            return
        ids = [int(t["id"]) for t in teams]
        tid_raw = session.get("active_team_id")
        try:
            tid = int(tid_raw) if tid_raw is not None else ids[0]
        except (TypeError, ValueError):
            tid = ids[0]
        if tid not in ids:
            tid = ids[0]
        session["active_team_id"] = tid
        trow = next((t for t in teams if int(t["id"]) == tid), teams[0])
        g.manager_team_id = int(trow["id"])
        g.manager_team_name = str(trow["name"])
        g.team_hub_mode = _normalize_hub_mode(str(trow["hub_mode"] or "leave"))
        g.manager_teams = [
            {"id": int(t["id"]), "name": t["name"], "hub_mode": _normalize_hub_mode(str(t["hub_mode"] or "leave"))}
            for t in teams
        ]
        tname = str(trow["name"] or "").strip()
        if tname.casefold() == "default":
            urows = conn.execute(
                "SELECT DISTINCT employee_name FROM team_roster ORDER BY employee_name COLLATE NOCASE"
            ).fetchall()
        else:
            urows = conn.execute(
                """
                SELECT employee_name FROM team_roster
                WHERE team_id = ?
                ORDER BY sort_order, employee_name COLLATE NOCASE
                """,
                (g.manager_team_id,),
            ).fetchall()
        conn.close()
        if urows:
            g.manager_roster = tuple(r[0] for r in urows)
        elif tname.casefold() == "default":
            g.manager_roster = EMPLOYEES
        else:
            g.manager_roster = ()

    def _csrf_form_ok() -> bool:
        if not app.config.get("WTF_CSRF_ENABLED", True):
            return True
        try:
            validate_csrf(request.form.get("csrf_token"))
            return True
        except Exception:
            return False

    @app.context_processor
    def _inject_nav() -> dict:
        manager_email = session.get("manager_user_email")
        if manager_email:
            manager_name = manager_email.split("@")[0].replace(".", " ").title()
        else:
            manager_name = "Manager"

        pu = _portal_session()
        portal_employee_sprints: list[dict[str, Any]] = []
        portal_employee_teams: list[dict[str, Any]] = []
        portal_employee_team_name: str | None = None
        portal_active_team_id: int | None = None
        current_portal_sprint_id: int | None = None
        if pu and not _manager_logged_in():
            roster = _portal_effective_roster_name(app, pu)
            portal_employee_teams = _portal_employee_team_rows(app, roster)
            raw_tid = request.args.get("team_id")
            team_hint: int | None = None
            if raw_tid is not None:
                try:
                    team_hint = int(raw_tid)
                except (TypeError, ValueError):
                    team_hint = None
            portal_active_team_id = _resolve_portal_selected_team_id(
                app, roster, team_hint, persist=True
            )
            portal_employee_team_name = _portal_team_name_for_id(
                portal_employee_teams, portal_active_team_id
            )
            portal_employee_sprints = _list_portal_employee_sprints(
                app, roster, portal_active_team_id
            )
            view_args = request.view_args or {}
            raw_sid = view_args.get("sprint_id")
            if raw_sid is not None:
                try:
                    current_portal_sprint_id = int(raw_sid)
                except (TypeError, ValueError):
                    current_portal_sprint_id = None

        return {
            "manager_logged_in": _manager_logged_in(),
            "manager_password_configured": _manager_password_configured(app),
            "lpo_sm_password_configured": _lpo_sm_password_configured(app),
            "manager_nav_label": (
                "LPO/SM" if (session.get("manager_role") or "").strip() == "lpo_sm" else manager_name
            ),
            "manager_teams": getattr(g, "manager_teams", []),
            "active_team_id": getattr(g, "manager_team_id", None),
            "active_team_name": getattr(g, "manager_team_name", None),
            "team_hub_mode": getattr(g, "team_hub_mode", "leave"),
            "portal_user": pu,
            "portal_employee_teams": portal_employee_teams,
            "portal_employee_team_name": portal_employee_team_name,
            "portal_active_team_id": portal_active_team_id,
            "portal_employee_sprints": portal_employee_sprints,
            "current_portal_sprint_id": current_portal_sprint_id,
            "microsoft_oauth_configured": bool(app.config.get("MICROSOFT_OAUTH_CLIENT_ID")),
            "email_otp_configured": _portal_otp_mail_configured(app),
            "portal_access_codes_configured": _portal_access_codes_configured(app),
        }

    @app.errorhandler(404)
    def not_found(_e):  # noqa: ANN001
        return render_template("errors/404.html"), 404

    @app.errorhandler(500)
    def server_error(_e):  # noqa: ANN001
        return render_template("errors/500.html"), 500

    @app.route("/api/employees")
    def api_employees():
        q = (request.args.get("q") or "").strip()
        union = get_union_roster_for_app(app)
        matches = fuzzy_employee_matches(q, 5, roster=union)
        return jsonify({"matches": matches})

    @app.route("/manager/login", methods=["GET", "POST"])
    def manager_login():
        """Legacy URL — forwards to /dashboard (307 keeps POST body for old bookmarks)."""
        if request.method == "POST":
            return redirect(url_for("dashboard"), code=307)
        return redirect(url_for("dashboard", **request.args))

    @app.route("/manager")
    def manager_dashboard():
        """Legacy URL — redirect to unified dashboard."""
        return redirect(url_for("dashboard", **request.args))

    @app.route("/manager/logout", methods=["POST"])
    def manager_logout():
        session.pop("manager", None)
        session.pop("manager_role", None)
        session.pop("active_team_id", None)
        session.pop("manager_user_email", None)
        flash("Signed out.", "success")
        return redirect(url_for("login_page"))

    @app.route("/dashboard", methods=["GET", "POST"])
    def dashboard():
        """Secret code on /dashboard when not signed in; full leave tracker after success."""
        pw_configured = _manager_password_configured(app)
        lpo_configured = _lpo_sm_password_configured(app)
        any_gate = pw_configured or lpo_configured

        def _gate_render(next_url_val: str | None, **extra):
            return render_template(
                "dashboard.html",
                gate_only=True,
                next_url=next_url_val,
                password_configured=pw_configured,
                lpo_sm_password_configured=lpo_configured,
                **extra,
            )

        if request.method == "POST" and _manager_logged_in():
            return redirect(url_for("dashboard", **request.args))

        if request.method == "POST" and not _manager_logged_in():
            if not any_gate:
                flash("Manager and LPO/SM access are not configured.", "error")
                return _gate_render((request.form.get("next") or request.args.get("next")))
            gate = (request.form.get("gate_kind") or "manager").strip().lower()
            next_raw = (request.form.get("next") or request.args.get("next") or "").strip()
            try:
                ry = int(request.form.get("return_year") or 0)
                rm = int(request.form.get("return_month") or 0)
            except ValueError:
                ry, rm = 0, 0

            def _finish_signin(role: str):
                session["manager"] = True
                session["manager_role"] = role
                session.pop("my_employee_name", None)
                flash("Signed in.", "success")
                dest = _safe_next_path(next_raw)
                if dest:
                    return redirect(dest, code=303)
                if 1 <= rm <= 12 and ry > 2000:
                    return redirect(url_for("dashboard", year=ry, month=rm), code=303)
                return redirect(url_for("dashboard"), code=303)

            if gate == "lpo_sm":
                if not lpo_configured:
                    flash("LPO/SM access is not configured.", "error")
                    return _gate_render(next_raw or None)
                pw = (request.form.get("secret_code") or request.form.get("lpo_sm_code") or "").strip()
                if not pw:
                    flash("Code required.", "error")
                    return _gate_render(next_raw or None)
                if not _check_lpo_sm_password(app, pw):
                    flash("Invalid code.", "error")
                    return _gate_render(next_raw or None)
                return _finish_signin("lpo_sm")

            if not pw_configured:
                flash("Manager access is not configured.", "error")
                return _gate_render(next_raw or None)
            pw = (request.form.get("secret_code") or request.form.get("manager_password") or "").strip()
            if not pw:
                flash("Code required.", "error")
                return _gate_render(next_raw or None)
            if not _check_manager_password(app, pw):
                flash("Invalid code.", "error")
                return _gate_render(next_raw or None)
            return _finish_signin("manager")

        if not _manager_logged_in():
            return _gate_render(request.args.get("next"))

        today = date.today()
        try:
            year = int(request.args.get("year") or today.year)
            month = int(request.args.get("month") or today.month)
        except ValueError:
            year, month = today.year, today.month
        month = max(1, min(12, month))

        roster: tuple[str, ...] = g.manager_roster
        ctx = build_month_context(app, year, month, roster=roster)
        prev_month = date(year, month, 1) - timedelta(days=1)
        next_month = date(year, month, 28) + timedelta(days=4)
        next_month = next_month.replace(day=1)

        conn = get_db(app)
        ph = ",".join("?" * len(roster))
        pending = conn.execute(
            f"""
            SELECT * FROM leave_requests
            WHERE status = 'pending' AND employee_name IN ({ph})
            ORDER BY created_at DESC
            LIMIT 50
            """,
            roster,
        ).fetchall()
        all_leaves = conn.execute(
            f"""
            SELECT * FROM leave_requests
            WHERE employee_name IN ({ph})
            ORDER BY start_date DESC, employee_name ASC
            LIMIT 300
            """,
            roster,
        ).fetchall()
        dashboard_sprint_summaries = _dashboard_sprint_summaries(conn, int(g.manager_team_id))
        conn.close()

        reason_labels = dict(LEAVE_REASONS)
        duration_labels = dict(DURATION_CHOICES)

        _rq = (request.args.get("records_q") or "").strip()
        if len(_rq) > 120:
            _rq = _rq[:120]

        return render_template(
            "dashboard.html",
            gate_only=False,
            pending=pending,
            all_leaves=all_leaves,
            employees=roster,
            reasons=LEAVE_REASONS,
            durations=DURATION_CHOICES,
            reason_labels=reason_labels,
            duration_labels=duration_labels,
            prev_y=prev_month.year,
            prev_m=prev_month.month,
            next_y=next_month.year,
            next_m=next_month.month,
            password_configured=pw_configured,
            dashboard_sprint_summaries=dashboard_sprint_summaries,
            today_iso=date.today().isoformat(),
            all_records_filter_q=_rq,
            **ctx,
        )

    @app.route("/dashboard/leave-tracker-month.xlsx")
    def dashboard_leave_tracker_month_xlsx():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        today = date.today()
        try:
            year = int(request.args.get("year") or today.year)
            month = int(request.args.get("month") or today.month)
        except ValueError:
            year, month = today.year, today.month
        month = max(1, min(12, month))
        roster: tuple[str, ...] = g.manager_roster
        ctx = build_month_context(app, year, month, roster=roster)
        data, err = build_leave_tracker_month_xlsx_bytes(ctx)
        if err or not data:
            flash(err or "Could not build Excel export.", "error")
            return redirect(url_for("dashboard", year=year, month=month))
        safe_team = re.sub(r"[^\w.\-]+", "_", (getattr(g, "manager_team_name", None) or "team").strip()) or "team"
        fname = f"leave_tracker_{safe_team}_{year:04d}-{month:02d}.xlsx"
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.route("/dashboard/leave-tracker-year.xlsx")
    def dashboard_leave_tracker_year_xlsx():
        """Export all months Jan → current month for the active year as a multi-sheet .xlsx."""
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        today = date.today()
        try:
            year = int(request.args.get("year") or today.year)
        except ValueError:
            year = today.year
        current_month = today.month if year == today.year else 12
        roster: tuple[str, ...] = g.manager_roster
        data, err = build_leave_tracker_year_xlsx_bytes(app, year, current_month, roster)
        if err or not data:
            flash(err or "Could not build Excel export.", "error")
            return redirect(url_for("dashboard"))
        safe_team = re.sub(r"[^\w.\-]+", "_", (getattr(g, "manager_team_name", None) or "team").strip()) or "team"
        fname = f"leave_tracker_{safe_team}_{year:04d}_Jan-{date(year, current_month, 1).strftime('%b')}.xlsx"
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.post("/dashboard/leave-tracker-year-upload")
    def dashboard_leave_tracker_year_upload():
        """
        Import leave data from a multi-sheet .xlsx previously exported by
        dashboard_leave_tracker_year_xlsx.

        Each sheet must have:
          Row 1 (hidden meta): col A = "LEAVE_IMPORT_META", col B = year, col C = month,
                                cols F+ = day numbers of the month
          Row 2: headers (skipped)
          Row 3+: col B = employee name, cols F+ = leave code (pl/ul/sl/ll/compoff/half_am/half_pm or empty)

        Existing pending/approved single-day leave entries for the covered month/employee
        combinations are replaced by what is in the sheet.  Only roster members are touched.
        """
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("dashboard"))

        f = request.files.get("leave_xlsx")
        if not f or not f.filename:
            flash("No file selected.", "error")
            return redirect(url_for("dashboard"))

        try:
            from openpyxl import load_workbook as _load_wb
        except ImportError:
            flash("openpyxl is required for import.", "error")
            return redirect(url_for("dashboard"))

        try:
            wb = _load_wb(f, data_only=True)
        except Exception:
            flash("Could not read the uploaded file. Make sure it is a valid .xlsx.", "error")
            return redirect(url_for("dashboard"))

        roster_set = set(g.manager_roster)

        def _normalise_upload_code(raw: str) -> str:
            """
            Normalise a cell value from the uploaded xlsx into a canonical leave
            code understood by the importer.  Returns "" when the value should be
            ignored.
            """
            v = raw.strip().lower()
            # Nokia e-tool approved — keep as-is (starts with "a")
            if v.startswith("a"):
                return v
            # CompOff aliases: "co", "c", "compoff", "comp off", "comp-off"
            if v in ("co", "c", "compoff", "comp off", "comp-off", "comp_off"):
                return "compoff"
            # half-day planned leave
            if v in ("pl1", "half_am", "am", "pl_am"):
                return "half_am"
            if v in ("pl2", "half_pm", "pm", "pl_pm"):
                return "half_pm"
            # half-day unplanned leave  (store as ul, half-day flag added in insert)
            if v in ("ul1",):
                return "ul1"
            if v in ("ul2",):
                return "ul2"
            # full planned leave (N → PL mapping already applied in the xlsx)
            if v in ("pl", "plpl", "ppl", "full", "fl"):
                return "pl"
            # standard codes
            if v in ("ul", "sl", "ll"):
                return v
            return ""

        # "a" and "a½" are Nokia e-tool approved cells (shown green)
        valid_codes = {c[0] for c in LEAVE_REASONS} | {
            "half_am", "half_pm", "full", "a", "a½", "a\u00bd",
            "ul1", "ul2",  # half-day unplanned
        }
        conn = get_db(app)
        sheets_imported = 0
        rows_written = 0
        errors: list[str] = []
        created_ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        ip = client_ip()

        for ws in wb.worksheets:
            # ── read meta row ────────────────────────────────────────────────
            meta_marker = ws.cell(row=1, column=1).value
            if str(meta_marker or "").strip() != "LEAVE_IMPORT_META":
                continue  # not our format — skip silently
            try:
                yr = int(ws.cell(row=1, column=2).value or 0)
                mo = int(ws.cell(row=1, column=3).value or 0)
            except (TypeError, ValueError):
                errors.append(f"Sheet '{ws.title}': invalid year/month in meta row.")
                continue
            if not (yr >= 2000 and 1 <= mo <= 12):
                errors.append(f"Sheet '{ws.title}': out-of-range year/month ({yr}-{mo:02d}).")
                continue

            day0 = 6  # column F
            # build day-number → column index map from meta row
            day_col_map: dict[int, int] = {}  # day_number → col index (1-based)
            max_col = ws.max_column or (day0 + 31)
            for col in range(day0, max_col + 1):
                v = ws.cell(row=1, column=col).value
                try:
                    dn = int(v)
                    if 1 <= dn <= 31:
                        day_col_map[dn] = col
                except (TypeError, ValueError):
                    break

            if not day_col_map:
                errors.append(f"Sheet '{ws.title}': no day columns found in meta row.")
                continue

            from calendar import monthrange as _mrange
            _, last_day = _mrange(yr, mo)

            # ── collect all leave data from the sheet ────────────────────────
            # blank cells are stored as "" — they will clear any existing "A" on that day.
            sheet_data: dict[str, dict[int, str]] = {}  # emp → {day: normalised code or ""}
            for row in ws.iter_rows(min_row=3, values_only=True):
                emp_raw = str(row[1] or "").strip() if len(row) > 1 else ""
                if not emp_raw or emp_raw not in roster_set:
                    continue
                emp_days: dict[int, str] = {}
                for day_num, col_idx in day_col_map.items():
                    raw_val = str(row[col_idx - 1] or "").strip() if len(row) >= col_idx else ""
                    if not raw_val:
                        emp_days[day_num] = ""  # explicit blank → clears any existing A
                        continue
                    cell_val = _normalise_upload_code(raw_val)
                    emp_days[day_num] = cell_val if (cell_val and cell_val in valid_codes) else ""
                sheet_data[emp_raw] = emp_days

            if not sheet_data:
                continue

            # ── upsert: the uploaded sheet is the full source of truth for the month.
            #    Delete ALL pending/approved leaves (including Nokia "A" rows) for the
            #    covered month, then re-insert only days that have a non-blank code. ──
            for emp, emp_days in sheet_data.items():
                month_start = date(yr, mo, 1).isoformat()
                month_end = date(yr, mo, last_day).isoformat()
                # Remove ALL pending/approved leaves for this employee in the month
                # (including Nokia e-tool approved — blank cells override them).
                conn.execute(
                    """
                    DELETE FROM leave_requests
                    WHERE employee_name = ?
                      AND start_date <= ? AND end_date >= ?
                      AND status IN ('pending', 'approved')
                    """,
                    (emp, month_end, month_start),
                )
                # Insert each day that has a non-blank code in the sheet.
                for day_num, code in emp_days.items():
                    if not code or day_num > last_day:
                        continue
                    work_date = date(yr, mo, day_num).isoformat()
                    # Nokia-approved "A" / "A½" cells
                    is_nokia = code.startswith("a")
                    half_nokia = code in ("a½", "a\u00bd")
                    if is_nokia:
                        reason = "pl"
                        duration_type = "half_am" if half_nokia else "full"
                        desc = f"Imported from xlsx ({yr}-{mo:02d}) {NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER}"
                    elif code in ("half_am", "half_pm"):
                        reason = "pl"
                        duration_type = code
                        desc = f"Imported from xlsx ({yr}-{mo:02d})"
                    elif code == "full":
                        reason = "pl"
                        duration_type = "full"
                        desc = f"Imported from xlsx ({yr}-{mo:02d})"
                    elif code == "ul1":
                        reason = "ul"
                        duration_type = "half_am"
                        desc = f"Imported from xlsx ({yr}-{mo:02d})"
                    elif code == "ul2":
                        reason = "ul"
                        duration_type = "half_pm"
                        desc = f"Imported from xlsx ({yr}-{mo:02d})"
                    elif code in {c[0] for c in LEAVE_REASONS}:
                        # covers: pl, ul, sl, ll, compoff
                        reason = code
                        duration_type = "full"
                        desc = f"Imported from xlsx ({yr}-{mo:02d})"
                    else:
                        continue
                    conn.execute(
                        """
                        INSERT INTO leave_requests
                        (employee_name, reason, description, start_date, end_date,
                         duration_type, status, created_at, submitted_ip)
                        VALUES (?, ?, ?, ?, ?, ?, 'approved', ?, ?)
                        """,
                        (emp, reason, desc, work_date, work_date, duration_type, created_ts, ip),
                    )
                    rows_written += 1

            sheets_imported += 1

        conn.commit()
        conn.close()

        if errors:
            flash("Import completed with warnings: " + "; ".join(errors[:3]), "warning")
        if sheets_imported == 0:
            flash("No valid leave sheets found in the uploaded file.", "error")
        else:
            flash(
                f"Imported {sheets_imported} month(s), {rows_written} leave day(s) written.",
                "success",
            )
        return redirect(url_for("dashboard"))

    @app.post("/dashboard/api/leave-tracker-eleave")
    def dashboard_api_leave_tracker_eleave():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 401
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        team_id = getattr(g, "manager_team_id", None)
        if team_id is None:
            return jsonify({"ok": False, "error": "team"}), 400
        payload = request.get_json(silent=True) or {}
        try:
            y = int(payload.get("year"))
            mo = int(payload.get("month"))
            emp = str(payload.get("employee_name") or "").strip()
            raw = payload.get("days")
            if raw is None or str(raw).strip() == "":
                days = 0.0
            else:
                days = float(raw)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "bad_input"}), 400
        if not (1 <= mo <= 12):
            return jsonify({"ok": False, "error": "bad_month"}), 400
        roster_t = tuple(g.manager_roster)
        if emp not in roster_t:
            return jsonify({"ok": False, "error": "roster"}), 400
        conn = get_db(app)
        if not conn.execute("SELECT 1 FROM teams WHERE id = ?", (int(team_id),)).fetchone():
            conn.close()
            return jsonify({"ok": False, "error": "team"}), 404
        ts = _utc_stamp()
        conn.execute(
            """
            INSERT INTO leave_tracker_eleaves (team_id, year, month, employee_name, days, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(team_id, year, month, employee_name) DO UPDATE SET
                days = excluded.days,
                updated_at = excluded.updated_at
            """,
            (int(team_id), y, mo, emp, days, ts),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "days": days})

    @app.post("/dashboard/api/leave-tracker-meet-quick-leave")
    def dashboard_api_leave_tracker_meet_quick_leave():
        """Manager: add a single-day leave on the month worksheet (same persistence as DSM /meet quick-leave)."""
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_form_ok():
            return jsonify({"ok": False, "error": "csrf"}), 400

        emp = (request.form.get("employee_name") or "").strip()
        work_date = (request.form.get("work_date") or "").strip()
        reason = (request.form.get("reason") or "pl").strip()
        duration_type = (request.form.get("duration_type") or "full").strip()
        try:
            y = int(request.form.get("year") or 0)
            mo = int(request.form.get("month") or 0)
        except ValueError:
            return jsonify({"ok": False, "error": "bad_input"}), 400

        if emp not in g.manager_roster or not work_date:
            return jsonify({"ok": False, "error": "bad_input"}), 400
        if reason not in {c[0] for c in LEAVE_REASONS}:
            return jsonify({"ok": False, "error": "reason"}), 400
        if duration_type not in {c[0] for c in DURATION_CHOICES} or duration_type == "multi":
            return jsonify({"ok": False, "error": "dur"}), 400
        if not (1 <= mo <= 12) or y < 2000:
            return jsonify({"ok": False, "error": "bad_month"}), 400
        try:
            wd = date.fromisoformat(work_date[:10])
        except ValueError:
            return jsonify({"ok": False, "error": "bad_date"}), 400
        first = date(y, mo, 1)
        _, ld = monthrange(y, mo)
        last = date(y, mo, ld)
        if not (first <= wd <= last):
            return jsonify({"ok": False, "error": "out_of_month"}), 400

        created = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        ip = client_ip()
        conn = get_db(app)
        conn.execute(
            """
            INSERT INTO leave_requests
            (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
            VALUES (?, ?, '', ?, ?, ?, 'approved', ?, ?)
            """,
            (emp, reason, work_date[:10], work_date[:10], duration_type, created, ip),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/dashboard/api/leave-tracker-meet-approve-day")
    def dashboard_api_leave_tracker_meet_approve_day():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_form_ok():
            return jsonify({"ok": False, "error": "csrf"}), 400
        try:
            y = int(request.form.get("year") or 0)
            mo = int(request.form.get("month") or 0)
        except ValueError:
            return jsonify({"ok": False, "error": "bad_input"}), 400
        row, err = _dashboard_validate_leave_for_month(
            app,
            request.form.get("leave_id"),
            (request.form.get("work_date") or "").strip(),
            y,
            mo,
            g.manager_roster,
        )
        if err or not row:
            return jsonify({"ok": False, "error": err or "bad_input"}), 400

        leave_id = int(row["id"])
        work_date = (request.form.get("work_date") or "").strip()[:10]
        ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        conn = get_db(app)
        conn.execute(
            """
            INSERT INTO meet_leave_day (leave_id, work_date, decision, updated_at)
            VALUES (?, ?, 'approved', ?)
            ON CONFLICT(leave_id, work_date) DO UPDATE SET
                decision = excluded.decision,
                updated_at = excluded.updated_at
            """,
            (leave_id, work_date, ts),
        )
        conn.commit()
        conn.close()
        _maybe_promote_leave_after_meet_days(app, leave_id)
        return jsonify({"ok": True})

    @app.post("/dashboard/api/leave-tracker-meet-remove-day")
    def dashboard_api_leave_tracker_meet_remove_day():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_form_ok():
            return jsonify({"ok": False, "error": "csrf"}), 400
        try:
            y = int(request.form.get("year") or 0)
            mo = int(request.form.get("month") or 0)
        except ValueError:
            return jsonify({"ok": False, "error": "bad_input"}), 400
        row, err = _dashboard_validate_leave_for_month(
            app,
            request.form.get("leave_id"),
            (request.form.get("work_date") or "").strip(),
            y,
            mo,
            g.manager_roster,
        )
        if err or not row:
            return jsonify({"ok": False, "error": err or "bad_input"}), 400

        leave_id = int(row["id"])
        work_date = (request.form.get("work_date") or "").strip()[:10]
        if not _apply_leave_tracker_day_removal(app, leave_id, work_date):
            return jsonify({"ok": False, "error": "not_found"}), 404
        return jsonify({"ok": True})

    @app.route("/manager/audit.csv")
    def manager_audit_csv():
        if not _manager_logged_in():
            return redirect(url_for("dashboard", next=request.path))
        conn = get_db(app)
        rows = conn.execute(
            """
            SELECT id, employee_name, reason, start_date, end_date, duration_type,
                   status, created_at, submitted_ip
            FROM leave_requests
            ORDER BY created_at DESC
            """
        ).fetchall()
        conn.close()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(
            [
                "id",
                "employee_name",
                "reason",
                "start_date",
                "end_date",
                "duration_type",
                "status",
                "created_at",
                "submitted_ip",
            ]
        )
        for row in rows:
            w.writerow(
                [
                    row["id"],
                    row["employee_name"],
                    row["reason"],
                    row["start_date"],
                    row["end_date"],
                    row["duration_type"],
                    row["status"],
                    row["created_at"],
                    row["submitted_ip"] if "submitted_ip" in row.keys() else "",
                ]
            )
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=leave_audit_log.csv"},
        )

    @app.route("/dashboard/leave/<int:leave_id>/edit", methods=["GET", "POST"])
    def dashboard_leave_edit(leave_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))

        conn = get_db(app)
        row = conn.execute("SELECT * FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()
        conn.close()
        if not row:
            flash("Not found.", "error")
            return redirect(url_for("dashboard"))

        if row["employee_name"] not in g.manager_roster:
            flash("That leave is not on your team roster.", "error")
            return redirect(url_for("dashboard"))

        reason_labels = dict(LEAVE_REASONS)
        duration_labels = dict(DURATION_CHOICES)

        if request.method == "POST":
            employee_name = (request.form.get("employee_name") or "").strip()
            reason = (request.form.get("reason") or "").strip()
            description = ""
            start_date = (request.form.get("start_date") or "").strip()
            end_date = (request.form.get("end_date") or "").strip()
            duration_type = (request.form.get("duration_type") or "").strip()
            status = (request.form.get("status") or "").strip()

            errors: list[str] = []
            if employee_name not in g.manager_roster:
                errors.append("Employee must be one of the roster names.")
            if reason not in {c[0] for c in LEAVE_REASONS}:
                errors.append("Select a valid leave reason.")
            if not start_date or not end_date:
                errors.append("Start and end dates are required.")
            if duration_type not in {c[0] for c in DURATION_CHOICES}:
                errors.append("Select a valid duration.")
            if status not in {"pending", "approved", "rejected"}:
                errors.append("Select a valid status.")
            else:
                try:
                    if date.fromisoformat(end_date) < date.fromisoformat(start_date):
                        errors.append("End date must be on or after start date.")
                except ValueError:
                    errors.append("Invalid date format.")

            if errors:
                for e in errors:
                    flash(e, "error")
                return render_template(
                    "leave_edit.html",
                    row=row,
                    reasons=LEAVE_REASONS,
                    durations=DURATION_CHOICES,
                    employees=g.manager_roster,
                    reason_labels=reason_labels,
                    duration_labels=duration_labels,
                    form=request.form,
                    return_year=request.form.get("return_year") or "",
                    return_month=request.form.get("return_month") or "",
                    return_records_q=(request.form.get("return_records_q") or "").strip()[:120],
                )

            conn = get_db(app)
            conn.execute(
                """
                UPDATE leave_requests
                SET employee_name = ?, reason = ?, description = ?, start_date = ?, end_date = ?,
                    duration_type = ?, status = ?
                WHERE id = ?
                """,
                (employee_name, reason, description, start_date, end_date, duration_type, status, leave_id),
            )
            if status == "rejected":
                _purge_meet_leave_day_rows_for_leave(conn, leave_id)
            conn.commit()
            conn.close()
            flash("Saved.", "success")
            try:
                ry = int(request.form.get("return_year") or 0)
                rm = int(request.form.get("return_month") or 0)
                rq = (request.form.get("return_records_q") or "").strip()[:120]
                if 1 <= rm <= 12 and ry > 2000:
                    if rq:
                        return redirect(url_for("dashboard", year=ry, month=rm, records_q=rq), code=303)
                    return redirect(url_for("dashboard", year=ry, month=rm), code=303)
            except ValueError:
                pass
            return redirect(url_for("dashboard"))

        return render_template(
            "leave_edit.html",
            row=row,
            reasons=LEAVE_REASONS,
            durations=DURATION_CHOICES,
            employees=g.manager_roster,
            reason_labels=reason_labels,
            duration_labels=duration_labels,
            form={k: row[k] for k in row.keys()},
            return_year=request.args.get("year") or "",
            return_month=request.args.get("month") or "",
            return_records_q=(request.args.get("records_q") or "").strip()[:120],
        )

    @app.route("/meet")
    def meet():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        anchor = _parse_meet_anchor(request.args.get("d"))
        ctx = build_meet_context(app, anchor, roster=g.manager_roster)
        return render_template("meet.html", reasons=LEAVE_REASONS, **ctx)

    @app.route("/meet/approve-day", methods=["POST"])
    def meet_approve_day():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_form_ok():
            return jsonify({"ok": False, "error": "csrf"}), 400

        row, _anchor, err = _meet_validate_leave_in_window(
            app,
            request.form.get("leave_id"),
            (request.form.get("work_date") or "").strip(),
            request.form.get("anchor"),
            g.manager_roster,
        )
        if err or not row:
            return jsonify({"ok": False, "error": err or "bad_input"}), 400

        leave_id = int(row["id"])
        work_date = (request.form.get("work_date") or "").strip()[:10]
        ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        conn = get_db(app)
        conn.execute(
            """
            INSERT INTO meet_leave_day (leave_id, work_date, decision, updated_at)
            VALUES (?, ?, 'approved', ?)
            ON CONFLICT(leave_id, work_date) DO UPDATE SET
                decision = excluded.decision,
                updated_at = excluded.updated_at
            """,
            (leave_id, work_date, ts),
        )
        conn.commit()
        conn.close()
        _maybe_promote_leave_after_meet_days(app, leave_id)
        return jsonify({"ok": True})

    @app.route("/meet/remove-day", methods=["POST"])
    def meet_remove_day():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_form_ok():
            return jsonify({"ok": False, "error": "csrf"}), 400

        row, _anchor, err = _meet_validate_leave_in_window(
            app,
            request.form.get("leave_id"),
            (request.form.get("work_date") or "").strip(),
            request.form.get("anchor"),
            g.manager_roster,
        )
        if err or not row:
            return jsonify({"ok": False, "error": err or "bad_input"}), 400

        leave_id = int(row["id"])
        work_date = (request.form.get("work_date") or "").strip()[:10]
        if not _apply_leave_tracker_day_removal(app, leave_id, work_date):
            return jsonify({"ok": False, "error": "not_found"}), 404
        return jsonify({"ok": True})

    @app.route("/meet/quick-leave", methods=["POST"])
    def meet_quick_leave():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_form_ok():
            return jsonify({"ok": False, "error": "csrf"}), 400

        emp = (request.form.get("employee_name") or "").strip()
        work_date = (request.form.get("work_date") or "").strip()
        reason = (request.form.get("reason") or "pl").strip()
        duration_type = (request.form.get("duration_type") or "full").strip()
        anchor = _parse_meet_anchor(request.form.get("anchor"))

        if emp not in g.manager_roster or not work_date:
            return jsonify({"ok": False, "error": "bad_input"}), 400
        if reason not in {c[0] for c in LEAVE_REASONS}:
            return jsonify({"ok": False, "error": "reason"}), 400
        if duration_type not in {c[0] for c in DURATION_CHOICES} or duration_type == "multi":
            return jsonify({"ok": False, "error": "dur"}), 400
        try:
            wd = date.fromisoformat(work_date)
        except ValueError:
            return jsonify({"ok": False, "error": "bad_date"}), 400
        d0, d4 = _meet_window(anchor)
        if not (d0 <= wd <= d4):
            return jsonify({"ok": False, "error": "out_of_window"}), 400

        created = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        ip = client_ip()
        conn = get_db(app)
        conn.execute(
            """
            INSERT INTO leave_requests
            (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
            VALUES (?, ?, '', ?, ?, ?, 'approved', ?, ?)
            """,
            (emp, reason, work_date, work_date, duration_type, created, ip),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.route("/scrum", methods=["GET"])
    def scrum():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        sprints = list(
            conn.execute(
                """
                SELECT id, name, start_date, end_date, goal, COALESCE(is_closed, 0) AS is_closed
                FROM scrum_sprint
                WHERE team_id = ?
                ORDER BY end_date DESC, start_date DESC, id DESC
                """,
                (team_id,),
            )
        )
        for sp in sprints:
            _maybe_auto_close_scrum_sprint(conn, int(sp["id"]))
        conn.commit()
        sprints = list(
            conn.execute(
                """
                SELECT id, name, start_date, end_date, goal, COALESCE(is_closed, 0) AS is_closed
                FROM scrum_sprint
                WHERE team_id = ?
                ORDER BY end_date DESC, start_date DESC, id DESC
                """,
                (team_id,),
            )
        )
        portal_proposals_pending = _count_pending_scrum_portal_proposals(conn, team_id)
        next_sd = _next_sprint_start_after_latest_end(conn, team_id)
        next_suggested_sprint_name = None
        if next_sd is not None:
            prev_nm = conn.execute(
                """
                SELECT name FROM scrum_sprint
                WHERE team_id = ?
                ORDER BY end_date DESC, start_date DESC, id DESC
                LIMIT 1
                """,
                (team_id,),
            ).fetchone()
            if prev_nm and (prev_nm["name"] or "").strip():
                next_suggested_sprint_name = suggest_next_sprint_name_from_previous(str(prev_nm["name"]))
        conn.close()
        td = date.today()
        next_iso = next_sd.isoformat() if next_sd else None
        return render_template(
            "scrum_hub.html",
            sprints=sprints,
            today_iso=td.isoformat(),
            next_sprint_start_iso=next_iso,
            next_suggested_sprint_name=next_suggested_sprint_name,
            portal_proposals_pending=portal_proposals_pending,
        )

    @app.get("/scrum/api/sprint-team-capacity")
    def scrum_api_sprint_team_capacity():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        try:
            sd = date.fromisoformat((request.args.get("start_date") or "").strip()[:10])
            ed = date.fromisoformat((request.args.get("end_date") or "").strip()[:10])
        except ValueError:
            return jsonify({"ok": False, "error": "bad_date"}), 400
        if ed < sd:
            sd, ed = ed, sd
        h = compute_team_sprint_capacity_leave_hours(app, g.manager_roster, sd, ed)
        return jsonify(
            {"ok": True, "hours": round(h, 1), "roster_count": len(g.manager_roster)}
        )

    @app.get("/scrum/team-bundle.json")
    def scrum_team_bundle_download():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        try:
            bundle = build_manager_team_bundle_dict(conn, team_id)
        except ValueError as e:
            conn.close()
            flash(str(e), "error")
            return redirect(url_for("scrum"))
        conn.close()
        raw = json.dumps(bundle, indent=2, ensure_ascii=False).encode("utf-8")
        safe = re.sub(r"[^\w.\-]+", "_", (getattr(g, "manager_team_name", None) or "team").strip()) or "team"
        fname = f"{safe}-team-bundle.json"
        return Response(
            raw,
            mimetype="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.post("/scrum/team-bundle/import")
    def scrum_team_bundle_import():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        f = request.files.get("bundle_file")
        if not f or not (getattr(f, "filename", None) or "").strip():
            flash("Choose a JSON bundle file to import.", "error")
            return redirect(url_for("scrum"))
        try:
            raw = f.read().decode("utf-8")
            bundle = json.loads(raw)
        except (UnicodeDecodeError, json.JSONDecodeError):
            flash("Could not read that file as JSON.", "error")
            return redirect(url_for("scrum"))
        if not isinstance(bundle, dict):
            flash("Bundle must be a JSON object.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        try:
            import_manager_team_bundle(conn, team_id, bundle)
            conn.commit()
        except ValueError as e:
            conn.rollback()
            conn.close()
            flash(str(e), "error")
            return redirect(url_for("scrum"))
        except Exception:
            conn.rollback()
            conn.close()
            _log.exception("team bundle import failed")
            flash("Import failed due to a database error.", "error")
            return redirect(url_for("scrum"))
        conn.close()
        flash(
            "Imported team bundle for this team: roster, hub mode, task kinds, sprints, stickies, "
            "activity, goals, daily tasks, appreciation, and portal proposals were replaced.",
            "success",
        )
        return redirect(url_for("scrum"))

    @app.route("/scrum/portal-proposals", methods=["GET"])
    def scrum_portal_proposals():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        rows = list(
            conn.execute(
                """
                SELECT p.*, s.name AS sprint_name,
                  (SELECT title FROM scrum_sprint_item WHERE id = p.item_id) AS item_title
                FROM scrum_portal_proposal p
                JOIN scrum_sprint s ON s.id = p.sprint_id AND s.team_id = p.team_id
                WHERE p.team_id = ? AND p.status = 'pending'
                ORDER BY p.created_at ASC
                """,
                (team_id,),
            )
        )
        conn.close()
        display: list[dict] = []
        for r in rows:
            try:
                payload = json.loads(r["payload_json"] or "{}")
            except json.JSONDecodeError:
                payload = {}
            it = None
            if "item_title" in r.keys():
                it = r["item_title"]
            display.append(
                {
                    "id": int(r["id"]),
                    "created_at": r["created_at"],
                    "sprint_name": r["sprint_name"],
                    "sprint_id": int(r["sprint_id"]),
                    "proposer": r["proposer_roster_name"],
                    "action": r["action"],
                    "summary": _portal_proposal_summary_line(str(r["action"]), payload, it),
                    "payload_pretty": json.dumps(payload, indent=2, ensure_ascii=False)[:12000],
                }
            )
        return render_template("scrum_portal_proposals.html", proposals=display)

    @app.post("/scrum/portal-proposal/resolve")
    def scrum_portal_proposal_resolve():
        next_raw = request.form.get("next")
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return _redirect_after_portal_proposal(next_raw)
        proposal_id = _parse_optional_int(request.form.get("proposal_id"))
        decision = (request.form.get("decision") or "").strip().lower()
        note = (request.form.get("note") or "").strip()[:500]
        if not proposal_id or decision not in ("approve", "reject"):
            flash("Missing proposal or decision.", "error")
            return _redirect_after_portal_proposal(next_raw)
        if decision == "approve" and request.form.get("manager_attest") != "1":
            flash(
                "Check the confirmation box: you must review the change (including DoD checklist when shown) before approving.",
                "error",
            )
            return _redirect_after_portal_proposal(next_raw)
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        try:
            conn.execute("BEGIN IMMEDIATE")
            prop = conn.execute(
                """
                SELECT * FROM scrum_portal_proposal
                WHERE id = ? AND team_id = ? AND status = 'pending'
                """,
                (proposal_id, team_id),
            ).fetchone()
            if not prop:
                conn.rollback()
                conn.close()
                flash("That proposal was not found or is already resolved.", "error")
                return _redirect_after_portal_proposal(next_raw)
            ts = _utc_stamp()
            if decision == "reject":
                conn.execute(
                    """
                    UPDATE scrum_portal_proposal
                    SET status = 'rejected', resolved_at = ?, resolution_note = ?
                    WHERE id = ? AND team_id = ? AND status = 'pending'
                    """,
                    (ts, note, proposal_id, team_id),
                )
                conn.commit()
                conn.close()
                flash("Proposal rejected.", "success")
                return _redirect_after_portal_proposal(next_raw)
            ok, err = _apply_scrum_portal_proposal_core(conn, prop, team_id)
            if not ok:
                conn.rollback()
                conn.close()
                msg = _sprint_board_frozen_flash_for_reason(err) if err in ("sprint_closed", "sprint_ended") else None
                flash(msg or f"Could not apply change: {err}", "error")
                return _redirect_after_portal_proposal(next_raw)
            conn.execute(
                """
                UPDATE scrum_portal_proposal
                SET status = 'approved', resolved_at = ?, resolution_note = ?
                WHERE id = ? AND team_id = ? AND status = 'pending'
                """,
                (ts, note, proposal_id, team_id),
            )
            conn.commit()
        except Exception as ex:
            conn.rollback()
            conn.close()
            _log.exception("portal proposal approve failed: %s", ex)
            flash("Something went wrong while applying the change.", "error")
            return _redirect_after_portal_proposal(next_raw)
        conn.close()
        flash("Change applied to the sprint board.", "success")
        return _redirect_after_portal_proposal(next_raw)

    @app.route("/scrum/sprint/<int:sprint_id>", methods=["GET"])
    def scrum_sprint_team(sprint_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        _maybe_auto_close_scrum_sprint(conn, int(sprint_id))
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        session["active_sprint_id"] = int(sprint_id)
        members = _sprint_team_overview_rows(app, conn, team_id, int(sprint_id), g.manager_roster)
        pending_by = _pending_portal_proposals_grouped_for_sprint(conn, team_id, int(sprint_id))
        for m in members:
            m["pending_portal"] = pending_by.get(m["name"], [])
        bd_ctx = build_sprint_burndown_chart_context(
            conn, int(sprint_id), str(sprint["start_date"])[:10], str(sprint["end_date"])[:10]
        )
        kind_stack_ctx = build_sprint_task_kind_stack_chart_context(conn, team_id, int(sprint_id))
        portal_proposals_pending = _count_pending_scrum_portal_proposals(conn, team_id, int(sprint_id))
        sprint_team_capacity_hours: float | None = None
        try:
            sd_cap = date.fromisoformat(str(sprint["start_date"])[:10])
            ed_cap = date.fromisoformat(str(sprint["end_date"])[:10])
            sprint_team_capacity_hours = round(
                compute_team_sprint_capacity_leave_hours(app, g.manager_roster, sd_cap, ed_cap), 1
            )
        except ValueError:
            sprint_team_capacity_hours = None
        meter_ctx = build_sprint_team_meter_context(
            sprint_team_capacity_hours=sprint_team_capacity_hours,
            kind_stack_total_est=float(kind_stack_ctx.get("kind_stack_total_est") or 0),
            kind_stack_total_burnt=float(kind_stack_ctx.get("kind_stack_total_burnt") or 0),
        )
        if sprint_team_capacity_hours is not None and not _sprint_board_frozen_reason_for_row(sprint):
            ts = _utc_stamp()
            conn.execute(
                """
                UPDATE scrum_sprint
                SET team_capacity_hours = ?, updated_at = ?
                WHERE id = ? AND team_id = ?
                """,
                (float(sprint_team_capacity_hours), ts, int(sprint_id), team_id),
            )
            conn.commit()
        tf = _sprint_team_page_template_flags(conn, team_id, int(sprint_id))
        conn.commit()
        conn.close()
        return render_template(
            "scrum_sprint_team.html",
            sprint_id=int(sprint_id),
            sprint_name=sprint["name"],
            sprint_start=str(sprint["start_date"])[:10],
            sprint_end=str(sprint["end_date"])[:10],
            sprint_team_capacity_hours=sprint_team_capacity_hours,
            sprint_board_readonly=tf["sprint_board_readonly"],
            sprint_manually_closed=tf["sprint_manually_closed"],
            sprint_status_label=tf["sprint_status_label"],
            sprint_freeze_mode=tf["sprint_freeze_mode"],
            members=members,
            portal_proposals_pending=portal_proposals_pending,
            **bd_ctx,
            **kind_stack_ctx,
            **meter_ctx,
        )

    @app.route("/scrum/sprint/<int:sprint_id>/team-detailed", methods=["GET"])
    def scrum_sprint_team_detailed(sprint_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        _maybe_auto_close_scrum_sprint(conn, int(sprint_id))
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        session["active_sprint_id"] = int(sprint_id)
        members = _sprint_team_overview_detailed_rows(app, conn, team_id, int(sprint_id), g.manager_roster)
        pending_by = _pending_portal_proposals_grouped_for_sprint(conn, team_id, int(sprint_id))
        for m in members:
            m["pending_portal"] = pending_by.get(m["name"], [])
        portal_proposals_pending = _count_pending_scrum_portal_proposals(conn, team_id, int(sprint_id))
        roster_t = tuple(g.manager_roster)
        member_filter = (request.args.get("member") or "").strip()
        if member_filter and member_filter not in roster_t:
            member_filter = ""
        members_display = [m for m in members if (not member_filter or m.get("name") == member_filter)]
        sprint_team_capacity_hours: float | None = None
        try:
            sd_cap = date.fromisoformat(str(sprint["start_date"])[:10])
            ed_cap = date.fromisoformat(str(sprint["end_date"])[:10])
            sprint_team_capacity_hours = round(
                compute_team_sprint_capacity_leave_hours(app, g.manager_roster, sd_cap, ed_cap), 1
            )
        except ValueError:
            sprint_team_capacity_hours = None
        if sprint_team_capacity_hours is not None and not _sprint_board_frozen_reason_for_row(sprint):
            ts = _utc_stamp()
            conn.execute(
                """
                UPDATE scrum_sprint
                SET team_capacity_hours = ?, updated_at = ?
                WHERE id = ? AND team_id = ?
                """,
                (float(sprint_team_capacity_hours), ts, int(sprint_id), team_id),
            )
            conn.commit()
        summary_one_liner = _sprint_team_detail_one_liner(
            members_display, str(sprint["name"] or ""), sprint_team_capacity_hours
        )
        tf = _sprint_team_page_template_flags(conn, team_id, int(sprint_id))
        conn.commit()
        conn.close()
        return render_template(
            "scrum_sprint_team_detailed.html",
            sprint_id=int(sprint_id),
            sprint_name=sprint["name"],
            sprint_start=str(sprint["start_date"])[:10],
            sprint_end=str(sprint["end_date"])[:10],
            sprint_team_capacity_hours=sprint_team_capacity_hours,
            sprint_readonly=tf["sprint_board_readonly"],
            members=members_display,
            member_filter=member_filter,
            team_detailed_roster_names=list(roster_t),
            portal_proposals_pending=portal_proposals_pending,
            summary_one_liner=summary_one_liner,
            kb_form_urls={
                "attachment_upload": url_for("scrum_sprint_item_attachment_upload"),
                "attachment_delete": url_for("scrum_sprint_item_attachment_delete"),
            },
        )

    @app.route("/scrum/hppm", methods=["GET"])
    def scrum_hppm_entry():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        sid = session.get("active_sprint_id")
        if not sid:
            flash("Open a sprint from the hub first, then use HPPM view.", "error")
            return redirect(url_for("scrum"))
        return redirect(url_for("scrum_sprint_hppm_view", sprint_id=int(sid)))

    @app.route("/scrum/sprint/<int:sprint_id>/hppm", methods=["GET"])
    def scrum_sprint_hppm_view(sprint_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        _maybe_auto_close_scrum_sprint(conn, int(sprint_id))
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        session["active_sprint_id"] = int(sprint_id)
        hppm_ctx = build_hppm_sprint_page_extra_context(
            conn,
            app,
            team_id=team_id,
            sprint_id=int(sprint_id),
            roster=g.manager_roster,
            sprint_start=str(sprint["start_date"])[:10],
            sprint_end=str(sprint["end_date"])[:10],
        )
        area_stack_ctx = build_sprint_task_area_stack_chart_context(conn, team_id, int(sprint_id), chart_palette="hppm")
        match_kind_h = None
        if area_stack_ctx.get("area_stack_has_chart"):
            try:
                match_kind_h = int(area_stack_ctx.get("area_stack_svg_h") or 0)
            except (TypeError, ValueError):
                match_kind_h = None
        kind_stack_ctx = build_sprint_task_kind_stack_chart_context(
            conn,
            team_id,
            int(sprint_id),
            horizontal=True,
            chart_palette="hppm",
            match_svg_height=match_kind_h,
        )
        tf = _sprint_team_page_template_flags(conn, team_id, int(sprint_id))
        conn.commit()
        conn.close()
        return render_template(
            "scrum_sprint_hppm.html",
            sprint_id=int(sprint_id),
            sprint_name=sprint["name"],
            sprint_start=str(sprint["start_date"])[:10],
            sprint_end=str(sprint["end_date"])[:10],
            sprint_readonly=tf["sprint_board_readonly"],
            sprint_freeze_mode=tf["sprint_freeze_mode"],
            **kind_stack_ctx,
            **area_stack_ctx,
            **hppm_ctx,
        )

    @app.route("/scrum/sprint/<int:sprint_id>/leave-tracker", methods=["GET"])
    def scrum_sprint_leave_tracker(sprint_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        try:
            sd = date.fromisoformat(str(sprint["start_date"])[:10])
            ed = date.fromisoformat(str(sprint["end_date"])[:10])
        except ValueError:
            conn.close()
            flash("Invalid sprint dates.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        if sd > ed:
            sd, ed = ed, sd
        days_meta, grid_rows = build_sprint_leave_tracker_context(conn, app, sd, ed, g.manager_roster)
        conn.close()
        if sd.year == ed.year:
            if sd.month == ed.month:
                date_range_label = f"{sd.strftime('%b')} {sd.day}–{ed.day}, {sd.year}"
            else:
                date_range_label = f"{sd.strftime('%b %d')} – {ed.strftime('%b %d, %Y')}"
        else:
            date_range_label = f"{sd.strftime('%b %d, %Y')} – {ed.strftime('%b %d, %Y')}"
        return render_template(
            "scrum_sprint_leave_tracker.html",
            sprint_id=int(sprint_id),
            sprint_name=sprint["name"],
            sprint_start=sd.isoformat(),
            sprint_end=ed.isoformat(),
            date_range_label=date_range_label,
            days=days_meta,
            grid_rows=grid_rows,
            back_url=url_for("scrum_sprint_team", sprint_id=sprint_id),
        )

    @app.route("/scrum/sprint/<int:sprint_id>/board", methods=["GET"])
    def scrum_member_board(sprint_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        raw = (request.args.get("assignee") or "").strip()
        assignee = resolve_employee_name(raw, roster=g.manager_roster) or raw[:200]
        if not assignee.strip():
            flash("Pick a team member (assignee).", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        _maybe_auto_close_scrum_sprint(conn, int(sprint_id))
        sprint = conn.execute(
            "SELECT * FROM scrum_sprint WHERE id = ? AND team_id = ?", (int(sprint_id), team_id)
        ).fetchone()
        if not sprint:
            conn.close()
            flash("Sprint not found.", "error")
            return redirect(url_for("scrum"))
        cards_by_col = _load_kanban_cards(conn, int(sprint_id), assignee, team_id)
        task_kinds_rows = _list_team_task_kinds(conn, team_id)
        tf = _sprint_team_page_template_flags(conn, team_id, int(sprint_id))
        conn.commit()
        conn.close()
        kb_leave_ctx = build_kanban_leave_worksheet_context(
            app,
            int(sprint_id),
            str(sprint["start_date"])[:10],
            str(sprint["end_date"])[:10],
            assignee,
        )
        return render_template(
            "scrum_kanban.html",
            sprint_id=int(sprint_id),
            sprint_name=sprint["name"],
            assignee=assignee,
            columns=SCRUM_KANBAN_COLUMNS,
            cards_by_col=cards_by_col,
            task_kinds_rows=task_kinds_rows,
            kb_back_url=url_for("scrum_sprint_team", sprint_id=sprint_id),
            kb_api_urls={
                "item_move": url_for("scrum_api_item_move"),
                "item_note": url_for("scrum_api_item_note"),
                "item_activity_update": url_for("scrum_api_activity_update"),
                "item_add": url_for("scrum_api_item_add"),
                "item_appreciation": url_for("scrum_api_item_appreciation"),
                "item_appreciation_delete_all": url_for("scrum_api_item_appreciation_delete_all"),
                "area_suggest": url_for("scrum_api_areas"),
            },
            kb_form_urls={
                "item_update": url_for("scrum_sprint_item_update"),
                "item_delete": url_for("scrum_sprint_item_delete"),
                "attachment_upload": url_for("scrum_sprint_item_attachment_upload"),
                "attachment_delete": url_for("scrum_sprint_item_attachment_delete"),
            },
            kb_checklist_urls={
                "add": url_for("scrum_api_item_checklist_add"),
                "update": url_for("scrum_api_item_checklist_update"),
                "delete": url_for("scrum_api_item_checklist_delete"),
            },
            kb_item_detail_url=url_for("scrum_api_item_detail"),
            portal_kanban=False,
            sprint_readonly=tf["sprint_board_readonly"],
            sprint_freeze_mode=tf["sprint_freeze_mode"],
            **kb_leave_ctx,
        )

    @app.post("/scrum/api/item/move")
    def scrum_api_item_move():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        assignee_raw = (data.get("assignee") or "").strip()
        assignee = resolve_employee_name(assignee_raw, roster=g.manager_roster) or assignee_raw[:200]
        to_col = _normalize_kanban_column(data.get("to_column"))
        body = (data.get("note") or "").strip()[:2000]
        if not item_id or not sprint_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, assignee):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        row = conn.execute(
            "SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (item_id,)
        ).fetchone()
        from_col = _normalize_kanban_column(row["kanban_column"] if row else None)
        est_new, est_err = _estimate_hours_for_backlog_to_do_move(data, from_col=from_col, to_col=to_col)
        if est_err:
            conn.close()
            return jsonify({"ok": False, "error": est_err}), 400
        if from_col == "backlog" and to_col == "do":
            body = ""
            ch = 0.0
        elif from_col == "do" and to_col == "doing":
            conn.execute("DELETE FROM scrum_item_activity WHERE item_id = ?", (item_id,))
            ch = _parse_committed_hours_for_kanban_move(
                data.get("committed_hours"), from_col=from_col, to_col=to_col
            )
        elif from_col == "doing" and to_col == "do":
            conn.execute("DELETE FROM scrum_item_activity WHERE item_id = ?", (item_id,))
            ch = 0.0
        else:
            ch = _parse_committed_hours_for_kanban_move(
                data.get("committed_hours"), from_col=from_col, to_col=to_col
            )
        ts = _utc_stamp()
        st = _status_for_kanban_column(to_col)
        artifacts_json: str | None = None
        if to_col == "done":
            artifacts_json, aerr = _normalize_done_artifacts_from_api(data.get("artifacts"))
            if aerr:
                conn.close()
                return jsonify({"ok": False, "error": aerr}), 400
        if to_col == "done" and artifacts_json is not None:
            conn.execute(
                """
                UPDATE scrum_sprint_item
                SET kanban_column = ?, status = ?, updated_at = ?, done_artifacts = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (to_col, st, ts, artifacts_json, item_id, sprint_id),
            )
        elif est_new is not None:
            conn.execute(
                """
                UPDATE scrum_sprint_item
                SET kanban_column = ?, status = ?, updated_at = ?, estimate_hours = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (to_col, st, ts, est_new, item_id, sprint_id),
            )
        else:
            conn.execute(
                """
                UPDATE scrum_sprint_item
                SET kanban_column = ?, status = ?, updated_at = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (to_col, st, ts, item_id, sprint_id),
            )
        conn.execute(
            """
            INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (item_id, body, ch, from_col, to_col, ts),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/scrum/api/item/note")
    def scrum_api_item_note():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        assignee_raw = (data.get("assignee") or "").strip()
        assignee = resolve_employee_name(assignee_raw, roster=g.manager_roster) or assignee_raw[:200]
        body = (data.get("note") or "").strip()[:2000]
        ch = _parse_hours_field(str(data.get("committed_hours")))
        if not item_id or not sprint_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        if not body and ch <= SCRUM_HOUR_EPS:
            return jsonify({"ok": False, "error": "missing"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, assignee):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        row = conn.execute(
            "SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (item_id,)
        ).fetchone()
        if not row or _normalize_kanban_column(row["kanban_column"]) != "doing":
            conn.close()
            return jsonify({"ok": False, "error": "only_doing"}), 400
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ?",
            (ts, item_id),
        )
        conn.execute(
            """
            INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
            VALUES (?, ?, ?, 'doing', 'doing', ?)
            """,
            (item_id, body, ch, ts),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/scrum/api/item/activity_update")
    def scrum_api_activity_update():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        activity_id = _parse_optional_int(data.get("activity_id"))
        assignee_raw = (data.get("assignee") or "").strip()
        assignee = resolve_employee_name(assignee_raw, roster=g.manager_roster) or assignee_raw[:200]
        body = (data.get("note") or "").strip()[:2000]
        ch = _parse_hours_field(str(data.get("committed_hours")))
        if not item_id or not sprint_id or not activity_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, assignee):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_item_activity SET body = ?, committed_hours = ? WHERE id = ? AND item_id = ?",
            (body, ch, activity_id, item_id),
        )
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ?",
            (ts, item_id),
        )
        burn = _kanban_single_item_burn_payload(conn, int(item_id))
        ch_row = conn.execute(
            "SELECT committed_hours FROM scrum_item_activity WHERE id = ? AND item_id = ?",
            (activity_id, item_id),
        ).fetchone()
        ch_out = float(ch_row["committed_hours"] or 0) if ch_row else ch
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "burn": burn, "activity_committed_hours": ch_out})

    @app.post("/scrum/api/item/appreciation")
    def scrum_api_item_appreciation():
        """Record a manager 'Well Done' appreciation comment on a sticky."""
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        assignee_raw = (data.get("assignee") or "").strip()
        assignee = resolve_employee_name(assignee_raw, roster=g.manager_roster) or assignee_raw[:200]
        comment = (data.get("comment") or "").strip()[:SCRUM_APPRECIATION_BODY_MAX]
        if not item_id or not sprint_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        if not comment:
            return jsonify({"ok": False, "error": "missing_comment"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, assignee):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        author = str(getattr(g, "manager_team_name", "") or "").strip() or "Manager"
        ts = _utc_stamp()
        conn.execute(
            """
            INSERT INTO scrum_item_appreciation (item_id, author, comment, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (item_id, author[:200], comment, ts),
        )
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ? AND sprint_id = ?",
            (ts, item_id, sprint_id),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/scrum/api/item/appreciation/delete-all")
    def scrum_api_item_appreciation_delete_all():
        """Remove all manager appreciation rows for one sticky (manager kanban)."""
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        assignee_raw = (data.get("assignee") or "").strip()
        assignee = resolve_employee_name(assignee_raw, roster=g.manager_roster) or assignee_raw[:200]
        if not item_id or not sprint_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _item_belongs_to_sprint_team(conn, item_id, sprint_id, team_id, assignee):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        cur = conn.execute("DELETE FROM scrum_item_appreciation WHERE item_id = ?", (item_id,))
        deleted = int(cur.rowcount or 0)
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_sprint_item SET updated_at = ? WHERE id = ? AND sprint_id = ?",
            (ts, item_id, sprint_id),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "deleted": deleted})

    @app.get("/scrum/api/areas")
    def scrum_api_areas():
        if not _manager_logged_in():
            return jsonify({"matches": []}), 403
        q = (request.args.get("q") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        if len(q) < 1:
            return jsonify({"matches": []})
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        try:
            out = _scrum_distinct_area_suggestions(conn, team_id, q)
        finally:
            conn.close()
        return jsonify({"matches": out})

    @app.post("/scrum/api/task-kind/add")
    def scrum_api_task_kind_add():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        return jsonify({"ok": False, "error": "fixed_kinds_only"}), 400

    @app.post("/scrum/api/item/add")
    def scrum_api_item_add():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        sprint_id = _parse_optional_int(data.get("sprint_id"))
        assignee_raw = (data.get("assignee") or "").strip()
        assignee = resolve_employee_name(assignee_raw, roster=g.manager_roster) or assignee_raw[:200]
        title = (data.get("title") or "").strip()[:500]
        notes = (data.get("notes") or "").strip()[:2000]
        dod = (data.get("dod") or "").strip()[:4000]
        area = (data.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        est = _parse_hours_field(str(data.get("estimate_hours")))
        if not sprint_id or not title or not assignee.strip():
            return jsonify({"ok": False, "error": "missing"}), 400
        if est <= SCRUM_HOUR_EPS:
            return jsonify({"ok": False, "error": "estimate_required"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            return jsonify({"ok": False, "error": fr}), 403
        _ensure_team_task_kinds(conn, team_id)
        new_l = (data.get("new_kind_label") or "").strip()
        if new_l:
            conn.close()
            return jsonify({"ok": False, "error": "no_new_kinds"}), 400
        raw = str(data.get("task_kind_code") or "ndy").strip().lower()
        if raw in SCRUM_LEGACY_TASK_KIND_MAP:
            raw = SCRUM_LEGACY_TASK_KIND_MAP[raw]
        tcode = raw if raw in SCRUM_TASK_KIND_CODES else "ndy"
        ts = _utc_stamp()
        st = _status_for_kanban_column("do")
        mx_row = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) + 1 AS n FROM scrum_sprint_item WHERE sprint_id = ?",
            (sprint_id,),
        ).fetchone()
        mx = int(mx_row["n"] if mx_row is not None else 0)
        conn.execute(
            """
            INSERT INTO scrum_sprint_item
            (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind, sticky_color_hex, area)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'do', ?, ?, ?)
            """,
            (sprint_id, assignee, title, est, st, notes, dod, mx, ts, ts, tcode, None, area),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/scrum/sprint/export-xlsx")
    def scrum_sprint_export_xlsx():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        try:
            sprint_id = int((request.form.get("sprint_id") or "").strip())
        except ValueError:
            flash("Pick a valid sprint.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        blob, meta = build_sprint_export_xlsx_bytes(
            conn,
            team_id,
            sprint_id,
            app=app,
            manager_roster=g.manager_roster,
        )
        conn.close()
        if blob is None:
            flash(meta, "error")
            return redirect(url_for("scrum"))
        return send_file(
            io.BytesIO(blob),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=meta,
        )

    @app.post("/team/hub-mode")
    def team_hub_mode_set():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("dashboard"))
        mode = _normalize_hub_mode(request.form.get("hub_mode"))
        tid = int(g.manager_team_id)
        conn = get_db(app)
        conn.execute("UPDATE teams SET hub_mode = ? WHERE id = ?", (mode, tid))
        conn.commit()
        conn.close()
        label = "Scrum dashboard" if mode == "scrum" else "Leave tracker"
        flash(f"Saved: this team’s hub is set to {label}.", "success")
        dest = _safe_next_path(request.form.get("next"))
        if dest:
            return redirect(dest)
        return redirect(url_for("scrum" if mode == "scrum" else "dashboard"))

    @app.post("/scrum/sprint/create")
    def scrum_sprint_create():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        name = (request.form.get("name") or "").strip() or "Sprint"
        goal = ""
        create_next = (request.form.get("create_next") or "").strip().lower() in ("1", "true", "yes", "on")
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if create_next:
            sd = _next_sprint_start_after_latest_end(conn, team_id)
            if sd is None:
                conn.close()
                flash(
                    "There is no previous sprint yet. Create your first sprint with a start date above.",
                    "error",
                )
                return redirect(url_for("scrum"))
        else:
            try:
                sd = date.fromisoformat((request.form.get("start_date") or "").strip()[:10])
            except ValueError:
                conn.close()
                flash("Invalid sprint dates (use YYYY-MM-DD).", "error")
                return redirect(url_for("scrum"))
        ed = scrum_sprint_default_end_date(sd)
        ts = _utc_stamp()
        cap_h = compute_team_sprint_capacity_leave_hours(app, g.manager_roster, sd, ed)
        if _sprint_name_exists_for_team(conn, team_id, name):
            conn.close()
            flash(
                f'A sprint named "{name}" already exists for this team. Choose a different name.',
                "error",
            )
            return redirect(url_for("scrum"))
        cur = conn.execute(
            """
            INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, team_capacity_hours, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (team_id, name, sd.isoformat(), ed.isoformat(), goal, cap_h, ts, ts),
        )
        new_id = int(cur.lastrowid)
        n_carried = _carry_forward_do_doing_to_new_sprint(conn, team_id, new_id, sd, ts)
        conn.commit()
        conn.close()
        session["active_sprint_id"] = new_id
        if n_carried:
            flash(
                f"Sprint created — carried {n_carried} sticky(ies) from To DO / In progress on the prior sprint into this sprint’s backlog "
                "(same assignees, details, and original estimate hours; Burnt starts at 0 on each new sticky).",
                "success",
            )
        else:
            flash("Sprint created — open it to see the team and sticky-note boards.", "success")
        return redirect(url_for("scrum_sprint_team", sprint_id=new_id))

    @app.post("/scrum/sprint/delete")
    def scrum_sprint_delete():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        sid = _parse_optional_int(request.form.get("sprint_id"))
        if not sid:
            flash("Missing sprint.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _sprint_row_for_team(conn, sid, team_id):
            conn.close()
            flash("Unknown sprint.", "error")
            return redirect(url_for("scrum"))
        # Deleting is allowed even when the sprint is closed or past end (hub + team cleanup).
        conn.execute("DELETE FROM scrum_sprint WHERE id = ? AND team_id = ?", (sid, team_id))
        conn.commit()
        conn.close()
        if session.get("active_sprint_id") == sid:
            session.pop("active_sprint_id", None)
        flash("Sprint deleted.", "success")
        return redirect(url_for("scrum"))

    @app.post("/scrum/sprint/update")
    def scrum_sprint_update():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        sid = _parse_optional_int(request.form.get("sprint_id"))
        if not sid:
            flash("Missing sprint.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _sprint_row_for_team(conn, sid, team_id):
            conn.close()
            flash("Unknown sprint.", "error")
            return redirect(url_for("scrum"))
        fr = _sprint_board_frozen_reason(conn, sid)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sid))
        try:
            sd = date.fromisoformat((request.form.get("start_date") or "").strip()[:10])
        except ValueError:
            conn.close()
            flash("Invalid sprint dates (use YYYY-MM-DD).", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sid))
        ed = scrum_sprint_default_end_date(sd)
        ts = _utc_stamp()
        cap_h = compute_team_sprint_capacity_leave_hours(app, g.manager_roster, sd, ed)
        conn.execute(
            """
            UPDATE scrum_sprint
            SET start_date = ?, end_date = ?, team_capacity_hours = ?, updated_at = ?
            WHERE id = ? AND team_id = ?
            """,
            (sd.isoformat(), ed.isoformat(), cap_h, ts, sid, team_id),
        )
        conn.commit()
        conn.close()
        flash("Sprint start updated — end date is always 14 calendar days from start (inclusive).", "success")
        return redirect(url_for("scrum_sprint_team", sprint_id=sid))

    @app.post("/scrum/sprint/rename")
    def scrum_sprint_rename():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        sid = _parse_optional_int(request.form.get("sprint_id"))
        if not sid:
            flash("Missing sprint.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        name = (request.form.get("name") or "").strip()[:200]
        if not name:
            flash("Sprint name is required.", "error")
            return redirect(url_for("scrum"))
        conn = get_db(app)
        row = _sprint_row_for_team(conn, sid, team_id)
        if not row:
            conn.close()
            flash("Unknown sprint.", "error")
            return redirect(url_for("scrum"))
        frn = _sprint_board_frozen_reason(conn, sid)
        if frn:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(frn) or "This sprint is read-only.", "error")
            return redirect(url_for("scrum"))
        old = (str(row["name"]) or "").strip()
        if name.lower() == old.lower():
            conn.close()
            return redirect(url_for("scrum"))
        if _sprint_name_taken_by_other(conn, team_id, name, sid):
            conn.close()
            flash(
                f'A sprint named "{name}" already exists for this team. Choose a different name.',
                "error",
            )
            return redirect(url_for("scrum"))
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_sprint SET name = ?, updated_at = ? WHERE id = ? AND team_id = ?",
            (name, ts, sid, team_id),
        )
        conn.commit()
        conn.close()
        flash("Sprint name updated.", "success")
        return redirect(url_for("scrum"))

    @app.post("/scrum/sprint/close")
    def scrum_sprint_close():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        sid = _parse_optional_int(request.form.get("sprint_id"))
        if not sid:
            flash("Missing sprint.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _sprint_row_for_team(conn, sid, team_id):
            conn.close()
            flash("Unknown sprint.", "error")
            return redirect(url_for("scrum"))
        if _sprint_is_closed_by_id(conn, sid):
            conn.close()
            flash("This sprint is already closed.", "info")
            return redirect(url_for("scrum_sprint_team", sprint_id=sid))
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_sprint SET is_closed = 1, updated_at = ? WHERE id = ? AND team_id = ?",
            (ts, sid, team_id),
        )
        conn.commit()
        conn.close()
        flash("Sprint closed — the board and sprint details are read-only until you open it again.", "success")
        return redirect(url_for("scrum_sprint_team", sprint_id=sid))

    @app.post("/scrum/sprint/open")
    def scrum_sprint_open():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        sid = _parse_optional_int(request.form.get("sprint_id"))
        if not sid:
            flash("Missing sprint.", "error")
            return redirect(url_for("scrum"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        row = _sprint_row_for_team(conn, sid, team_id)
        if not row:
            conn.close()
            flash("Unknown sprint.", "error")
            return redirect(url_for("scrum"))
        if not _sprint_is_closed_by_id(conn, sid):
            conn.close()
            flash("This sprint is already open.", "info")
            return redirect(url_for("scrum_sprint_team", sprint_id=sid))
        ts = _utc_stamp()
        conn.execute(
            "UPDATE scrum_sprint SET is_closed = 0, updated_at = ? WHERE id = ? AND team_id = ?",
            (ts, sid, team_id),
        )
        conn.commit()
        conn.close()
        if _sprint_row_past_end(row):
            flash(
                "Sprint opened for records, but this sprint’s end date has already passed — the board stays read-only "
                "and burndown is frozen at the last sprint day.",
                "info",
            )
        else:
            flash("Sprint opened — editing is enabled again.", "success")
        return redirect(url_for("scrum_sprint_team", sprint_id=sid))

    @app.post("/scrum/sprint/item/add")
    def scrum_sprint_item_add():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not sprint_id or not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("scrum"))
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        title = (request.form.get("title") or "").strip()
        if not title:
            conn.close()
            flash("Task title is required.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        assign_raw = (request.form.get("assignee") or "").strip()
        assignee = resolve_employee_name(assign_raw, roster=g.manager_roster) or assign_raw[:200]
        est = _parse_hours_field(request.form.get("estimate_hours"))
        if est <= SCRUM_HOUR_EPS:
            conn.close()
            flash("Estimated hours are required (must be greater than zero).", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        notes = (request.form.get("notes") or "").strip()[:2000]
        dod = (request.form.get("dod") or "").strip()[:4000]
        area = (request.form.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        kcol = _normalize_kanban_column(request.form.get("kanban_column") or "do")
        raw_tk = request.form.get("task_kind")
        if raw_tk is None:
            raw_tk = "ndy"
        tkind = _coerce_sprint_item_task_kind(conn, team_id, raw_tk)
        sticky_hex: str | None = None
        st = _status_for_kanban_column(kcol)
        ts = _utc_stamp()
        mx_row = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) + 1 AS n FROM scrum_sprint_item WHERE sprint_id = ?",
            (sprint_id,),
        ).fetchone()
        mx = int(mx_row["n"] if mx_row is not None else 0)
        conn.execute(
            """
            INSERT INTO scrum_sprint_item
            (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind, sticky_color_hex, area)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (sprint_id, assignee, title, est, st, notes, dod, mx, ts, ts, kcol, tkind, sticky_hex, area),
        )
        conn.commit()
        conn.close()
        flash("Sticky added.", "success")
        return redirect(
            url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee)
        )

    @app.post("/scrum/sprint/item/update")
    def scrum_sprint_item_update():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        item_id = _parse_optional_int(request.form.get("item_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        team_id = int(g.manager_team_id)
        if not item_id or not sprint_id:
            flash("Missing backlog item.", "error")
            return redirect(url_for("scrum"))
        conn = get_db(app)
        if not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("scrum"))
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        row = conn.execute(
            "SELECT assignee, kanban_column, sticky_color_hex FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not row:
            conn.close()
            flash("Item not found.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        assignee = (row["assignee"] or "").strip()
        cur_col = _normalize_kanban_column(row["kanban_column"] if "kanban_column" in row.keys() else None)
        ts = _utc_stamp()

        if cur_col == "done":
            notes = (request.form.get("notes") or "").strip()[:2000]
            da_lines = request.form.get("done_artifacts_lines") or ""
            da_json, da_err = _normalize_done_artifacts_from_lines(da_lines)
            if da_err:
                conn.close()
                if da_err == "artifacts_limit":
                    msg = "Too many artifact links (maximum 20)."
                else:
                    msg = "Artifact links must be valid http(s) URLs (one per line, optional label before |)."
                flash(msg, "error")
                return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))
            conn.execute(
                """
                UPDATE scrum_sprint_item SET notes = ?, done_artifacts = ?, updated_at = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (notes, da_json, ts, item_id, sprint_id),
            )
            conn.commit()
            conn.close()
            flash("Done sticky updated (details and artifacts).", "success")
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))

        if cur_col == "doing":
            est = _parse_hours_field(request.form.get("estimate_hours"))
            if est <= SCRUM_HOUR_EPS:
                conn.close()
                flash("Estimated hours must be greater than zero while this sticky is In progress.", "error")
                return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))
            st = _status_for_kanban_column("doing")
            conn.execute(
                """
                UPDATE scrum_sprint_item SET estimate_hours = ?, status = ?, updated_at = ?
                WHERE id = ? AND sprint_id = ?
                """,
                (est, st, ts, item_id, sprint_id),
            )
            conn.commit()
            conn.close()
            flash("Plan estimate updated for this In progress sticky.", "success")
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))

        if cur_col != "do":
            conn.close()
            flash("Planning edits are only in To DO; details and artifacts can be edited in Done.", "error")
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))
        title = (request.form.get("title") or "").strip()
        if not title:
            conn.close()
            flash("Title is required.", "error")
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))
        est = _parse_hours_field(request.form.get("estimate_hours"))
        if est <= SCRUM_HOUR_EPS:
            conn.close()
            flash("Estimated hours must be greater than zero while planning in To DO.", "error")
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))
        tkind = _coerce_sprint_item_task_kind(conn, team_id, request.form.get("task_kind"))
        notes = (request.form.get("notes") or "").strip()[:2000]
        dod = (request.form.get("dod") or "").strip()[:4000]
        area = (request.form.get("area") or "").strip()[:SCRUM_STICKY_AREA_MAX_LEN]
        sticky_hex = None
        st = _status_for_kanban_column(cur_col)
        conn.execute(
            """
            UPDATE scrum_sprint_item SET
                title = ?, estimate_hours = ?, task_kind = ?, notes = ?, dod = ?, status = ?, updated_at = ?, sticky_color_hex = ?, area = ?
            WHERE id = ? AND sprint_id = ?
            """,
            (title, est, tkind, notes, dod, st, ts, sticky_hex, area, item_id, sprint_id),
        )
        conn.commit()
        conn.close()
        flash("Plan saved — you can edit again while this sticky stays in To DO.", "success")
        return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))

    @app.post("/scrum/sprint/item/delete")
    def scrum_sprint_item_delete():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        item_id = _parse_optional_int(request.form.get("item_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        team_id = int(g.manager_team_id)
        if not item_id or not sprint_id:
            flash("Missing item.", "error")
            return redirect(url_for("scrum"))
        conn = get_db(app)
        if not _sprint_row_for_team(conn, sprint_id, team_id):
            conn.close()
            flash("Invalid sprint.", "error")
            return redirect(url_for("scrum"))
        fr = _sprint_board_frozen_reason(conn, sprint_id)
        if fr:
            conn.close()
            flash(_sprint_board_frozen_flash_for_reason(fr) or "This sprint is read-only.", "error")
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        ar = conn.execute(
            "SELECT assignee, kanban_column FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?",
            (item_id, sprint_id),
        ).fetchone()
        if not ar:
            conn.close()
            flash("Item not found.", "error")
            return redirect(url_for("scrum"))
        if _normalize_kanban_column(ar["kanban_column"] if "kanban_column" in ar.keys() else None) != "do":
            assignee_err = (ar["assignee"] or "").strip()
            conn.close()
            flash("Stickies can only be removed while they are in To DO.", "error")
            if assignee_err:
                return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee_err))
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        assignee = (ar["assignee"] or "").strip()
        conn.execute("DELETE FROM scrum_sprint_item WHERE id = ? AND sprint_id = ?", (item_id, sprint_id))
        conn.commit()
        conn.close()
        flash("Sticky removed.", "success")
        if assignee:
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee))
        return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))

    @app.post("/scrum/sprint/item/attachment")
    def scrum_sprint_item_attachment_upload():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        item_id = _parse_optional_int(request.form.get("item_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        assignee_redir = (request.form.get("assignee") or "").strip()
        team_id = int(g.manager_team_id)
        if not item_id or not sprint_id:
            flash("Missing item.", "error")
            return redirect(url_for("scrum"))
        conn = get_db(app)
        cat, msg, assignee = _scrum_attachment_upload_run_batch(
            app, conn, team_id=team_id, sprint_id=sprint_id, item_id=item_id, roster_gate=None
        )
        conn.close()
        flash(msg, cat)
        if (request.form.get("return_to") or "").strip() == "team_detailed":
            return redirect(url_for("scrum_sprint_team_detailed", sprint_id=sprint_id))
        assignee_out = assignee_redir or assignee
        if assignee_out:
            return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee_out))
        return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))

    @app.post("/scrum/sprint/item/attachment/delete")
    def scrum_sprint_item_attachment_delete():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("scrum"))
        aid = _parse_optional_int(request.form.get("attachment_id"))
        sprint_id = _parse_optional_int(request.form.get("sprint_id"))
        assignee_redir = (request.form.get("assignee") or "").strip()
        team_id = int(g.manager_team_id)
        if not aid or not sprint_id:
            flash("Missing attachment.", "error")
            return redirect(url_for("scrum"))
        conn = get_db(app)
        ok, msg, assignee = _scrum_attachment_delete_core(
            app, conn, team_id=team_id, sprint_id=sprint_id, attachment_id=int(aid), roster_gate=None
        )
        conn.close()
        flash(msg, "success" if ok else "error")
        if (request.form.get("return_to") or "").strip() == "team_detailed":
            return redirect(url_for("scrum_sprint_team_detailed", sprint_id=sprint_id))
        if not ok and not assignee:
            return redirect(url_for("scrum_sprint_team", sprint_id=sprint_id))
        return redirect(url_for("scrum_member_board", sprint_id=sprint_id, assignee=assignee_redir or assignee))

    @app.get("/scrum/sprint/item/attachment/<int:attachment_id>/file")
    def scrum_sprint_item_attachment_download(attachment_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        ar = _scrum_attachment_auth_row(conn, int(attachment_id))
        conn.close()
        if not ar or int(ar["team_id"]) != team_id:
            return Response("Not found.", status=404, mimetype="text/plain")
        rel = (str(ar["rel_path"]) or "").strip().replace("\\", "/")
        if ".." in rel or rel.startswith("/"):
            return Response("Bad path.", status=400, mimetype="text/plain")
        path = _scrum_attachment_root_dir(app) / rel
        if not path.is_file():
            return Response("File missing on server.", status=404, mimetype="text/plain")
        ct = str(ar["content_type"] or "application/octet-stream").strip()[:200] or "application/octet-stream"
        inline_req = (request.args.get("inline") or "").strip().lower() in ("1", "true", "yes")
        inline_ok = inline_req and (
            ct.startswith("image/")
            or ct == "application/pdf"
            or ct.startswith("text/plain")
            or ct in ("text/csv", "text/markdown", "application/json")
        )
        return send_file(
            path,
            as_attachment=not inline_ok,
            download_name=str(ar["original_filename"] or "download").strip()[:200] or "download",
            mimetype=ct,
        )

    @app.get("/scrum/sprint/item/attachment/<int:attachment_id>/preview-html")
    def scrum_sprint_item_attachment_preview_html(attachment_id: int):
        if not _manager_logged_in():
            return Response("Unauthorized.", status=401, mimetype="text/plain")
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        ar = _scrum_attachment_auth_row(conn, int(attachment_id))
        conn.close()
        if not ar or int(ar["team_id"]) != team_id:
            return Response("Not found.", status=404, mimetype="text/plain")
        rel = (str(ar["rel_path"]) or "").strip().replace("\\", "/")
        if ".." in rel or rel.startswith("/"):
            return Response("Bad path.", status=400, mimetype="text/plain")
        path = _scrum_attachment_root_dir(app) / rel
        if not path.is_file():
            return Response("File missing on server.", status=404, mimetype="text/plain")
        low = (str(ar["original_filename"]) or "").lower()
        if not (low.endswith((".xlsx", ".xlsm", ".xltx", ".csv"))):
            return Response("Preview not available for this file type.", status=415, mimetype="text/plain")
        doc = _scrum_attachment_build_preview_html(path, str(ar["original_filename"] or ""))
        if not doc:
            return Response("Could not build preview (install openpyxl for Excel files).", status=500, mimetype="text/plain")
        return Response(doc, mimetype="text/html; charset=utf-8")

    @app.route("/")
    def home():
        if _portal_session() and not _manager_logged_in():
            return _portal_employee_landing_response(app)
        return redirect(url_for("login_page"))

    @app.route("/login", methods=["GET", "POST"])
    def login_page():
        if request.method == "GET":
            if _manager_logged_in():
                return redirect(url_for("dashboard"))
            if _portal_session() and not _manager_logged_in():
                return _portal_employee_landing_response(app)
            return render_template(
                "login.html",
                hide_nav=True,
            )

        gate = (request.form.get("gate_kind") or "").strip().lower()

        if gate == "employee":
            flash("Email sign-in is disabled. Use your manager access code.", "error")
            return redirect(url_for("login_page"))

        if gate == "manager":
            email = (request.form.get("email") or "").strip().lower()
            password = (request.form.get("password") or "")
            if not email or not password:
                flash("Email and password are required.", "error")
                return redirect(url_for("login_page"))
            conn = get_db(app)
            row = conn.execute(
                "SELECT * FROM managers WHERE email = ? COLLATE NOCASE", (email,)
            ).fetchone()
            if not row:
                conn.close()
                pw_configured = _manager_password_configured(app)
                if pw_configured and _check_manager_password(app, password):
                    session["manager"] = True
                    session["manager_role"] = "manager"
                    session.pop("my_employee_name", None)
                    flash("Signed in.", "success")
                    return redirect(url_for("dashboard"))
                flash("Invalid email or password.", "error")
                return redirect(url_for("login_page"))
            if not _verify_manager_password(str(row["password_hash"]), password):
                conn.close()
                flash("Invalid email or password.", "error")
                return redirect(url_for("login_page"))
            team_name = str(row["team_name"] or "Default").strip()
            conn.close()
            session["manager"] = True
            session["manager_role"] = "manager"
            session["manager_user_email"] = email
            session.pop("my_employee_name", None)
            conn2 = get_db(app)
            trow = conn2.execute("SELECT id FROM teams WHERE name = ? COLLATE NOCASE", (team_name,)).fetchone()
            if trow:
                session["active_team_id"] = int(trow["id"])
            conn2.close()
            flash("Signed in.", "success")
            return redirect(url_for("dashboard"))

        flash("Invalid request.", "error")
        return redirect(url_for("login_page"))

    @app.route("/register", methods=["GET", "POST"])
    def register_page():
        if request.method == "GET":
            if _manager_logged_in():
                return redirect(url_for("dashboard"))
            return render_template("register.html", hide_nav=True)

        email = (request.form.get("email") or "").strip().lower()
        password = (request.form.get("password") or "")
        team_name = (request.form.get("team_name") or "").strip()
        if not team_name and email:
            team_name = f"{email.split('@')[0].capitalize()}'s Team"
        if not email or not password:
            flash("Email and password are required.", "error")
            return redirect(url_for("register_page"))
        if len(password) < 4:
            flash("Password must be at least 4 characters.", "error")
            return redirect(url_for("register_page"))
        if len(team_name) > 200:
            team_name = team_name[:200]
        conn = get_db(app)
        existing = conn.execute("SELECT 1 FROM managers WHERE email = ? COLLATE NOCASE", (email,)).fetchone()
        if existing:
            conn.close()
            flash("An account with that email already exists.", "error")
            return redirect(url_for("register_page"))
        pw_hash = _hash_manager_password(password)
        ts = _utc_stamp()
        conn.execute(
            "INSERT INTO managers (email, password_hash, team_name, created_at) VALUES (?, ?, ?, ?)",
            (email, pw_hash, team_name, ts),
        )
        conn.execute("INSERT OR IGNORE INTO teams (name, created_at, owner_email) VALUES (?, ?, ?)", (team_name, ts, email))
        conn.commit()
        conn.close()
        session["manager"] = True
        session["manager_role"] = "manager"
        session["manager_user_email"] = email
        session.pop("my_employee_name", None)
        conn2 = get_db(app)
        trow = conn2.execute("SELECT id FROM teams WHERE name = ? COLLATE NOCASE", (team_name,)).fetchone()
        if trow:
            session["active_team_id"] = int(trow["id"])
        conn2.close()
        flash("Account created. Set up your team roster below.", "success")
        return redirect(url_for("dashboard"))

    @app.route("/my-requests", methods=["GET", "POST"])
    def my_requests():
        pu = _portal_session()
        if not pu:
            flash("Sign in to view your leave requests.", "error")
            return redirect(url_for("home"))
        if request.method == "POST":
            flash("You can only view your own requests.", "error")
            return redirect(url_for("my_requests"))

        name = _portal_effective_roster_name(app, pu)
        today = date.today()
        try:
            year = int(request.args.get("year") or today.year)
            month = int(request.args.get("month") or today.month)
        except (TypeError, ValueError):
            year, month = today.year, today.month
        year = max(2000, min(2100, year))
        month = max(1, min(12, month))

        canonical_name = name
        if name:
            conn = get_db(app)
            cn_row = conn.execute(
                """
                SELECT employee_name FROM leave_requests
                WHERE employee_name = ? COLLATE NOCASE
                ORDER BY id DESC LIMIT 1
                """,
                (name,),
            ).fetchone()
            if not cn_row:
                cn_row = conn.execute(
                    "SELECT employee_name FROM team_roster WHERE employee_name = ? COLLATE NOCASE LIMIT 1",
                    (name,),
                ).fetchone()
            conn.close()
            if cn_row:
                canonical_name = str(cn_row["employee_name"]).strip()

        month_tabs = [
            {"num": m, "name": calendar.month_abbr[m], "full": calendar.month_name[m]}
            for m in range(1, 13)
        ]

        last_day_num = monthrange(year, month)[1]
        month_start_iso = date(year, month, 1).isoformat()
        month_end_iso = date(year, month, last_day_num).isoformat()
        base_reasons = dict(LEAVE_REASONS)

        raw_leave_rows: list[sqlite3.Row] = []
        if canonical_name:
            conn = get_db(app)
            raw_leave_rows = conn.execute(
                """
                SELECT id, reason, description, start_date, end_date, duration_type,
                       status, created_at, portal_leave_label
                FROM leave_requests
                WHERE employee_name = ? COLLATE NOCASE
                  AND status IN ('pending', 'approved')
                  AND start_date <= ? AND end_date >= ?
                ORDER BY start_date ASC, id ASC
                """,
                (canonical_name, month_end_iso, month_start_iso),
            ).fetchall()
            conn.close()

        day_leave_map: dict[str, sqlite3.Row] = {}
        for r in raw_leave_rows:
            try:
                sd = date.fromisoformat(str(r["start_date"])[:10])
                ed = date.fromisoformat(str(r["end_date"])[:10])
            except ValueError:
                continue
            for d in _daterange_inclusive(sd, ed):
                d_iso = d.isoformat()
                if d_iso < month_start_iso or d_iso > month_end_iso:
                    continue
                existing = day_leave_map.get(d_iso)
                if existing is None or _leave_row_display_tiebreak(r) > _leave_row_display_tiebreak(existing):
                    day_leave_map[d_iso] = r

        days_meta: list[dict] = []
        for d in _daterange_inclusive(date(year, month, 1), date(year, month, last_day_num)):
            wd = d.weekday()
            days_meta.append({
                "date": d,
                "iso": d.isoformat(),
                "weekday": calendar.day_abbr[wd],
                "day": d.day,
                "is_weekend": wd >= 5,
            })

        cells: list[dict | None] = []
        leave_days_total = 0.0
        nokia_a_day_units = 0.0
        for dm in days_meta:
            row = day_leave_map.get(dm["iso"])
            cell: dict | None = None
            if row is not None:
                dtype = (row["duration_type"] or "full").strip()
                desc = str(row["description"] or "")
                day_unit = 0.5 if dtype in ("half_am", "half_pm") else 1.0
                if _leave_reason_counts_toward_day_units(row["reason"]):
                    leave_days_total += day_unit
                if NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in desc:
                    nokia_a_day_units += day_unit
                eff = str(row["status"] or "approved")
                code = leave_cell_code(row["reason"], dtype, eff, desc)
                css = cell_css_class(row["reason"], eff, desc)
                cell = {
                    "code": code,
                    "css": css,
                    "status": eff,
                    "leave_id": int(row["id"]),
                    "work_date": dm["iso"],
                    "title": f"{base_reasons.get(row['reason'], row['reason'])} · {eff}",
                }
            cells.append(cell)

        grid_rows: list[dict] = []
        if canonical_name:
            grid_rows = [{
                "employee": canonical_name,
                "leave_days_total": round(leave_days_total, 2),
                "eleave_days": round(nokia_a_day_units, 2),
                "gap_days": round(nokia_a_day_units - leave_days_total, 2),
                "cells": cells,
            }]

        leave_records: list[dict] = []
        for r in raw_leave_rows:
            try:
                sd = date.fromisoformat(str(r["start_date"])[:10])
                ed = date.fromisoformat(str(r["end_date"])[:10])
            except ValueError:
                sd = ed = date.today()
            dtype = (r["duration_type"] or "full").strip()
            if dtype in ("half_am", "half_pm"):
                days_val: float = 0.5
            elif dtype == "full":
                days_val = 1.0
            else:
                days_val = float(_working_weekdays_count(sd, ed))
            plab = (r["portal_leave_label"] if "portal_leave_label" in r.keys() else "") or ""
            leave_type = plab.strip() if plab.strip() else base_reasons.get(r["reason"], r["reason"].upper())
            half_str = " (½ AM)" if dtype == "half_am" else " (½ PM)" if dtype == "half_pm" else ""
            desc_raw = (r["description"] or "").strip()
            leave_records.append({
                "leave_id": int(r["id"]),
                "can_withdraw": NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER not in desc_raw,
                "leave_type": leave_type + half_str,
                "from_date": str(r["start_date"])[:10],
                "to_date": str(r["end_date"])[:10],
                "days": days_val,
                "status": (r["status"] or "").capitalize(),
                "description": desc_raw,
                "nokia_approved": NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in desc_raw,
            })

        return render_template(
            "my_requests.html",
            hide_nav=True,
            viewer=canonical_name,
            portal_employee_view=True,
            year=year,
            month=month,
            month_name=calendar.month_name[month],
            month_tabs=month_tabs,
            days=days_meta,
            grid_rows=grid_rows,
            leave_records=leave_records,
        )

    @app.route("/my-requests/withdraw", methods=["POST"])
    def portal_my_requests_withdraw():
        pu = _portal_session()
        if not pu:
            flash("Sign in to manage your leave requests.", "error")
            return redirect(url_for("home"))
        if not _csrf_form_ok():
            flash("Session expired. Try again.", "error")
            return redirect(url_for("my_requests"))
        roster_name = _portal_effective_roster_name(app, pu)
        try:
            leave_id = int(request.form.get("leave_id") or 0)
        except (TypeError, ValueError):
            leave_id = 0
        try:
            year = int(request.form.get("year") or date.today().year)
            month = int(request.form.get("month") or date.today().month)
        except (TypeError, ValueError):
            year, month = date.today().year, date.today().month
        year = max(2000, min(2100, year))
        month = max(1, min(12, month))
        err = _portal_withdraw_leave_request(app, roster_name, leave_id)
        if err == "nokia_locked":
            flash("Nokia e-tool approved leave cannot be withdrawn here.", "error")
        elif err:
            flash("Could not withdraw that leave request.", "error")
        else:
            flash("Leave withdrawn and removed from the tracker.", "success")
        return redirect(url_for("my_requests", year=year, month=month))

    @app.route("/leave", methods=["GET", "POST"])
    def leave_apply():
        if _portal_session():
            return redirect(url_for("portal_leave_apply"))
        if not _manager_logged_in():
            flash("Sign in to apply for leave.", "error")
            return redirect(url_for("home"))
        today_iso = date.today().isoformat()
        if request.method == "POST":
            raw_name = (request.form.get("employee_name") or "").strip()
            canonical = resolve_employee_name(raw_name, roster=get_union_roster_for_app(app))
            reason = (request.form.get("reason") or "").strip()
            start_date = (request.form.get("start_date") or "").strip()
            end_date = (request.form.get("end_date") or "").strip()
            day_part = (request.form.get("day_part") or "full").strip()

            errors: list[str] = []
            if not canonical:
                errors.append("Name must match a team roster entry. Use the suggestions while typing.")
            if reason not in {c[0] for c in LEAVE_REASONS}:
                errors.append("Select a valid leave type.")
            if not start_date:
                errors.append("First day is required.")
            if not end_date:
                errors.append("Last day is required.")

            duration_type = "full"
            if not errors:
                try:
                    sd = date.fromisoformat(start_date)
                    ed = date.fromisoformat(end_date)
                except ValueError:
                    errors.append("Invalid date format.")
                else:
                    if ed < sd:
                        errors.append("Last day must be on or after the first day.")
                    elif ed > sd:
                        duration_type = "multi"
                    else:
                        if day_part not in ("full", "half_am", "half_pm"):
                            errors.append("Select how much of that day is off.")
                        else:
                            duration_type = day_part

            if errors:
                for e in errors:
                    flash(e, "error")
                return render_template(
                    "leave.html",
                    reasons=LEAVE_REASONS,
                    day_parts=DAY_PART_CHOICES,
                    form=request.form,
                    default_start=today_iso,
                    default_end=today_iso,
                )

            created = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
            ip = client_ip()
            conn = get_db(app)
            conn.execute(
                """
                INSERT INTO leave_requests
                (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
                VALUES (?, ?, '', ?, ?, ?, 'pending', ?, ?)
                """,
                (canonical, reason, start_date, end_date, duration_type, created, ip),
            )
            conn.commit()
            conn.close()
            session["my_employee_name"] = canonical
            flash("Request submitted.", "success")
            return redirect(url_for("my_requests"))

        return render_template(
            "leave.html",
            reasons=LEAVE_REASONS,
            day_parts=DAY_PART_CHOICES,
            form={},
            default_start=today_iso,
            default_end=today_iso,
        )

    @app.route("/attendance", methods=["GET", "POST"])
    def attendance():
        return redirect(url_for("home")), 302

    @app.route("/reports/team-select", methods=["POST"])
    def reports_team_select():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("reports"))
        try:
            tid = int((request.form.get("team_id") or "").strip())
        except ValueError:
            tid = 0
        conn = get_db(app)
        manager_email = session.get("manager_user_email")
        if manager_email:
            row = conn.execute("SELECT id, name FROM teams WHERE id = ? AND (owner_email = ? OR owner_email = '')", (tid, manager_email)).fetchone()
        else:
            row = conn.execute("SELECT id, name FROM teams WHERE id = ?", (tid,)).fetchone()
        conn.close()
        if not row:
            flash("Unknown team.", "error")
            return redirect(_safe_next_path(request.form.get("next")) or url_for("reports"))
        session["active_team_id"] = int(row["id"])
        flash(f"Active team: {row['name']}.", "success")
        dest = _safe_next_path(request.form.get("next"))
        return redirect(dest or url_for("reports"))

    @app.route("/delete_team", methods=["POST"])
    def delete_team():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("dashboard"))
            
        try:
            tid = int((request.form.get("team_id") or "").strip())
        except ValueError:
            tid = 0
            
        conn = get_db(app)
        manager_email = session.get("manager_user_email")
        if manager_email:
            row = conn.execute("SELECT id, name FROM teams WHERE id = ? AND (owner_email = ? OR owner_email = '')", (tid, manager_email)).fetchone()
        else:
            row = conn.execute("SELECT id, name FROM teams WHERE id = ?", (tid,)).fetchone()
            
        if not row:
            conn.close()
            flash("Unknown team or access denied.", "error")
            return redirect(_safe_next_path(request.form.get("next")) or url_for("home"))
            
        team_name = row["name"]
        conn.execute("DELETE FROM teams WHERE id = ?", (tid,))
        
        # Basic cleanup of associated data
        conn.execute("DELETE FROM team_roster WHERE team_id = ?", (tid,))
        conn.execute("DELETE FROM scrum_sprint WHERE team_id = ?", (tid,))
        conn.commit()
        conn.close()
        
        if session.get("active_team_id") == tid:
            session.pop("active_team_id", None)
            
        flash(f"Team '{team_name}' has been permanently deleted.", "success")
        return redirect(_safe_next_path(request.form.get("next")) or url_for("home"))


    @app.route("/reports/roster-upload", methods=["POST"])
    def reports_roster_upload():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        if not _csrf_form_ok():
            flash("Invalid security token. Try again.", "error")
            return redirect(url_for("reports"))
        f = request.files.get("roster_csv")
        if not f or not f.filename:
            flash("Choose a roster file (.xlsx, CSV, or JSON).", "error")
            return redirect(url_for("reports"))
        raw = f.read()
        if len(raw) > 2_000_000:
            flash("File too large (max 2 MB).", "error")
            return redirect(url_for("reports"))
        fn = (f.filename or "").lower()
        try:
            if fn.endswith(".json"):
                n_teams, n_rows, warns = ingest_team_roster_json(app, raw)
            elif fn.endswith(".xlsx") or (len(raw) >= 4 and raw[:4] == b"PK\x03\x04"):
                n_teams, n_rows, warns = ingest_team_roster_xlsx(app, raw)
            else:
                try:
                    text = raw.decode("utf-8-sig")
                except UnicodeDecodeError:
                    try:
                        text = raw.decode("latin-1")
                    except Exception:  # noqa: BLE001
                        flash("CSV could not be decoded (try UTF-8).", "error")
                        return redirect(url_for("reports"))
                n_teams, n_rows, warns = ingest_team_roster_csv(app, text)
        except ValueError as ex:
            flash(str(ex), "error")
            return redirect(url_for("reports"))
        for w in warns:
            flash(w, "error")
        flash(f"Roster updated: {n_teams} team(s), {n_rows} member row(s).", "success")
        return redirect(url_for("reports"))

    @app.route("/reports/roster-template.xlsx")
    def reports_roster_template_xlsx():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Roster Template"
        ws.append(["Team Name", "Employee Name", "Employee Email"])
        ws.append(["Alpha", "Anas P", "anas.p@nokia.com"])
        buf = io.BytesIO()
        wb.save(buf)
        body = buf.getvalue()
        return Response(
            body,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="roster_template.xlsx"'},
        )

    @app.route("/reports/roster-export.json")
    def reports_roster_export_json():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        conn = get_db(app)
        rows = build_team_roster_export_rows(conn)
        conn.close()
        payload = [
            {"team_name": team, "employee_name": name, "employee_email": email}
            for team, name, email in rows
        ]
        return jsonify({"rows": payload})

    @app.route("/reports/roster-export.xlsx")
    def reports_roster_export_xlsx():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        from openpyxl import Workbook

        conn = get_db(app)
        export_pairs = build_team_roster_export_rows(conn)
        conn.close()
        wb = Workbook()
        ws = wb.active
        ws.append(["TeamName", "EmployeeName", "EmployeeEmail"])
        for team_name, emp_name, emp_email in export_pairs:
            ws.append([team_name, emp_name, emp_email])
        buf = io.BytesIO()
        wb.save(buf)
        body = buf.getvalue()
        fname = "team_roster_mapping.xlsx"
        return Response(
            body,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.route("/reports/roster-export.csv")
    def reports_roster_export_csv_legacy():
        """Old bookmark: roster export is now Excel (.xlsx)."""
        return redirect(url_for("reports_roster_export_xlsx"), code=301)

    @app.route("/reports")
    def reports():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))

        conn = get_db(app)
        lpo_raw = conn.execute(
            "SELECT id, email, created_at FROM lpo_manager_emails ORDER BY email COLLATE NOCASE"
        ).fetchall()
        all_teams = conn.execute("SELECT id, name FROM teams ORDER BY name COLLATE NOCASE").fetchall()
        # Build per-LPO assigned team ids set
        lpo_team_map: dict[str, set[int]] = {}
        for r in conn.execute("SELECT lpo_email, team_id FROM lpo_manager_team_access").fetchall():
            key = normalize_email(str(r["lpo_email"] or ""))
            lpo_team_map.setdefault(key, set()).add(int(r["team_id"]))
        # Enrich lpo_rows with assigned teams info
        lpo_rows = []
        for r in lpo_raw:
            em = normalize_email(str(r["email"] or ""))
            assigned = lpo_team_map.get(em, set())
            lpo_rows.append({
                "id": r["id"],
                "email": r["email"],
                "created_at": r["created_at"],
                "assigned_team_ids": assigned,
            })
        conn.close()

        roster = getattr(g, "manager_roster", EMPLOYEES)
        employee_access_codes = _portal_employee_access_code_status_for_roster(app, roster)
        roster_table_rows: list[dict[str, str]] = []
        tid = getattr(g, "manager_team_id", None)
        if tid is not None:
            conn = get_db(app)
            rrows = conn.execute(
                """
                SELECT tr.employee_name, tr.employee_email, t.name AS team_name
                FROM team_roster tr
                JOIN teams t ON t.id = tr.team_id
                WHERE tr.team_id = ?
                ORDER BY tr.sort_order, tr.employee_name COLLATE NOCASE
                """,
                (int(tid),),
            ).fetchall()
            conn.close()
            roster_table_rows = [
                {
                    "team_name": str(r["team_name"] or "").strip(),
                    "employee_name": str(r["employee_name"] or "").strip(),
                    "employee_email": normalize_email(str(r["employee_email"] or "")),
                }
                for r in rrows
            ]

        return render_template(
            "reports.html",
            lpo_manager_emails=lpo_rows,
            all_teams=all_teams,
            employee_access_codes=employee_access_codes,
            roster_table_rows=roster_table_rows,
            active_team_name=getattr(g, "manager_team_name", None),
        )

    @app.post("/reports/lpo-emails/add")
    def reports_lpo_emails_add():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("reports"))
        raw = (request.form.get("lpo_emails") or request.form.get("lpo_email") or "").strip()
        if not raw:
            flash("Enter at least one Nokia email address.", "error")
            return redirect(url_for("reports"))
        candidates = [normalize_email(line) for line in raw.replace(",", "\n").splitlines()]
        candidates = [c for c in candidates if c]
        if not candidates:
            flash("Enter at least one valid email address.", "error")
            return redirect(url_for("reports"))
        added = 0
        skipped = 0
        invalid = 0
        ts = _utc_stamp()
        conn = get_db(app)
        for em in candidates:
            if "@" not in em:
                invalid += 1
                continue
            try:
                conn.execute(
                    "INSERT INTO lpo_manager_emails (email, created_at) VALUES (?, ?)",
                    (em, ts),
                )
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
        conn.commit()
        conn.close()
        parts: list[str] = []
        if added:
            parts.append(f"{added} LPO email(s) added")
        if skipped:
            parts.append(f"{skipped} already listed")
        if invalid:
            parts.append(f"{invalid} invalid")
        flash(". ".join(parts) + ".", "success" if added else "info")
        return redirect(url_for("reports"))

    @app.post("/reports/lpo-emails/<int:lpo_id>/delete")
    def reports_lpo_emails_delete(lpo_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("reports"))
        conn = get_db(app)
        row = conn.execute("SELECT email FROM lpo_manager_emails WHERE id = ?", (lpo_id,)).fetchone()
        if row:
            conn.execute(
                "DELETE FROM lpo_manager_team_access WHERE lpo_email = ? COLLATE NOCASE",
                (str(row["email"]),),
            )
        conn.execute("DELETE FROM lpo_manager_emails WHERE id = ?", (lpo_id,))
        conn.commit()
        conn.close()
        flash("LPO email removed.", "success")
        return redirect(url_for("reports"))

    @app.post("/reports/lpo-emails/<int:lpo_id>/teams")
    def reports_lpo_emails_set_teams(lpo_id: int):
        """Save which teams this LPO email can access (checkbox list)."""
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("reports"))
        conn = get_db(app)
        row = conn.execute("SELECT email FROM lpo_manager_emails WHERE id = ?", (lpo_id,)).fetchone()
        if not row:
            conn.close()
            flash("LPO email not found.", "error")
            return redirect(url_for("reports"))
        email = normalize_email(str(row["email"]))
        selected_ids = [int(v) for v in request.form.getlist("team_ids") if v.strip().isdigit()]
        valid_ids = {
            int(r["id"])
            for r in conn.execute("SELECT id FROM teams").fetchall()
        }
        selected_ids = [tid for tid in selected_ids if tid in valid_ids]
        ts = _utc_stamp()
        conn.execute(
            "DELETE FROM lpo_manager_team_access WHERE lpo_email = ? COLLATE NOCASE", (email,)
        )
        for tid in selected_ids:
            try:
                conn.execute(
                    "INSERT INTO lpo_manager_team_access (lpo_email, team_id, created_at) VALUES (?, ?, ?)",
                    (email, tid, ts),
                )
            except sqlite3.IntegrityError:
                pass
        conn.commit()
        conn.close()
        flash(f"Teams updated for {email} ({len(selected_ids)} assigned).", "success")
        return redirect(url_for("reports"))

    @app.post("/reports/employee-access-code/set")
    def reports_employee_access_code_set():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("reports"))
        roster = getattr(g, "manager_roster", ())
        raw_name = (request.form.get("employee_name") or "").strip()
        canonical = _canonical_roster_name_on_manager_roster(roster, raw_name)
        if not canonical:
            flash("That employee is not on the active team roster.", "error")
            return redirect(url_for("reports"))
        code = (request.form.get("access_code") or "").strip()
        err = _set_portal_employee_access_code(
            app,
            canonical,
            code,
            manager_email=str(session.get("manager_user_email") or ""),
        )
        if err:
            flash(err, "error")
            return redirect(url_for("reports"))
        flash(
            f"Access code saved for {canonical}. Share the 6-digit code with the employee; "
            "it is not shown again here.",
            "success",
        )
        return redirect(url_for("reports"))

    @app.post("/reports/employee-access-code/generate")
    def reports_employee_access_code_generate():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("reports"))
        roster = getattr(g, "manager_roster", ())
        raw_name = (request.form.get("employee_name") or "").strip()
        canonical = _canonical_roster_name_on_manager_roster(roster, raw_name)
        if not canonical:
            flash("That employee is not on the active team roster.", "error")
            return redirect(url_for("reports"))
        code = f"{secrets.randbelow(900000) + 100000:06d}"
        err = _set_portal_employee_access_code(
            app,
            canonical,
            code,
            manager_email=str(session.get("manager_user_email") or ""),
        )
        if err:
            flash(err, "error")
            return redirect(url_for("reports"))
        flash(
            f"New access code for {canonical}: {code}. Give this to the employee once; "
            "it will not be displayed again.",
            "success",
        )
        return redirect(url_for("reports"))

    @app.post("/reports/employee-access-code/clear")
    def reports_employee_access_code_clear():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))
        if not _csrf_form_ok():
            flash("CSRF validation failed. Please try again.", "error")
            return redirect(url_for("reports"))
        roster = getattr(g, "manager_roster", ())
        raw_name = (request.form.get("employee_name") or "").strip()
        canonical = _canonical_roster_name_on_manager_roster(roster, raw_name)
        if not canonical:
            flash("That employee is not on the active team roster.", "error")
            return redirect(url_for("reports"))
        _clear_portal_employee_access_code(app, canonical)
        flash(f"Access code cleared for {canonical}.", "success")
        return redirect(url_for("reports"))

    @app.route("/worksheet/nokia-audit-compare.xlsx")
    @app.route("/reports/nokia-audit-compare.xlsx")
    def nokia_audit_compare_xlsx():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))
        today = date.today()
        try:
            year = int(request.args.get("year") or today.year)
            month = int(request.args.get("month") or today.month)
        except ValueError:
            year, month = today.year, today.month
        year = max(2000, min(2100, year))
        month = max(1, min(12, month))
        employee_name = (request.args.get("employee_name") or "").strip()
        if not employee_name or employee_name not in g.manager_roster:
            flash("Select an employee from your team roster.", "error")
            return redirect(url_for("nokia_audit", year=year, month=month))
        compare_rows_result, cmp_err = _nokia_audit_compare_rows_from_session(employee_name, year)
        if cmp_err:
            flash(cmp_err, "error")
            return redirect(url_for("nokia_audit", year=year, month=month, employee_name=employee_name))
        rows = compare_rows_result or []
        data, err = build_nokia_compare_xlsx_bytes(rows, employee_name=employee_name, year=year)
        if err or not data:
            flash(err or "Could not build Excel export.", "error")
            return redirect(url_for("nokia_audit", year=year, month=month, employee_name=employee_name))
        safe_emp = _sanitize_xlsx_base_filename(employee_name)
        fname = f"nokia_eleave_compare_{safe_emp}_{year}.xlsx"
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.route("/worksheet/nokia-audit", methods=["GET", "POST"])
    @app.route("/reports/nokia-audit", methods=["GET", "POST"])
    def nokia_audit():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))

        today = date.today()
        try:
            year = int(request.values.get("year") or today.year)
            month = int(request.values.get("month") or today.month)
        except ValueError:
            year, month = today.year, today.month
        year = max(2000, min(2100, year))
        month = max(1, min(12, month))

        employee_name = (request.values.get("employee_name") or "").strip()

        marked_rows: list[dict[str, Any]] = []
        tracker_approved_rows: list[dict[str, Any]] = []
        tracker_approved_submitted = False
        dsm_tracker_rows: list[dict[str, Any]] = []
        dsm_tracker_submitted = False
        compare_rows: list[dict[str, Any]] = []
        compare_submitted = False
        parse_error: str | None = None
        csv_paste_value = ""

        if request.method == "POST":
            if not _csrf_form_ok():
                flash("Invalid security token. Try again.", "error")
            else:
                csv_paste_value = request.form.get("nokia_csv") or ""
                csv_text_input = (csv_paste_value or "").strip()

                if app.config.get("TESTING"):
                    ot = (request.form.get("nokia_ocr_text") or "").strip()
                    if ot:
                        csv_text_input = ot
                        if not (request.form.get("nokia_csv") or "").strip():
                            csv_paste_value = ot

                audit_action = (request.form.get("audit_action") or "show_approved").strip()
                month_lab = date(year, month, 1).strftime("%B %Y")

                if audit_action == "show_approved":
                    if not employee_name or employee_name not in g.manager_roster:
                        parse_error = "Select an employee from your team roster."
                    elif not csv_text_input.strip():
                        parse_error = "Paste Nokia eLeave text first, then click Show Approved Leaves."
                    else:
                        name_err = _nokia_paste_employee_must_match_selected_or_error(
                            employee_name,
                            csv_text_input,
                            g.manager_roster,
                            require_roster_name_in_paste=True,
                        )
                        if name_err:
                            parse_error = name_err
                        else:
                            tracker_approved_submitted = True
                            preview_rows, prev_err, preview_segs = _nokia_paste_approved_preview_rows(
                                app, csv_text_input.strip(), employee_name
                            )
                            if prev_err:
                                parse_error = prev_err
                                tracker_approved_rows = []
                                session.pop(NOKIA_AUDIT_SESSION_APPROVED_SEGS, None)
                                session.pop(NOKIA_AUDIT_SESSION_LAST_APPROVED_PREVIEW, None)
                            else:
                                tracker_approved_rows = preview_rows
                                session[NOKIA_AUDIT_SESSION_LAST_APPROVED_PREVIEW] = {
                                    "employee_name": employee_name,
                                    "rows": _nokia_audit_session_safe_row_dicts(tracker_approved_rows),
                                }
                                session.modified = True
                                if preview_segs is not None:
                                    session[NOKIA_AUDIT_SESSION_APPROVED_SEGS] = {
                                        "employee_name": employee_name,
                                        "fingerprint": _nokia_paste_fingerprint(csv_text_input.strip()),
                                        "segments": _nokia_segs_to_store(preview_segs),
                                    }
                                    session.modified = True
                                flash(
                                    f"Parsed {len(tracker_approved_rows)} approved row(s) from pasted text for "
                                    f"{employee_name} ({month_lab}).",
                                    "success",
                                )
                elif audit_action == "mark_approved":
                    if not employee_name or employee_name not in g.manager_roster:
                        parse_error = "Select an employee from your team roster."
                    elif not csv_text_input.strip():
                        parse_error = "Paste Nokia approved leave text into the text area first."
                    else:
                        name_err = _nokia_paste_employee_must_match_selected_or_error(
                            employee_name,
                            csv_text_input,
                            g.manager_roster,
                            require_roster_name_in_paste=True,
                        )
                        if name_err:
                            parse_error = name_err
                        else:
                            payload = session.get(NOKIA_AUDIT_SESSION_APPROVED_SEGS)
                            fp = _nokia_paste_fingerprint(csv_text_input.strip())
                            segs_override: list[tuple[date, date, str, str, bool]] | None = None
                            if (
                                isinstance(payload, dict)
                                and payload.get("employee_name") == employee_name
                                and payload.get("fingerprint") == fp
                            ):
                                segs_override = _nokia_segs_from_store(payload.get("segments") or [])
                                if segs_override is None:
                                    session.pop(NOKIA_AUDIT_SESSION_APPROVED_SEGS, None)
                            elif payload is not None:
                                session.pop(NOKIA_AUDIT_SESSION_APPROVED_SEGS, None)
                            marked_rows, tagged_existing, mk_err = _nokia_mark_approved_leaves(
                                app,
                                employee_name,
                                year,
                                month,
                                csv_text_input.strip(),
                                precomputed_segments=segs_override,
                            )
                            if mk_err:
                                parse_error = mk_err
                                marked_rows = []
                            elif not marked_rows and not tagged_existing:
                                flash(
                                    f"No new leave was added — Nokia dates in range were already on the tracker for "
                                    f"{employee_name} ({month_lab}).",
                                    "info",
                                )
                            else:
                                session.pop(NOKIA_AUDIT_SESSION_APPROVED_SEGS, None)
                                session.pop(NOKIA_AUDIT_SESSION_LAST_APPROVED_PREVIEW, None)
                                parts: list[str] = []
                                if marked_rows:
                                    parts.append(
                                        f"Added {len(marked_rows)} approved leave record(s) to the tracker for "
                                        f"{employee_name} ({month_lab})."
                                    )
                                if tagged_existing:
                                    parts.append(
                                        f"Updated {tagged_existing} existing tracker record(s) with the Nokia-approved "
                                        f"indicator (green A) for {employee_name} ({month_lab})."
                                    )
                                flash(" ".join(parts), "success")
                elif audit_action == "show_dsm":
                    if not employee_name or employee_name not in g.manager_roster:
                        parse_error = "Select an employee from your team roster."
                    else:
                        dsm_name_err: str | None = None
                        if csv_text_input.strip() and _nokia_paste_has_nokia_approved_leave_date_line(
                            csv_text_input
                        ):
                            dsm_name_err = _nokia_paste_employee_must_match_selected_or_error(
                                employee_name,
                                csv_text_input,
                                g.manager_roster,
                                require_roster_name_in_paste=True,
                            )
                        if dsm_name_err:
                            parse_error = dsm_name_err
                        else:
                            dsm_tracker_submitted = True
                            dsm_tracker_rows = _nokia_audit_dsm_leave_rows(app, employee_name, year)
                            n_dsm = len(dsm_tracker_rows)
                            session[NOKIA_AUDIT_SESSION_LAST_DSM] = {
                                "employee_name": employee_name,
                                "year": int(year),
                                "rows": _nokia_audit_session_safe_row_dicts(dsm_tracker_rows),
                            }
                            session.modified = True
                            if n_dsm:
                                flash(
                                    f"Loaded {n_dsm} leave record(s) from Leave tracker for "
                                    f"{employee_name} (calendar year {year}).",
                                    "success",
                                )
                            else:
                                flash(
                                    f"No pending or approved leave in Leave tracker for {employee_name} "
                                    f"in calendar year {year}.",
                                    "info",
                                )
                elif audit_action == "compare":
                    if not employee_name or employee_name not in g.manager_roster:
                        parse_error = "Select an employee from your team roster."
                    else:
                        compare_rows_result, cmp_err = _nokia_audit_compare_rows_from_session(
                            employee_name, year
                        )
                        if cmp_err:
                            parse_error = cmp_err
                        else:
                            compare_submitted = True
                            compare_rows = compare_rows_result or []
                            if compare_rows:
                                flash(f"Compare: {len(compare_rows)} row(s) merged.", "success")
                            else:
                                flash("Compare: no rows to show (both sources were empty).", "info")
                else:
                    parse_error = "Unknown action."

        month_label = date(year, month, 1).strftime("%B %Y")
        tracker_approved_days_total = _sum_nokia_audit_day_column(tracker_approved_rows)
        marked_rows_days_total = _sum_nokia_audit_day_column(marked_rows)
        dsm_tracker_days_total = _sum_nokia_audit_day_column(dsm_tracker_rows)
        compare_days_elv_total = _sum_nokia_audit_compare_days(compare_rows, "days_elv")
        compare_days_dsm_total = _sum_nokia_audit_compare_days(compare_rows, "days_dsm")
        return render_template(
            "nokia_audit.html",
            year=year,
            month=month,
            month_label=month_label,
            parse_error=parse_error,
            csv_paste_value=csv_paste_value,
            employee_name=employee_name,
            roster_names=g.manager_roster,
            marked_rows=marked_rows,
            tracker_approved_rows=tracker_approved_rows,
            tracker_approved_submitted=tracker_approved_submitted,
            tracker_approved_days_total=tracker_approved_days_total,
            marked_rows_days_total=marked_rows_days_total,
            dsm_tracker_rows=dsm_tracker_rows,
            dsm_tracker_submitted=dsm_tracker_submitted,
            dsm_tracker_days_total=dsm_tracker_days_total,
            compare_rows=compare_rows,
            compare_submitted=compare_submitted,
            compare_days_elv_total=compare_days_elv_total,
            compare_days_dsm_total=compare_days_dsm_total,
        )

    @app.route("/reports/leave/<int:leave_id>/status", methods=["POST"])
    def update_leave_status(leave_id: int):
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard"))

        new_status = (request.form.get("status") or "").strip()
        if new_status not in {"pending", "approved", "rejected"}:
            flash("Invalid status.", "error")
            return redirect(url_for("reports"))

        conn = get_db(app)
        prow = conn.execute(
            "SELECT employee_name FROM leave_requests WHERE id = ?",
            (leave_id,),
        ).fetchone()
        if not prow:
            conn.close()
            flash("Leave request not found.", "error")
            return redirect(url_for("reports"))
        if prow["employee_name"] not in g.manager_roster:
            conn.close()
            flash("That leave is not on your team roster.", "error")
            return redirect(url_for("reports"))

        conn.execute(
            "UPDATE leave_requests SET status = ? WHERE id = ?",
            (new_status, leave_id),
        )
        if new_status == "rejected":
            _purge_meet_leave_day_rows_for_leave(conn, leave_id)
        conn.commit()
        conn.close()
        flash("Saved.", "success")
        if request.form.get("redirect_manager") == "1" and _manager_logged_in():
            try:
                y = int(request.form.get("mgr_year") or 0)
                m = int(request.form.get("mgr_month") or 0)
                if 1 <= m <= 12 and y > 2000:
                    return redirect(url_for("dashboard", year=y, month=m))
            except ValueError:
                pass
        rs = (request.form.get("report_start") or "").strip()
        re = (request.form.get("report_end") or "").strip()
        return redirect(url_for("reports", start=rs or None, end=re or None))

    @app.route("/reports/export-leaves.xlsx")
    def reports_leave_export_xlsx():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))

        start_raw = (request.args.get("start") or "").strip() or date.today().replace(day=1).isoformat()
        end_raw = (request.args.get("end") or "").strip() or date.today().isoformat()
        try:
            sd = date.fromisoformat(start_raw[:10])
            ed = date.fromisoformat(end_raw[:10])
        except ValueError:
            flash("Invalid date range.", "error")
            return redirect(url_for("reports"))
        if ed < sd:
            sd, ed = ed, sd

        roster = tuple(g.manager_roster)
        detail_rows, totals = _leave_tracker_day_rows_for_range(app, roster, sd, ed)

        req_units: dict[int, float] = {}
        req_date_sets: dict[int, set[str]] = {}
        for r in detail_rows:
            lid = int(r["leave_request_id"])
            req_units[lid] = req_units.get(lid, 0.0) + float(r["day_units"])
            req_date_sets.setdefault(lid, set()).add(str(r["leave_date"]))
        req_dates_csv = {k: ",".join(sorted(v)) for k, v in req_date_sets.items()}

        reason_labels = dict(LEAVE_REASONS)
        duration_labels = dict(DURATION_CHOICES)

        ph = ",".join("?" * len(roster))
        conn = get_db(app)
        if roster:
            all_leaves = conn.execute(
                f"""
                SELECT employee_name, reason, start_date, end_date, duration_type, status, created_at, submitted_ip, id
                FROM leave_requests
                WHERE start_date <= ? AND end_date >= ? AND employee_name IN ({ph})
                ORDER BY start_date ASC, employee_name ASC, id ASC
                """,
                (ed.isoformat(), sd.isoformat(), *roster),
            ).fetchall()
        else:
            all_leaves = []
        conn.close()

        from openpyxl import Workbook
        from sprint_hub_snapshot_png import sanitize_excel_sheet_title

        def request_detail_row(row: sqlite3.Row) -> list[Any]:
            lid = int(row["id"])
            reason = str(row["reason"] or "").strip()
            leave_type = str(reason_labels.get(reason, reason))
            dur_code = str(row["duration_type"] or "").strip()
            dur_label = str(duration_labels.get(dur_code, dur_code))
            try:
                ls = date.fromisoformat(str(row["start_date"])[:10])
                le = date.fromisoformat(str(row["end_date"])[:10])
            except ValueError:
                ls = le = sd
            if le < ls:
                ls, le = le, ls
            cal_full = (le - ls).days + 1
            ov = _leave_report_overlap_window(str(row["start_date"]), str(row["end_date"]), sd, ed)
            if ov:
                wdays_ov = _working_weekdays_count(ov[0], ov[1])
            else:
                wdays_ov = 0
            taken = float(round(req_units.get(lid, 0.0), 4))
            leaves_dates = req_dates_csv.get(lid, "")
            return [
                lid,
                row["employee_name"],
                leave_type,
                row["status"],
                cal_full,
                wdays_ov,
                taken,
                row["created_at"],
                str(row["start_date"])[:10],
                str(row["end_date"])[:10],
                dur_code,
                dur_label,
                leaves_dates,
                reason,
                row["submitted_ip"] if "submitted_ip" in row.keys() else "",
            ]

        wb = Workbook()
        ws0 = wb.active
        ws0.title = "Summary"
        team_nm = getattr(g, "manager_team_name", None) or ""
        ws0.append(["Team", team_nm])
        ws0.append(["From", sd.isoformat(), "To", ed.isoformat()])
        ws0.append([])
        ws0.append(
            [
                "Totals use the same rules as the leave tracker: pending and approved requests only; "
                "DSM meet per-day approve/remove applies; when several requests overlap the same day, "
                "the newest request wins; half-day (AM/PM) counts as 0.5.",
            ]
        )
        ws0.append([])
        ws0.append(["Employee", "Total leave days (tracker units in range)"])
        for emp in roster:
            ws0.append([emp, float(round(totals.get(emp, 0.0), 4))])

        ws_detail = wb.create_sheet("Request detail", 1)
        hdr = [
            "leave_request_id",
            "employee_name",
            "leave_type",
            "leave_status",
            "no_of_calendar_days_full_request",
            "working_days_in_report_overlap",
            "days_leaves_taken_in_range_tracker",
            "dates_applied_submitted",
            "date_from",
            "date_to",
            "duration_type",
            "duration_label",
            "leaves_taken_dates_in_range",
            "reason_code",
            "submitted_ip",
        ]
        ws_detail.append(hdr)
        for row in all_leaves:
            ws_detail.append(request_detail_row(row))

        ws_leave = wb.create_sheet("Leave dates", 2)
        ws_leave.append(
            [
                "Employee",
                "Leave date",
                "Day units",
                "Reason code",
                "Reason",
                "Duration code",
                "Duration",
                "Status",
                "Leave request id",
                "Request start",
                "Request end",
            ]
        )
        for r in detail_rows:
            ws_leave.append(
                [
                    r["employee_name"],
                    r["leave_date"],
                    r["day_units"],
                    r["reason"],
                    r["reason_label"],
                    r["duration_type"],
                    r["duration_label"],
                    r["status"],
                    r["leave_request_id"],
                    r["request_start"],
                    r["request_end"],
                ]
            )

        used_person_tabs: set[str] = set()
        for emp in roster:
            tab = sanitize_excel_sheet_title(emp)
            base_tab = tab
            n = 2
            while tab in used_person_tabs:
                tab = sanitize_excel_sheet_title(f"{base_tab[:22]}_{n}")
                n += 1
            used_person_tabs.add(tab)
            wsp = wb.create_sheet(tab)
            wsp.append(["Team", team_nm])
            wsp.append(["Employee", emp])
            wsp.append(["Report from", sd.isoformat(), "Report to", ed.isoformat()])
            wsp.append([])
            tot_u = float(round(totals.get(emp, 0.0), 4))
            emp_reqs = [row for row in all_leaves if str(row["employee_name"]) == emp]
            wsp.append(["Summary — total leave days in range (tracker units)", tot_u])
            wsp.append(["Summary — requests overlapping date range (any status)", len(emp_reqs)])
            wsp.append([])
            wsp.append(hdr)
            for row in emp_reqs:
                wsp.append(request_detail_row(row))

        buf = io.BytesIO()
        wb.save(buf)
        body = buf.getvalue()
        fname = f"leave_report_{sd.isoformat()}_to_{ed.isoformat()}.xlsx"
        return Response(
            body,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.route("/reports/export.csv")
    def export_csv():
        if not _manager_logged_in():
            flash("Manager sign-in required.", "error")
            return redirect(url_for("dashboard", next=request.path))

        start = (request.args.get("start") or date.today().replace(day=1).isoformat()).strip()
        end = (request.args.get("end") or date.today().isoformat()).strip()

        roster = g.manager_roster
        ph = ",".join("?" * len(roster))
        conn = get_db(app)
        leaves = conn.execute(
            f"""
            SELECT employee_name, reason, start_date, end_date, duration_type, status, created_at, submitted_ip
            FROM leave_requests
            WHERE start_date <= ? AND end_date >= ? AND employee_name IN ({ph})
            ORDER BY start_date ASC
            """,
            (end, start, *roster),
        ).fetchall()
        conn.close()

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(
            [
                "employee_name",
                "reason",
                "start_date",
                "end_date",
                "duration_type",
                "status",
                "created_at",
                "submitted_ip",
            ]
        )
        for row in leaves:
            w.writerow(
                [
                    row["employee_name"],
                    row["reason"],
                    row["start_date"],
                    row["end_date"],
                    row["duration_type"],
                    row["status"],
                    row["created_at"],
                    row["submitted_ip"] if "submitted_ip" in row.keys() else "",
                ]
            )

        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=leave_report_{start}_to_{end}.csv"
            },
        )

    # ── Linked Files API ───────────────────────────────────────────────────────

    @app.post("/scrum/api/item/linked-file/add")
    def scrum_api_item_linked_file_add():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        url_val = (data.get("url") or "").strip()[:2000]
        display_name = (data.get("display_name") or "").strip()[:300]
        auth_user = (data.get("auth_user") or "").strip()[:200]
        auth_pass = (data.get("auth_pass") or "").strip()[:200]
        if not item_id or not url_val:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = int(g.manager_team_id)
        item = conn.execute(
            """
            SELECT i.id FROM scrum_sprint_item i
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE i.id = ? AND s.team_id = ?
            """,
            (item_id, team_id),
        ).fetchone()
        if not item:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM scrum_item_linked_file WHERE item_id = ?", (item_id,)
        ).fetchone()[0]
        ts = _utc_stamp()
        cur = conn.execute(
            """
            INSERT INTO scrum_item_linked_file
                (item_id, sort_order, display_name, url, auth_user, auth_pass, updated_at, updated_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (item_id, int(max_order) + 1, display_name or url_val[:80], url_val,
             auth_user, auth_pass, ts, session.get("manager", "")),
        )
        new_id = cur.lastrowid
        saved_name = display_name or url_val[:80]
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id, "display_name": saved_name, "url": url_val,
                        "has_auth": bool(auth_user)})

    @app.post("/scrum/api/item/linked-file/update")
    def scrum_api_item_linked_file_update():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        row_id = _parse_optional_int(data.get("id"))
        item_id = _parse_optional_int(data.get("item_id"))
        display_name = (data.get("display_name") or "").strip()[:300]
        url_val = (data.get("url") or "").strip()[:2000]
        # credentials: send empty strings to clear, send values to store
        auth_user = data.get("auth_user")   # None means "don't touch"
        auth_pass = data.get("auth_pass")
        if not row_id or not item_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = int(g.manager_team_id)
        row = conn.execute(
            """
            SELECT lf.id FROM scrum_item_linked_file lf
            JOIN scrum_sprint_item i ON i.id = lf.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE lf.id = ? AND lf.item_id = ? AND s.team_id = ?
            """,
            (row_id, item_id, team_id),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        ts = _utc_stamp()
        actor = session.get("manager", "")
        if display_name:
            conn.execute(
                "UPDATE scrum_item_linked_file SET display_name = ?, updated_at = ?, updated_by = ? WHERE id = ?",
                (display_name, ts, actor, row_id),
            )
        if url_val:
            conn.execute(
                "UPDATE scrum_item_linked_file SET url = ?, updated_at = ?, updated_by = ? WHERE id = ?",
                (url_val, ts, actor, row_id),
            )
        if auth_user is not None:
            conn.execute(
                "UPDATE scrum_item_linked_file SET auth_user = ?, updated_at = ?, updated_by = ? WHERE id = ?",
                (str(auth_user).strip()[:200], ts, actor, row_id),
            )
        if auth_pass is not None:
            conn.execute(
                "UPDATE scrum_item_linked_file SET auth_pass = ?, updated_at = ?, updated_by = ? WHERE id = ?",
                (str(auth_pass).strip()[:200], ts, actor, row_id),
            )
        new_user = str(auth_user or "").strip() if auth_user is not None else None
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "has_auth": bool(new_user) if new_user is not None else None})

    @app.post("/scrum/api/item/linked-file/delete")
    def scrum_api_item_linked_file_delete():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        row_id = _parse_optional_int(data.get("id"))
        item_id = _parse_optional_int(data.get("item_id"))
        if not row_id or not item_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = int(g.manager_team_id)
        row = conn.execute(
            """
            SELECT lf.id FROM scrum_item_linked_file lf
            JOIN scrum_sprint_item i ON i.id = lf.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE lf.id = ? AND lf.item_id = ? AND s.team_id = ?
            """,
            (row_id, item_id, team_id),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        conn.execute("DELETE FROM scrum_item_linked_file WHERE id = ?", (row_id,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.get("/scrum/api/item/linked-file/proxy/<int:file_id>")
    def scrum_api_item_linked_file_proxy(file_id: int):
        """Server-side proxy: fetch a linked file using stored credentials and stream back.

        Only available to logged-in managers/portal users for their team's files.
        Credentials are stored per linked-file row; we support NTLM (SharePoint/Windows)
        with a fallback to HTTP Basic auth.
        """
        import urllib.request
        import urllib.error

        authed = _manager_logged_in() or bool(_portal_session())
        if not authed:
            return jsonify({"ok": False, "error": "auth"}), 403

        conn = get_db(app)
        if _manager_logged_in():
            team_id = int(g.manager_team_id)
        else:
            team_id = int(g.portal_team_id)

        row = conn.execute(
            """
            SELECT lf.url, lf.auth_user, lf.auth_pass
            FROM scrum_item_linked_file lf
            JOIN scrum_sprint_item i ON i.id = lf.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE lf.id = ? AND s.team_id = ?
            """,
            (file_id, team_id),
        ).fetchone()
        conn.close()

        if not row:
            return jsonify({"ok": False, "error": "not_found"}), 404

        target_url = str(row["url"] or "")
        auth_user = str(row["auth_user"] or "").strip()
        auth_pass = str(row["auth_pass"] or "").strip()

        if not target_url:
            return jsonify({"ok": False, "error": "no_url"}), 400

        try:
            import requests as _req  # type: ignore[import]
            _req_available = True
        except ImportError:
            _req_available = False

        try:
            if _req_available and auth_user:
                # Try NTLM first (SharePoint / Windows-auth), fall back to Basic
                try:
                    from requests_ntlm import HttpNtlmAuth  # type: ignore[import]
                    auth = HttpNtlmAuth(auth_user, auth_pass)
                except ImportError:
                    from requests.auth import HTTPBasicAuth  # type: ignore[import]
                    auth = HTTPBasicAuth(auth_user, auth_pass)
                resp = _req.get(target_url, auth=auth, timeout=20, stream=True,
                                headers={"User-Agent": "Mozilla/5.0"})
                content_type = resp.headers.get("Content-Type", "application/octet-stream")
                status = resp.status_code
                body = resp.content
            elif _req_available:
                resp = _req.get(target_url, timeout=20,
                                headers={"User-Agent": "Mozilla/5.0"})
                content_type = resp.headers.get("Content-Type", "application/octet-stream")
                status = resp.status_code
                body = resp.content
            else:
                req = urllib.request.Request(target_url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=20) as resp:
                    content_type = resp.headers.get("Content-Type", "application/octet-stream")
                    status = resp.status
                    body = resp.read()

            from flask import Response as _Response  # noqa: F811 – already imported
            return _Response(body, status=status, content_type=content_type)

        except Exception as exc:  # noqa: BLE001
            return jsonify({"ok": False, "error": str(exc)}), 502

    # ── Task Checklist API ─────────────────────────────────────────────────────

    @app.get("/scrum/api/item/detail")
    def scrum_api_item_detail():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        item_id = _parse_optional_int(request.args.get("item_id"))
        sprint_id = _parse_optional_int(request.args.get("sprint_id"))
        if not item_id or not sprint_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        team_id = int(g.manager_team_id)
        conn = get_db(app)
        if not _scrum_checklist_item_on_team(conn, item_id, team_id):
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        payload = _build_scrum_item_detail_payload(conn, team_id, sprint_id, item_id)
        conn.close()
        if not payload:
            return jsonify({"ok": False, "error": "not_found"}), 404
        return jsonify({"ok": True, **payload})

    @app.get("/scrum/api/item/checklist")
    def scrum_api_item_checklist_get():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        item_id = _parse_optional_int(request.args.get("item_id"))
        if not item_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        rows = _fetch_item_checklist_map(conn, [item_id]).get(item_id, [])
        conn.close()
        return jsonify({"ok": True, "rows": rows})

    @app.post("/scrum/api/item/checklist/add")
    def scrum_api_item_checklist_add():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        item_id = _parse_optional_int(data.get("item_id"))
        if not item_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        initial_item_name = (data.get("items_to_finish") or "").strip()[:500]
        conn = get_db(app)
        team_id = int(g.manager_team_id)
        item = conn.execute(
            """
            SELECT i.id FROM scrum_sprint_item i
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE i.id = ? AND s.team_id = ?
            """,
            (item_id, team_id),
        ).fetchone()
        if not item:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM scrum_item_checklist WHERE item_id = ?", (item_id,)
        ).fetchone()[0]
        ts = _utc_stamp()
        cur = conn.execute(
            """
            INSERT INTO scrum_item_checklist (item_id, sort_order, items_to_finish, status, le_to_complete, done_till_date, updated_at, updated_by)
            VALUES (?, ?, ?, '', '', '', ?, ?)
            """,
            (item_id, int(max_order) + 1, initial_item_name, ts, session.get("manager", "")),
        )
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": new_id, "sort_order": int(max_order) + 1})

    @app.post("/scrum/api/item/checklist/update")
    def scrum_api_item_checklist_update():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        row_id = _parse_optional_int(data.get("id"))
        item_id = _parse_optional_int(data.get("item_id"))
        field = (data.get("field") or "").strip()
        value = (data.get("value") or "")[:2000]
        allowed_fields = {"items_to_finish", "status", "le_to_complete", "done_till_date"}
        if not row_id or not item_id or field not in allowed_fields:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = int(g.manager_team_id)
        row = conn.execute(
            """
            SELECT c.id FROM scrum_item_checklist c
            JOIN scrum_sprint_item i ON i.id = c.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE c.id = ? AND c.item_id = ? AND s.team_id = ?
            """,
            (row_id, item_id, team_id),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        ts = _utc_stamp()
        conn.execute(
            f"UPDATE scrum_item_checklist SET {field} = ?, updated_at = ?, updated_by = ? WHERE id = ?",
            (value, ts, session.get("manager", ""), row_id),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @app.post("/scrum/api/item/checklist/delete")
    def scrum_api_item_checklist_delete():
        if not _manager_logged_in():
            return jsonify({"ok": False, "error": "auth"}), 403
        if not _csrf_api_ok(app):
            return jsonify({"ok": False, "error": "csrf"}), 400
        data = request.get_json(silent=True) or {}
        row_id = _parse_optional_int(data.get("id"))
        item_id = _parse_optional_int(data.get("item_id"))
        if not row_id or not item_id:
            return jsonify({"ok": False, "error": "missing"}), 400
        conn = get_db(app)
        team_id = int(g.manager_team_id)
        row = conn.execute(
            """
            SELECT c.id FROM scrum_item_checklist c
            JOIN scrum_sprint_item i ON i.id = c.item_id
            JOIN scrum_sprint s ON s.id = i.sprint_id
            WHERE c.id = ? AND c.item_id = ? AND s.team_id = ?
            """,
            (row_id, item_id, team_id),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": "forbidden"}), 403
        conn.execute("DELETE FROM scrum_item_checklist WHERE id = ?", (row_id,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    with app.app_context():
        init_db(app)

    import team_tracker_backup

    team_tracker_backup.maybe_start_backup_scheduler(app)

    return app
