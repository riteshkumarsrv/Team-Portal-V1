"""Tests for the team tracker website."""

from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime, timedelta, timezone

import pytest

from app import (
    EMPLOYEES,
    NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER,
    SCRUM_KANBAN_WEEKDAY_HOURS,
    create_app,
    fuzzy_employee_matches,
    get_db,
    parse_nokia_grid_combined,
    parse_nokia_employee_summary_leave_dates,
    parse_nokia_grid_tsv,
    parse_nokia_grid_whitespace,
    resolve_employee_name,
    build_month_context,
    _available_hours_for_assignee_sprint_window,
    _nokia_reason_and_label_from_line,
    _nokia_eleavetool_type_display,
    _nokia_paste_approved_preview_rows,
    _parse_nokia_audit_dd_mm_range,
)


def _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path: str) -> None:
    """After ``load_dotenv``, restore ``TEAM_TRACKER_DB_PATH`` so project ``.env`` cannot override test DB."""
    import config as config_module

    _real_load_dotenv = config_module.load_dotenv

    def _load_dotenv_then_restore_test_db(*args, **kwargs):
        _real_load_dotenv(*args, **kwargs)
        monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)

    monkeypatch.setattr(config_module, "load_dotenv", _load_dotenv_then_restore_test_db)


def test_roster_fuzzy():
    assert "Shaishta Anjum" in fuzzy_employee_matches("shais", 5)
    assert resolve_employee_name("shaishta anjum") == "Shaishta Anjum"


@pytest.fixture()
def app(tmp_path, monkeypatch):
    """Isolated SQLite under tmp_path; pin TEAM_TRACKER_DB_PATH after create_app's dotenv load."""
    db_path = str(tmp_path / "test.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.setenv("MICROSOFT_OAUTH_CLIENT_ID", "test-ms-client-id")
    monkeypatch.setenv("MICROSOFT_OAUTH_CLIENT_SECRET", "test-ms-client-secret")

    import app as app_module

    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)

    application = app_module.create_app()
    application.config["TESTING"] = True
    monkeypatch.setenv("TEAM_TRACKER_AUTO_TESSERACT", "0")
    return application


@pytest.fixture()
def client(app):
    return app.test_client()


def _manager_login(c):
    res = c.post("/dashboard", data={"gate_kind": "manager", "secret_code": "mgrpw"}, follow_redirects=False)
    assert res.status_code in (302, 303)


def test_healthz(client):
    res = client.get("/healthz")
    assert res.status_code == 200
    assert res.get_json() == {"status": "ok"}


def test_api_employees_requires_query(client):
    assert client.get("/api/employees?q=").get_json()["matches"] == []


def test_api_employees_match(client):
    data = client.get("/api/employees?q=var").get_json()
    assert "Varshini Raj" in data["matches"]


def test_home_ok(client):
    res = client.get("/", follow_redirects=True)
    assert res.status_code == 200
    assert b"TEAM PORTAL" in res.data
    assert b"Employee Portal" in res.data or b"Verify Code" in res.data
    assert b"Manager Access" in res.data


def test_home_shows_email_otp_when_smtp_only(tmp_path, monkeypatch):
    db_path = str(tmp_path / "otp_home.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    monkeypatch.setenv("PORTAL_OTP_SMTP_HOST", "localhost")
    monkeypatch.setenv("PORTAL_OTP_FROM", "noreply@example.test")
    application = create_app()
    application.config["TESTING"] = True
    c = application.test_client()
    r = c.get("/", follow_redirects=True)
    assert r.status_code == 200
    assert b"Employee Portal" in r.data or b"Continue as Employee" in r.data


def test_home_auto_enables_dev_otp_when_unconfigured(tmp_path, monkeypatch):
    db_path = str(tmp_path / "auto_otp_home.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    monkeypatch.delenv("PORTAL_OTP_SMTP_HOST", raising=False)
    monkeypatch.delenv("PORTAL_OTP_FROM", raising=False)
    monkeypatch.delenv("PORTAL_OTP_DEV_CONSOLE", raising=False)
    monkeypatch.delenv("TEAM_TRACKER_PRODUCTION", raising=False)
    application = create_app()
    application.config["TESTING"] = True
    r = application.test_client().get("/", follow_redirects=True)
    assert r.status_code == 200
    assert b"Employee Portal" in r.data or b"Continue as Employee" in r.data


def test_home_employee_signin_off_when_production_without_auth(tmp_path, monkeypatch):
    db_path = str(tmp_path / "prod_signin_home.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.setenv("TEAM_TRACKER_PRODUCTION", "1")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    monkeypatch.delenv("PORTAL_OTP_SMTP_HOST", raising=False)
    monkeypatch.delenv("PORTAL_OTP_FROM", raising=False)
    monkeypatch.setenv("PORTAL_OTP_DEV_CONSOLE", "0")
    application = create_app()
    application.config["TESTING"] = True
    r = application.test_client().get("/", follow_redirects=True)
    assert r.status_code == 200
    assert b"Employee Portal" in r.data


def test_portal_email_otp_send_verify(tmp_path, monkeypatch):
    db_path = str(tmp_path / "otp_flow.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    monkeypatch.setenv("PORTAL_OTP_SMTP_HOST", "localhost")
    monkeypatch.setenv("PORTAL_OTP_FROM", "noreply@example.test")
    application = create_app()
    application.config["TESTING"] = True
    c = application.test_client()

    captured: list[tuple[str, str]] = []

    def fake_send(app, to_addr, code):
        captured.append((to_addr, code))

    monkeypatch.setattr("app.send_portal_otp_smtp", fake_send)

    r = c.post("/auth/email-otp/send", data={"email": "akanksha.jha@nokia.com"})
    assert r.status_code in (302, 303)
    assert captured and captured[0][0] == "akanksha.jha@nokia.com"
    code = captured[0][1]

    r2 = c.post("/auth/email-otp/verify", data={"code": code}, follow_redirects=False)
    assert r2.status_code in (302, 303)
    assert "/portal" in (r2.headers.get("Location") or "")

    dash = c.get("/portal")
    assert dash.status_code == 200
    assert b"My sprint work" in dash.data


def test_portal_email_otp_dev_console_flow(tmp_path, monkeypatch):
    db_path = str(tmp_path / "otp_dev.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    monkeypatch.delenv("PORTAL_OTP_SMTP_HOST", raising=False)
    monkeypatch.delenv("PORTAL_OTP_FROM", raising=False)
    monkeypatch.delenv("PORTAL_OTP_DEV_CONSOLE", raising=False)
    monkeypatch.delenv("TEAM_TRACKER_PRODUCTION", raising=False)
    application = create_app()
    application.config["TESTING"] = True
    c = application.test_client()
    r = c.post(
        "/auth/email-otp/send",
        data={"email": "akanksha.jha@nokia.com"},
        follow_redirects=True,
    )
    assert r.status_code == 200
    m = re.search(rb"your code is (\d{6})", r.data)
    assert m
    code = m.group(1).decode()
    r2 = c.post("/auth/email-otp/verify", data={"code": code}, follow_redirects=False)
    assert r2.status_code in (302, 303)
    assert "/portal" in (r2.headers.get("Location") or "")


def test_leave_submit_via_portal_uses_roster_and_ip(client, app):
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "akanksha.jha@nokia.com",
            "name": "Akanksha Jha",
            "roster_name": "Akanksha Jha",
            "role": "employee",
        }
    res = client.post(
        "/portal/leave/apply",
        data={
            "leave_kind": "pl",
            "start_date": "2026-07-01",
            "end_date": "2026-07-01",
            "reason": "Test day off",
        },
        environ_base={"REMOTE_ADDR": "203.0.113.9"},
        follow_redirects=False,
    )
    assert res.status_code in (302, 303)
    assert "/portal" in (res.headers.get("Location") or "")
    conn = get_db(app)
    row = conn.execute(
        "SELECT employee_name, submitted_ip FROM leave_requests ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    assert row["employee_name"] == "Akanksha Jha"
    assert row["submitted_ip"] == "203.0.113.9"


def test_dashboard_gate_shows_without_password_config(tmp_path, monkeypatch):
    db_path = str(tmp_path / "nogate.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    app = create_app()
    app.config["TESTING"] = True
    c = app.test_client()
    r = c.get("/dashboard")
    assert r.status_code == 200
    assert b"Manager access is not configured" in r.data or b"Manager access" in r.data
    assert b"Access code" in r.data or b"gate" in r.data.lower()


def test_portal_flow_with_fake_session(client, app):
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "akanksha.jha@nokia.com",
            "name": "Akanksha Jha",
            "roster_name": "Akanksha Jha",
            "role": "employee",
        }
    dash = client.get("/portal")
    assert dash.status_code == 200
    assert b"My sprint work" in dash.data
    rv = client.post(
        "/portal/leave/apply",
        data={
            "leave_kind": "pl",
            "start_date": "2026-06-02",
            "end_date": "2026-06-04",
            "reason": "Time off",
        },
        follow_redirects=True,
    )
    assert rv.status_code == 200
    hist = client.get("/my-requests")
    assert hist.status_code == 200
    assert b"Time off" in hist.data


def test_portal_leave_rejects_weekend_start(client, app):
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "sumit.patra@nokia.com",
            "name": "Sumit Patra",
            "roster_name": "Sumit Patra",
            "role": "employee",
        }
    res = client.post(
        "/portal/leave/apply",
        data={
            "leave_kind": "sl",
            "start_date": "2026-06-06",
            "end_date": "2026-06-08",
            "reason": "Bad range",
        },
    )
    assert res.status_code == 200
    assert b"weekend" in res.data.lower()


def test_portal_requires_login(client):
    assert client.get("/portal", follow_redirects=False).status_code == 302


def test_portal_my_requests_blocks_name_change_post(client, app):
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "akanksha.jha@nokia.com",
            "name": "Akanksha Jha",
            "roster_name": "Akanksha Jha",
            "role": "employee",
        }
    res = client.post(
        "/my-requests",
        data={"employee_name": "Sumit Patra"},
        follow_redirects=True,
    )
    assert res.status_code == 200
    assert b"only view" in res.data.lower() or b"own requests" in res.data.lower()


def test_portal_leave_redirects_legacy_leave(client, app):
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "akanksha.jha@nokia.com",
            "name": "Akanksha Jha",
            "roster_name": "Akanksha Jha",
            "role": "employee",
        }
    r = client.get("/leave", follow_redirects=False)
    assert r.status_code in (302, 303)
    assert "/portal/leave/apply" in (r.headers.get("Location") or "")


def test_leave_submit_manager_session_uses_roster_and_ip(client, app):
    _manager_login(client)
    name = EMPLOYEES[0]
    res = client.post(
        "/leave",
        data={
            "employee_name": name.lower(),
            "reason": "pl",
            "start_date": "2026-07-01",
            "end_date": "2026-07-01",
            "day_part": "full",
        },
        environ_base={"REMOTE_ADDR": "203.0.113.9"},
        follow_redirects=False,
    )
    assert res.status_code == 302
    assert "/my-requests" in (res.headers.get("Location") or "")
    conn = get_db(app)
    row = conn.execute(
        "SELECT employee_name, submitted_ip FROM leave_requests ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    assert row["employee_name"] == name
    assert row["submitted_ip"] == "203.0.113.9"


def test_dashboard_wrong_secret(client):
    res = client.post("/dashboard", data={"gate_kind": "manager", "secret_code": "nope"})
    assert res.status_code == 200
    assert b"Invalid code" in res.data


def test_lpo_sm_signin_opens_leave_tracker(tmp_path, monkeypatch):
    db_path = str(tmp_path / "lpo_signin.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrx")
    monkeypatch.setenv("LPO_SM_DASHBOARD_PASSWORD", "lponsecret")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    application = create_app()
    application.config["TESTING"] = True
    c = application.test_client()
    rv = c.post("/dashboard", data={"gate_kind": "lpo_sm", "secret_code": "lponsecret"}, follow_redirects=False)
    assert rv.status_code in (302, 303)
    with c.session_transaction() as sess:
        assert sess.get("manager") is True
        assert sess.get("manager_role") == "lpo_sm"
    dash = c.get("/dashboard?year=2026&month=7")
    assert dash.status_code == 200
    assert b"Leave tracker" in dash.data


def test_lpo_sm_wrong_secret(tmp_path, monkeypatch):
    db_path = str(tmp_path / "lpo_wrong.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrx")
    monkeypatch.setenv("LPO_SM_DASHBOARD_PASSWORD", "goodlpo")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    application = create_app()
    application.config["TESTING"] = True
    c = application.test_client()
    res = c.post("/dashboard", data={"gate_kind": "lpo_sm", "secret_code": "badlpo"})
    assert res.status_code == 200
    assert b"Invalid code" in res.data


def test_lpo_sm_rejects_manager_code(tmp_path, monkeypatch):
    """Manager password must not unlock the LPO/SM gate."""
    db_path = str(tmp_path / "lpo_mix.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "onlymgr")
    monkeypatch.setenv("LPO_SM_DASHBOARD_PASSWORD", "onlylpo")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("MICROSOFT_OAUTH_CLIENT_SECRET", raising=False)
    application = create_app()
    application.config["TESTING"] = True
    c = application.test_client()
    res = c.post("/dashboard", data={"gate_kind": "lpo_sm", "secret_code": "onlymgr"})
    assert res.status_code == 200
    assert b"Invalid code" in res.data


def test_manager_reports_and_export(client):
    c = client
    assert c.get("/reports").status_code == 302

    _manager_login(c)

    r = c.get("/reports?start=2026-07-01&end=2026-07-31")
    assert r.status_code == 200
    assert b"Leave Reports" in r.data

    exp = c.get("/reports/export.csv?start=2026-07-01&end=2026-07-31")
    assert exp.status_code == 200
    assert b"submitted_ip" in exp.data

    xlsx = c.get("/reports/export-leaves.xlsx?start=2026-07-01&end=2026-07-31")
    assert xlsx.status_code == 200
    assert xlsx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert xlsx.data[:2] == b"PK"
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx.data))
    assert "Summary" in wb.sheetnames
    assert "Request detail" in wb.sheetnames
    assert "Leave dates" in wb.sheetnames
    assert len(wb.sheetnames) >= 4
    ws_det = wb["Request detail"]
    assert ws_det.cell(row=1, column=3).value == "leave_type"
    assert ws_det.cell(row=1, column=13).value == "leaves_taken_dates_in_range"

    audit = c.get("/manager/audit.csv")
    assert audit.status_code == 200
    assert b"submitted_ip" in audit.data


def test_dashboard_leave_tracker_renders(client):
    c = client
    _manager_login(c)
    r = c.get("/dashboard?year=2026&month=7")
    assert r.status_code == 200
    assert b"worksheet-wrap" in r.data
    assert b"ws-month-fit" in r.data
    assert b"ws-lead-5" in r.data
    assert b"ws-has-rownum" in r.data
    assert b"ws-rownum-col" in r.data
    assert b"ws-total-col" in r.data
    assert b"ws-eleave-col" in r.data
    assert b"ws-eleave-display" in r.data
    assert b"ws-eleave-input" not in r.data
    assert b"ws-gap-col" in r.data
    assert b"ws-cell-interactive-add" in r.data
    assert b"dashboard-meet-add-dialog" in r.data
    assert b"leave-tracker-meet-quick-leave" in r.data
    assert b"dashboard-all-leaves-name-filter" in r.data
    assert b"ws-weekend-head" in r.data
    assert b"Leave tracker" in r.data
    assert b"leave-tracker-month.xlsx" in r.data
    assert b"Export .xlsx" in r.data
    assert b"CompOFF" in r.data
    assert b"WFH" not in r.data

    xlsx = c.get("/dashboard/leave-tracker-month.xlsx?year=2026&month=7")
    assert xlsx.status_code == 200
    assert xlsx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert xlsx.data[:2] == b"PK"
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx.data))
    assert "July 2026" in wb.sheetnames
    ws = wb["July 2026"]
    assert ws.cell(row=1, column=1).value == "#"
    assert ws.cell(row=1, column=2).value == "July"
    assert ws.cell(row=1, column=3).value == "Total"
    assert ws.cell(row=1, column=4).value == "eLeaveCount"
    assert ws.cell(row=1, column=5).value == "GAP"
    assert ws.cell(row=1, column=6).value == "Wed"
    sat_cols = [col for col in range(6, 6 + 31) if ws.cell(1, col).value == "Sat"]
    sun_cols = [col for col in range(6, 6 + 31) if ws.cell(1, col).value == "Sun"]
    assert sat_cols and sun_cols
    for col in sat_cols:
        assert ws.cell(3, col).fill.fill_type == "solid"
        assert str(ws.cell(3, col).fill.start_color.rgb).upper().endswith("E0E7FF")
    for col in sun_cols:
        assert ws.cell(3, col).fill.fill_type == "solid"
        assert str(ws.cell(3, col).fill.start_color.rgb).upper().endswith("EDE9FE")


def test_dashboard_all_records_filter_survives_records_q_in_url(client):
    _manager_login(client)
    r = client.get("/dashboard?year=2026&month=7&records_q=anas")
    assert r.status_code == 200
    assert b'value="anas"' in r.data


def test_dashboard_leave_tracker_meet_quick_leave_creates_row(client):
    _manager_login(client)
    name = EMPLOYEES[0]
    rv = client.post(
        "/dashboard/api/leave-tracker-meet-quick-leave",
        data={
            "employee_name": name,
            "work_date": "2026-07-11",
            "year": "2026",
            "month": "7",
            "reason": "pl",
            "duration_type": "full",
        },
    )
    assert rv.status_code == 200
    assert rv.get_json() == {"ok": True}
    dash = client.get("/dashboard?year=2026&month=7")
    assert dash.status_code == 200
    assert name.encode("utf-8") in dash.data


def test_compoff_shown_on_leave_grid_but_not_in_leave_day_total(app):
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'compoff', '', '2026-07-01', '2026-07-01', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-07-02', '2026-07-02', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()

    ctx = build_month_context(app, 2026, 7, roster=(name,))
    row = next(r for r in ctx["grid_rows"] if r["employee"] == name)
    assert row["leave_days_total"] == 1.0
    day_cells = list(zip(ctx["days"], row["cells"]))
    c1 = next(c for d, c in day_cells if d["iso"] == "2026-07-01")
    c2 = next(c for d, c in day_cells if d["iso"] == "2026-07-02")
    assert c1 and c1.get("code") == "CO"
    assert c1.get("css", "").startswith("cell-compoff")
    assert c2 and c2.get("code") == "PL"


def test_compoff_does_not_reduce_sprint_capacity_hours(app):
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'compoff', '', '2026-07-14', '2026-07-14', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()
    sd = date(2026, 7, 13)
    ed = date(2026, 7, 17)
    h = _available_hours_for_assignee_sprint_window(app, name, sd, ed)
    assert abs(h - 5 * SCRUM_KANBAN_WEEKDAY_HOURS) < 1e-6


def test_nokia_reason_line_maps_comp_off_to_compoff():
    code, lab = _nokia_reason_and_label_from_line("approved compensatory off 01/07/2026".lower())
    assert code == "compoff"
    assert lab == "CompOFF"


def test_reject_leave_via_reports_clears_meet_leave_day_rows(client, app):
    """Rejecting a request must drop meet_leave_day rows so calendars/capacity cannot still see DSM state."""
    _manager_login(client)
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-06-01', '2026-06-05', 'multi', 'pending', ?, '')
        """,
        (name, ts),
    )
    lid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO meet_leave_day (leave_id, work_date, decision, updated_at)
        VALUES (?, '2026-06-03', 'approved', ?)
        """,
        (lid, ts),
    )
    conn.commit()
    conn.close()

    rv = client.post(
        f"/reports/leave/{lid}/status",
        data={
            "status": "rejected",
            "report_start": "2026-06-01",
            "report_end": "2026-06-30",
        },
        follow_redirects=False,
    )
    assert rv.status_code in (302, 303)

    conn = get_db(app)
    n_day = int(conn.execute("SELECT COUNT(*) AS c FROM meet_leave_day WHERE leave_id = ?", (lid,)).fetchone()["c"])
    st = conn.execute("SELECT status FROM leave_requests WHERE id = ?", (lid,)).fetchone()["status"]
    conn.close()
    assert n_day == 0
    assert st == "rejected"


def test_manager_main_nav_home_links_to_scrum_hub(client):
    c = client
    _manager_login(c)
    r = c.get("/dashboard?year=2026&month=7")
    assert r.status_code == 200
    html = r.data.decode("utf-8", errors="replace")
    start = html.find('<nav class="main-nav"')
    assert start != -1
    end = html.find("</nav>", start)
    assert end != -1
    nav = html[start:end]
    h = nav.find(">Home</a>")
    assert h != -1
    pre = nav[max(0, h - 160) : h]
    assert 'href="/scrum"' in pre


def test_meet_redirects_without_manager(client):
    assert client.get("/meet", follow_redirects=False).status_code == 302


def test_meet_renders_for_manager(client):
    c = client
    _manager_login(c)
    r = c.get("/meet?d=2026-07-10")
    assert r.status_code == 200
    assert b"Leave Plan/DSM Attendance" in r.data
    assert b"meet-table" in r.data


def test_meet_leave_day_approve_remove_promote_and_quick_leave(client, app):
    c = client
    _manager_login(c)
    name = EMPLOYEES[0]
    c.post(
        "/leave",
        data={
            "employee_name": name.lower(),
            "reason": "pl",
            "start_date": "2026-07-09",
            "end_date": "2026-07-11",
            "day_part": "full",
        },
        follow_redirects=False,
    )
    conn = get_db(app)
    leave_id = int(conn.execute("SELECT id FROM leave_requests ORDER BY id DESC LIMIT 1").fetchone()["id"])
    assert conn.execute("SELECT status FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()["status"] == "pending"
    conn.close()

    anchor = "2026-07-10"
    assert c.post(
        "/meet/approve-day",
        data={"leave_id": leave_id, "work_date": "2026-07-09", "anchor": anchor},
    ).get_json() == {"ok": True}

    conn = get_db(app)
    assert (
        conn.execute(
            "SELECT decision FROM meet_leave_day WHERE leave_id = ? AND work_date = ?",
            (leave_id, "2026-07-09"),
        ).fetchone()["decision"]
        == "approved"
    )
    assert conn.execute("SELECT status FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()["status"] == "pending"
    conn.close()

    assert c.post(
        "/meet/approve-day",
        data={"leave_id": leave_id, "work_date": "2026-07-10", "anchor": anchor},
    ).get_json() == {"ok": True}
    assert c.post(
        "/meet/approve-day",
        data={"leave_id": leave_id, "work_date": "2026-07-11", "anchor": anchor},
    ).get_json() == {"ok": True}
    conn = get_db(app)
    assert conn.execute("SELECT status FROM leave_requests WHERE id = ?", (leave_id,)).fetchone()["status"] == "approved"
    conn.close()

    assert c.post(
        "/meet/remove-day",
        data={"leave_id": leave_id, "work_date": "2026-07-10", "anchor": anchor},
    ).get_json() == {"ok": True}

    r = c.post(
        "/meet/quick-leave",
        data={
            "employee_name": name,
            "work_date": "2026-07-10",
            "reason": "pl",
            "duration_type": "full",
            "anchor": anchor,
        },
    )
    assert r.status_code == 200
    assert r.get_json() == {"ok": True}

    bad = c.post(
        "/meet/quick-leave",
        data={
            "employee_name": name,
            "work_date": "2020-01-01",
            "reason": "pl",
            "duration_type": "full",
            "anchor": anchor,
        },
    )
    assert bad.status_code == 400
    assert bad.get_json() == {"ok": False, "error": "out_of_window"}


def test_parse_nokia_grid_tsv_detects_leave():
    hdr = ["Emp ID", "Name", "Country"] + [str(d) for d in range(1, 32)]
    cells = ["99", "Raj, Varshini", "India"] + [""] * 31
    cells[12] = "L"
    text = "\t".join(hdr) + "\n" + "\t".join(cells)
    m, un, err = parse_nokia_grid_tsv(text, 2026, 5)
    assert err is None
    assert m is not None
    assert 10 in m["Varshini Raj"]
    assert not un


def test_parse_nokia_grid_combined_accepts_csv_export():
    """Quoted comma-separated rows (typical Excel CSV) normalize to TSV and parse."""
    import csv
    import io

    hdr = ["Emp ID", "Name", "Country"] + [str(d) for d in range(1, 32)]
    row = ["99", "Raj, Varshini", "India"] + [""] * 31
    row[12] = "L"
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(hdr)
    w.writerow(row)
    m, un, err = parse_nokia_grid_combined(buf.getvalue(), 2026, 5)
    assert err is None
    assert m is not None
    assert 10 in m["Varshini Raj"]
    assert not un


def test_parse_nokia_grid_whitespace_detects_leave():
    """Space-separated grid (OCR-style) after a header row with 1 2 3…"""
    hdr = "Emp Name " + " ".join(str(d) for d in range(1, 32))
    parts = ["99", "Raj,", "Varshini"] + ["."] * 31
    parts[3 + 9] = "L"  # day-of-month 10
    row = " ".join(parts)
    text = hdr + "\n" + row
    m, un, err = parse_nokia_grid_whitespace(text, 2026, 5)
    assert err is None
    assert m is not None
    assert 10 in m["Varshini Raj"]
    assert not un


def test_parse_nokia_employee_summary_leave_dates_may():
    txt = """Leave Entry 5845074 Approved 15-05-2026 15-05-2026
Leave Entry 5845071 Approved 29-05-2026 29-05-2026
"""
    s, err = parse_nokia_employee_summary_leave_dates(txt, 2026, 5)
    assert err is None
    assert s == {15, 29}


def test_parse_nokia_audit_dd_mm_range_accepts_arrow_and_ascii_arrow():
    assert _parse_nokia_audit_dd_mm_range("20-03-2026 → 20-03-2026") == (date(2026, 3, 20), date(2026, 3, 20))
    assert _parse_nokia_audit_dd_mm_range("20-03-2026 -> 20-03-2026") == (date(2026, 3, 20), date(2026, 3, 20))
    assert _parse_nokia_audit_dd_mm_range("20/03/2026 -> 21/03/2026") == (date(2026, 3, 20), date(2026, 3, 21))
    assert _parse_nokia_audit_dd_mm_range("29-05-2026") == (date(2026, 5, 29), date(2026, 5, 29))
    assert _parse_nokia_audit_dd_mm_range("") is None


def test_nokia_compare_dedupes_duplicate_elv_same_day_so_dsm_stays_paired():
    from app import _nokia_audit_build_compare_rows

    dsm = [
        {
            "name": "X",
            "days": "1",
            "type": "A",
            "status": "Approved",
            "dates_range": "20-03-2026 → 20-03-2026",
        }
    ]
    nokia_dup = [
        {
            "name": "X",
            "days": "1",
            "type": "Annual leave",
            "status": "Approved",
            "dates_range": "20-03-2026 → 20-03-2026",
        },
        {
            "name": "X",
            "days": "1",
            "type": "A",
            "status": "Approved",
            "dates_range": "20-03-2026 → 20-03-2026",
        },
    ]
    rows = _nokia_audit_build_compare_rows(nokia_dup, dsm)
    assert len(rows) == 1
    assert rows[0]["days_dsm"] == "1"
    assert rows[0]["dates_dsm"] == "20-03-2026 → 20-03-2026"


def test_nokia_audit_merge_contiguous_dsm_rows_then_compare_single_row():
    """Adjacent single-day DSM rows (same type/status) merge so one Nokia range pairs once."""
    from app import _merge_contiguous_nokia_audit_dsm_leave_rows, _nokia_audit_build_compare_rows

    dsm_raw = [
        {
            "name": "Emp X",
            "days": "1",
            "type": "A",
            "status": "Approved",
            "dates_range": "26-03-2026 → 26-03-2026",
        },
        {
            "name": "Emp X",
            "days": "1",
            "type": "A",
            "status": "Approved",
            "dates_range": "27-03-2026 → 27-03-2026",
        },
    ]
    dsm = _merge_contiguous_nokia_audit_dsm_leave_rows(dsm_raw)
    assert len(dsm) == 1
    assert dsm[0]["dates_range"] == "26-03-2026 → 27-03-2026"
    assert dsm[0]["days"] == "2"

    nokia = [
        {
            "name": "Emp X",
            "days": "2",
            "type": "A",
            "status": "Approved",
            "dates_range": "26-03-2026 → 27-03-2026",
        }
    ]
    rows = _nokia_audit_build_compare_rows(nokia, dsm)
    assert len(rows) == 1
    assert rows[0]["dates_elv"] == "26-03-2026 → 27-03-2026"
    assert rows[0]["dates_dsm"] == "26-03-2026 → 27-03-2026"


def test_nokia_compare_merges_contiguous_elv_fragments_one_dsm_row():
    """When eTool lists two adjacent chips but DSM is one span, Compare shows a single merged row."""
    from app import _nokia_audit_build_compare_rows

    dsm = [
        {
            "name": "X",
            "days": "3",
            "type": "A",
            "status": "Approved",
            "dates_range": "23-03-2026 → 25-03-2026",
        }
    ]
    nokia = [
        {
            "name": "X",
            "days": "1",
            "type": "A",
            "status": "Approved",
            "dates_range": "23-03-2026 → 23-03-2026",
        },
        {
            "name": "X",
            "days": "2",
            "type": "A",
            "status": "Approved",
            "dates_range": "24-03-2026 → 25-03-2026",
        },
    ]
    rows = _nokia_audit_build_compare_rows(nokia, dsm)
    assert len(rows) == 1
    assert rows[0]["dates_elv"] == "23-03-2026 → 25-03-2026"
    assert rows[0]["dates_dsm"] == "23-03-2026 → 25-03-2026"
    assert rows[0]["days_elv"] == "3"


def test_nokia_audit_show_approved_parses_paste_into_preview_table(client):
    _manager_login(client)
    name = EMPLOYEES[0]
    txt = f"Leave Entry 5845071 Approved 29-05-2026 29-05-2026 {name}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "29-05-2026" in html
    assert "Leave dates (From → To)" in html
    assert "No. of Days" in html
    assert "Annual leave" not in html
    assert ">A</td>" in html
    assert "TOTAL" in html
    assert "1.00" in html


def test_nokia_eleavetool_type_display_and_half_day_preview(app):
    assert _nokia_eleavetool_type_display("Annual leave", "1") == "A"
    assert _nokia_eleavetool_type_display("Annual leave", "0.5") == "A1/2"
    assert _nokia_eleavetool_type_display("Annual leave", 0.5) == "A1/2"
    assert _nokia_eleavetool_type_display("Sick leave", "1") == "Sick leave"
    assert _nokia_eleavetool_type_display("PL — Planned leave", "0.5") == "A1/2"
    assert _nokia_eleavetool_type_display("PL — Planned leave", "1") == "A"
    name = EMPLOYEES[0]
    txt = "Leave Entry 1 Approved 0.5 day 15-05-2026 15-05-2026\n"
    rows, err, segs = _nokia_paste_approved_preview_rows(app, txt, name)
    assert err is None
    assert segs is not None
    assert rows[0]["days"] == "0.5"
    assert rows[0]["type"] == "A1/2"


def test_nokia_audit_show_approved_includes_all_dates_not_only_form_month(client):
    """Paste preview lists every parseable row; not limited to year/month hidden fields."""
    _manager_login(client)
    name = EMPLOYEES[0]
    txt = (
        f"Leave Entry 1 Approved 29-05-2026 29-05-2026 {name}\n"
        f"Leave Entry 2 Approved 31-07-2026 31-07-2026 {name}\n"
    )
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "29-05-2026" in html
    assert "31-07-2026" in html
    assert "TOTAL" in html
    assert "2.00" in html


def test_nokia_approved_preview_sorted_chronologically_not_lexicographic_on_dd_mm_yyyy(app):
    """Leave dates column uses dd-mm-yyyy; table rows are ordered by real calendar start date."""
    name = EMPLOYEES[0]
    txt = (
        "Leave Entry 1 Approved 01-06-2026 02-06-2026\n"
        "Leave Entry 2 Approved 04-03-2026 05-03-2026\n"
        "Leave Entry 3 Approved 07-01-2026 09-01-2026\n"
    )
    rows, err, segs = _nokia_paste_approved_preview_rows(app, txt, name)
    assert err is None
    assert segs is not None
    assert [r["dates_range"] for r in rows] == [
        "07-01-2026 → 09-01-2026",
        "04-03-2026 → 05-03-2026",
        "01-06-2026 → 02-06-2026",
    ]


def test_nokia_audit_mark_approved_inserts_and_shows_table(client, app):
    _manager_login(client)
    name = EMPLOYEES[0]
    txt = f"Leave Entry 5845071 Approved 29-05-2026 29-05-2026 {name}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "mark_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "No. of days" in html
    assert "Approved in tracker" in html
    assert "Leave dates (From → To)" in html
    assert "29-05-2026 → 29-05-2026" in html
    assert "TOTAL" in html
    assert "1.00" in html
    conn = get_db(app)
    n = conn.execute(
        """
        SELECT COUNT(*) FROM leave_requests
        WHERE employee_name = ? AND status = 'approved'
          AND start_date = '2026-05-29' AND end_date = '2026-05-29'
        """,
        (name,),
    ).fetchone()[0]
    desc_row = conn.execute(
        """
        SELECT description FROM leave_requests
        WHERE employee_name = ? AND status = 'approved' AND start_date = '2026-05-29' LIMIT 1
        """,
        (name,),
    ).fetchone()
    conn.close()
    assert int(n) >= 1
    assert desc_row is not None
    assert NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in (desc_row["description"] or "")


def test_dashboard_month_grid_shows_green_a_for_nokia_marked_leave(client):
    _manager_login(client)
    name = EMPLOYEES[0]
    txt = f"Leave Entry 5845071 Approved 29-05-2026 29-05-2026 {name}\n"
    assert (
        client.post(
            "/worksheet/nokia-audit",
            data={
                "year": "2026",
                "month": "5",
                "employee_name": name,
                "nokia_ocr_text": txt,
                "audit_action": "mark_approved",
            },
        ).status_code
        == 200
    )
    dash = client.get("/dashboard?year=2026&month=5")
    assert dash.status_code == 200
    html = dash.get_data(as_text=True)
    assert "cell-nokia-approved" in html
    assert re.search(r"cell-nokia-approved[^>]*>\s*A\s*<", html)
    assert f'data-employee="{name}"' in html
    assert re.search(
        r'data-employee="' + re.escape(name) + r'"[^>]*data-eleave-days="1\.00"',
        html,
    )


def test_nokia_mark_appends_marker_when_day_already_on_tracker(client, app):
    """If the day already has approved leave without the Nokia marker, Mark Approved tags it so the grid shows green A."""
    _manager_login(client)
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-20', '2026-05-20', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()
    txt = f"Leave Entry 999 Approved 20-05-2026 20-05-2026 {name}\n"
    assert (
        client.post(
            "/worksheet/nokia-audit",
            data={
                "year": "2026",
                "month": "5",
                "employee_name": name,
                "nokia_ocr_text": txt,
                "audit_action": "mark_approved",
            },
        ).status_code
        == 200
    )
    conn = get_db(app)
    n_rows = int(
        conn.execute(
            "SELECT COUNT(*) AS c FROM leave_requests WHERE employee_name = ? AND start_date = '2026-05-20'",
            (name,),
        ).fetchone()["c"]
    )
    desc_row = conn.execute(
        "SELECT description FROM leave_requests WHERE employee_name = ? AND start_date = '2026-05-20' LIMIT 1",
        (name,),
    ).fetchone()
    conn.close()
    assert n_rows == 1
    assert desc_row is not None
    assert NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER in (desc_row["description"] or "")
    dash = client.get("/dashboard?year=2026&month=5")
    assert dash.status_code == 200
    html = dash.get_data(as_text=True)
    assert "cell-nokia-approved" in html
    assert re.search(r"cell-nokia-approved[^>]*>\s*A\s*<", html)
    assert re.search(
        r'data-employee="' + re.escape(name) + r'"[^>]*data-eleave-days="1\.00"',
        html,
    )


def test_dashboard_month_prefers_nokia_tagged_row_on_same_day_overlap(client, app):
    """When two approved rows overlap one day, the Nokia-tagged row must win display even if its id is lower."""
    _manager_login(client)
    name = EMPLOYEES[3]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    marker = NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', ?, '2026-05-14', '2026-05-14', 'full', 'approved', ?, '')
        """,
        (name, marker, ts),
    )
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-14', '2026-05-14', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()
    dash = client.get("/dashboard?year=2026&month=5")
    assert dash.status_code == 200
    html = dash.get_data(as_text=True)
    assert "cell-nokia-approved" in html
    assert re.search(r"cell-nokia-approved[^>]*>\s*A\s*<", html)
    assert re.search(
        r'data-employee="' + re.escape(name) + r'"[^>]*data-eleave-days="1\.00"',
        html,
    )


def test_nokia_audit_requires_manager(client):
    assert client.get("/worksheet/nokia-audit", follow_redirects=False).status_code == 302


def test_nokia_audit_paste_suffix_name_mismatch_blocks_show_approved(client):
    _manager_login(client)
    sel = EMPLOYEES[0]
    wrong = EMPLOYEES[1]
    txt = f"Leave Entry 1 Approved 29-05-2026 29-05-2026 Approver {wrong}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": sel,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "does not match" in html
    assert wrong in html
    assert sel in html
    assert "No. of Days" not in html


def test_nokia_audit_paste_suffix_name_mismatch_blocks_mark_approved(client, app):
    _manager_login(client)
    sel = EMPLOYEES[0]
    wrong = EMPLOYEES[2]
    txt = f"Leave Entry 1 Approved 29-05-2026 29-05-2026 {wrong}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": sel,
            "nokia_ocr_text": txt,
            "audit_action": "mark_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "does not match" in html
    conn = get_db(app)
    n = int(
        conn.execute(
            "SELECT COUNT(*) FROM leave_requests WHERE employee_name = ? AND start_date = '2026-05-29'",
            (sel,),
        ).fetchone()[0]
    )
    conn.close()
    assert n == 0


def test_nokia_audit_paste_suffix_matches_selection_allows_mark_approved(client, app):
    _manager_login(client)
    sel = EMPLOYEES[0]
    txt = f"Leave Entry 1 Approved 29-05-2026 29-05-2026 {sel}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": sel,
            "nokia_ocr_text": txt,
            "audit_action": "mark_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "does not match" not in html
    conn = get_db(app)
    n = int(
        conn.execute(
            "SELECT COUNT(*) FROM leave_requests WHERE employee_name = ? AND start_date = '2026-05-29'",
            (sel,),
        ).fetchone()[0]
    )
    conn.close()
    assert n >= 1


def test_nokia_audit_mark_approved_blocks_without_detectable_employee_in_paste(client, app):
    """Approved+date rows must include a roster name (suffix or TSV last column) before marking."""
    _manager_login(client)
    sel = EMPLOYEES[0]
    txt = "Leave Entry 5845071 Approved 29-05-2026 29-05-2026\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": sel,
            "nokia_ocr_text": txt,
            "audit_action": "mark_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Could not find a roster employee name" in html
    conn = get_db(app)
    n = int(
        conn.execute(
            "SELECT COUNT(*) FROM leave_requests WHERE employee_name = ? AND start_date = '2026-05-29'",
            (sel,),
        ).fetchone()[0]
    )
    conn.close()
    assert n == 0


def test_nokia_paste_require_roster_name_in_paste_helper():
    from app import EMPLOYEES, _nokia_paste_employee_must_match_selected_or_error

    name = EMPLOYEES[0]
    txt = "Leave Entry 5845071 Approved 29-05-2026 29-05-2026\n"
    assert _nokia_paste_employee_must_match_selected_or_error(name, txt, EMPLOYEES) is None
    err = _nokia_paste_employee_must_match_selected_or_error(
        name, txt, EMPLOYEES, require_roster_name_in_paste=True
    )
    assert err is not None
    assert "Could not find a roster employee name" in err
    assert (
        _nokia_paste_employee_must_match_selected_or_error(
            name, f"{txt.strip()} {name}\n", EMPLOYEES, require_roster_name_in_paste=True
        )
        is None
    )


def test_nokia_tsv_employee_last_column_used_for_name_match_and_segment_parses():
    """Tab-separated eLeave rows: employee is the last column (manager may appear before)."""
    from app import (
        EMPLOYEES,
        _nokia_paste_employee_must_match_selected_or_error,
        _nokia_paste_trailing_roster_name_hints,
        _nokia_segment_tuples_from_filtered_approved_lines,
    )

    roster = list(EMPLOYEES)
    sh = "Shaishta Anjum"
    mgr = "Gajendra Singh Thakur"
    line = f"5780137\tApproved\t\t24-03-2026\t25-03-2026\t-2\t\t{mgr}\t{sh}\n"
    hints = _nokia_paste_trailing_roster_name_hints(line, roster)
    assert hints == {sh}, hints
    assert _nokia_paste_employee_must_match_selected_or_error(sh, line, roster) is None
    assert _nokia_paste_employee_must_match_selected_or_error(mgr, line, roster) is not None
    filtered = "\n".join(L for L in line.splitlines() if "approved" in L.lower())
    segs = _nokia_segment_tuples_from_filtered_approved_lines(filtered.strip())
    assert len(segs) == 1
    assert segs[0][0] == date(2026, 3, 24) and segs[0][1] == date(2026, 3, 25)


def test_nokia_audit_paste_multiple_suffix_names_blocks(client):
    _manager_login(client)
    a, b = EMPLOYEES[0], EMPLOYEES[1]
    txt = (
        f"Leave Entry 1 Approved 29-05-2026 29-05-2026 {a}\n"
        f"Leave Entry 2 Approved 30-05-2026 30-05-2026 {b}\n"
    )
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": a,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "more than one employee name" in html


def test_nokia_mark_approved_three_consecutive_weekdays_one_multi_row(client, app):
    """Multi-day Nokia range that is Wed–Fri inserts one contiguous approved row (no Sat/Sun in range)."""
    _manager_login(client)
    name = EMPLOYEES[4]
    d0 = date(2026, 5, 6)
    d2 = d0 + timedelta(days=2)
    assert d0.weekday() == 2 and d2.weekday() == 4
    txt = f"5766761 Approved {d0.strftime('%d-%m-%Y')} {d2.strftime('%d-%m-%Y')} -3  Ritesh Kumar {name}\n"
    assert (
        client.post(
            "/worksheet/nokia-audit",
            data={
                "year": "2026",
                "month": "5",
                "employee_name": name,
                "nokia_ocr_text": txt,
                "audit_action": "mark_approved",
            },
        ).status_code
        == 200
    )
    conn = get_db(app)
    rows = conn.execute(
        """
        SELECT start_date, end_date, duration_type FROM leave_requests
        WHERE employee_name = ? AND description LIKE ?
        ORDER BY id ASC
        """,
        (name, "%" + NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER + "%"),
    ).fetchall()
    conn.close()
    assert len(rows) == 1
    assert str(rows[0]["start_date"]).startswith("2026-05-06")
    assert str(rows[0]["end_date"]).startswith("2026-05-08")
    assert rows[0]["duration_type"] == "multi"


def test_nokia_mark_approved_spanning_weekend_skips_sat_sun(client, app):
    """Fri–Tue range: only Fri, Mon, Tue get tracker rows (Sat/Sun omitted)."""
    _manager_login(client)
    name = EMPLOYEES[5]
    fri = date(2026, 5, 8)
    tue = date(2026, 5, 12)
    assert fri.weekday() == 4 and tue.weekday() == 1
    txt = f"5766761 Approved {fri.strftime('%d-%m-%Y')} {tue.strftime('%d-%m-%Y')} -5  Ritesh Kumar {name}\n"
    assert (
        client.post(
            "/worksheet/nokia-audit",
            data={
                "year": "2026",
                "month": "5",
                "employee_name": name,
                "nokia_ocr_text": txt,
                "audit_action": "mark_approved",
            },
        ).status_code
        == 200
    )
    conn = get_db(app)
    rows = list(
        conn.execute(
            """
            SELECT start_date, end_date FROM leave_requests
            WHERE employee_name = ? AND description LIKE ?
            ORDER BY start_date ASC
            """,
            (name, "%" + NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER + "%"),
        )
    )
    conn.close()
    assert len(rows) == 2
    spans = {(str(r["start_date"])[:10], str(r["end_date"])[:10]) for r in rows}
    assert ("2026-05-08", "2026-05-08") in spans
    assert ("2026-05-11", "2026-05-12") in spans


def test_nokia_mark_approved_single_weekend_day_inserts_nothing(client, app):
    """A Nokia line that is only a Saturday (or Sunday) must not create tracker rows (no weekend marking)."""
    _manager_login(client)
    name = EMPLOYEES[6]
    sat = date(2026, 5, 30)
    assert sat.weekday() == 5
    txt = f"5766761 Approved {sat.strftime('%d-%m-%Y')} {sat.strftime('%d-%m-%Y')} -1  Ritesh Kumar {name}\n"
    assert (
        client.post(
            "/worksheet/nokia-audit",
            data={
                "year": "2026",
                "month": "5",
                "employee_name": name,
                "nokia_ocr_text": txt,
                "audit_action": "mark_approved",
            },
        ).status_code
        == 200
    )
    conn = get_db(app)
    n = int(
        conn.execute(
            """
            SELECT COUNT(*) FROM leave_requests
            WHERE employee_name = ? AND description LIKE ?
            """,
            (name, "%" + NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER + "%"),
        ).fetchone()[0]
    )
    conn.close()
    assert n == 0


def test_nokia_audit_show_approved_preview_uses_weekday_count_for_range(client):
    """Preview 'No. of Days' for a multi-day Nokia row counts Mon–Fri only."""
    _manager_login(client)
    name = EMPLOYEES[6]
    fri = date(2026, 5, 8)
    mon = date(2026, 5, 11)
    txt = f"5766761 Approved {fri.strftime('%d-%m-%Y')} {mon.strftime('%d-%m-%Y')} -4  Ritesh Kumar {name}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert re.search(r"<strong>\s*2\.00\s*</strong>", html)
    assert "4.00" not in html


def test_nokia_show_then_mark_replays_same_plan_as_preview(client, app):
    """After Show Approved, Mark uses stored segments so tracker rows match the preview table."""
    _manager_login(client)
    name = EMPLOYEES[2]
    txt = f"Leave Entry 1 Approved 10-06-2026 12-06-2026 Approver {name}\n"
    common = {
        "year": "2026",
        "month": "6",
        "employee_name": name,
        "nokia_ocr_text": txt,
    }
    show = client.post("/worksheet/nokia-audit", data={**common, "audit_action": "show_approved"})
    assert show.status_code == 200
    prev_html = show.get_data(as_text=True)
    assert "Leave dates (From → To)" in prev_html
    mark = client.post("/worksheet/nokia-audit", data={**common, "audit_action": "mark_approved"})
    assert mark.status_code == 200
    conn = get_db(app)
    n = int(
        conn.execute(
            "SELECT COUNT(*) FROM leave_requests WHERE employee_name = ? AND description LIKE ?",
            (name, "%" + NOKIA_AUDIT_LEAVE_DESCRIPTION_MARKER + "%"),
        ).fetchone()[0]
    )
    conn.close()
    assert n >= 1


def test_nokia_audit_show_approved_preview_lists_each_line_even_same_calendar_day(client):
    """Preview lists one row per Nokia approved line; overlapping days are not collapsed in the table."""
    _manager_login(client)
    name = EMPLOYEES[0]
    txt = (
        f"Leave Entry 1 Approved 15-05-2026 15-05-2026 {name}\n"
        f"Leave Entry 2 Approved 15-05-2026 15-05-2026 {name}\n"
    )
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert html.count("15-05-2026 → 15-05-2026") == 2
    assert re.search(r"<strong>\s*2\.00\s*</strong>", html)


def test_nokia_reason_maps_medical_to_sick_leave():
    low = "5766761 approved medical 10-06-2026 10-06-2026"
    code, _lab = _nokia_reason_and_label_from_line(low)
    assert code == "sl"


def test_nokia_audit_year_month_query_and_month_names(client):
    c = client
    _manager_login(c)
    r = c.get("/worksheet/nokia-audit?year=2024&month=3")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert 'name="year"' in html and 'value="2024"' in html
    assert 'name="month"' in html and 'value="3"' in html


def test_team_roster_csv_upload_and_switch_worksheet(client, app):
    c = client
    _manager_login(c)
    csv_body = "Team,Name\nGamma,Only One\nGamma,Second Person\n"
    up = c.post(
        "/reports/roster-upload",
        data={"roster_csv": (io.BytesIO(csv_body.encode("utf-8")), "rosters.csv")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert up.status_code == 200
    assert b"Roster updated" in up.data
    with app.app_context():
        conn = get_db(app)
        row = conn.execute("SELECT id FROM teams WHERE name = ?", ("Gamma",)).fetchone()
        conn.close()
    assert row
    gid = int(row["id"])
    sw = c.post(
        "/reports/team-select",
        data={"team_id": str(gid), "next": "/dashboard?year=2026&month=5"},
        follow_redirects=False,
    )
    assert sw.status_code in (302, 303)
    dash = c.get("/dashboard?year=2026&month=5")
    assert dash.status_code == 200
    assert b"Only One" in dash.data
    assert b"Second Person" in dash.data


def test_team_roster_xlsx_upload(client, app):
    from openpyxl import Workbook

    c = client
    _manager_login(c)
    bio = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["TeamName", "EmployeeName"])
    ws.append(["XlsxOnly", "Only Xlsx Person"])
    wb.save(bio)
    bio.seek(0)
    up = c.post(
        "/reports/roster-upload",
        data={"roster_csv": (bio, "roster.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert up.status_code == 200
    assert b"Roster updated" in up.data
    conn = get_db(app)
    row = conn.execute(
        "SELECT tr.employee_name FROM team_roster tr JOIN teams t ON t.id = tr.team_id WHERE t.name = ?",
        ("XlsxOnly",),
    ).fetchone()
    conn.close()
    assert row and row[0] == "Only Xlsx Person"


def test_reports_roster_export_xlsx_current_team(client, app):
    c = client
    _manager_login(c)
    csv_body = "TeamName,EmployeeName\nExportTestZ,Person A\nExportTestZ,Person B\n"
    c.post(
        "/reports/roster-upload",
        data={"roster_csv": (io.BytesIO(csv_body.encode("utf-8")), "rosters.csv")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    conn = get_db(app)
    row = conn.execute("SELECT id FROM teams WHERE name = ?", ("ExportTestZ",)).fetchone()
    conn.close()
    assert row
    gid = int(row["id"])
    c.post(
        "/reports/team-select",
        data={"team_id": str(gid), "next": "/reports"},
        follow_redirects=True,
    )
    exp = c.get("/reports/roster-export.xlsx")
    assert exp.status_code == 200
    ct = exp.headers.get("Content-Type", "")
    assert "spreadsheetml" in ct or "officedocument" in ct
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(exp.data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = [tuple(c for c in r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    assert rows[0] == ("TeamName", "EmployeeName")
    body_rows = [r for r in rows[1:] if any(x for x in r if x)]
    assert ("ExportTestZ", "Person A") in body_rows
    assert ("ExportTestZ", "Person B") in body_rows


def test_default_team_shows_union_of_all_rosters(client, app):
    c = client
    _manager_login(c)
    csv_body = "TeamName,EmployeeName\nGammaUnion,Gam Person Only\n"
    c.post(
        "/reports/roster-upload",
        data={"roster_csv": (io.BytesIO(csv_body.encode("utf-8")), "r.csv")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    conn = get_db(app)
    default_id = int(conn.execute("SELECT id FROM teams WHERE lower(trim(name)) = 'default'").fetchone()["id"])
    conn.close()
    c.post(
        "/reports/team-select",
        data={"team_id": str(default_id), "next": "/dashboard?year=2026&month=1"},
        follow_redirects=True,
    )
    dash = c.get("/dashboard?year=2026&month=1")
    assert dash.status_code == 200
    html = dash.get_data(as_text=True)
    assert "Gam Person Only" in html
    assert EMPLOYEES[0] in html


def test_nokia_audit_show_approved_empty_or_table(client):
    """Show Approved Leaves lists approved rows or empty copy for that month."""
    c = client
    _manager_login(c)
    name = EMPLOYEES[0]
    r = c.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "audit_action": "show_approved",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Aprooved eLeave in Tool" in html
    assert "Approved leave list" not in html
    assert "No approved leave records for this person" not in html
    assert "Paste Nokia eLeave text first" in html
    assert "<th>No. of Days</th>" not in html


def test_nokia_audit_show_dsm_lists_tracker_with_grid_leave_codes(client, app):
    """DSM table lists overlapping leave for the full calendar year (not only the form month)."""
    _manager_login(client)
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-21', '2026-05-21', 'half_am', 'pending', ?, '')
        """,
        (name, ts),
    )
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'ul', '', '2026-06-10', '2026-06-10', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()
    r = client.post(
        "/worksheet/nokia-audit",
        data={"year": "2026", "month": "5", "employee_name": name, "audit_action": "show_dsm"},
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "nokia-dsm-leaves-table" in html
    assert "Leave tracker — DSM leaves" in html
    assert "calendar year 2026" in html
    assert "Pending" in html
    assert "Approved" in html
    assert "PL" in html
    assert "UL" in html
    assert "21-05-2026" in html
    assert "10-06-2026" in html


def test_nokia_audit_show_dsm_blocked_when_paste_employee_mismatch(client, app):
    """Show DSM Leaves skips the table if pasted Nokia approved rows name someone other than the dropdown."""
    _manager_login(client)
    sel = EMPLOYEES[0]
    wrong = EMPLOYEES[2]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-21', '2026-05-21', 'full', 'approved', ?, '')
        """,
        (sel, ts),
    )
    conn.commit()
    conn.close()
    txt = f"Leave Entry 1 Approved 29-05-2026 29-05-2026 {wrong}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": sel,
            "nokia_ocr_text": txt,
            "audit_action": "show_dsm",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "does not match" in html
    assert "nokia-dsm-leaves-table" not in html


def test_nokia_audit_show_dsm_shows_table_when_paste_matches_selection(client, app):
    _manager_login(client)
    sel = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-21', '2026-05-21', 'full', 'approved', ?, '')
        """,
        (sel, ts),
    )
    conn.commit()
    conn.close()
    txt = f"Leave Entry 1 Approved 29-05-2026 29-05-2026 {sel}\n"
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": sel,
            "nokia_ocr_text": txt,
            "audit_action": "show_dsm",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "does not match" not in html
    assert "nokia-dsm-leaves-table" in html
    assert "21-05-2026" in html


def test_nokia_audit_compare_merges_elv_and_dsm_types(client, app):
    _manager_login(client)
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-29', '2026-05-29', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()
    txt = f"Leave Entry 5845071 Approved 29-05-2026 29-05-2026 {name}\n"
    client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "audit_action": "show_dsm",
        },
    )
    r = client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "audit_action": "compare",
        },
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "nokia-compare-table" in html
    assert "eLeavetool Type" in html
    assert "Type of leave DSM" in html
    assert "PL" in html
    assert "Annual leave" not in html
    assert ">A</td>" in html
    assert "29-05-2026" in html
    assert "nokia-audit-compare.xlsx" in html


def test_nokia_audit_compare_xlsx_download(client, app):
    _manager_login(client)
    name = EMPLOYEES[0]
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn = get_db(app)
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-05-29', '2026-05-29', 'full', 'approved', ?, '')
        """,
        (name, ts),
    )
    conn.commit()
    conn.close()
    txt = f"Leave Entry 5845071 Approved 29-05-2026 29-05-2026 {name}\n"
    client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "nokia_ocr_text": txt,
            "audit_action": "show_approved",
        },
    )
    client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "audit_action": "show_dsm",
        },
    )
    client.post(
        "/worksheet/nokia-audit",
        data={
            "year": "2026",
            "month": "5",
            "employee_name": name,
            "audit_action": "compare",
        },
    )
    rx = client.get(
        "/worksheet/nokia-audit-compare.xlsx",
        query_string={"year": "2026", "month": "5", "employee_name": name},
    )
    assert rx.status_code == 200
    assert rx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert rx.data[:4] == b"PK\x03\x04"


def test_export_forbidden_without_manager(client):
    assert client.get("/reports/export.csv").status_code == 302
    assert client.get("/reports/export-leaves.xlsx").status_code == 302
    assert client.get("/reports/roster-export.xlsx").status_code == 302
    assert client.get("/dashboard/leave-tracker-month.xlsx").status_code == 302
    assert client.get("/reports/roster-export.csv", follow_redirects=False).status_code == 301


def test_scrum_sprint_default_two_week_end_inclusive():
    from datetime import date

    from app import scrum_sprint_default_end_date

    assert scrum_sprint_default_end_date(date(2026, 6, 5)) == date(2026, 6, 18)
    assert scrum_sprint_default_end_date(date(2026, 9, 1)) == date(2026, 9, 14)


def test_suggest_next_sprint_name_from_previous():
    from app import suggest_next_sprint_name_from_previous

    assert suggest_next_sprint_name_from_previous("FRONTIER-FB2612") == "FRONTIER-FB2613"
    assert suggest_next_sprint_name_from_previous("Sprint 12") == "Sprint 13"
    assert suggest_next_sprint_name_from_previous("  v2.008 ") == "v2.009"
    assert suggest_next_sprint_name_from_previous("id099") == "id100"
    assert suggest_next_sprint_name_from_previous("Alpha") is None
    assert suggest_next_sprint_name_from_previous("") is None


def test_scrm_requires_manager(client):
    assert client.get("/scrum", follow_redirects=False).status_code == 302


def test_scrm_renders(client):
    c = client
    _manager_login(c)
    r = c.get("/scrum")
    assert r.status_code == 200
    assert b"Sprint hub" in r.data
    assert b"14 calendar days" in r.data or b"two-week" in r.data


def test_scrum_hub_create_next_prefills_incremented_sprint_name(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={"name": "FRONTIER-FB2612", "start_date": "2026-06-01"},
        follow_redirects=True,
    )
    html = c.get("/scrum").get_data(as_text=True)
    assert 'id="sp-next-name"' in html
    assert 'value="FRONTIER-FB2613"' in html


def test_scrum_hub_no_bottom_leave_settings_buttons(client):
    """Sprint hub used to duplicate Leave tracker / Settings at the bottom; main nav still has them."""
    c = client
    _manager_login(c)
    html = c.get("/scrum").get_data(as_text=True)
    assert "Create sprint" in html
    assert "scrum-hub-create-next" not in html  # no sprints yet
    assert 'class="btn secondary" href="/dashboard"' not in html
    assert 'class="btn secondary" href="/reports"' not in html


def test_scrum_sprint_create_next_requires_prior_sprint(client, app):
    c = client
    _manager_login(c)
    r = c.post(
        "/scrum/sprint/create",
        data={"name": "Should Not Exist", "create_next": "1"},
        follow_redirects=True,
    )
    assert r.status_code == 200
    assert b"first sprint" in r.data or b"no previous sprint" in r.data
    conn = get_db(app)
    n = int(
        conn.execute(
            "SELECT COUNT(*) AS c FROM scrum_sprint WHERE name = ?", ("Should Not Exist",)
        ).fetchone()["c"]
    )
    conn.close()
    assert n == 0


def test_scrum_sprint_create_next_chains_start_and_carries_doing(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={"name": "Prior Sprint Next", "start_date": "2026-09-01"},
        follow_redirects=True,
    )
    conn = get_db(app)
    sid1 = int(
        conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Prior Sprint Next",)).fetchone()["id"]
    )
    conn.close()
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid1),
            "title": "Carry Me",
            "assignee": EMPLOYEES[0],
            "estimate_hours": "5",
            "kanban_column": "doing",
            "task_kind": "ndy",
        },
        follow_redirects=True,
    )
    hub = c.get("/scrum")
    assert hub.status_code == 200
    assert b"Create next sprint" in hub.data
    assert b"2026-09-15" in hub.data
    c.post(
        "/scrum/sprint/create",
        data={"name": "Chained Sprint", "create_next": "1"},
        follow_redirects=True,
    )
    conn = get_db(app)
    row = conn.execute(
        "SELECT id, start_date, end_date FROM scrum_sprint WHERE name = ?", ("Chained Sprint",)
    ).fetchone()
    assert row
    assert row["start_date"][:10] == "2026-09-15"
    assert row["end_date"][:10] == "2026-09-28"
    sid2 = int(row["id"])
    carried = conn.execute(
        "SELECT title, kanban_column FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
        (sid2, "Carry Me"),
    ).fetchone()
    conn.close()
    assert carried
    assert str(carried["kanban_column"]).lower().strip() == "backlog"


def test_scrum_sprint_rename_from_hub(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Rename Me Sprint",
            "start_date": "2026-09-01",
            "end_date": "2026-09-10",
            "goal": "",
        },
        follow_redirects=True,
    )
    hub = c.get("/scrum")
    assert hub.status_code == 200
    assert b"scrum-sprint-rename-form" in hub.data
    assert b"/scrum/sprint/rename" in hub.data
    assert b"SprintLeaveView" in hub.data
    assert b"/scrum/sprint/delete" in hub.data
    assert b"scrum-hub-sprint-delete-form" in hub.data
    assert b"Delete sprint" in hub.data
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Rename Me Sprint",)).fetchone()["id"])
    conn.close()
    assert f"/scrum/sprint/{sid}/leave-tracker".encode() in hub.data
    c.post(
        "/scrum/sprint/rename",
        data={"sprint_id": str(sid), "name": "FB0052"},
        follow_redirects=True,
    )
    conn = get_db(app)
    nm = conn.execute("SELECT name FROM scrum_sprint WHERE id = ?", (sid,)).fetchone()["name"]
    conn.close()
    assert nm == "FB0052"


def test_scrum_past_sprint_end_blocks_kanban_move_api(client, app, monkeypatch):
    """After the sprint window ends, the sprint is auto-closed and board mutations return ``sprint_closed``."""

    monkeypatch.setattr(
        "app._sprint_clock_utc",
        lambda: datetime(2026, 6, 10, 12, 0, 0, tzinfo=timezone.utc),
    )

    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={"name": "Frozen Sprint", "start_date": "2026-06-01", "goal": ""},
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Frozen Sprint",)).fetchone()["id"])
    conn.close()
    assignee = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "Task A",
            "assignee": assignee,
            "estimate_hours": "5",
            "kanban_column": "do",
            "task_kind": "ndy",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    iid = int(
        conn.execute("SELECT id FROM scrum_sprint_item WHERE sprint_id = ? LIMIT 1", (sid,)).fetchone()["id"]
    )
    conn.close()

    monkeypatch.setattr(
        "app._sprint_clock_utc",
        lambda: datetime(2026, 7, 1, 0, 0, 0, tzinfo=timezone.utc),
    )

    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": assignee,
            "to_column": "doing",
            "note": "start",
        },
    )
    assert r.status_code == 403
    assert r.get_json()["error"] == "sprint_closed"


def test_scrum_rename_blocked_after_sprint_window_ends(client, app, monkeypatch):
    monkeypatch.setattr(
        "app._sprint_clock_utc",
        lambda: datetime(2026, 6, 10, 12, 0, 0, tzinfo=timezone.utc),
    )
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={"name": "Old Sprint Name", "start_date": "2026-06-01", "goal": ""},
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Old Sprint Name",)).fetchone()["id"])
    conn.close()

    monkeypatch.setattr(
        "app._sprint_clock_utc",
        lambda: datetime(2026, 7, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    c.post("/scrum/sprint/rename", data={"sprint_id": str(sid), "name": "Try Rename"}, follow_redirects=True)
    conn = get_db(app)
    nm = conn.execute("SELECT name FROM scrum_sprint WHERE id = ?", (sid,)).fetchone()["name"]
    conn.close()
    assert nm == "Old Sprint Name"


def test_scrum_sprint_auto_close_sets_is_closed(client, app, monkeypatch):
    monkeypatch.setattr(
        "app._sprint_clock_utc",
        lambda: datetime(2026, 6, 10, 12, 0, 0, tzinfo=timezone.utc),
    )
    c = client
    _manager_login(c)
    sprint_name = "AutoClose Sprint " + uuid.uuid4().hex[:12]
    resp = c.post(
        "/scrum/sprint/create",
        data={"name": sprint_name, "start_date": "2026-06-01", "goal": ""},
        follow_redirects=False,
    )
    assert resp.status_code in (302, 303)
    loc = resp.headers.get("Location") or ""
    m = re.search(r"/scrum/sprint/(\d+)", loc)
    assert m, (loc, sprint_name)
    sid = int(m.group(1))
    conn = get_db(app)
    assert int(conn.execute("SELECT COALESCE(is_closed,0) FROM scrum_sprint WHERE id = ?", (sid,)).fetchone()[0]) == 0
    conn.close()
    monkeypatch.setattr(
        "app._sprint_clock_utc",
        lambda: datetime(2026, 7, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    r = c.get(f"/scrum/sprint/{sid}")
    assert r.status_code == 200
    conn = get_db(app)
    assert int(conn.execute("SELECT COALESCE(is_closed,0) FROM scrum_sprint WHERE id = ?", (sid,)).fetchone()[0]) == 1
    conn.close()


def test_scrum_sprint_rename_rejects_duplicate(client, app):
    c = client
    _manager_login(c)
    for nm in ("Sprint Alpha", "Sprint Beta"):
        c.post(
            "/scrum/sprint/create",
            data={
                "name": nm,
                "start_date": "2026-08-01",
                "end_date": "2026-08-14",
                "goal": "",
            },
            follow_redirects=True,
        )
    conn = get_db(app)
    sid_beta = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Sprint Beta",)).fetchone()["id"])
    conn.close()
    c.post(
        "/scrum/sprint/rename",
        data={"sprint_id": str(sid_beta), "name": "Sprint Alpha"},
        follow_redirects=True,
    )
    conn = get_db(app)
    row = conn.execute("SELECT name FROM scrum_sprint WHERE id = ?", (sid_beta,)).fetchone()
    conn.close()
    assert row["name"] == "Sprint Beta"


def test_scrum_sprint_team_capacity_api_and_create_persists(client, app):
    c = client
    _manager_login(c)
    r = c.get("/scrum/api/sprint-team-capacity", query_string={"start_date": "2026-06-09", "end_date": "2026-06-09"})
    assert r.status_code == 200
    j = r.get_json()
    assert j["ok"] is True
    n = len(EMPLOYEES)
    assert abs(float(j["hours"]) - 8.0 * n) < 0.05

    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Cap Store Sprint",
            "start_date": "2026-06-09",
            "end_date": "2026-06-09",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    row = conn.execute(
        "SELECT team_capacity_hours FROM scrum_sprint WHERE name = ?", ("Cap Store Sprint",)
    ).fetchone()
    conn.close()
    assert row
    assert row["team_capacity_hours"] is not None
    cap_r = c.get(
        "/scrum/api/sprint-team-capacity",
        query_string={"start_date": "2026-06-09", "end_date": "2026-06-22"},
    )
    assert cap_r.status_code == 200
    expected = float(cap_r.get_json()["hours"])
    assert abs(float(row["team_capacity_hours"]) - expected) < 0.05


def test_scrum_sprint_team_delete_ui_has_modal_and_delete_post_works(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Delete Modal Sprint",
            "start_date": "2026-10-01",
            "end_date": "2026-10-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(
        conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Delete Modal Sprint",)).fetchone()["id"]
    )
    conn.close()
    r = c.get("/scrum/sprint/%d" % sid)
    assert r.status_code == 200
    assert b'Sprint capacity' in r.data
    assert ("/scrum/sprint/%d/team-detailed" % sid).encode() in r.data
    assert b'id="scrum-delete-sprint-dialog"' in r.data
    assert b'id="scrum-delete-sprint-open"' in r.data
    assert b'id="scrum-sprint-delete-form"' in r.data
    assert b"Delete permanently" in r.data
    del_r = c.post("/scrum/sprint/delete", data={"sprint_id": str(sid)}, follow_redirects=True)
    assert del_r.status_code == 200
    conn = get_db(app)
    gone = conn.execute("SELECT id FROM scrum_sprint WHERE id = ?", (sid,)).fetchone() is None
    conn.close()
    assert gone


def test_scrum_attachment_build_preview_html_xlsx(tmp_path):
    from openpyxl import Workbook

    from app import _scrum_attachment_build_preview_html

    p = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["ColA", "ColB"])
    ws.append([1, 2])
    wb.save(p)
    wb.close()
    html = _scrum_attachment_build_preview_html(p, "sample.xlsx")
    assert html
    assert "<table>" in html
    assert "ColA" in html


def test_scrum_attachment_build_preview_html_csv(tmp_path):
    from app import _scrum_attachment_build_preview_html

    p = tmp_path / "t.csv"
    p.write_text("h1,h2\n3,4\n", encoding="utf-8")
    html = _scrum_attachment_build_preview_html(p, "t.csv")
    assert html and "h1" in html and "3" in html


def test_scrum_sprint_team_detailed_renders_summary_and_stickies(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "DetailedOverviewSp",
            "start_date": "2026-08-01",
            "end_date": "2026-08-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(
        conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("DetailedOverviewSp",)).fetchone()["id"]
    )
    emp = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "DetailOverviewSticky",
            "assignee": emp,
            "estimate_hours": "1",
            "kanban_column": "do",
            "task_kind": "ndy",
            "notes": "",
        },
        follow_redirects=True,
    )
    conn.close()
    r = c.get("/scrum/sprint/%d/team-detailed" % sid)
    assert r.status_code == 200
    assert b"scrum-team-detail-summary-line" in r.data
    assert b"scrum-team-detail-member" in r.data
    assert b"scrum-team-detail-sticky-attachments" in r.data
    assert b'name="return_to"' in r.data
    assert b"team_detailed" in r.data
    assert b"scrum-team-detail-member-filter" in r.data
    assert b'show tasks for' in r.data.lower()
    assert b"preview-html" in r.data
    n_blocks = r.data.count(b'<article class="scrum-team-detail-member"')
    r_one = c.get("/scrum/sprint/%d/team-detailed" % sid, query_string={"member": emp})
    assert r_one.status_code == 200
    assert r_one.data.count(b'<article class="scrum-team-detail-member"') == 1
    r_bad = c.get("/scrum/sprint/%d/team-detailed" % sid, query_string={"member": "__not_on_roster__"})
    assert r_bad.status_code == 200
    assert r_bad.data.count(b'<article class="scrum-team-detail-member"') == n_blocks


def test_scrum_team_detailed_attachment_upload_redirects_back(client, app):
    from io import BytesIO

    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "DetailAttachRedirectSp",
            "start_date": "2026-08-10",
            "end_date": "2026-08-24",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(
        conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("DetailAttachRedirectSp",)).fetchone()["id"]
    )
    emp = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "DetailAttachSticky",
            "assignee": emp,
            "estimate_hours": "2",
            "kanban_column": "doing",
            "task_kind": "ndy",
            "notes": "",
        },
        follow_redirects=True,
    )
    iid = int(
        conn.execute(
            "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
            (sid, "DetailAttachSticky"),
        ).fetchone()["id"]
    )
    conn.close()
    up = c.post(
        "/scrum/sprint/item/attachment",
        data={
            "item_id": str(iid),
            "sprint_id": str(sid),
            "assignee": emp,
            "return_to": "team_detailed",
            "file": (BytesIO(b"detailed-page-bytes"), "doc.txt"),
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )
    assert up.status_code in (302, 303)
    loc = up.headers.get("Location") or ""
    assert "/team-detailed" in loc
    det = c.get("/scrum/sprint/%d/team-detailed" % sid)
    assert det.status_code == 200
    assert b"detailed-page-bytes" not in det.data
    assert b"doc.txt" in det.data


def test_scrum_sprint_create_rejects_duplicate_name(client, app):
    c = client
    _manager_login(c)
    name = "Unique Dup Name Sprint 7f3a"
    c.post(
        "/scrum/sprint/create",
        data={
            "name": name,
            "start_date": "2026-09-01",
            "end_date": "2026-09-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    r_dup = c.post(
        "/scrum/sprint/create",
        data={
            "name": name,
            "start_date": "2026-10-01",
            "end_date": "2026-10-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    assert r_dup.status_code == 200
    assert b"already exists" in r_dup.data or b"different name" in r_dup.data
    r_case = c.post(
        "/scrum/sprint/create",
        data={
            "name": name.upper(),
            "start_date": "2026-11-01",
            "end_date": "2026-11-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    assert r_case.status_code == 200
    assert b"already exists" in r_case.data or b"different name" in r_case.data
    conn = get_db(app)
    n = conn.execute(
        "SELECT COUNT(*) AS c FROM scrum_sprint WHERE lower(trim(name)) = lower(trim(?))",
        (name,),
    ).fetchone()["c"]
    conn.close()
    assert int(n) == 1


def test_scrum_sprint_export_xlsx(client, app):
    from openpyxl import load_workbook

    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Export Sprint",
            "start_date": "2026-09-01",
            "end_date": "2026-09-14",
            "goal": "Excel check",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Export Sprint",)).fetchone()["id"])
    conn.close()
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "Sticky A",
            "assignee": EMPLOYEES[0],
            "estimate_hours": "4",
            "kanban_column": "doing",
            "task_kind": "ndy",
            "notes": "Details here",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    iid = int(
        conn.execute(
            "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
            (sid, "Sticky A"),
        ).fetchone()["id"]
    )
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, 'Logged work', 1.5, 'doing', 'doing', '2026-09-02T10:00:00Z')
        """,
        (iid,),
    )
    conn.execute(
        """
        INSERT INTO scrum_item_appreciation (item_id, author, comment, created_at)
        VALUES (?, 'TestMgr', 'Great sprint contribution', '2026-09-03T12:00:00Z')
        """,
        (iid,),
    )
    conn.commit()
    conn.close()

    bad = c.post("/scrum/sprint/export-xlsx", data={"sprint_id": "999999"})
    assert bad.status_code in (302, 303)

    res = c.post("/scrum/sprint/export-xlsx", data={"sprint_id": str(sid)})
    assert res.status_code == 200
    assert res.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert res.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(res.data))
    assert wb.sheetnames[0] == "Summary"
    assert "SprintStatus" in wb.sheetnames
    assert "HPPM" in wb.sheetnames
    wh = wb["HPPM"]
    assert wh.cell(row=1, column=1).value and "HPPM" in str(wh.cell(row=1, column=1).value)
    assert "Activity log" not in wb.sheetnames
    assert "Daily tasks Details" in wb.sheetnames
    assert "Appriciation" in wb.sheetnames
    ws0 = wb["Summary"]
    found_name_hdr = ws0.cell(row=2, column=4).value == "NAME"
    assert found_name_hdr
    found_sprint = False
    found_cap = False
    found_planned = False
    for r in range(1, 60):
        if ws0.cell(row=r, column=1).value == "Sprint name":
            assert ws0.cell(row=r, column=2).value == "Export Sprint"
            found_sprint = True
        if ws0.cell(row=r, column=1).value == "Total Sprint Capacity (h)":
            assert ws0.cell(row=r, column=2).value not in (None, "", "—")
            found_cap = True
        if ws0.cell(row=r, column=1).value == "Planned capacity ((sum estimates + Sprint leaves) ÷ sprint capacity, %)":
            found_planned = True
    assert found_sprint and found_cap and found_planned
    titles = [wb["SprintStatus"].cell(row=2, column=j).value for j in range(1, 5)]
    assert titles[2] == "Sticky A"
    hdr1 = [wb["SprintStatus"].cell(row=1, column=j).value for j in range(1, wb["SprintStatus"].max_column + 1)]
    dcol = hdr1.index("2026-09-02") + 1
    assert float(wb["SprintStatus"].cell(row=2, column=dcol).value or 0) >= 1.4
    wsd = wb["Daily tasks Details"]
    found_ov = False
    for row in range(1, 25):
        if wsd.cell(row=row, column=1).value and "Per-member sprint overview" in str(
            wsd.cell(row=row, column=1).value
        ):
            found_ov = True
            break
    assert found_ov
    wap = wb["Appriciation"]
    assert wap.cell(row=3, column=3).value == "Appriciation author"
    assert wap.cell(row=4, column=3).value == "TestMgr"
    assert wap.cell(row=4, column=4).value == "Great sprint contribution"
    assert wap.cell(row=4, column=7).value == "Sticky A"
    assert wap.cell(row=4, column=8).value == EMPLOYEES[0]


def test_team_hub_mode_updates_active_team(client, app):
    c = client
    _manager_login(c)
    res = c.post("/team/hub-mode", data={"hub_mode": "scrum", "next": "/scrum"}, follow_redirects=False)
    assert res.status_code in (302, 303)
    conn = get_db(app)
    row = conn.execute("SELECT hub_mode FROM teams ORDER BY id ASC LIMIT 1").fetchone()
    conn.close()
    assert row["hub_mode"] == "scrum"


def test_scrum_hppm_entry_redirects_and_page_loads(client, app):
    c = client
    _manager_login(c)
    r0 = c.get("/scrum/hppm", follow_redirects=False)
    assert r0.status_code == 302
    loc0 = r0.headers.get("Location") or ""
    assert "/scrum" in loc0

    c.post(
        "/scrum/sprint/create",
        data={
            "name": "HPPM Sprint",
            "start_date": "2026-10-01",
            "end_date": "2026-10-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sp = conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("HPPM Sprint",)).fetchone()
    assert sp
    sid = int(sp["id"])
    conn.close()

    team = c.get("/scrum/sprint/%d" % sid)
    assert team.status_code == 200
    assert b"HPPM view" in team.data
    assert ("/scrum/sprint/%d/hppm" % sid).encode() in team.data
    assert b"Type stack" in team.data

    hppm = c.get("/scrum/sprint/%d/hppm" % sid)
    assert hppm.status_code == 200
    assert b"Type stack" in hppm.data
    assert b"scrum-team-kind-stack--horizontal" in hppm.data
    assert b"Area stack" in hppm.data
    assert b"HPPM summary" in hppm.data
    assert b"Total" in hppm.data
    assert b"Estimate sprint" in hppm.data
    assert b"Absences" in hppm.data
    assert b"Feature Support (Y)" in hppm.data

    r1 = c.get("/scrum/hppm", follow_redirects=False)
    assert r1.status_code == 302
    loc1 = r1.headers.get("Location") or ""
    assert str(sid) in loc1 and "/hppm" in loc1


def test_scrm_sprint_team_and_sticky(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Sprint A",
            "start_date": "2026-09-01",
            "end_date": "2026-09-14",
            "goal": "Ship widgets",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sp = conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Sprint A",)).fetchone()
    assert sp
    sid = int(sp["id"])
    conn.close()
    team = c.get("/scrum/sprint/%d" % sid)
    assert team.status_code == 200
    assert b"Team" in team.data
    assert b"SprintLeaveView" in team.data
    assert ("/scrum/sprint/%d/leave-tracker" % sid).encode() in team.data
    lv = c.get("/scrum/sprint/%d/leave-tracker" % sid)
    assert lv.status_code == 200
    assert b"ws-table" in lv.data
    assert b"ws-month-fit" in lv.data
    assert b"Total</th>" in lv.data
    assert b"Leave tracker" in lv.data
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "Story 1",
            "assignee": EMPLOYEES[0],
            "estimate_hours": "5",
            "kanban_column": "backlog",
            "task_kind": "story",
            "notes": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    n = int(conn.execute("SELECT COUNT(*) AS c FROM scrum_sprint_item WHERE sprint_id = ?", (sid,)).fetchone()["c"])
    conn.close()
    assert n >= 1
    board = c.get("/scrum/sprint/%d/board" % sid, query_string={"assignee": EMPLOYEES[0]})
    assert board.status_code == 200
    assert b"Sticky board" in board.data
    assert b"SprintLeaveView" in board.data
    assert ("/scrum/sprint/%d/leave-tracker" % sid).encode() in board.data
    assert b"ws-table" in board.data
    assert b"Available" in board.data


def test_scrum_sticky_attachment_upload_and_download(client, app):
    from io import BytesIO

    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "AttachSprint",
            "start_date": "2026-10-01",
            "end_date": "2026-10-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("AttachSprint",)).fetchone()["id"])
    emp = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "WithAttach",
            "assignee": emp,
            "estimate_hours": "3",
            "kanban_column": "do",
            "task_kind": "ndy",
            "notes": "",
        },
        follow_redirects=True,
    )
    row = conn.execute(
        "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?", (sid, "WithAttach")
    ).fetchone()
    assert row
    iid = int(row["id"])
    conn.close()
    up = c.post(
        "/scrum/sprint/item/attachment",
        data={
            "item_id": str(iid),
            "sprint_id": str(sid),
            "assignee": emp,
            "file": (BytesIO(b"hello-attachment-bytes"), "notes.txt"),
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )
    assert up.status_code in (302, 303)
    conn = get_db(app)
    aid = int(conn.execute("SELECT id FROM scrum_sprint_item_attachment WHERE item_id = ?", (iid,)).fetchone()["id"])
    conn.close()
    dl = c.get("/scrum/sprint/item/attachment/%d/file" % aid)
    assert dl.status_code == 200
    assert dl.data == b"hello-attachment-bytes"


def test_scrum_sticky_attachment_multi_upload(client, app):
    from io import BytesIO

    from werkzeug.datastructures import MultiDict

    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "MultiAttachSprint",
            "start_date": "2026-10-01",
            "end_date": "2026-10-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("MultiAttachSprint",)).fetchone()["id"])
    emp = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "MultiAttach",
            "assignee": emp,
            "estimate_hours": "1",
            "kanban_column": "do",
            "task_kind": "ndy",
            "notes": "",
        },
        follow_redirects=True,
    )
    iid = int(
        conn.execute(
            "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?", (sid, "MultiAttach")
        ).fetchone()["id"]
    )
    conn.close()
    up = c.post(
        "/scrum/sprint/item/attachment",
        data=MultiDict(
            [
                ("item_id", str(iid)),
                ("sprint_id", str(sid)),
                ("assignee", emp),
                ("files", (BytesIO(b"one"), "one.txt")),
                ("files", (BytesIO(b"two"), "two.txt")),
            ]
        ),
        content_type="multipart/form-data",
        follow_redirects=False,
    )
    assert up.status_code in (302, 303)
    conn = get_db(app)
    n = int(
        conn.execute("SELECT COUNT(*) AS c FROM scrum_sprint_item_attachment WHERE item_id = ?", (iid,)).fetchone()[
            "c"
        ]
    )
    rows = conn.execute(
        "SELECT original_filename FROM scrum_sprint_item_attachment WHERE item_id = ? ORDER BY id",
        (iid,),
    ).fetchall()
    conn.close()
    assert n == 2
    assert {r["original_filename"] for r in rows} == {"one.txt", "two.txt"}


def test_portal_sticky_attachment_upload_delete_and_download(client, app):
    name = EMPLOYEES[0]
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "portal.attach@nokia.com",
            "name": name,
            "roster_name": name,
            "role": "employee",
        }
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-12-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Portal Attach Sprint', '2026-12-01', '2026-12-14', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'PortalAttachCard', 1, 'open', '', 0, ?, ?, 'do', 'ndy')
        """,
        (sid, name, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()

    board = client.get("/portal/sprint/%d/board" % sid)
    assert board.status_code == 200
    assert b"/portal/scrum/sprint/item/attachment" in board.data

    up = client.post(
        "/portal/scrum/sprint/item/attachment",
        data={
            "item_id": str(iid),
            "sprint_id": str(sid),
            "file": (io.BytesIO(b"portal-bytes"), "note.pdf"),
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )
    assert up.status_code in (302, 303)
    conn = get_db(app)
    aid = int(conn.execute("SELECT id FROM scrum_sprint_item_attachment WHERE item_id = ?", (iid,)).fetchone()["id"])
    conn.close()

    dl = client.get("/portal/scrum/sprint/item/attachment/%d/file" % aid)
    assert dl.status_code == 200
    assert dl.data == b"portal-bytes"

    rm = client.post(
        "/portal/scrum/sprint/item/attachment/delete",
        data={"attachment_id": str(aid), "sprint_id": str(sid)},
        follow_redirects=False,
    )
    assert rm.status_code in (302, 303)
    conn = get_db(app)
    n = int(conn.execute("SELECT COUNT(*) AS c FROM scrum_sprint_item_attachment WHERE id = ?", (aid,)).fetchone()["c"])
    conn.close()
    assert n == 0


def test_portal_sticky_attachment_multi_upload_one_post(client, app):
    from werkzeug.datastructures import MultiDict

    name = EMPLOYEES[0]
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "portal.multi@nokia.com",
            "name": name,
            "roster_name": name,
            "role": "employee",
        }
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-12-02T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Portal Multi Sprint', '2026-12-02', '2026-12-12', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'PortalMulti', 1, 'open', '', 0, ?, ?, 'do', 'ndy')
        """,
        (sid, name, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    up = client.post(
        "/portal/scrum/sprint/item/attachment",
        data=MultiDict(
            [
                ("item_id", str(iid)),
                ("sprint_id", str(sid)),
                ("files", (io.BytesIO(b"a"), "a.txt")),
                ("files", (io.BytesIO(b"b"), "b.txt")),
            ]
        ),
        content_type="multipart/form-data",
        follow_redirects=False,
    )
    assert up.status_code in (302, 303)
    conn = get_db(app)
    n = int(
        conn.execute("SELECT COUNT(*) AS c FROM scrum_sprint_item_attachment WHERE item_id = ?", (iid,)).fetchone()[
            "c"
        ]
    )
    conn.close()
    assert n == 2


def test_scrum_kanban_capacity_strip_shows_daily_burnt_hours_including_weekend(client, app):
    """Sprint capacity grid shows per-calendar-day burnt hours (sticky activity), including weekends."""
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "DailyHoursStrip",
            "start_date": "2026-09-01",
            "end_date": "2026-09-10",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("DailyHoursStrip",)).fetchone()["id"])
    emp = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "StripCard",
            "assignee": emp,
            "estimate_hours": "4",
            "kanban_column": "doing",
            "task_kind": "ndy",
            "notes": "",
        },
        follow_redirects=True,
    )
    row = conn.execute(
        "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
        (sid, "StripCard"),
    ).fetchone()
    assert row
    iid = int(row["id"])
    # 2026-09-06 is Sunday — still show logged hours in that column.
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, '', 3.0, 'doing', 'doing', '2026-09-06T14:00:00Z')
        """,
        (iid,),
    )
    conn.commit()
    conn.close()
    board = c.get("/scrum/sprint/%d/board" % sid, query_string={"assignee": emp})
    assert board.status_code == 200
    assert b"Total burnt" in board.data
    assert b"scrum-bar-cap-burnt" in board.data
    assert b'name="files"' in board.data
    assert b"multiple" in board.data
    assert b"scrum-ws-cell-effort" in board.data
    assert b"3.0h" in board.data
    assert b"Improvement" in board.data
    assert b"process_tools" in board.data


def test_scrum_sprint_team_shows_ndy_fsy_burn_when_overall_below_67(client, app):
    """Team overview shows NDY/FSY/CODE % burnt when overall Sprint burnt is under 67%."""
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Kind Burn Sprint",
            "start_date": "2026-09-01",
            "end_date": "2026-09-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sp = conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Kind Burn Sprint",)).fetchone()
    assert sp
    sid = int(sp["id"])
    conn.close()
    emp = EMPLOYEES[0]
    for title, kind in (("Ndy work", "ndy"), ("Fsy work", "fsy"), ("Code work", "code")):
        c.post(
            "/scrum/sprint/item/add",
            data={
                "sprint_id": str(sid),
                "title": title,
                "assignee": emp,
                "estimate_hours": "5",
                "kanban_column": "backlog",
                "task_kind": kind,
                "notes": "",
            },
            follow_redirects=True,
        )
    conn = get_db(app)
    ts = "2026-09-02T12:00:00Z"
    for title, hours in (("Ndy work", 1.0), ("Fsy work", 2.0), ("Code work", 0.5)):
        row = conn.execute(
            "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
            (sid, title),
        ).fetchone()
        assert row
        iid = int(row["id"])
        conn.execute(
            """
            INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
            VALUES (?, '', ?, 'backlog', 'backlog', ?)
            """,
            (iid, hours, ts),
        )
    conn.commit()
    conn.close()
    team = c.get("/scrum/sprint/%d" % sid)
    assert team.status_code == 200
    html = team.get_data(as_text=True)
    assert "Cap:" in html
    assert "SCAPACITY" in html
    assert "scrum-member-kind-burn" in html
    assert "Sprint burnt" in html
    assert ">20%<" in html or "20%" in html  # NDY: 1/5
    assert ">40%<" in html or "40%" in html  # FSY: 2/5
    assert ">10%<" in html or "10%" in html  # CODE: 0.5/5
    assert "NDY" in html and "FSY" in html and "CODE" in html


def test_manager_team_bundle_export_import_roundtrip(client, app):
    import json
    from io import BytesIO

    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id LIMIT 1").fetchone()["id"])
    ts = "2026-10-01T12:00:00Z"
    cur = conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'BundleSprint', '2026-10-01', '2026-10-14', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(cur.lastrowid)
    cur2 = conn.execute(
        """
        INSERT INTO scrum_sprint_item
          (sprint_id, assignee, title, estimate_hours, status, notes, dod, done_artifacts, sort_order, created_at, updated_at, kanban_column, task_kind, area)
        VALUES (?, ?, 'Sticky A', 3.0, 'open', '', '', '[]', 0, ?, ?, 'backlog', 'ndy', '')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(cur2.lastrowid)
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, 'hello', 0.5, 'backlog', 'backlog', ?)
        """,
        (iid, ts),
    )
    conn.commit()
    conn.close()

    r = c.get("/scrum/team-bundle.json")
    assert r.status_code == 200
    bundle = json.loads(r.get_data(as_text=True))
    assert bundle["format"] == "team_tracker_team_bundle_v1"
    assert any(s.get("name") == "BundleSprint" for s in bundle["scrum_sprint"])

    conn = get_db(app)
    conn.execute("DELETE FROM scrum_sprint WHERE team_id = ? AND name = ?", (tid, "BundleSprint"))
    conn.commit()
    conn.close()

    buf = BytesIO(json.dumps(bundle).encode("utf-8"))
    buf.seek(0)
    r2 = c.post(
        "/scrum/team-bundle/import",
        data={"bundle_file": (buf, "team-bundle.json")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert r2.status_code == 200

    conn = get_db(app)
    row_sp = conn.execute(
        "SELECT id FROM scrum_sprint WHERE team_id = ? AND name = ?", (tid, "BundleSprint")
    ).fetchone()
    assert row_sp
    n_act = conn.execute(
        """
        SELECT COUNT(*) AS c FROM scrum_item_activity
        WHERE item_id IN (SELECT id FROM scrum_sprint_item WHERE sprint_id = ?)
        """,
        (int(row_sp["id"]),),
    ).fetchone()
    assert int(n_act["c"]) >= 1
    conn.close()


def test_scrum_sprint_team_highlights_activity_within_last_24h(client, app):
    """Latest activity lines from the last 24h get a distinct highlight class."""
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Recent24h Sprint",
            "start_date": "2026-09-01",
            "end_date": "2026-09-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Recent24h Sprint",)).fetchone()["id"])
    emp = EMPLOYEES[0]
    c.post(
        "/scrum/sprint/item/add",
        data={
            "sprint_id": str(sid),
            "title": "ActivityCard",
            "assignee": emp,
            "estimate_hours": "2",
            "kanban_column": "doing",
            "task_kind": "ndy",
            "notes": "",
        },
        follow_redirects=True,
    )
    row = conn.execute(
        "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
        (sid, "ActivityCard"),
    ).fetchone()
    assert row
    iid = int(row["id"])
    fresh_ts = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat(timespec="seconds").replace("+00:00", "Z")
    old_ts = (datetime.now(timezone.utc) - timedelta(hours=50)).isoformat(timespec="seconds").replace("+00:00", "Z")
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, 'older sync', 0.5, 'doing', 'doing', ?)
        """,
        (iid, old_ts),
    )
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, 'fresh sync', 1.0, 'doing', 'doing', ?)
        """,
        (iid, fresh_ts),
    )
    conn.commit()
    conn.close()
    team = c.get("/scrum/sprint/%d" % sid)
    assert team.status_code == 200
    html = team.get_data(as_text=True)
    assert "scrum-note-recent-24h" in html
    assert "scrum-note-card-recent-24h" in html


def test_unified_login_register_and_empty_workspace(tmp_path, monkeypatch):
    """Unified login: register manager account → sign in → empty workspace banner."""
    db_path = str(tmp_path / "login_flow.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "mgrpw")
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.setenv("TEAM_TRACKER_AUTO_TESSERACT", "0")
    app = create_app()
    app.config["TESTING"] = True
    c = app.test_client()

    r = c.get("/login")
    assert r.status_code == 200
    assert b"Employee Portal" in r.data
    assert b"Manager Access" in r.data
    assert b"Create one now" in r.data

    r = c.get("/register")
    assert r.status_code == 200
    assert b"Create Manager Account" in r.data

    r = c.post(
        "/register",
        data={"email": "new@test.com", "password": "secret1", "team_name": "TestTeam"},
        follow_redirects=True,
    )
    assert r.status_code == 200
    assert b"Account created" in r.data
    assert b"Leave tracker" in r.data

    r2 = c.get("/dashboard")
    assert r2.status_code == 200
    assert b"Download roster template" in r2.data or b"Leave tracker" in r2.data

    c.post("/manager/logout", follow_redirects=True)

    r3 = c.post(
        "/login",
        data={"gate_kind": "manager", "email": "new@test.com", "password": "secret1"},
        follow_redirects=True,
    )
    assert r3.status_code == 200
    assert b"Signed in" in r3.data or b"Leave tracker" in r3.data

    r4 = c.post(
        "/login",
        data={"gate_kind": "manager", "email": "new@test.com", "password": "wrong"},
        follow_redirects=True,
    )
    assert b"Invalid email or password" in r4.data

    r5 = c.get("/reports/roster-template.xlsx")
    assert r5.status_code == 200
    assert b"PK" in r5.data[:4]


def test_primary_owner_manager_seeded_from_master_pin(tmp_path, monkeypatch):
    seed_email = "owner-seed-test@example.com"
    db_path = str(tmp_path / "owner_seed.db")
    monkeypatch.setenv("TEAM_TRACKER_DB_PATH", db_path)
    _monkeypatch_dotenv_restore_db_path(monkeypatch, db_path)
    monkeypatch.setenv("MANAGER_DASHBOARD_PASSWORD", "my_master_99")
    monkeypatch.setenv("PRIMARY_OWNER_MANAGER_EMAIL", seed_email)
    monkeypatch.setenv("WTF_CSRF_ENABLED", "false")
    monkeypatch.setenv("TEAM_TRACKER_AUTO_TESSERACT", "0")
    application = create_app()
    application.config["TESTING"] = True
    conn = get_db(application)
    row = conn.execute(
        "SELECT email, team_name FROM managers WHERE email = ? COLLATE NOCASE",
        (seed_email,),
    ).fetchone()
    conn.close()
    assert row
    assert row["email"].lower() == seed_email.lower()
    c = application.test_client()
    r = c.post(
        "/login",
        data={
            "gate_kind": "manager",
            "email": seed_email,
            "password": "my_master_99",
        },
        follow_redirects=True,
    )
    assert r.status_code == 200
    assert b"Leave tracker" in r.data or b"Signed in" in r.data


def test_scrum_sprint_update_dates(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Date Edit Sprint",
            "start_date": "2026-10-01",
            "end_date": "2026-10-05",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    row = conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Date Edit Sprint",)).fetchone()
    assert row
    sid = int(row["id"])
    conn.close()
    r = c.post(
        "/scrum/sprint/update",
        data={"sprint_id": str(sid), "start_date": "2026-10-08", "end_date": "2026-10-22"},
        follow_redirects=True,
    )
    assert r.status_code == 200
    conn = get_db(app)
    sp = conn.execute("SELECT start_date, end_date FROM scrum_sprint WHERE id = ?", (sid,)).fetchone()
    conn.close()
    assert str(sp["start_date"]).startswith("2026-10-08")
    assert str(sp["end_date"]).startswith("2026-10-21")


def test_scrum_sprint_carry_forward_do_and_doing(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-01-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Carry Old', '2026-01-01', '2026-01-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    old_sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    emp = EMPLOYEES[0]
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind, sticky_color_hex)
        VALUES (?, ?, 'Do task', 3, 'open', 'n1', 'd1', 0, ?, ?, 'do', 'ndy', NULL)
        """,
        (old_sid, emp, ts, ts),
    )
    do_iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind, sticky_color_hex)
        VALUES (?, ?, 'Doing task', 4, 'doing', 'n2', 'd2', 1, ?, ?, 'doing', 'code', '#112233')
        """,
        (old_sid, emp, ts, ts),
    )
    doing_iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, '', 1.5, 'do', 'do', ?)
        """,
        (do_iid, ts),
    )
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, '', 2.0, 'doing', 'doing', ?)
        """,
        (doing_iid, ts),
    )
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/sprint/create",
        data={"name": "Carry New", "start_date": "2026-01-15", "end_date": "2026-01-31", "goal": ""},
        follow_redirects=True,
    )
    assert r.status_code == 200
    assert b"carried" in r.data.lower()
    conn = get_db(app)
    new_sid = int(conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Carry New",)).fetchone()[0])
    rows = list(
        conn.execute(
            """
            SELECT title, kanban_column, assignee, notes, dod, task_kind, sticky_color_hex, estimate_hours
            FROM scrum_sprint_item WHERE sprint_id = ? ORDER BY title
            """,
            (new_sid,),
        )
    )
    assert len(rows) == 2
    by_title = {x["title"]: x for x in rows}
    assert by_title["Do task"]["kanban_column"] == "backlog"
    assert by_title["Doing task"]["kanban_column"] == "backlog"
    assert by_title["Do task"]["notes"] == "n1"
    assert by_title["Doing task"]["notes"] == "n2"
    assert by_title["Do task"]["dod"] == "d1"
    assert by_title["Doing task"]["dod"] == "d2"
    assert by_title["Do task"]["task_kind"] == "ndy"
    assert by_title["Doing task"]["task_kind"] == "code"
    assert (by_title["Doing task"]["sticky_color_hex"] or "").lower() == "#112233"
    assert by_title["Do task"]["assignee"] == emp
    assert by_title["Doing task"]["assignee"] == emp
    assert abs(float(by_title["Do task"]["estimate_hours"]) - 3.0) < 0.001
    assert abs(float(by_title["Doing task"]["estimate_hours"]) - 4.0) < 0.001
    for title in ("Do task", "Doing task"):
        iid = int(
            conn.execute(
                "SELECT id FROM scrum_sprint_item WHERE sprint_id = ? AND title = ?",
                (new_sid, title),
            ).fetchone()["id"]
        )
        burn = float(
            conn.execute(
                "SELECT COALESCE(SUM(committed_hours), 0) AS h FROM scrum_item_activity WHERE item_id = ?",
                (iid,),
            ).fetchone()["h"]
            or 0
        )
        assert burn == 0.0
    conn.close()


def test_scrum_board_leave_strip_shows_stretch_when_overallocated(client, app):
    c = client
    _manager_login(c)
    c.post(
        "/scrum/sprint/create",
        data={
            "name": "Sprint Cap",
            "start_date": "2026-09-01",
            "end_date": "2026-09-14",
            "goal": "",
        },
        follow_redirects=True,
    )
    conn = get_db(app)
    sp = conn.execute("SELECT id FROM scrum_sprint WHERE name = ?", ("Sprint Cap",)).fetchone()
    assert sp
    sid = int(sp["id"])
    ts = "2026-08-15T12:00:00Z"
    conn.execute(
        """
        INSERT INTO leave_requests
        (employee_name, reason, description, start_date, end_date, duration_type, status, created_at, submitted_ip)
        VALUES (?, 'pl', '', '2026-09-08', '2026-09-10', 'multi', 'approved', ?, '')
        """,
        (EMPLOYEES[0], ts),
    )
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, dod, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Big', 90, 'open', '', '', 0, ?, ?, 'do', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    conn.commit()
    conn.close()
    board = c.get("/scrum/sprint/%d/board" % sid, query_string={"assignee": EMPLOYEES[0]})
    assert board.status_code == 200
    assert b"Stretch" in board.data
    assert b"PL" in board.data


def test_scrm_api_move_updates_column(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S Move', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Card', 1, 'open', '', 0, ?, ?, 'backlog', 'task')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "doing",
            "committed_hours": "1.5",
            "note": "Started work",
        },
    )
    assert r.status_code == 200
    assert r.get_json()["ok"] is True
    conn = get_db(app)
    col = conn.execute("SELECT kanban_column, status FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()
    conn.close()
    assert col["kanban_column"] == "doing"
    assert col["status"] == "doing"


def test_scrm_api_appreciation_add_and_delete_all(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S Appr', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Appr card', 1, 'open', '', 0, ?, ?, 'doing', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r_add = c.post(
        "/scrum/api/item/appreciation",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "comment": "Nice work",
        },
    )
    assert r_add.status_code == 200
    assert r_add.get_json()["ok"] is True
    conn = get_db(app)
    n1 = int(conn.execute("SELECT COUNT(*) AS c FROM scrum_item_appreciation WHERE item_id = ?", (iid,)).fetchone()["c"])
    conn.close()
    assert n1 == 1

    r_del = c.post(
        "/scrum/api/item/appreciation/delete-all",
        json={"item_id": iid, "sprint_id": sid, "assignee": EMPLOYEES[0]},
    )
    assert r_del.status_code == 200
    j = r_del.get_json()
    assert j["ok"] is True
    assert j["deleted"] == 1
    conn = get_db(app)
    n2 = int(conn.execute("SELECT COUNT(*) AS c FROM scrum_item_appreciation WHERE item_id = ?", (iid,)).fetchone()["c"])
    conn.close()
    assert n2 == 0


def test_scrm_api_move_to_done_saves_artifacts(client, app):
    import json

    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S Done Art', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Ship it', 1, 'open', '', 0, ?, ?, 'doing', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "done",
            "committed_hours": "0",
            "note": "wrapped",
            "artifacts": [{"label": "PR", "url": "https://example.com/pr/99"}],
        },
    )
    assert r.status_code == 200
    assert r.get_json()["ok"] is True
    conn = get_db(app)
    raw = conn.execute(
        "SELECT kanban_column, done_artifacts FROM scrum_sprint_item WHERE id = ?", (iid,)
    ).fetchone()
    conn.close()
    assert raw["kanban_column"] == "done"
    data = json.loads(raw["done_artifacts"])
    assert len(data) == 1
    assert data[0]["url"] == "https://example.com/pr/99"
    assert data[0]["label"] == "PR"


def test_scrm_api_move_to_done_rejects_bad_artifact_url(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S Bad Art', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'X', 1, 'open', '', 0, ?, ?, 'doing', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "done",
            "committed_hours": "0",
            "note": "",
            "artifacts": [{"url": "javascript:alert(1)"}],
        },
    )
    assert r.status_code == 400
    assert r.get_json()["ok"] is False


def test_scrm_api_move_backlog_to_doing_keeps_prior_burnt(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-02T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S Burnt', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Card2', 1, 'open', '', 0, ?, ?, 'backlog', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, '', 2.0, 'backlog', 'backlog', ?)
        """,
        (iid, ts),
    )
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "doing",
            "committed_hours": "1",
            "note": "move",
        },
    )
    assert r.status_code == 200
    conn = get_db(app)
    total = float(
        conn.execute(
            "SELECT COALESCE(SUM(committed_hours), 0) AS s FROM scrum_item_activity WHERE item_id = ?",
            (iid,),
        ).fetchone()["s"]
    )
    conn.close()
    assert abs(total - 3.0) < 0.001


def test_scrum_api_activity_update_returns_burn_totals(client, app):
    """Editing stand-up hours updates DB and JSON includes refreshed task burn (sum of activity hours)."""
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY name COLLATE NOCASE LIMIT 1").fetchone()["id"])
    name_row = conn.execute(
        "SELECT employee_name FROM team_roster WHERE team_id = ? ORDER BY employee_name COLLATE NOCASE LIMIT 1",
        (tid,),
    ).fetchone()
    assignee = str(name_row["employee_name"]).strip() if name_row else EMPLOYEES[0]
    ts = "2026-11-05T10:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S ActUpd', '2026-11-01', '2026-11-14', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Burn card', 8, 'open', '', 0, ?, ?, 'doing', 'ndy')
        """,
        (sid, assignee, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, 'work', 2.5, 'doing', 'doing', ?)
        """,
        (iid, ts),
    )
    aid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/activity_update",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": assignee,
            "activity_id": aid,
            "note": "revised",
            "committed_hours": "1",
        },
    )
    assert r.status_code == 200
    j = r.get_json()
    assert j.get("ok") is True
    assert j.get("activity_committed_hours") == 1.0
    br = j.get("burn") or {}
    assert abs(float(br.get("committed_logged_hours", 0)) - 1.0) < 0.001
    assert br.get("burn_pct") is not None
    conn = get_db(app)
    total = float(
        conn.execute(
            "SELECT COALESCE(SUM(committed_hours), 0) AS s FROM scrum_item_activity WHERE item_id = ?",
            (iid,),
        ).fetchone()["s"]
    )
    conn.close()
    assert abs(total - 1.0) < 0.001


def test_scrm_api_move_do_to_doing_resets_burnt(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-03T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S Reset', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Card3', 1, 'open', '', 0, ?, ?, 'do', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, '', 10.0, 'do', 'do', ?)
        """,
        (iid, ts),
    )
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "doing",
            "committed_hours": "0.5",
            "note": "into doing",
        },
    )
    assert r.status_code == 200
    conn = get_db(app)
    nrows = int(
        conn.execute("SELECT COUNT(*) AS c FROM scrum_item_activity WHERE item_id = ?", (iid,)).fetchone()["c"]
    )
    total = float(
        conn.execute(
            "SELECT COALESCE(SUM(committed_hours), 0) AS s FROM scrum_item_activity WHERE item_id = ?",
            (iid,),
        ).fetchone()["s"]
    )
    conn.close()
    assert nrows == 1
    assert abs(total - 0.5) < 0.001


def test_scrm_api_move_doing_to_do_resets_burnt(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-03T01:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S DoingToDo', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'DoingCard', 8, 'doing', '', 0, ?, ?, 'doing', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, 'standup', 2.5, 'doing', 'doing', ?)
        """,
        (iid, ts),
    )
    conn.execute(
        """
        INSERT INTO scrum_item_activity (item_id, body, committed_hours, from_column, to_column, created_at)
        VALUES (?, '', 1.0, 'do', 'doing', ?)
        """,
        (iid, ts),
    )
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "do",
            "committed_hours": "99",
            "note": "ignored hours",
        },
    )
    assert r.status_code == 200
    conn = get_db(app)
    col = conn.execute("SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()["kanban_column"]
    nrows = int(
        conn.execute("SELECT COUNT(*) AS c FROM scrum_item_activity WHERE item_id = ?", (iid,)).fetchone()["c"]
    )
    total = float(
        conn.execute(
            "SELECT COALESCE(SUM(committed_hours), 0) AS s FROM scrum_item_activity WHERE item_id = ?",
            (iid,),
        ).fetchone()["s"]
    )
    last_h = float(
        conn.execute(
            "SELECT committed_hours FROM scrum_item_activity WHERE item_id = ? ORDER BY id DESC LIMIT 1",
            (iid,),
        ).fetchone()["committed_hours"]
        or 0
    )
    conn.close()
    assert col == "do"
    assert nrows == 1
    assert abs(total) < 0.001
    assert abs(last_h) < 0.001


def test_scrm_api_move_do_to_doing_empty_hours_stores_null(client, app):
    c = client
    _manager_login(c)
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-04T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'S NullH', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'CardNull', 2, 'open', '', 0, ?, ?, 'do', 'ndy')
        """,
        (sid, EMPLOYEES[0], ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = c.post(
        "/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": EMPLOYEES[0],
            "to_column": "doing",
            "committed_hours": None,
            "note": "",
        },
    )
    assert r.status_code == 200
    conn = get_db(app)
    ch = conn.execute(
        "SELECT committed_hours FROM scrum_item_activity WHERE item_id = ? ORDER BY id DESC LIMIT 1",
        (iid,),
    ).fetchone()["committed_hours"]
    total = float(
        conn.execute(
            "SELECT COALESCE(SUM(committed_hours), 0) AS s FROM scrum_item_activity WHERE item_id = ?",
            (iid,),
        ).fetchone()["s"]
    )
    conn.close()
    assert ch is None
    assert abs(total) < 0.001
    name = EMPLOYEES[0]
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "portal.test@nokia.com",
            "name": name,
            "roster_name": name,
            "role": "employee",
        }
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Portal Board Sprint', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Portal Kanban Card', 1, 'open', '', 0, ?, ?, 'backlog', 'task')
        """,
        (sid, name, ts, ts),
    )
    conn.commit()
    conn.close()
    board = client.get("/portal/sprint/%d/board" % sid)
    assert board.status_code == 200
    assert b"Backlog" in board.data
    assert b"Portal Kanban Card" in board.data
    assert b"/portal/scrum/api/item/move" in board.data
    assert b"kb-doing-reset-dialog" in board.data
    assert b"Dashboard" in board.data
    hop = client.get("/portal/sprint", follow_redirects=False)
    assert hop.status_code in (302, 303)
    assert "/portal/my-sprint/board" in (hop.headers.get("Location") or "")
    hop2 = client.get("/portal/my-sprint/board", follow_redirects=False)
    assert hop2.status_code in (302, 303)
    assert "/portal/sprint/%d/board" % sid in (hop2.headers.get("Location") or "")


def test_portal_scrum_api_move_queues_proposal_manager_approve_writes_activity(client, app):
    name = EMPLOYEES[0]
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "portal.test@nokia.com",
            "name": name,
            "roster_name": name,
            "role": "employee",
        }
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Portal Move Sprint', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Move me', 1, 'open', '', 0, ?, ?, 'backlog', 'task')
        """,
        (sid, name, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = client.post(
        "/portal/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": name,
            "to_column": "doing",
            "committed_hours": "0.25",
            "note": "picked up",
        },
    )
    assert r.status_code == 200
    j = r.get_json()
    assert j["ok"] is True
    assert j.get("pending_approval") is True
    conn = get_db(app)
    n_act = int(
        conn.execute("SELECT COUNT(*) AS c FROM scrum_item_activity WHERE item_id = ?", (iid,)).fetchone()["c"]
    )
    assert n_act == 0
    col = conn.execute("SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()["kanban_column"]
    assert col == "backlog"
    pid = int(
        conn.execute(
            "SELECT id FROM scrum_portal_proposal WHERE item_id = ? AND status = 'pending'",
            (iid,),
        ).fetchone()["id"]
    )
    conn.close()

    _manager_login(client)
    res = client.post(
        "/scrum/portal-proposal/resolve",
        data={"proposal_id": str(pid), "decision": "approve", "note": "ok", "manager_attest": "1"},
        follow_redirects=False,
    )
    assert res.status_code in (302, 303)

    conn = get_db(app)
    body = conn.execute(
        "SELECT body FROM scrum_item_activity WHERE item_id = ? ORDER BY id DESC LIMIT 1",
        (iid,),
    ).fetchone()["body"]
    col2 = conn.execute("SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()["kanban_column"]
    conn.close()
    assert "[approved employee change]" in body
    assert "picked up" in body
    assert col2 == "doing"


def test_portal_proposal_approve_without_attest_does_not_apply(client, app):
    name = EMPLOYEES[0]
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "portal.test@nokia.com",
            "name": name,
            "roster_name": name,
            "role": "employee",
        }
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Attest Sprint', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Attest card', 1, 'open', '', 0, ?, ?, 'backlog', 'task')
        """,
        (sid, name, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    assert client.post(
        "/portal/scrum/api/item/move",
        json={
            "item_id": iid,
            "sprint_id": sid,
            "assignee": name,
            "to_column": "doing",
            "committed_hours": "0",
            "note": "x",
        },
    ).get_json()["ok"]
    conn = get_db(app)
    pid = int(
        conn.execute(
            "SELECT id FROM scrum_portal_proposal WHERE item_id = ? AND status = 'pending'",
            (iid,),
        ).fetchone()["id"]
    )
    conn.close()
    _manager_login(client)
    client.post(
        "/scrum/portal-proposal/resolve",
        data={"proposal_id": str(pid), "decision": "approve", "note": "no checkbox"},
        follow_redirects=False,
    )
    conn = get_db(app)
    col = conn.execute("SELECT kanban_column FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()["kanban_column"]
    st = conn.execute("SELECT status FROM scrum_portal_proposal WHERE id = ?", (pid,)).fetchone()["status"]
    conn.close()
    assert col == "backlog"
    assert st == "pending"


def test_manager_scrum_item_update_doing_column_updates_estimate_only(client, app):
    _manager_login(client)
    name = EMPLOYEES[0]
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Est Doing Sprint', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Doing est', 4, 'doing', '', 0, ?, ?, 'doing', 'task')
        """,
        (sid, name, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    res = client.post(
        "/scrum/sprint/item/update",
        data={"item_id": str(iid), "sprint_id": str(sid), "estimate_hours": "12.5"},
        follow_redirects=False,
    )
    assert res.status_code in (302, 303)
    conn = get_db(app)
    est = float(conn.execute("SELECT estimate_hours FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()[0])
    title = conn.execute("SELECT title FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()["title"]
    conn.close()
    assert abs(est - 12.5) < 0.001
    assert title == "Doing est"


def test_portal_doing_estimate_change_queues_proposal_and_approve_updates_db(client, app):
    name = EMPLOYEES[0]
    with client.session_transaction() as sess:
        sess["portal_user"] = {
            "email": "portal.test@nokia.com",
            "name": name,
            "roster_name": name,
            "role": "employee",
        }
    conn = get_db(app)
    tid = int(conn.execute("SELECT id FROM teams ORDER BY id ASC LIMIT 1").fetchone()["id"])
    ts = "2026-11-01T00:00:00Z"
    conn.execute(
        """
        INSERT INTO scrum_sprint (team_id, name, start_date, end_date, goal, created_at, updated_at)
        VALUES (?, 'Portal Est Sprint', '2026-11-01', '2026-11-10', '', ?, ?)
        """,
        (tid, ts, ts),
    )
    sid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.execute(
        """
        INSERT INTO scrum_sprint_item
        (sprint_id, assignee, title, estimate_hours, status, notes, sort_order, created_at, updated_at, kanban_column, task_kind)
        VALUES (?, ?, 'Portal doing est', 3, 'doing', '', 0, ?, ?, 'doing', 'task')
        """,
        (sid, name, ts, ts),
    )
    iid = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
    conn.commit()
    conn.close()
    r = client.post(
        "/portal/scrum/sprint/item/update",
        data={"item_id": str(iid), "sprint_id": str(sid), "estimate_hours": "8"},
        follow_redirects=False,
    )
    assert r.status_code in (302, 303)
    conn = get_db(app)
    est_before = float(conn.execute("SELECT estimate_hours FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()[0])
    prow = conn.execute(
        "SELECT id, action FROM scrum_portal_proposal WHERE item_id = ? AND status = 'pending'",
        (iid,),
    ).fetchone()
    conn.close()
    assert abs(est_before - 3.0) < 0.001
    assert prow is not None
    assert prow["action"] == "item_update_doing_estimate"
    pid = int(prow["id"])
    _manager_login(client)
    res = client.post(
        "/scrum/portal-proposal/resolve",
        data={"proposal_id": str(pid), "decision": "approve", "note": "ok", "manager_attest": "1"},
        follow_redirects=False,
    )
    assert res.status_code in (302, 303)
    conn = get_db(app)
    est_after = float(conn.execute("SELECT estimate_hours FROM scrum_sprint_item WHERE id = ?", (iid,)).fetchone()[0])
    st = conn.execute("SELECT status FROM scrum_portal_proposal WHERE id = ?", (pid,)).fetchone()["status"]
    conn.close()
    assert abs(est_after - 8.0) < 0.001
    assert st == "approved"


def test_attendance_url_redirects_home(client):
    assert client.get("/attendance", follow_redirects=False).status_code == 302


def test_legacy_manager_paths_redirect(client):
    c = client
    assert c.get("/manager", follow_redirects=False).status_code == 302
    assert c.get("/manager/login", follow_redirects=False).status_code == 302
    _manager_login(c)
    assert c.get("/manager?year=2026&month=5", follow_redirects=False).status_code == 302
